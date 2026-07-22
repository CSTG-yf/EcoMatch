#!/usr/bin/env python3
"""Build DATA-01 JSONL files from the competition workbook."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


METRICS = {
    "ZB001": ("各项存款余额", ["存款余额", "存款规模", "存款总额", "存款"]),
    "ZB002": ("各项贷款余额", ["贷款余额", "贷款规模", "贷款总额", "贷款"]),
    "ZB003": ("对公存款余额", ["对公存款", "公司存款"]),
    "ZB004": ("个人存款余额", ["个人存款", "储蓄存款", "零售存款"]),
    "ZB005": ("对公贷款余额", ["对公贷款", "公司贷款"]),
    "ZB006": ("个人贷款余额", ["个人贷款", "零售贷款"]),
    "ZB007": ("中间业务收入", ["中收", "中间收入", "手续费收入"]),
    "ZB008": ("净利息收入", ["净息收", "利息净收入"]),
    "ZB009": ("营业收入", ["营收"]),
    "ZB010": ("营业支出", ["营业成本"]),
    "ZB011": ("净利润", ["利润"]),
    "ZB012": ("成本收入比", ["成本收支比", "成本收人比"]),
    "ZB013": ("不良贷款率", ["不良率", "不良货款率"]),
    "ZB014": ("不良贷款余额", ["不良余额", "不良货款余额"]),
    "ZB015": ("拨备覆盖率", ["拨备率", "拨备覆盖", "拨备"]),
    "ZB016": ("资本充足率", ["资本充足"]),
    "ZB017": ("逾期贷款率", ["逾期率", "逾期货款率"]),
    "ZB018": ("员工人数", ["员工数", "人数"]),
    "ZB019": ("网点数量", ["网点数", "营业网点"]),
    "ZB020": ("个人客户数", ["个人客户", "零售客户数", "零售客户"]),
    "ZB021": ("对公客户数", ["对公客户", "公司客户数", "企业客户数"]),
}

TYPO_TERMS = ["不良货款率", "不良货款余额", "逾期货款率", "成本收人比"]
BROAD_METRIC = re.compile(r"(贷款|存款|经营|风险)(情况|指标|表现|怎么样|如何)")
PROVINCE_SCOPE = re.compile(r"全省|13家|十三家|哪家|各家|所有(?:机构|农商行)")
DATE_PATTERN = re.compile(
    r"20\d{2}[-/]\d{1,2}[-/]\d{1,2}|20\d{2}年\d{1,2}月\d{1,2}日|"
    r"20\d{2}年\d{1,2}月(?:末|底)|20\d{2}年(?:末|底|年末|年底|全年|年度)|"
    r"20\d{2}年?(?:第)?[一二三四1-4]季度末?|20\d{2}年[上下]半年末|"
    r"年初|本月|上月|今年|去年|今天|当前|现在|最近|近期"
)


def source_split(question_id: str) -> str:
    if question_id.startswith("TRAIN"):
        return "train"
    if question_id.startswith("VAL"):
        return "dev"
    return "test"


def classify_intent(text: str, organization_count: int) -> str:
    scores = {"POINT_QUERY": 0.72}
    if any(word in text for word in ["趋势", "走势", "逐月", "逐日", "每天", "连续", "全年变化"]):
        scores["TREND"] = 0.99
    if (any(word in text for word in ["环比", "同比", "较年初", "较上季", "较上月", "较同期", "增幅", "增量", "增长", "下降", "变动", "变化", "从"])
            and any(word in text for word in ["到", "比", "较", "变化", "增长", "下降", "变动"])):
        scores["CHANGE"] = 0.98
    if any(word in text for word in ["占比", "比例", "比重", "除以", "存贷比", "净利润率"]):
        scores["RATIO"] = 0.98
    if any(word in text for word in ["超过", "高于", "大于", "低于", "小于", "不低于", "不高于", "达标", "满足", "监管要求"]):
        scores["THRESHOLD"] = 0.99
    if any(word in text for word in ["排名", "第几", "第一", "最后", "最高", "最低", "最多", "最少", "前三", "后三", "后四", "表现较好", "表现较差"]):
        scores["RANKING"] = 0.98
    if any(word in text for word in ["平均", "均值", "合计", "总和", "加起来", "多少家", "有几家", "多少天"]):
        scores["AGGREGATION"] = 0.91
    if ((organization_count >= 2 and any(word in text for word in ["谁", "更", "比", "差"]))
            or any(word in text for word in ["两家相比", "机构间比较"])):
        scores["COMPARISON"] = 0.97
    return max(scores, key=scores.get)


def extract_metrics(text: str) -> list[dict]:
    hits = []
    for code, (name, aliases) in METRICS.items():
        for alias in [name, *aliases]:
            for match in re.finditer(re.escape(alias), text):
                hits.append((-(match.end() - match.start()), match.start(), match.end(), alias, code, name))
    occupied = []
    matches = {}
    for _, start, end, alias, code, name in sorted(hits):
        if any(start < right and end > left for left, right in occupied):
            continue
        occupied.append((start, end))
        matches.setdefault(code, {"code": code, "name": name, "matchedText": alias})
    composites = {
        "净利润率": ["ZB011", "ZB009"],
        "存贷比": ["ZB002", "ZB001"],
        "风险指标": ["ZB013", "ZB015", "ZB017", "ZB016"],
    }
    for phrase, codes in composites.items():
        if phrase in text:
            for code in codes:
                matches.setdefault(code, {"code": code, "name": METRICS[code][0], "matchedText": phrase})
    return list(matches.values())


def extract_organizations(text: str, organizations: list[tuple[str, str]]) -> list[dict]:
    result = []
    for code, name in organizations:
        city = name.removeprefix("江苏省").removesuffix("市农商行")
        aliases = [name, f"{city}行", f"{city}市农商行", f"{city}农商行"]
        matched = next((alias for alias in aliases if alias in text), None)
        if matched:
            result.append({"code": code, "name": name, "matchedText": matched})
    return result


def scene(metrics: list[dict]) -> str:
    codes = {metric["code"] for metric in metrics}
    if codes & {"ZB013", "ZB014", "ZB015", "ZB016", "ZB017"}:
        return "RISK_CONTROL"
    if codes & {"ZB020", "ZB021"}:
        return "CUSTOMER_MARKETING"
    return "OPERATION_ANALYSIS"


def linguistic_features(text: str) -> list[str]:
    features = []
    if any(term in text for term in TYPO_TERMS):
        features.append("TYPO")
    if re.search(r"[A-M](?:行|市农商行|农商行)", text) or any(term in text for term in ["不良率", "逾期率", "拨备率", "营收"]):
        features.append("ABBREVIATION")
    if any(term in text for term in ["帮忙", "查一下", "谁家", "咋", "有没有", "达标没"]):
        features.append("COLLOQUIAL")
    if any(term in text for term in ["最近", "近期", "当前", "现在", "今年", "去年", "本月", "上月"]):
        features.append("VAGUE_TIME")
    if BROAD_METRIC.search(text) or any(term in text for term in ["大不大", "怎么样", "如何"]):
        features.append("AMBIGUOUS")
    return features or ["STANDARD"]


def template(text: str, organizations: list[tuple[str, str]]) -> str:
    normalized = text
    for _, name in sorted(organizations, key=lambda item: len(item[1]), reverse=True):
        normalized = normalized.replace(name, "<ORG>")
    for _, (name, aliases) in METRICS.items():
        for value in sorted([name, *aliases], key=len, reverse=True):
            normalized = normalized.replace(value, "<METRIC>")
    normalized = DATE_PATTERN.sub("<DATE>", normalized)
    normalized = re.sub(r"\d+(?:\.\d+)?%?", "<NUM>", normalized)
    return hashlib.sha256(normalized.encode("utf-8")).hexdigest()[:16]


def clarification_expected(text: str, metrics: list[dict], organizations: list[dict], times: list[str]) -> bool:
    return (not metrics or bool(BROAD_METRIC.search(text)) or not times
            or any(term in text for term in ["最近", "近期"])
            or (not organizations and not PROVINCE_SCOPE.search(text))
            or any(term in text for term in ["大不大", "怎么样", "如何"]))


def annotate(question_id: str, source: str, difficulty: str, text: str, answer: str,
             organizations: list[tuple[str, str]]) -> dict:
    metric_slots = extract_metrics(text)
    organization_slots = extract_organizations(text, organizations)
    times = DATE_PATTERN.findall(text)
    intent = classify_intent(text, len(organization_slots))
    filters = []
    for operator, value in re.findall(r"(不低于|不高于|至少|至多|超过|高于|大于|低于|小于)(\d+(?:\.\d+)?%?)", text):
        filters.append({"operator": operator, "value": value})
    if any(term in text for term in ["全省平均", "全省均值", "平均水平"]):
        filters.append({"operator": "COMPARE", "value": "PROVINCE_AVERAGE"})
    return {
        "id": question_id,
        "source": source,
        "sourceSplit": source_split(question_id) if source == "competition_workbook" else None,
        "split": source_split(question_id),
        "difficulty": difficulty,
        "question": text,
        "answer": answer,
        "scene": scene(metric_slots),
        "intent": intent,
        "metrics": metric_slots,
        "dimensions": ["bank_data_date", "bank_organization"],
        "time": {"expressions": times},
        "organizations": organization_slots,
        "filters": filters,
        "linguisticFeatures": linguistic_features(text),
        "clarificationExpected": clarification_expected(text, metric_slots, organization_slots, times),
        "templateGroup": template(text, organizations),
        "referenceDate": "2026-07-22",
    }


def augmentations() -> list[tuple[str, str, str, str, str]]:
    return [
        ("AUG-TRAIN-01", "train", "普通", "帮我看下A行今年的不良货款率", ""),
        ("AUG-TRAIN-02", "train", "普通", "B市农商行上月存款规模是多少", ""),
        ("AUG-TRAIN-03", "train", "复杂", "最近C行贷款情况如何", ""),
        ("AUG-TRAIN-04", "train", "普通", "2026年一季度末全省哪家拨备率最高", ""),
        ("AUG-TRAIN-05", "train", "简单", "D行当前个人客户有多少", ""),
        ("AUG-TRAIN-06", "train", "普通", "E农商行去年营收同比增长多少", ""),
        ("AUG-DEV-01", "dev", "复杂", "F行近期风险指标表现怎么样", ""),
        ("AUG-DEV-02", "dev", "普通", "今年G行成本收人比超过30%吗", ""),
        ("AUG-DEV-03", "dev", "简单", "H市农商行本月网点数", ""),
        ("AUG-TEST-01", "test", "普通", "I行上月不良率达标没", ""),
        ("AUG-TEST-02", "test", "复杂", "最近贷款情况怎么样", ""),
        ("AUG-TEST-03", "test", "普通", "今年全省个人客户数最高的是哪家", ""),
    ]


def write_jsonl(path: Path, records: list[dict]) -> None:
    path.write_text("".join(json.dumps(record, ensure_ascii=False) + "\n" for record in records), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--output", type=Path, default=Path(__file__).parent)
    args = parser.parse_args()
    args.output.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(args.workbook, read_only=True, data_only=True)
    organizations = [(str(row[0]), str(row[1])) for row in workbook.worksheets[0].iter_rows(min_row=2, values_only=True)]
    questions = list(workbook.worksheets[4].iter_rows(min_row=2, values_only=True))
    records = [annotate(str(row[0]), "competition_workbook", str(row[2]), str(row[3]), str(row[4]), organizations)
               for row in questions]
    for question_id, split, difficulty, text, answer in augmentations():
        record = annotate(question_id, "curated_augmentation", difficulty, text, answer, organizations)
        record["split"] = split
        record["sourceSplit"] = split
        records.append(record)

    groups = defaultdict(list)
    for record in records:
        groups[record["templateGroup"]].append(record)
    priority = {"train": 0, "dev": 1, "test": 2}
    reassigned = 0
    for group_records in groups.values():
        target = max((record["split"] for record in group_records), key=priority.get)
        for record in group_records:
            if record["split"] != target:
                record["split"] = target
                reassigned += 1

    split_records = {split: [record for record in records if record["split"] == split]
                     for split in ["train", "dev", "test"]}
    for split, items in split_records.items():
        write_jsonl(args.output / f"{split}.jsonl", items)

    split_templates = {split: {record["templateGroup"] for record in items}
                       for split, items in split_records.items()}
    overlaps = {
        "trainDev": len(split_templates["train"] & split_templates["dev"]),
        "trainTest": len(split_templates["train"] & split_templates["test"]),
        "devTest": len(split_templates["dev"] & split_templates["test"]),
    }
    source_hash = hashlib.sha256(args.workbook.read_bytes()).hexdigest()
    manifest = {
        "version": "1.0.0",
        "generatedAt": "2026-07-22",
        "sourceWorkbook": args.workbook.name,
        "sourceSha256": source_hash,
        "sourceQuestionCount": len(questions),
        "augmentationCount": len(augmentations()),
        "reassignedForTemplateIsolation": reassigned,
        "counts": {split: len(items) for split, items in split_records.items()},
        "templateCounts": {split: len(values) for split, values in split_templates.items()},
        "templateOverlap": overlaps,
        "intentCounts": dict(Counter(record["intent"] for record in records)),
        "sceneCounts": dict(Counter(record["scene"] for record in records)),
        "featureCounts": dict(Counter(feature for record in records for feature in record["linguisticFeatures"])),
        "clarificationCount": sum(record["clarificationExpected"] for record in records),
    }
    (args.output / "manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(manifest, ensure_ascii=False))


if __name__ == "__main__":
    main()
