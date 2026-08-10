#!/usr/bin/env python3
"""Create an immutable answer-only child release from an official package.

The parent package is never modified.  Corrections are accepted only for
train/dev answerText cells whose old SHA-256 matches the parent workbook.
Question text, split assignment, test rows and all fact sheets are preserved.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import shutil
import sys
import tempfile
from itertools import zip_longest
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from build_dataset import QUESTION_HEADERS, QUESTION_SHEET, SOURCE_SPLITS
from clarify_ground_truth_contracts import _fix_zip_timestamps, fact_region_digest


GENERATOR_NAME = "amend_official_ground_truth"
GENERATOR_VERSION = "1.0.0"
AMENDMENT_LEDGER = "answer-amendment-ledger.json"
AMENDMENT_AUDIT = "answer-amendment-audit-summary.json"
MANIFEST_NAME = "official-manifest.json"
_SHA256 = re.compile(r"^[0-9A-F]{64}$")


class OfficialAmendmentError(ValueError):
    """The requested child release violates the immutable package contract."""


def sha256_text(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest().upper()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for block in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest().upper()


def _read_json(path: Path, label: str) -> dict[str, Any]:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise OfficialAmendmentError(f"{label} 无法解析: {exc}") from exc
    if not isinstance(payload, dict):
        raise OfficialAmendmentError(f"{label} 必须为 JSON 对象")
    return payload


def _write_json(path: Path, payload: dict[str, Any]) -> None:
    path.write_bytes(
        (json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n").encode(
            "utf-8"
        )
    )


def _write_sidecar(path: Path) -> None:
    path.with_name(path.name + ".sha256").write_bytes(
        f"{sha256_file(path)}  {path.name}\n".encode("ascii")
    )


def _verify_sidecar(path: Path) -> str:
    sidecar = path.with_name(path.name + ".sha256")
    if not path.is_file() or not sidecar.is_file():
        raise OfficialAmendmentError(f"父正式包缺少文件或 sidecar: {path.name}")
    parts = sidecar.read_text(encoding="ascii").strip().split()
    if len(parts) != 2 or parts[1] != path.name or not _SHA256.fullmatch(parts[0]):
        raise OfficialAmendmentError(f"父正式包 sidecar 格式非法: {sidecar.name}")
    actual = sha256_file(path)
    if parts[0] != actual:
        raise OfficialAmendmentError(
            f"父正式包 sidecar 哈希不一致: {path.name} ({actual} != {parts[0]})"
        )
    return actual


def _relative_artifact(parent_dir: Path, value: Any, field: str) -> Path:
    if not isinstance(value, str) or not value:
        raise OfficialAmendmentError(f"父 manifest {field} 非法")
    relative = Path(value)
    if relative.is_absolute() or ".." in relative.parts:
        raise OfficialAmendmentError(f"父 manifest {field} 不能越出正式包")
    return parent_dir / relative


def _validate_parent(parent_dir: Path) -> tuple[dict[str, Any], Path, Path, Path, str]:
    manifest_path = parent_dir / MANIFEST_NAME
    manifest_sha = _verify_sidecar(manifest_path)
    manifest = _read_json(manifest_path, "父 official manifest")
    if manifest.get("canonicalReady") is not True:
        raise OfficialAmendmentError("父 official manifest canonicalReady 必须为 true")
    parent_version = manifest.get("datasetVersion")
    if not isinstance(parent_version, str) or not parent_version:
        raise OfficialAmendmentError("父 official manifest 缺少 datasetVersion")

    workbook_path = _relative_artifact(
        parent_dir, manifest.get("groundTruthWorkbook"), "groundTruthWorkbook"
    )
    ledger_path = _relative_artifact(parent_dir, manifest.get("changeLedger"), "changeLedger")
    audit_path = _relative_artifact(
        parent_dir, manifest.get("finalAuditSummary"), "finalAuditSummary"
    )
    hashes = manifest.get("artifactSha256")
    if not isinstance(hashes, dict):
        raise OfficialAmendmentError("父 official manifest artifactSha256 非法")
    for field, path in (
        ("groundTruthWorkbook", workbook_path),
        ("changeLedger", ledger_path),
        ("finalAuditSummary", audit_path),
    ):
        actual = _verify_sidecar(path)
        if hashes.get(field) != actual:
            raise OfficialAmendmentError(
                f"父 official manifest artifactSha256.{field} 与文件不一致"
            )
    fact_sha = fact_region_digest(workbook_path)
    if manifest.get("factRegionSha256") != fact_sha:
        raise OfficialAmendmentError("父正式工作簿事实区域哈希与 manifest 不一致")
    return manifest, workbook_path, ledger_path, audit_path, manifest_sha


def _load_spec(spec_path: Path, parent_version: str) -> tuple[dict[str, Any], list[dict[str, str]]]:
    spec = _read_json(spec_path, "答案修正清单")
    if spec.get("schemaVersion") != "1.0":
        raise OfficialAmendmentError("答案修正清单 schemaVersion 必须为 1.0")
    if spec.get("parentVersion") != parent_version:
        raise OfficialAmendmentError("答案修正清单 parentVersion 与父正式包不一致")
    target_version = spec.get("targetVersion")
    if not isinstance(target_version, str) or not re.fullmatch(r"\d+\.\d+\.\d+", target_version):
        raise OfficialAmendmentError("答案修正清单 targetVersion 非法")
    entries = spec.get("entries")
    if not isinstance(entries, list) or not entries:
        raise OfficialAmendmentError("答案修正清单 entries 必须为非空列表")

    normalized: list[dict[str, str]] = []
    seen: set[str] = set()
    for raw in entries:
        if not isinstance(raw, dict):
            raise OfficialAmendmentError("答案修正清单 entry 必须为对象")
        sample_id = raw.get("id")
        split = raw.get("split")
        old_hash = raw.get("oldAnswerSha256")
        corrected = raw.get("correctedAnswerText")
        reason = raw.get("reason")
        if not isinstance(sample_id, str) or not sample_id:
            raise OfficialAmendmentError("答案修正清单 id 非法")
        if sample_id in seen:
            raise OfficialAmendmentError(f"答案修正清单存在重复 ID: {sample_id}")
        seen.add(sample_id)
        if split == "test":
            raise OfficialAmendmentError(f"禁止修正 test split: {sample_id}")
        if split not in {"train", "dev"}:
            raise OfficialAmendmentError(f"答案修正清单 split 非法: {sample_id}")
        if not isinstance(old_hash, str) or not _SHA256.fullmatch(old_hash):
            raise OfficialAmendmentError(f"答案修正清单 oldAnswerSha256 非法: {sample_id}")
        if not isinstance(corrected, str) or not corrected.strip():
            raise OfficialAmendmentError(f"答案修正清单 correctedAnswerText 非法: {sample_id}")
        if not isinstance(reason, str) or not reason.strip():
            raise OfficialAmendmentError(f"答案修正清单 reason 非法: {sample_id}")
        normalized.append(
            {
                "id": sample_id,
                "split": split,
                "oldAnswerSha256": old_hash,
                "correctedAnswerText": corrected,
                "reason": reason,
            }
        )
    return spec, normalized


def _apply_answers(
    parent_workbook: Path,
    child_workbook: Path,
    corrections: list[dict[str, str]],
) -> list[dict[str, str]]:
    workbook = load_workbook(parent_workbook)
    try:
        if QUESTION_SHEET not in workbook.sheetnames:
            raise OfficialAmendmentError(f"父工作簿缺少 {QUESTION_SHEET}")
        sheet = workbook[QUESTION_SHEET]
        header = tuple(cell.value for cell in sheet[1][: len(QUESTION_HEADERS)])
        if header != QUESTION_HEADERS:
            raise OfficialAmendmentError(f"父工作簿 {QUESTION_SHEET} 表头非法")
        by_id: dict[str, tuple[int, str, str]] = {}
        for row_number in range(2, sheet.max_row + 1):
            sample_id = sheet.cell(row_number, 1).value
            question_type = sheet.cell(row_number, 2).value
            answer = sheet.cell(row_number, 5).value
            if sample_id is None:
                continue
            sample_id = str(sample_id).strip()
            if sample_id in by_id:
                raise OfficialAmendmentError(f"父工作簿存在重复 ID: {sample_id}")
            split = SOURCE_SPLITS.get(str(question_type).strip())
            if split is None:
                raise OfficialAmendmentError(f"父工作簿问题类型非法: {sample_id}")
            by_id[sample_id] = (row_number, split, "" if answer is None else str(answer))

        ledger_entries: list[dict[str, str]] = []
        for correction in corrections:
            sample_id = correction["id"]
            current = by_id.get(sample_id)
            if current is None:
                raise OfficialAmendmentError(f"答案修正 ID 不存在于父工作簿: {sample_id}")
            row_number, actual_split, old_answer = current
            if actual_split != correction["split"]:
                raise OfficialAmendmentError(f"答案修正 split 与父工作簿不一致: {sample_id}")
            actual_old_hash = sha256_text(old_answer)
            if actual_old_hash != correction["oldAnswerSha256"]:
                raise OfficialAmendmentError(
                    f"答案修正 oldAnswerSha256 与父工作簿不一致: {sample_id}"
                )
            new_answer = correction["correctedAnswerText"]
            if new_answer == old_answer:
                raise OfficialAmendmentError(f"答案修正没有产生变化: {sample_id}")
            sheet.cell(row_number, 5, new_answer)
            ledger_entries.append(
                {
                    "id": sample_id,
                    "split": actual_split,
                    "oldAnswerSha256": actual_old_hash,
                    "newAnswerSha256": sha256_text(new_answer),
                    "reason": correction["reason"],
                }
            )
        workbook.save(child_workbook)
    finally:
        workbook.close()
    _fix_zip_timestamps(child_workbook)
    return ledger_entries


def _verify_only_answers_changed(
    parent_workbook: Path,
    child_workbook: Path,
    corrections: list[dict[str, str]],
) -> None:
    expected = {item["id"]: item["correctedAnswerText"] for item in corrections}
    changed: set[str] = set()
    parent = load_workbook(parent_workbook, read_only=True, data_only=True)
    child = load_workbook(child_workbook, read_only=True, data_only=True)
    try:
        if parent.sheetnames != child.sheetnames:
            raise OfficialAmendmentError("子工作簿 sheet 名称或顺序发生变化")
        for sheet_name in parent.sheetnames:
            parent_rows = parent[sheet_name].iter_rows(values_only=True)
            child_rows = child[sheet_name].iter_rows(values_only=True)
            for row_number, (before, after) in enumerate(
                zip_longest(parent_rows, child_rows, fillvalue=None), start=1
            ):
                if before is None or after is None or len(before) != len(after):
                    raise OfficialAmendmentError(f"子工作簿 {sheet_name} 行结构发生变化")
                if before == after:
                    continue
                sample_id = str(before[0]) if before and before[0] is not None else ""
                for column, (old_value, new_value) in enumerate(zip(before, after), start=1):
                    if old_value == new_value:
                        continue
                    if (
                        sheet_name != QUESTION_SHEET
                        or row_number == 1
                        or column != 5
                        or sample_id not in expected
                        or new_value != expected[sample_id]
                    ):
                        raise OfficialAmendmentError(
                            f"子工作簿出现未授权变化: {sheet_name}!R{row_number}C{column}"
                        )
                    changed.add(sample_id)
        if changed != set(expected):
            raise OfficialAmendmentError(
                f"子工作簿答案变化集合不完整: {sorted(changed)} != {sorted(expected)}"
            )
    finally:
        parent.close()
        child.close()


def amend_official_package(
    parent_dir: Path | str,
    spec_path: Path | str,
    output_dir: Path | str,
    *,
    update_current: bool = False,
) -> dict[str, Any]:
    parent_dir = Path(parent_dir).resolve()
    spec_path = Path(spec_path).resolve()
    output_dir = Path(output_dir).resolve()
    if output_dir.exists():
        raise OfficialAmendmentError(f"输出目录已存在，拒绝覆盖: {output_dir}")
    if not parent_dir.is_dir():
        raise OfficialAmendmentError(f"父正式包不存在: {parent_dir}")
    if not spec_path.is_file():
        raise OfficialAmendmentError(f"答案修正清单不存在: {spec_path}")

    parent_manifest, parent_workbook, parent_ledger, parent_audit, parent_manifest_sha = (
        _validate_parent(parent_dir)
    )
    spec, corrections = _load_spec(spec_path, str(parent_manifest["datasetVersion"]))
    target_version = str(spec["targetVersion"])
    if output_dir.name != target_version:
        raise OfficialAmendmentError("输出目录名必须等于 targetVersion")
    output_dir.parent.mkdir(parents=True, exist_ok=True)

    with tempfile.TemporaryDirectory(prefix="official-amendment-", dir=output_dir.parent) as temp_dir:
        stage = Path(temp_dir) / target_version
        stage.mkdir()
        child_workbook_name = f"bank-nl2sql-ground-truth-v{target_version}.xlsx"
        child_workbook = stage / child_workbook_name
        ledger_entries = _apply_answers(parent_workbook, child_workbook, corrections)
        _verify_only_answers_changed(parent_workbook, child_workbook, corrections)
        child_fact_sha = fact_region_digest(child_workbook)
        if child_fact_sha != parent_manifest["factRegionSha256"]:
            raise OfficialAmendmentError("答案修正意外改变了事实区域")

        copied_ledger = stage / parent_ledger.name
        shutil.copy2(parent_ledger, copied_ledger)
        _write_sidecar(copied_ledger)

        amendment_ledger_path = stage / AMENDMENT_LEDGER
        amendment_ledger = {
            "schemaVersion": "1.0",
            "generator": {"name": GENERATOR_NAME, "version": GENERATOR_VERSION},
            "parentDatasetVersion": parent_manifest["datasetVersion"],
            "parentOfficialManifestSha256": parent_manifest_sha,
            "targetDatasetVersion": target_version,
            "correctionSpecSha256": sha256_file(spec_path),
            "count": len(ledger_entries),
            "entries": ledger_entries,
        }
        _write_json(amendment_ledger_path, amendment_ledger)
        _write_sidecar(amendment_ledger_path)

        audit_path = stage / AMENDMENT_AUDIT
        audit = {
            "schemaVersion": "1.0",
            "datasetVersion": target_version,
            "parentDatasetVersion": parent_manifest["datasetVersion"],
            "answerAmendmentCount": len(ledger_entries),
            "verifiedIds": [item["id"] for item in ledger_entries],
            "parentWorkbookSha256": sha256_file(parent_workbook),
            "candidateWorkbookSha256": sha256_file(child_workbook),
            "factRegionSha256": child_fact_sha,
            "onlyDeclaredAnswerCellsChanged": True,
            "testSplitChanged": False,
            "canonicalReady": True,
            "auditErrors": 0,
        }
        _write_json(audit_path, audit)
        _write_sidecar(audit_path)
        _write_sidecar(child_workbook)

        artifact_hashes = dict(parent_manifest["artifactSha256"])
        artifact_hashes.update(
            {
                "groundTruthWorkbook": sha256_file(child_workbook),
                "changeLedger": sha256_file(copied_ledger),
                "answerAmendmentLedger": sha256_file(amendment_ledger_path),
                "finalAuditSummary": sha256_file(audit_path),
            }
        )
        manifest = {
            "schemaVersion": "2.1",
            "releaseMode": "INCREMENTAL_ANSWER_AMENDMENT",
            "datasetVersion": target_version,
            "parent": {
                "datasetVersion": parent_manifest["datasetVersion"],
                "officialManifestSha256": parent_manifest_sha,
                "groundTruthWorkbookSha256": sha256_file(parent_workbook),
                "finalAuditSummarySha256": sha256_file(parent_audit),
            },
            "canonicalReady": True,
            "officialCount": parent_manifest["officialCount"],
            "sourceSplitCounts": parent_manifest["sourceSplitCounts"],
            "removedIds": parent_manifest["removedIds"],
            "sourceSha256": parent_manifest.get("sourceSha256"),
            "factRegionSha256": child_fact_sha,
            "generator": parent_manifest["generator"],
            "changeLedger": copied_ledger.name,
            "changeCounts": parent_manifest["changeCounts"],
            "answerAmendmentLedger": amendment_ledger_path.name,
            "answerAmendmentCount": len(ledger_entries),
            "amendmentGenerator": {"name": GENERATOR_NAME, "version": GENERATOR_VERSION},
            "groundTruthWorkbook": child_workbook_name,
            "finalAuditSummary": audit_path.name,
            "artifactSha256": artifact_hashes,
            "auditStatus": {
                "auditErrors": 0,
                "candidateReady": True,
                "onlyDeclaredAnswerCellsChanged": True,
                "testSplitChanged": False,
            },
        }
        manifest_path = stage / MANIFEST_NAME
        _write_json(manifest_path, manifest)
        _write_sidecar(manifest_path)
        current_temp = Path(temp_dir) / "CURRENT.json"
        if update_current:
            _write_json(
                current_temp,
                {
                    "currentVersion": target_version,
                    "directory": target_version,
                    "groundTruthWorkbook": f"{target_version}/{child_workbook_name}",
                    "officialManifest": f"{target_version}/{MANIFEST_NAME}",
                },
            )
        stage.replace(output_dir)
        if update_current:
            current_temp.replace(output_dir.parent / "CURRENT.json")
    return manifest


def _amend_jsonl_answers(
    source_path: Path,
    target_path: Path,
    corrections: dict[str, dict[str, str]],
) -> set[str]:
    if not source_path.is_file():
        raise OfficialAmendmentError(f"父派生数据集缺少 {source_path.name}")
    changed: set[str] = set()
    output_lines: list[str] = []
    for raw_line in source_path.read_text(encoding="utf-8").splitlines():
        if not raw_line.strip():
            continue
        try:
            record = json.loads(raw_line)
        except json.JSONDecodeError as exc:
            raise OfficialAmendmentError(f"父派生数据集 {source_path.name} 无法解析") from exc
        sample_id = record.get("id") if isinstance(record, dict) else None
        correction = corrections.get(str(sample_id))
        if correction is None:
            output_lines.append(raw_line)
            continue
        expected = record.get("expected")
        if not isinstance(expected, dict) or not isinstance(expected.get("answerText"), str):
            raise OfficialAmendmentError(f"父派生记录缺少 expected.answerText: {sample_id}")
        if sha256_text(expected["answerText"]) != correction["oldAnswerSha256"]:
            raise OfficialAmendmentError(
                f"父派生记录 oldAnswerSha256 与修正清单不一致: {sample_id}"
            )
        expected["answerText"] = correction["correctedAnswerText"]
        output_lines.append(json.dumps(record, ensure_ascii=False, sort_keys=True))
        changed.add(str(sample_id))
    target_path.write_bytes(
        ("\n".join(output_lines) + ("\n" if output_lines else "")).encode("utf-8")
    )
    return changed


def amend_derived_dataset(
    parent_dataset_dir: Path | str,
    child_official_dir: Path | str,
    spec_path: Path | str,
    output_dir: Path | str,
) -> dict[str, Any]:
    """Copy a frozen derived dataset and alter only declared train/dev answers."""

    parent_dataset_dir = Path(parent_dataset_dir).resolve()
    child_official_dir = Path(child_official_dir).resolve()
    spec_path = Path(spec_path).resolve()
    output_dir = Path(output_dir).resolve()
    if output_dir.exists():
        raise OfficialAmendmentError(f"派生数据输出目录已存在，拒绝覆盖: {output_dir}")
    parent_manifest_path = parent_dataset_dir / "manifest.json"
    parent_gold_path = parent_dataset_dir / "gold_manifest.json"
    parent_manifest = _read_json(parent_manifest_path, "父派生 manifest")
    parent_gold = _read_json(parent_gold_path, "父派生 gold manifest")
    parent_version = parent_manifest.get("version")
    if not isinstance(parent_version, str) or parent_gold.get("version") != parent_version:
        raise OfficialAmendmentError("父派生 manifest/gold_manifest 版本不一致")

    child_manifest, child_workbook, _, _, child_manifest_sha = _validate_parent(
        child_official_dir
    )
    spec, entries = _load_spec(spec_path, parent_version)
    target_version = str(spec["targetVersion"])
    if child_manifest.get("datasetVersion") != target_version:
        raise OfficialAmendmentError("子正式包版本与修正清单 targetVersion 不一致")
    amendment_path = _relative_artifact(
        child_official_dir,
        child_manifest.get("answerAmendmentLedger"),
        "answerAmendmentLedger",
    )
    amendment_sha = _verify_sidecar(amendment_path)
    if child_manifest.get("artifactSha256", {}).get("answerAmendmentLedger") != amendment_sha:
        raise OfficialAmendmentError("子正式包 answerAmendmentLedger 哈希不一致")
    amendment = _read_json(amendment_path, "子正式包答案修正账本")
    if amendment.get("correctionSpecSha256") != sha256_file(spec_path):
        raise OfficialAmendmentError("子正式包答案修正账本未绑定当前修正清单")

    corrections_by_split: dict[str, dict[str, dict[str, str]]] = {"train": {}, "dev": {}}
    for entry in entries:
        corrections_by_split[entry["split"]][entry["id"]] = entry

    output_dir.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(prefix="derived-amendment-", dir=output_dir.parent) as temp_dir:
        stage = Path(temp_dir) / "dataset"
        stage.mkdir()
        changed: set[str] = set()
        for split in ("train", "dev"):
            changed.update(
                _amend_jsonl_answers(
                    parent_dataset_dir / f"{split}.jsonl",
                    stage / f"{split}.jsonl",
                    corrections_by_split[split],
                )
            )
        expected_ids = {entry["id"] for entry in entries}
        if changed != expected_ids:
            raise OfficialAmendmentError(
                f"派生数据答案变化集合不完整: {sorted(changed)} != {sorted(expected_ids)}"
            )

        for name in ("test.jsonl", "augmentation.jsonl", "schema.json"):
            source = parent_dataset_dir / name
            if not source.is_file():
                raise OfficialAmendmentError(f"父派生数据集缺少 {name}")
            shutil.copy2(source, stage / name)

        derived_manifest = json.loads(json.dumps(parent_manifest))
        derived_manifest["version"] = target_version
        derived_manifest["parentVersion"] = parent_version
        derived_manifest["answerAmendment"] = {
            "count": len(entries),
            "officialManifestSha256": child_manifest_sha,
            "ledgerSha256": amendment_sha,
            "canonicalWorkbook": child_manifest["groundTruthWorkbook"],
            "canonicalWorkbookSha256": sha256_file(child_workbook),
        }
        _write_json(stage / "manifest.json", derived_manifest)

        derived_gold = json.loads(json.dumps(parent_gold))
        derived_gold["version"] = target_version
        derived_gold["parentVersion"] = parent_version
        derived_gold["answerAmendmentLedgerSha256"] = amendment_sha
        _write_json(stage / "gold_manifest.json", derived_gold)
        stage.replace(output_dir)
    return derived_manifest


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--parent-dir", type=Path, required=True)
    parser.add_argument("--corrections", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--update-current", action="store_true")
    args = parser.parse_args(argv)
    try:
        manifest = amend_official_package(
            args.parent_dir,
            args.corrections,
            args.output,
            update_current=args.update_current,
        )
    except OfficialAmendmentError as exc:
        parser.error(str(exc))
    print(
        json.dumps(
            {
                "datasetVersion": manifest["datasetVersion"],
                "answerAmendmentCount": manifest["answerAmendmentCount"],
                "canonicalReady": manifest["canonicalReady"],
            },
            ensure_ascii=False,
            sort_keys=True,
        )
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
