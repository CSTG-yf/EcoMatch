#!/usr/bin/env python3
"""第二版银行 NL2SQL Ground Truth 全量审查器。

唯一证据：冻结工作簿的机构信息表 / 指标清单表 / 衍生维度说明 / 指标数据表。
旧答案（问题结果列）只作为待核对对象，不参与规则推断。

输出（.local-dev/gt-audit/output/）：
- audit-summary.json        计数、划分、fullEvidence/auditErrors、源与输出 SHA-256、规则版本
- review.ndjson             每题一行：id/状态/类别/claims/核对记录
- correction-ledger.json    所有 CORRECTED 题的旧答案与完整新答案
- candidate-reviewed.xlsx   全量审查表（canonicalReady=false 时）
- canonical-corrected.xlsx  canonical 数据集（canonicalReady=true 时）

终端只打印计数、状态与 ID，不打印题目、答案、事实行或 SQL。
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
import zipfile
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

import gt_answer_rules as rules

SOURCE_SHA256_EXPECTED = rules.SOURCE_SHA256_EXPECTED

# --------------------------------------------------------------------------- 加载


class WorkbookData:
    def __init__(self, workbook_path: Path) -> None:
        self.workbook_path = workbook_path
        self.source_sha256 = ""
        self.organizations: dict[str, rules.Organization] = {}
        self.metrics: dict[str, rules.Metric] = {}
        self.derived_specs: dict[str, str] = {}
        self.facts: dict[tuple[str, str, str], float] = {}
        self.dates: list[str] = []
        self.questions: list[dict[str, Any]] = []
        self._load()

    def _load(self) -> None:
        digest = hashlib.sha256()
        with self.workbook_path.open("rb") as source:
            for block in iter(lambda: source.read(1024 * 1024), b""):
                digest.update(block)
        self.source_sha256 = digest.hexdigest().upper()
        workbook = load_workbook(self.workbook_path, read_only=True, data_only=True)
        try:
            self._load_organizations(workbook)
            self._load_metrics(workbook)
            self._load_derived(workbook)
            self._load_facts(workbook)
            self._load_questions(workbook)
        finally:
            workbook.close()

    def _rows(self, workbook: Any, sheet_name: str, headers: tuple[str, ...]) -> list[tuple[int, tuple[Any, ...]]]:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"缺少工作表: {sheet_name}")
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        actual = tuple(str(value).strip() if value is not None else "" for value in rows[0])
        if actual[: len(headers)] != headers:
            raise ValueError(f"{sheet_name} 表头不匹配: {actual!r}")
        result: list[tuple[int, tuple[Any, ...]]] = []
        for index, row in enumerate(rows[1:], start=2):
            if all(value is None for value in row):
                continue
            result.append((index, tuple(row[: len(headers)])))
        return result

    def _load_organizations(self, workbook: Any) -> None:
        for _row_number, row in self._rows(workbook, rules.SHEET_ORG, rules.ORG_HEADERS):
            code, name = str(row[0]).strip(), str(row[1]).strip()
            self.organizations[code] = rules.Organization(code=code, name=name)

    def _load_metrics(self, workbook: Any) -> None:
        for _row_number, row in self._rows(workbook, rules.SHEET_METRIC, rules.METRIC_HEADERS):
            code, name, meaning, unit = (str(value).strip() for value in row)
            self.metrics[code] = rules.Metric(code=code, name=name, meaning=meaning, unit=unit)

    def _load_derived(self, workbook: Any) -> None:
        for _row_number, row in self._rows(workbook, rules.SHEET_DERIVED, rules.DERIVED_HEADERS):
            name, spec = str(row[0]).strip(), str(row[1]).strip()
            self.derived_specs[name] = spec

    def _load_facts(self, workbook: Any) -> None:
        date_set: set[str] = set()
        for _row_number, row in self._rows(workbook, rules.SHEET_FACT, rules.FACT_HEADERS):
            data_date, metric_code, metric_name, org_code, metric_value = row
            data_date = str(data_date).strip()
            metric_code = str(metric_code).strip()
            org_code = str(org_code).strip()
            if metric_name.strip() != self.metrics[metric_code].name:
                raise ValueError(f"事实表指标名称不一致: {metric_code} -> {metric_name}")
            self.facts[(data_date, org_code, metric_code)] = float(metric_value)
            date_set.add(data_date)
        self.dates = sorted(date_set)

    def _load_questions(self, workbook: Any) -> None:
        seen: set[str] = set()
        for row_number, row in self._rows(workbook, rules.SHEET_QUESTION, rules.QUESTION_HEADERS):
            qid, question_type, difficulty, question, answer = (str(value).strip() for value in row)
            if qid in seen:
                raise ValueError(f"重复问题编号: {qid}")
            seen.add(qid)
            self.questions.append(
                {
                    "id": qid,
                    "split": rules.SPLIT_MAP[question_type],
                    "difficulty": difficulty,
                    "question": question,
                    "answerText": answer,
                    "rowNumber": row_number,
                }
            )

    # ---- 查询原语（唯一数值证据） ----

    def value(self, metric: str, org: str, date_text: str) -> float | None:
        return self.facts.get((date_text, org, metric))

    def org_values(self, metric: str, date_text: str) -> dict[str, float]:
        return {org: value for (d, org, m), value in self.facts.items() if d == date_text and m == metric}

    def province_mean(self, metric: str, date_text: str) -> float | None:
        values = self.org_values(metric, date_text)
        if len(values) != len(self.organizations):
            return None
        return sum(values.values()) / len(values)

    def series(self, metric: str, org: str, dates: list[str]) -> list[float | None]:
        return [self.value(metric, org, date_text) for date_text in dates]

    def range_mean(self, metric: str, org: str, start: str, end: str) -> float | None:
        values = [value for (d, o, m), value in self.facts.items() if start <= d <= end and o == org and m == metric]
        return sum(values) / len(values) if values else None

    def range_stats(self, metric: str, org: str, start: str, end: str) -> dict[str, float] | None:
        values = [value for (d, o, m), value in self.facts.items() if start <= d <= end and o == org and m == metric]
        if not values:
            return None
        return {"mean": sum(values) / len(values), "min": min(values), "max": max(values)}

    def count_days_above_mean(self, metric: str, org: str, start: str, end: str) -> dict[str, Any] | None:
        total = 0
        above = 0
        for day in self.dates:
            if not (start <= day <= end):
                continue
            values = self.org_values(metric, day)
            if len(values) != len(self.organizations):
                continue
            mean = sum(values.values()) / len(values)
            org_value = values.get(org)
            if org_value is None:
                continue
            total += 1
            if org_value > mean:
                above += 1
        if total == 0:
            return None
        return {"above": above, "total": total, "pct": above * 100.0 / total}

    def exists_date(self, date_text: str) -> bool:
        return date_text in self.dates

    def exists_fact(self, metric: str, org: str, date_text: str) -> bool:
        return (date_text, org, metric) in self.facts


# --------------------------------------------------------------------------- claims


def make_claim(
    key: str,
    label: str,
    *,
    kind: str,
    value: float | None,
    unit: str | None = None,
    rounding: int = 2,
    metric: str | None = None,
    org: str | None = None,
    date: str | None = None,
    baseline: str | None = None,
    comparison_type: str | None = None,
    direction: str | None = None,
    role: str | None = None,
    must_appear: bool = True,
    extras: list[dict[str, Any]] | None = None,
    note: str | None = None,
    evidence: dict[str, Any] | None = None,
) -> dict[str, Any]:
    metric_name = data_metrics.get(metric) if metric else None
    org_name = data_org_names.get(org) if org else None
    return {
        "key": key,
        "label": label,
        "kind": kind,
        "value": value,
        "unit": unit,
        "rounding": rounding,
        "metric": metric,
        "metricName": metric_name,
        "org": org,
        "orgName": org_name,
        "date": date,
        "baseline": baseline,
        "comparisonType": comparison_type,
        "direction": direction,
        "role": role,
        "mustAppear": must_appear,
        "extras": extras or [],
        "note": note,
        "evidence": evidence,
        "matched": False,
        "matchNote": None,
    }


data_metrics: dict[str, str] = {}
data_org_names: dict[str, str] = {}


def claim_display(claim: dict[str, Any]) -> str:
    value = claim["value"]
    if value is None:
        return "缺失"
    if claim["kind"] == "TREND":
        return {"up": "上升", "down": "下降", "flat": "持平"}.get(claim.get("direction") or "", str(value))
    text = rules.display_value(value, claim["rounding"])
    unit = claim["unit"] or ""
    return f"{text}{unit}"


# --------------------------------------------------------------------------- 证据契约
#
# 每个 claim 携带结构化 evidence：operation（可机器复核的重算操作）、输入范围
# （metric/org/date 或 dateRange/baseline/scope.orgs；派生声明用 dependsOn 引用
# 本 review 内其他 claim 的 key）、formula/rounding、sourceFactCount 与
# sourceFactsSha256（按稳定顺序序列化底层事实键与值后的 SHA-256，与 answerText
# 无关）。fullEvidence 由 validate_review_evidence 的结果决定，绝不从 status 推导。

EVIDENCE_RULES_VERSION = "1.0.0"

# operation -> 必填字段（scope/dependsOn/dateRange 有专属检查）
_EVIDENCE_REQUIRED: dict[str, tuple[str, ...]] = {
    "VALUE_LOOKUP": ("metric", "org", "date"),
    "RATIO": ("numerator", "denominator", "org", "date"),
    "SUM": ("metric", "date", "scope"),
    "SUM_DEPENDS": ("dependsOn",),
    "DELTA": ("metric", "org", "date", "baseline"),
    "DELTA_DEPENDS": ("dependsOn",),
    "DELTA_ABS": ("dependsOn",),
    "PCT_CHANGE": ("metric", "org", "date", "baseline"),
    "TREND_DELTA": ("dependsOn",),
    "TREND_SLOPE": ("dependsOn",),
    "EXTREME_DEPENDS": ("dependsOn",),
    "EXTREME_RANGE": ("metric", "org", "dateRange"),
    "EXTREME_ALL": ("metric", "dateRange", "scope"),
    "RANGE_MEAN": ("metric", "org", "dateRange"),
    "RANGE_STATS": ("metric", "org", "dateRange"),
    "PROVINCE_MEAN": ("metric", "date", "scope"),
    "PROVINCE_DIFF": ("dependsOn",),
    "DAYS_ABOVE_MEAN": ("metric", "org", "dateRange", "scope"),
    "COUNT_ABOVE_MEAN": ("metric", "date", "scope"),
    "COUNT_CONDITIONS": ("metricA", "metricB", "date", "scope"),
    "RANK_POSITION": ("org", "date", "scope"),  # metric 或 derived 二选一，见 validate_claim_evidence
    "TOP_N": ("metric", "date", "scope"),
    "TOP_N_MEAN": ("metric", "dateRange", "scope"),
    "TOP_N_PCT_CHANGE": ("metric", "date", "baseline", "scope"),
    "TOP_N_DECLINE": ("metric", "date", "baseline", "scope"),
    "RANK_EXTREME": ("metric", "date", "scope"),
    "COMPARE_WINNER": ("metric", "date", "scope"),
    "MEETS": ("dependsOn", "threshold"),
    "SET_GOOD": ("dependsOn",),
    "SET_BAD": ("dependsOn",),
}


def fact_digest(facts: dict[tuple[str, str, str], float]) -> str:
    """按 (日期, 机构, 指标) 稳定排序后序列化事实键与值，返回 SHA-256。"""
    payload = "\n".join(f"{d}|{o}|{m}={v!r}" for (d, o, m), v in sorted(facts.items()))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest().upper()


def _evidence_factory(
    operation: str,
    formula: str,
    rounding: int,
    data: "WorkbookData",
    **inputs: Any,
) -> dict[str, Any]:
    """构造 evidence 并填充直接事实摘要；dependsOn 类由 finalize_evidence 递归填充。"""
    ev: dict[str, Any] = {"operation": operation, "formula": formula, "rounding": rounding}
    ev.update(inputs)
    if "dependsOn" in ev:
        return ev
    facts = _collect_facts(data, ev)
    ev["sourceFactCount"] = len(facts)
    ev["sourceFactsSha256"] = fact_digest(facts)
    return ev


def _dates_in_range(data: "WorkbookData", start: str, end: str) -> list[str]:
    return [day for day in data.dates if start <= day <= end]


def _collect_facts(data: "WorkbookData", ev: dict[str, Any]) -> dict[tuple[str, str, str], float]:
    """按 evidence 输入范围收集直接底层事实（稳定键序由 fact_digest 处理）。"""
    op = ev["operation"]
    facts: dict[tuple[str, str, str], float] = {}
    if op in ("VALUE_LOOKUP",):
        value = data.value(ev["metric"], ev["org"], ev["date"])
        if value is not None:
            facts[(ev["date"], ev["org"], ev["metric"])] = value
    elif op in ("RATIO",):
        for metric in (ev["numerator"], ev["denominator"]):
            value = data.value(metric, ev["org"], ev["date"])
            if value is not None:
                facts[(ev["date"], ev["org"], metric)] = value
    elif op in ("DELTA", "PCT_CHANGE"):
        for date_text in (ev.get("date"), ev.get("baseline")):
            value = data.value(ev["metric"], ev["org"], date_text)
            if value is not None:
                facts[(date_text, ev["org"], ev["metric"])] = value
    elif op in ("SUM", "RANK_POSITION", "TOP_N", "RANK_EXTREME", "COMPARE_WINNER", "COUNT_ABOVE_MEAN", "PROVINCE_MEAN"):
        date_text = ev["date"]
        if op == "RANK_POSITION" and isinstance(ev.get("derived"), dict):
            # 衍生指标排名：收集分子/分母两个指标的机构 × 日期事实
            for metric in (ev["derived"]["numerator"], ev["derived"]["denominator"]):
                for org in ev["scope"]["orgs"]:
                    value = data.value(metric, org, date_text)
                    if value is not None:
                        facts[(date_text, org, metric)] = value
        else:
            for org in ev["scope"]["orgs"]:
                value = data.value(ev["metric"], org, date_text)
                if value is not None:
                    facts[(date_text, org, ev["metric"])] = value
    elif op == "COUNT_CONDITIONS":
        date_text = ev["date"]
        for metric in (ev["metricA"], ev["metricB"]):
            for org in ev["scope"]["orgs"]:
                value = data.value(metric, org, date_text)
                if value is not None:
                    facts[(date_text, org, metric)] = value
    elif op in ("RANGE_MEAN", "RANGE_STATS", "EXTREME_RANGE"):
        start, end = ev.get("dateRange") or ("", "")
        for day in _dates_in_range(data, start, end):
            value = data.value(ev["metric"], ev["org"], day)
            if value is not None:
                facts[(day, ev["org"], ev["metric"])] = value
    elif op in ("EXTREME_ALL", "DAYS_ABOVE_MEAN", "TOP_N_MEAN"):
        start, end = ev.get("dateRange") or ("", "")
        for day in _dates_in_range(data, start, end):
            for org in ev["scope"]["orgs"]:
                value = data.value(ev["metric"], org, day)
                if value is not None:
                    facts[(day, org, ev["metric"])] = value
    elif op in ("TOP_N_PCT_CHANGE", "TOP_N_DECLINE"):
        for date_text in (ev["date"], ev["baseline"]):
            for org in ev["scope"]["orgs"]:
                value = data.value(ev["metric"], org, date_text)
                if value is not None:
                    facts[(date_text, org, ev["metric"])] = value
    return facts


def _all_facts(
    data: "WorkbookData",
    ev: dict[str, Any],
    claims_by_key: dict[str, dict[str, Any]],
    _visiting: frozenset[str] = frozenset(),
) -> dict[tuple[str, str, str], float]:
    """dependsOn 类 operation：递归合并依赖链上全部直接事实（无环前提下）。"""
    if "dependsOn" not in ev:
        return _collect_facts(data, ev)
    facts: dict[tuple[str, str, str], float] = {}
    for key in ev.get("dependsOn", []):
        if key in _visiting:
            continue
        target = claims_by_key.get(key)
        if target is None:
            continue
        target_ev = target.get("evidence")
        if not isinstance(target_ev, dict):
            continue
        facts.update(_all_facts(data, target_ev, claims_by_key, _visiting | {key}))
    return facts


def _rank_values(data: "WorkbookData", metric: str, date_text: str, orgs: list[str]) -> dict[str, float]:
    values: dict[str, float] = {}
    for org in orgs:
        value = data.value(metric, org, date_text)
        if value is None:
            return {}
        values[org] = value
    return values


def _rank_positions_for(data: "WorkbookData", metric: str, date_text: str, orgs: list[str]) -> list[tuple[str, str, float, int]]:
    values = _rank_values(data, metric, date_text, orgs)
    if not values:
        return []
    return rules.rank_positions([(code, data.organizations[code].name, value) for code, value in values.items()], metric)


def _dep_value(claims_by_key: dict[str, dict[str, Any]], key: str) -> float | None:
    target = claims_by_key.get(key)
    return target["value"] if target else None


def _recompute(data: "WorkbookData", ev: dict[str, Any], claims_by_key: dict[str, dict[str, Any]]) -> Any:
    """按 operation 从证据输入重算值；与 claim value 比对由 _computed_matches 完成。"""
    op = ev["operation"]
    if op == "VALUE_LOOKUP":
        return data.value(ev["metric"], ev["org"], ev["date"])
    if op == "RATIO":
        nv = data.value(ev["numerator"], ev["org"], ev["date"])
        dv = data.value(ev["denominator"], ev["org"], ev["date"])
        if nv is None or dv is None or dv == 0:
            return None
        value = nv * ev.get("scale", 1.0) / dv
        if ev.get("percent"):
            value *= 100.0
        return value
    if op == "SUM":
        total = 0.0
        for org in ev["scope"]["orgs"]:
            value = data.value(ev["metric"], org, ev["date"])
            if value is None:
                return None
            total += value
        return total
    if op == "SUM_DEPENDS":
        total = 0.0
        for key in ev["dependsOn"]:
            value = _dep_value(claims_by_key, key)
            if value is None:
                return None
            total += value
        return total
    if op in ("DELTA", "PCT_CHANGE"):
        current = data.value(ev["metric"], ev["org"], ev.get("date"))
        base = data.value(ev["metric"], ev["org"], ev.get("baseline"))
        if current is None or base is None:
            return None
        if op == "DELTA":
            return current - base
        if base == 0:
            return None
        return (current - base) * 100.0 / base
    if op == "DELTA_DEPENDS":
        first = _dep_value(claims_by_key, ev["dependsOn"][0])
        second = _dep_value(claims_by_key, ev["dependsOn"][1])
        if first is None or second is None:
            return None
        diff = first - second
        return abs(diff) if ev.get("abs") else diff
    if op == "DELTA_ABS":
        first = _dep_value(claims_by_key, ev["dependsOn"][0])
        second = _dep_value(claims_by_key, ev["dependsOn"][1])
        if first is None or second is None:
            return None
        return abs(first - second)
    if op == "TREND_DELTA":
        first = _dep_value(claims_by_key, ev["dependsOn"][0])
        second = _dep_value(claims_by_key, ev["dependsOn"][1])
        if first is None or second is None:
            return None
        diff = second - first
        return 1.0 if diff > 0.005 else -1.0 if diff < -0.005 else 0.0
    if op == "TREND_SLOPE":
        values = [_dep_value(claims_by_key, key) for key in ev["dependsOn"]]
        if any(value is None for value in values):
            return None
        xs = list(range(len(values)))
        mean_x = sum(xs) / len(xs)
        mean_y = sum(values) / len(values)  # type: ignore[arg-type]
        slope = sum((x - mean_x) * (values[i] - mean_y) for i, x in enumerate(xs))
        return 1.0 if slope > 1e-9 else -1.0 if slope < -1e-9 else 0.0
    if op == "EXTREME_DEPENDS":
        values = [_dep_value(claims_by_key, key) for key in ev["dependsOn"]]
        if any(value is None for value in values):
            return None
        return max(values) if ev["extreme"] == "max" else min(values)  # type: ignore[arg-type]
    if op == "RANGE_MEAN":
        facts = _collect_facts(data, ev)
        return sum(facts.values()) / len(facts) if facts else None
    if op == "RANGE_STATS":
        facts = _collect_facts(data, ev)
        if not facts:
            return None
        values = list(facts.values())
        return {"mean": sum(values) / len(values), "min": min(values), "max": max(values)}
    if op == "EXTREME_RANGE":
        facts = _collect_facts(data, ev)
        if not facts:
            return None
        return max(facts.values()) if ev["extreme"] == "max" else min(facts.values())
    if op == "EXTREME_ALL":
        start, end = ev.get("dateRange") or ("", "")
        extremes: dict[str, float] = {}
        for org in ev["scope"]["orgs"]:
            values = [data.value(ev["metric"], org, day) for day in _dates_in_range(data, start, end)]
            values = [value for value in values if value is not None]
            if not values:
                return None
            extremes[org] = max(values) if ev["extreme"] == "max" else min(values)
        target_org = max(extremes, key=extremes.get) if ev["extreme"] == "max" else min(extremes, key=extremes.get)
        return (target_org, extremes[target_org])
    if op == "PROVINCE_MEAN":
        facts = _collect_facts(data, ev)
        return sum(facts.values()) / len(facts) if facts else None
    if op == "PROVINCE_DIFF":
        first = _dep_value(claims_by_key, ev["dependsOn"][0])
        second = _dep_value(claims_by_key, ev["dependsOn"][1])
        if first is None or second is None:
            return None
        return first - second
    if op == "DAYS_ABOVE_MEAN":
        start, end = ev.get("dateRange") or ("", "")
        metric, org = ev["metric"], ev["org"]
        orgs = ev["scope"]["orgs"]
        total = above = 0
        for day in _dates_in_range(data, start, end):
            values = _rank_values(data, metric, day, orgs)
            if not values:
                continue
            mean = sum(values.values()) / len(values)
            org_value = values.get(org)
            if org_value is None:
                continue
            total += 1
            if org_value > mean:
                above += 1
        if total == 0:
            return None
        stat = ev["stat"]
        return {"above": float(above), "total": float(total), "pct": above * 100.0 / total}[stat]
    if op == "COUNT_ABOVE_MEAN":
        values = _rank_values(data, ev["metric"], ev["date"], ev["scope"]["orgs"])
        if not values:
            return None
        mean = sum(values.values()) / len(values)
        if ev["direction"] == "above":
            return float(sum(1 for value in values.values() if value > mean))
        return float(sum(1 for value in values.values() if value < mean))
    if op == "COUNT_CONDITIONS":
        values_a = _rank_values(data, ev["metricA"], ev["date"], ev["scope"]["orgs"])
        values_b = _rank_values(data, ev["metricB"], ev["date"], ev["scope"]["orgs"])
        if not values_a or not values_b:
            return None
        mean_a = sum(values_a.values()) / len(values_a)
        mean_b = sum(values_b.values()) / len(values_b)
        return float(sum(1 for org in ev["scope"]["orgs"] if values_a[org] < mean_a and values_b[org] > mean_b))
    if op == "RANK_POSITION":
        if isinstance(ev.get("derived"), dict):
            values: dict[str, float] = {}
            for org in ev["scope"]["orgs"]:
                nv = data.value(ev["derived"]["numerator"], org, ev["date"])
                dv = data.value(ev["derived"]["denominator"], org, ev["date"])
                if nv is None or dv is None or dv == 0:
                    return None
                derived_value = nv * ev["derived"].get("scale", 1.0) / dv
                if ev["derived"].get("percent"):
                    derived_value *= 100.0
                values[org] = derived_value
            positions = rules.rank_positions(
                [(code, data.organizations[code].name, value) for code, value in values.items()],
                ev["derived"]["numerator"],
            )
        else:
            positions = _rank_positions_for(data, ev["metric"], ev["date"], ev["scope"]["orgs"])
        if not positions:
            return None
        target = next((item for item in positions if item[0] == ev["org"]), None)
        return float(target[3]) if target else None
    if op in ("TOP_N", "TOP_N_MEAN", "TOP_N_PCT_CHANGE", "TOP_N_DECLINE"):
        if op == "TOP_N":
            positions = _rank_positions_for(data, ev["metric"], ev["date"], ev["scope"]["orgs"])
            if not positions:
                return None
            ordered = sorted(positions, key=lambda item: item[3], reverse=True) if ev.get("side") == "back" else sorted(positions, key=lambda item: item[3])
        elif op == "TOP_N_MEAN":
            start, end = ev.get("dateRange") or ("", "")
            means: dict[str, float] = {}
            for org in ev["scope"]["orgs"]:
                values = [data.value(ev["metric"], org, day) for day in _dates_in_range(data, start, end)]
                values = [value for value in values if value is not None]
                if not values:
                    return None
                means[org] = sum(values) / len(values)
            positions = rules.rank_positions([(code, data.organizations[code].name, value) for code, value in means.items()], ev["metric"])
            ordered = sorted(positions, key=lambda item: item[3], reverse=True) if ev.get("side") == "back" else sorted(positions, key=lambda item: item[3])
        else:
            changes: list[tuple[str, float]] = []
            for org in ev["scope"]["orgs"]:
                current = data.value(ev["metric"], org, ev.get("date"))
                base = data.value(ev["metric"], org, ev.get("baseline"))
                if current is None or base is None:
                    return None
                change = (current - base) * 100.0 / base if op == "TOP_N_PCT_CHANGE" else base - current
                changes.append((org, change))
            changes.sort(key=lambda item: (-item[1], item[0]))
            return [(org, value, index) for index, (org, value) in enumerate(changes[:3], start=1)]
        return [(item[0], item[2], item[3]) for item in ordered[:3]]
    if op == "RANK_EXTREME":
        values = _rank_values(data, ev["metric"], ev["date"], ev["scope"]["orgs"])
        if not values:
            return None
        if ev.get("mode") == "rank":  # 最后一名：按名次取极值（rank 最大）
            positions = rules.rank_positions([(code, data.organizations[code].name, value) for code, value in values.items()], ev["metric"])
            target = max(positions, key=lambda item: item[3])
            return (target[0], target[2])
        if ev["extreme"] == "max":
            target_org = max(values, key=values.get)
        else:
            target_org = min(values, key=values.get)
        return (target_org, values[target_org])
    if op == "COMPARE_WINNER":
        values = _rank_values(data, ev["metric"], ev["date"], ev["scope"]["orgs"])
        if not values:
            return None
        if ev.get("ascending"):
            target_org = min(values, key=values.get)
        else:
            target_org = max(values, key=values.get)
        return (target_org, values[target_org])
    if op == "MEETS":
        value = _dep_value(claims_by_key, ev["dependsOn"][0])
        if value is None:
            return None
        return 1.0 if value > ev["threshold"] else 0.0
    if op == "SET_GOOD":
        values = [_dep_value(claims_by_key, key) for key in ev["dependsOn"]]
        if any(value is None for value in values):
            return None
        return float(sum(1 for value in values if value <= rules.GOOD_RANK_CUTOFF))  # type: ignore[arg-type]
    if op == "SET_BAD":
        values = [_dep_value(claims_by_key, key) for key in ev["dependsOn"]]
        if any(value is None for value in values):
            return None
        cutoff = ev.get("total", 13) - rules.BAD_TAIL_SIZE + 1
        return float(sum(1 for value in values if value >= cutoff))  # type: ignore[arg-type]
    return None


def _computed_matches(claim: dict[str, Any], computed: Any) -> bool:
    rounding = claim.get("rounding", 2)
    claim_value = claim.get("value")
    if isinstance(computed, list):  # TOP_N 系列条目列表
        target = next((item for item in computed if item[0] == claim.get("org")), None)
        if target is None:
            return False
        return target[2] == claim.get("position") and rules.values_match(claim_value, target[1], rounding)
    if isinstance(computed, tuple):  # (org, value)
        return computed[0] == claim.get("org") and rules.values_match(claim_value, computed[1], rounding)
    if isinstance(computed, dict):  # RANGE_STATS / DAYS_ABOVE_MEAN 统计字典
        if claim["kind"] == "EXTREME":
            key = claim.get("evidence", {}).get("extreme")
        else:
            key = {"COUNT": "above", "TOTAL": "total", "PCT_RATIO": "pct", "MEAN": "mean"}.get(claim["kind"])
        value = computed.get(key) if key else None
        return value is not None and rules.values_match(claim_value, value, rounding)
    return rules.values_match(claim_value, computed, rounding)


def finalize_evidence(claims: list[dict[str, Any]], data: "WorkbookData") -> None:
    """dependsOn 类 evidence 的摘要依赖链在全部 claim 生成后填充。"""
    claims_by_key = {claim["key"]: claim for claim in claims}
    for claim in claims:
        ev = claim.get("evidence")
        if not isinstance(ev, dict) or "dependsOn" not in ev:
            continue
        facts = _all_facts(data, ev, claims_by_key)
        ev["sourceFactCount"] = len(facts)
        ev["sourceFactsSha256"] = fact_digest(facts)


def _find_depends_cycle(claims_by_key: dict[str, dict[str, Any]]) -> list[str] | None:
    """返回第一个 dependsOn 环（节点 key 序列）或 None。"""
    state: dict[str, int] = {}
    for start in claims_by_key:
        if state.get(start) == 1:
            continue
        stack: list[tuple[str, list[str]]] = [(start, list((claims_by_key[start].get("evidence") or {}).get("dependsOn", [])))]
        path = [start]
        state[start] = 0
        while stack:
            node, deps = stack[-1]
            if deps:
                dep = deps.pop(0)
                if dep not in claims_by_key or state.get(dep) == 1:
                    continue
                if state.get(dep) == 0:
                    cycle = path[path.index(dep):] + [dep]
                    return cycle
                state[dep] = 0
                path.append(dep)
                stack.append((dep, list((claims_by_key[dep].get("evidence") or {}).get("dependsOn", []))))
            else:
                state[node] = 1
                stack.pop()
                path.pop()
    return None


def validate_claim_evidence(claim: dict[str, Any], data: "WorkbookData", claims_by_key: dict[str, dict[str, Any]] | None = None) -> tuple[bool, list[str]]:
    """逐 claim 验证：必填字段、事实摘要可重算、派生依赖存在、computed 值与 claim value 相符。"""
    errors: list[str] = []
    if claim.get("value") is None:
        return True, errors  # 缺值由审计流程按“声明无法重算”处理
    ev = claim.get("evidence")
    if not isinstance(ev, dict):
        return False, ["缺少 evidence"]
    op = ev.get("operation")
    if not op:
        return False, ["evidence 缺少 operation"]
    required = _EVIDENCE_REQUIRED.get(op)
    if required is None:
        return False, [f"未知证据 operation: {op}"]
    for field in required:
        if field == "scope":
            if not ev.get("scope") or not ev["scope"].get("orgs"):
                errors.append("evidence 缺少 scope.orgs")
        elif field == "dependsOn":
            if not ev.get("dependsOn"):
                errors.append("evidence 缺少 dependsOn")
        elif field == "dateRange":
            if not (isinstance(ev.get("dateRange"), (list, tuple)) and len(ev["dateRange"]) == 2):
                errors.append("evidence 缺少 dateRange")
        else:
            if ev.get(field) in (None, ""):
                errors.append(f"evidence 缺少 {field}")
    if op == "RANK_POSITION":
        derived = ev.get("derived")
        if ev.get("metric") not in (None, "") and isinstance(derived, dict):
            errors.append("RANK_POSITION 不应同时设置 metric 与 derived")
        if ev.get("metric") in (None, "") and not isinstance(derived, dict):
            errors.append("RANK_POSITION evidence 缺少 metric 或 derived")
        if isinstance(derived, dict):
            for field in ("numerator", "denominator"):
                if derived.get(field) in (None, ""):
                    errors.append(f"evidence 缺少 derived.{field}")
    if not ev.get("formula"):
        errors.append("evidence 缺少 formula")
    if ev.get("rounding") != claim.get("rounding"):
        errors.append("evidence.rounding 与 claim.rounding 不一致")
    by_key = claims_by_key or {}
    for key in ev.get("dependsOn", []):
        target = by_key.get(key)
        if target is None:
            errors.append(f"dependsOn 目标缺失: {key}")
        elif target.get("value") is None:
            errors.append(f"dependsOn 目标无值: {key}")
    try:
        facts = _all_facts(data, ev, by_key)
    except RecursionError:
        errors.append("dependsOn 递归过深（疑似成环）")
        facts = {}
    if ev.get("sourceFactCount") != len(facts):
        errors.append(f"sourceFactCount 不符（{ev.get('sourceFactCount')} != 重算 {len(facts)}）")
    if ev.get("sourceFactsSha256") != fact_digest(facts):
        errors.append("sourceFactsSha256 与按输入范围重算的事实集不一致")
    computed = _recompute(data, ev, by_key)
    if computed is None:
        errors.append("computed value 重算失败（输入事实缺失）")
    elif not _computed_matches(claim, computed):
        errors.append("computed value 与 claim value 在指定舍入下不符")
    return not errors, errors


def validate_review_evidence(claims: list[dict[str, Any]], data: "WorkbookData") -> dict[str, Any]:
    """review 级证据验证：全部 claim 逐项验证 + dependsOn 存在性与无环。

    返回 {"valid", "claimResults": {key: {"ok", "errors"}}, "errors"}。
    """
    claims_by_key = {claim["key"]: claim for claim in claims}
    claim_results: dict[str, Any] = {}
    review_errors: list[str] = []
    cycle = _find_depends_cycle(claims_by_key)
    if cycle is not None:
        review_errors.append(f"dependsOn 成环: {' -> '.join(cycle)}")
    for claim in claims:
        ok, errors = validate_claim_evidence(claim, data, claims_by_key)
        if cycle is not None and claim["key"] in cycle:
            ok = False
            errors = list(errors) + ["参与 dependsOn 成环"]
        claim_results[claim["key"]] = {"ok": ok, "errors": errors}
    valid = not review_errors and all(result["ok"] for result in claim_results.values())
    return {"valid": valid, "claimResults": claim_results, "errors": review_errors}


# --------------------------------------------------------------------------- 日期工具


def extract_all_dates(question: str) -> list[tuple[int, str]]:
    """返回 [(位置, ISO 日期)]，按出现顺序。"""
    found: list[tuple[int, str]] = []
    for pattern in (rules._ISO_DATE_RE, rules._CN_DATE_RE, rules._CN_MONTH_END_RE, rules._CN_Q_RE, rules._CN_QUARTER_RE, rules._CN_HALF_RE, rules._CN_YEAR_END_RE):
        for match in pattern.finditer(question):
            iso = rules.parse_absolute_date(question[match.start():match.end()])
            if iso:
                found.append((match.start(), iso))
    # 无年份的“年末/年底”：取该位置之前最近的年份
    for match in re.finditer(r"年\s*(?:底|末)", question):
        year_match = re.search(r"(20\d{2})\s*年", question[:match.start()])
        if year_match:
            found.append((match.start(), f"{int(year_match.group(1)):04d}-12-31"))
    found.sort(key=lambda item: item[0])
    # 去重（同一位置同一日期）
    unique: list[tuple[int, str]] = []
    seen: set[str] = set()
    for position, iso in found:
        if (position, iso) not in seen:
            seen.add((position, iso))
            unique.append((position, iso))
    return unique


def main_date_of(question: str) -> str | None:
    all_dates = extract_all_dates(question)
    if not all_dates:
        return None
    for marker in ("截至", "到", "在"):
        index = question.find(marker)
        if index >= 0:
            for position, iso in all_dates:
                if position >= index:
                    return iso
    candidates = []
    for position, iso in all_dates:
        nearby = question[max(0, position - 3):position + 3]
        if iso == rules.YEAR_START and ("年末" in nearby or "年初" in nearby):
            continue
        candidates.append(iso)
    return candidates[-1] if candidates else all_dates[-1][1]


# --------------------------------------------------------------------------- 解析与 claim 生成


class ParseResult:
    def __init__(self, question: str, data: WorkbookData) -> None:
        self.question = question
        self.data = data
        self.orgs: list[rules.Organization] = []
        self.metrics: list[rules.MetricHit] = []
        self.claims: list[dict[str, Any]] = []
        self.category: str | None = None
        self.unresolved_reason: str | None = None
        self.ambiguous = False


def _unresolved(result: ParseResult, reason: str) -> None:
    result.unresolved_reason = reason
    result.ambiguous = True


def _main_date(result: ParseResult) -> str | None:
    date_text = main_date_of(result.question)
    if date_text is None:
        _unresolved(result, "缺少主日期")
    return date_text


def parse_question(question: str, data: WorkbookData) -> ParseResult:
    result = ParseResult(question, data)
    result.orgs = rules.extract_organizations(question, list(data.organizations.values()))
    result.metrics = rules.extract_metrics(question)
    text = question
    if "逐季" in text or "季度变化趋势" in text or re.search(r"从\s*20\d{2}\s*[Qq]\s*[1-4]", text) or re.search(r"从\s*20\d{2}\s*年\s*第?[一二三四1-4]\s*季度", text):
        _template_trend_quarters(result)
    elif "从规模" in text and "三个维度" in text:
        if "维度与指标映射" in text:
            _template_dimension_3_clarified(result)
        else:
            result.category = "DIMENSION_3"
            _unresolved(result, "各维度对应指标无法从题意唯一确定")
    elif "待评价指标集合" in text and ("表现较好" in text or "表现较差" in text):
        _template_performance_explicit(result)
    elif ("环比" in text and "同比" in text) or ("较上月" in text and "较去年同期" in text):
        _template_mom_yoy(result)
    elif "盈利能力" in text:
        _template_profit_eval(result)
    elif "多少天" in text:
        _template_days_above_avg(result)
    elif "日均值" in text or "最高日" in text:
        _template_daily_mean_extreme(result)
    elif "日均" in text:
        _template_daily_mean(result)
    elif "均值排名" in text or "均值前" in text or re.search(r"全年[^，。]*均值[^，。]*前3", text):
        _template_annual_mean_top_bottom(result)
    elif "单日最高" in text or "单日最低" in text:
        _template_annual_day_extreme(result)
    elif "增幅排名前三" in text:
        _template_top3_pct_change(result)
    elif "下降幅度最大" in text:
        _template_top3_decline(result)
    elif "排名分别变化" in text:
        _template_rank_change(result)
    elif "同时满足" in text:
        _template_simultaneous(result)
    elif "风险指标" in text:
        _template_risk_list(result)
    elif "哪些排名较好" in text or "排名中，哪些排名较好" in text or ("排名较好" in text and "排名较差" in text):
        _template_performance_explicit(result)
    elif "表现较好" in text or "表现较差" in text:
        result.category = "PERFORMANCE"
        _unresolved(result, "指标全集无法从题意唯一确定（表现较好/较差需在确定的指标集合上排名）")
    elif "主要经营指标" in text:
        result.category = "MAIN_INDICATORS"
        _unresolved(result, "主要经营指标的指标全集无法从题意唯一确定")
    elif "逾期贷款率" in text and "不良贷款率" in text and "高多少" in text:
        _template_extreme_diff(result)
    elif "排第几" in text and "是多少" in text:
        _template_value_plus_rank(result)
    elif "排第几" in text:
        _template_rank_org(result)
    elif re.search(r"排名(前三|前3|后三|后3|最后)", text) or re.search(r"(前三|前3|后三|后3)名", text):
        _template_rank_top_n(result)
    elif re.search(r"排(第一|最后一名|第1名)", text) or re.search(r"(最高|最低|最少|最多)的是哪家", text) or re.search(r"哪家农商行[^？]*?(最高|最低|最少|最多|排第一|最后一名)", text):
        _template_rank_extreme(result)
    elif ("全省均值" in text or "省均值" in text or "全省平均" in text) and "多少家" not in text and "几家" not in text:
        if re.search(r"存款|贷款|不良率|净利润", text) and re.search(r"四项|逐一", text):
            _template_province_multi(result)
        else:
            _template_province_single(result)
    elif re.search(r"有多少家|有几家|超过全省均值的有几家|低于全省均值的有几家", text):
        _template_threshold_count(result)
    elif re.search(r"监管要求|最低要求|有没有超过|满足.*要求", text):
        _template_threshold_reg(result)
    elif "不良" in text and "逾期" in text and ("合计" in text or "占贷款比" in text):
        _template_ratio_sum(result)
    elif re.search(r"(分别)?占比|占各项贷款|占营业收入|占营业收入的比重|占贷款总额", text) or "存贷比" in text or "净利润率" in text or "人均利润" in text or "网点平均存款" in text:
        _template_ratio(result)
    elif "合计" in text or "加起来" in text or "分别是多少" in text:
        _template_sum_or_multi(result)
    elif "是不是等于" in text or "差额" in text:
        _template_identity(result)
    elif ("比" in text and re.search(r"多多少|少多少|差多少|更高|更低|最多|最好|最高|最差|高多少", text)) or "差多少" in text or re.search(r"谁[^？]*?(更多|最多|最好|最高|最低|更低|更高|最差)", text) or "多多少" in text:
        _template_compare(result)
    elif "个百分点" in text or re.search(r"变动了几个百分点", text):
        _template_change_point(result)
    elif re.search(r"变化了百分之多少|增幅是多少|增幅", text):
        _template_change_pct(result)
    elif re.search(r"变化了多少|变动了多少|变化情况", text):
        _template_change_delta(result)
    elif "变动方向" in text:
        _template_direction_multi(result)
    elif "排名" in text and re.search(r"在.*方面", text):
        _template_rank_3dim(result)
    else:
        _template_point(result)
    return result


def _value_claim(result: ParseResult, key: str, label: str, metric: str, org_code: str, date_text: str, must_appear: bool = True) -> None:
    data = result.data
    value = data.value(metric, org_code, date_text)
    evidence = _evidence_factory("VALUE_LOOKUP", "直接取事实值 fact[date][org][metric]", 2, data, metric=metric, org=org_code, date=date_text)
    result.claims.append(
        make_claim(
            key, label, kind="VALUE", value=value, unit=data.metrics[metric].unit, metric=metric,
            org=org_code, date=date_text, must_appear=must_appear, evidence=evidence,
        )
    )


def _derived_unit_scale(denominator: str) -> tuple[str, float]:
    if denominator == "ZB018":
        return "万元/人", 1.0
    if denominator == "ZB019":
        return "万元/网点", 10000.0
    return "%", 1.0


def _derived_claim(result: ParseResult, hit: rules.MetricHit, org_code: str, date_text: str) -> None:
    data = result.data
    numerator, denominator = hit.derived  # type: ignore[misc]
    unit, scale = _derived_unit_scale(denominator)
    nv = data.value(numerator, org_code, date_text)
    dv = data.value(denominator, org_code, date_text)
    value = nv * scale / dv if (nv is not None and dv) else None
    if value is not None and unit == "%":
        value = value * 100.0
    evidence = _evidence_factory(
        "RATIO", f"{numerator}{'*%d' % scale if scale != 1.0 else ''}/{denominator}" + ("×100%" if unit == "%" else ""), 2, data,
        numerator=numerator, denominator=denominator, org=org_code, date=date_text,
        scale=scale, percent=unit == "%",
    )
    result.claims.append(
        make_claim(
            hit.matched_text, f"{data.organizations[org_code].name}{date_text}的{hit.matched_text}",
            kind="RATIO_PERCENT", value=value, unit=unit, metric=numerator, org=org_code, date=date_text,
            role="ratio" if unit == "%" else None, note=None if value is not None else "分子或分母缺失",
            evidence=evidence,
        )
    )


# ---- 单点 ----

def _template_point(result: ParseResult) -> None:
    result.category = "POINT"
    data = result.data
    date_text = _main_date(result)
    if date_text is None:
        return
    if not result.orgs:
        _unresolved(result, "机构解析失败")
        return
    org = result.orgs[0]
    hits = [hit for hit in result.metrics if hit.code is not None or hit.derived]
    if not hits:
        _unresolved(result, "指标解析失败")
        return
    for hit in hits:
        if hit.derived is not None:
            _derived_claim(result, hit, org.code, date_text)
        else:
            _value_claim(result, hit.matched_text, f"{org.name}{date_text}的{hit.matched_text}", hit.code, org.code, date_text)  # type: ignore[arg-type]


# ---- 占比/比率 ----

def _template_ratio(result: ParseResult) -> None:
    data = result.data
    date_text = _main_date(result)
    if date_text is None:
        return
    if not result.orgs:
        _unresolved(result, "缺少机构")
        return
    org = result.orgs[0]
    text = result.question

    if "对公" in text and "个人" in text and "分别" in text:
        if "存款" in text:
            dual = ("ZB003", "ZB004", "ZB001", "对公存款占比", "个人存款占比")
        elif "贷款" in text:
            dual = ("ZB005", "ZB006", "ZB002", "对公贷款占比", "个人贷款占比")
        else:
            dual = None
        if dual is not None:
            result.category = "DUAL_RATIO"
            num1, num2, den, label1, label2 = dual
            dv = data.value(den, org.code, date_text)
            if not dv:
                _unresolved(result, "双占比分母缺失")
                return
            v1 = data.value(num1, org.code, date_text)
            v2 = data.value(num2, org.code, date_text)
            result.claims.append(make_claim("ratio1", f"{org.name}{date_text}的{label1}", kind="RATIO_PERCENT", value=v1 * 100.0 / dv if v1 is not None else None, unit="%", metric=num1, org=org.code, date=date_text, role="ratio", evidence=_evidence_factory("RATIO", f"{num1}/{den}×100%", 2, data, numerator=num1, denominator=den, org=org.code, date=date_text, scale=1.0, percent=True)))
            result.claims.append(make_claim("ratio2", f"{org.name}{date_text}的{label2}", kind="RATIO_PERCENT", value=v2 * 100.0 / dv if v2 is not None else None, unit="%", metric=num2, org=org.code, date=date_text, role="ratio", evidence=_evidence_factory("RATIO", f"{num2}/{den}×100%", 2, data, numerator=num2, denominator=den, org=org.code, date=date_text, scale=1.0, percent=True)))
            return

    ratio_hits = [hit for hit in result.metrics if hit.derived is not None]
    base_hits = [hit for hit in result.metrics if hit.code is not None]
    if ratio_hits:
        result.category = "RATIO_PERCENT"
        for hit in ratio_hits:
            _derived_claim(result, hit, org.code, date_text)
        return
    if len(base_hits) >= 2 and ("占" in text or "比重" in text):
        result.category = "RATIO_PERCENT"
        numerator = base_hits[0]
        denominator = base_hits[1]
        nv = data.value(numerator.code, org.code, date_text)  # type: ignore[union-attr]
        dv = data.value(denominator.code, org.code, date_text)  # type: ignore[union-attr]
        result.claims.append(make_claim("ratio", f"{org.name}{date_text}的{numerator.matched_text}占{denominator.matched_text}比例", kind="RATIO_PERCENT", value=nv * 100.0 / dv if (nv is not None and dv) else None, unit="%", metric=numerator.code, org=org.code, date=date_text, role="ratio", evidence=_evidence_factory("RATIO", f"{numerator.code}/{denominator.code}×100%", 2, data, numerator=numerator.code, denominator=denominator.code, org=org.code, date=date_text, scale=1.0, percent=True)))  # type: ignore[index]
        return
    _unresolved(result, "占比/比率解析失败")


# ---- 求和与多指标 ----

def _template_sum_or_multi(result: ParseResult) -> None:
    data = result.data
    date_text = _main_date(result)
    if date_text is None:
        return
    text = result.question
    hits = [hit for hit in result.metrics if hit.code is not None]
    orgs = result.orgs
    if not hits:
        _unresolved(result, "指标解析失败")
        return

    if "加起来" in text or ("两家" in text and "总额" in text):
        result.category = "SUM_ORGS"
        metric = hits[0].code  # type: ignore[assignment]
        total = 0.0
        org_codes = [org.code for org in orgs]
        for org in orgs:
            v = data.value(metric, org.code, date_text)
            result.claims.append(make_claim(f"org_{org.code}", f"{org.name}{date_text}的{data.metrics[metric].name}", kind="VALUE", value=v, unit=data.metrics[metric].unit, metric=metric, org=org.code, date=date_text, must_appear=False, evidence=_evidence_factory("VALUE_LOOKUP", "直接取事实值 fact[date][org][metric]", 2, data, metric=metric, org=org.code, date=date_text)))
            total += v or 0.0
        result.claims.append(make_claim("sum", f"{'、'.join(org.name for org in orgs)}{date_text}的{data.metrics[metric].name}合计", kind="SUM", value=total, unit=data.metrics[metric].unit, metric=metric, role="sum", evidence=_evidence_factory("SUM", "Σ 各机构事实值", 2, data, metric=metric, date=date_text, scope={"orgs": org_codes})))
        return

    if "合计" in text and len(hits) >= 2:
        result.category = "SUM_METRICS"
        total = 0.0
        sum_keys: list[str] = []
        for hit in hits[:2]:
            v = data.value(hit.code, orgs[0].code, date_text)  # type: ignore[union-attr]
            result.claims.append(make_claim(hit.matched_text, f"{orgs[0].name}{date_text}的{hit.matched_text}", kind="VALUE", value=v, unit=data.metrics[hit.code].unit, metric=hit.code, org=orgs[0].code, date=date_text, evidence=_evidence_factory("VALUE_LOOKUP", "直接取事实值 fact[date][org][metric]", 2, data, metric=hit.code, org=orgs[0].code, date=date_text)))  # type: ignore[index]
            total += v or 0.0
            sum_keys.append(hit.matched_text)
        result.claims.append(make_claim("sum", f"{orgs[0].name}{date_text}的{'与'.join(h.matched_text for h in hits[:2])}合计", kind="SUM", value=total, unit=data.metrics[hits[0].code].unit, metric=hits[0].code, role="sum", evidence=_evidence_factory("SUM_DEPENDS", "Σ 依赖声明值", 2, data, dependsOn=sum_keys)))  # type: ignore[index]
        return

    if len(hits) >= 2:
        result.category = "POINT_MULTI"
        for hit in hits:
            _value_claim(result, hit.matched_text, f"{orgs[0].name}{date_text}的{hit.matched_text}", hit.code, orgs[0].code, date_text)  # type: ignore[arg-type]
        return
    _unresolved(result, "求和/多指标解析失败")


# ---- 恒等式 ----

def _template_identity(result: ParseResult) -> None:
    data = result.data
    date_text = _main_date(result)
    if date_text is None:
        return
    if not result.orgs:
        _unresolved(result, "缺少机构")
        return
    org = result.orgs[0]
    text = result.question
    result.category = "IDENTITY"
    if "对公" in text and "个人" in text and "各项存款" in text:
        v_corp = data.value("ZB003", org.code, date_text)
        v_pers = data.value("ZB004", org.code, date_text)
        v_total = data.value("ZB001", org.code, date_text)
        result.claims.append(make_claim("corp", f"{org.name}{date_text}的对公存款余额", kind="VALUE", value=v_corp, unit="亿元", metric="ZB003", org=org.code, date=date_text, evidence=_evidence_factory("VALUE_LOOKUP", "直接取事实值 fact[date][org][metric]", 2, data, metric="ZB003", org=org.code, date=date_text)))
        result.claims.append(make_claim("pers", f"{org.name}{date_text}的个人存款余额", kind="VALUE", value=v_pers, unit="亿元", metric="ZB004", org=org.code, date=date_text, evidence=_evidence_factory("VALUE_LOOKUP", "直接取事实值 fact[date][org][metric]", 2, data, metric="ZB004", org=org.code, date=date_text)))
        result.claims.append(make_claim("sum", "对公+个人合计", kind="SUM", value=(v_corp + v_pers) if (v_corp is not None and v_pers is not None) else None, unit="亿元", role="sum", evidence=_evidence_factory("SUM_DEPENDS", "corp+pers", 2, data, dependsOn=["corp", "pers"])))
        result.claims.append(make_claim("total", f"{org.name}{date_text}的各项存款余额", kind="VALUE", value=v_total, unit="亿元", metric="ZB001", org=org.code, date=date_text, evidence=_evidence_factory("VALUE_LOOKUP", "直接取事实值 fact[date][org][metric]", 2, data, metric="ZB001", org=org.code, date=date_text)))
        result.claims.append(make_claim("diff", "差额", kind="DIFF", value=(v_corp + v_pers - v_total) if (v_corp is not None and v_pers is not None and v_total is not None) else None, unit="亿元", direction="flat", role="diff", evidence=_evidence_factory("DELTA_DEPENDS", "sum-total（恒等式）", 2, data, dependsOn=["sum", "total"])))
    else:
        _unresolved(result, "恒等式结构不匹配")


# ---- 比较 ----

def _template_compare(result: ParseResult) -> None:
    data = result.data
    date_text = _main_date(result)
    if date_text is None:
        return
    text = result.question
    orgs = result.orgs
    hits = [hit for hit in result.metrics if hit.code is not None]
    if not orgs or len(orgs) < 2 or not hits:
        _unresolved(result, "比较题机构或指标不足")
        return
    metric = hits[0].code  # type: ignore[assignment]
    org_values = {org.code: data.value(metric, org.code, date_text) for org in orgs}
    if any(value is None for value in org_values.values()):
        _unresolved(result, "比较题事实缺失")
        return
    ascending = data.metrics[metric].rank_ascending

    if "多多少" in text or "少多少" in text or "差多少" in text or "高多少" in text:
        result.category = "COMPARE_DIFF"
        anchor = orgs[0]
        other = orgs[1]
        diff = org_values[anchor.code] - org_values[other.code]
        if "差多少" in text:
            diff = abs(diff)
            direction = None
        else:
            direction = "up" if diff > 0.005 else "down" if diff < -0.005 else "flat"
        result.claims.append(make_claim("anchor", f"{anchor.name}{date_text}的{data.metrics[metric].name}", kind="VALUE", value=org_values[anchor.code], unit=data.metrics[metric].unit, metric=metric, org=anchor.code, date=date_text, evidence=_evidence_factory("VALUE_LOOKUP", "直接取事实值 fact[date][org][metric]", 2, data, metric=metric, org=anchor.code, date=date_text)))
        result.claims.append(make_claim("other", f"{other.name}{date_text}的{data.metrics[metric].name}", kind="VALUE", value=org_values[other.code], unit=data.metrics[metric].unit, metric=metric, org=other.code, date=date_text, must_appear=False, evidence=_evidence_factory("VALUE_LOOKUP", "直接取事实值 fact[date][org][metric]", 2, data, metric=metric, org=other.code, date=date_text)))
        result.claims.append(make_claim("diff", f"{anchor.name}比{other.name}的差额", kind="DELTA_SIGNED", value=diff, unit=data.metrics[metric].unit, direction=direction, evidence=_evidence_factory("DELTA_DEPENDS", "anchor-other" + ("（取绝对值）" if "差多少" in text else ""), 2, data, dependsOn=["anchor", "other"], abs="差多少" in text)))
        return

    result.category = "COMPARE_WINNER"
    if ascending:
        winner = min(orgs, key=lambda org: org_values[org.code])
    else:
        winner = max(orgs, key=lambda org: org_values[org.code])
    scope_orgs = [org.code for org in orgs]
    result.claims.append(make_claim("winner", f"更优的是{winner.name}", kind="WINNER", value=org_values[winner.code], unit=data.metrics[metric].unit, metric=metric, org=winner.code, date=date_text, evidence=_evidence_factory("COMPARE_WINNER", "按排名方向取最值机构", 2, data, metric=metric, date=date_text, scope={"orgs": scope_orgs}, ascending=ascending)))
    if len(orgs) == 2:
        other = [org for org in orgs if org.code != winner.code][0]
        result.claims.append(make_claim("other", f"{other.name}{date_text}的{data.metrics[metric].name}", kind="VALUE", value=org_values[other.code], unit=data.metrics[metric].unit, metric=metric, org=other.code, date=date_text, must_appear=False, evidence=_evidence_factory("VALUE_LOOKUP", "直接取事实值 fact[date][org][metric]", 2, data, metric=metric, org=other.code, date=date_text)))
        diff = abs(org_values[winner.code] - org_values[other.code])
        result.claims.append(make_claim("diff", "两者的差值", kind="DIFF", value=diff, unit=data.metrics[metric].unit, must_appear=False, evidence=_evidence_factory("DELTA_DEPENDS", "|winner-other|", 2, data, dependsOn=["winner", "other"], abs=True)))


# ---- 全省均值 ----

def _province_diff_claims(result: ParseResult, metric: str, org_code: str, date_text: str, label: str) -> None:
    data = result.data
    value = data.value(metric, org_code, date_text)
    mean = data.province_mean(metric, date_text)
    diff = value - mean if (value is not None and mean is not None) else None
    scope_orgs = sorted(data.organizations)
    result.claims.append(make_claim("value", label, kind="VALUE", value=value, unit=data.metrics[metric].unit, metric=metric, org=org_code, date=date_text, evidence=_evidence_factory("VALUE_LOOKUP", "直接取事实值 fact[date][org][metric]", 2, data, metric=metric, org=org_code, date=date_text)))
    result.claims.append(make_claim("mean", f"{date_text}全省均值", kind="PROVINCE_MEAN", value=mean, unit=data.metrics[metric].unit, metric=metric, role="mean", must_appear=False, evidence=_evidence_factory("PROVINCE_MEAN", "Σ 全省机构事实值 / 机构数", 2, data, metric=metric, date=date_text, scope={"orgs": scope_orgs})))
    result.claims.append(make_claim("diff", f"{label}与全省均值之差", kind="PROVINCE_DIFF", value=diff, unit="个百分点" if data.metrics[metric].is_rate else data.metrics[metric].unit, direction=("up" if (diff or 0) > 0.005 else "down" if (diff or 0) < -0.005 else "flat"), evidence=_evidence_factory("PROVINCE_DIFF", "value-mean", 2, data, dependsOn=["value", "mean"])))


def _template_province_single(result: ParseResult) -> None:
    data = result.data
    date_text = _main_date(result)
    if date_text is None:
        return
    if not result.orgs:
        _unresolved(result, "缺少机构")
        return
    hits = [hit for hit in result.metrics if hit.code is not None]
    if not hits:
        _unresolved(result, "指标解析失败")
        return
    result.category = "PROVINCE_SINGLE"
    metric = hits[0].code  # type: ignore[assignment]
    org = result.orgs[0]
    _province_diff_claims(result, metric, org.code, date_text, f"{org.name}{date_text}的{data.metrics[metric].name}")


def _template_province_multi(result: ParseResult) -> None:
    data = result.data
    date_text = _main_date(result)
    if date_text is None:
        return
    if not result.orgs:
        _unresolved(result, "缺少机构")
        return
    org = result.orgs[0]
    text = result.question
    result.category = "PROVINCE_MULTI"
    mapping = {"存款": "ZB001", "贷款": "ZB002", "不良率": "ZB013", "净利润": "ZB011"}
    selected = [code for keyword, code in mapping.items() if keyword in text]
    if not selected:
        _unresolved(result, "四项指标解析失败")
        return
    for metric in selected:
        _province_diff_claims(result, metric, org.code, date_text, f"{org.name}{date_text}的{data.metrics[metric].name}")


# ---- 阈值家数 / 监管阈值 ----

def _template_threshold_count(result: ParseResult) -> None:
    data = result.data
    date_text = _main_date(result)
    if date_text is None:
        return
    hits = [hit for hit in result.metrics if hit.code is not None]
    if not hits:
        _unresolved(result, "指标解析失败")
        return
    metric = hits[0].code  # type: ignore[assignment]
    text = result.question
    result.category = "THRESHOLD_COUNT"
    mean = data.province_mean(metric, date_text)
    if mean is None:
        _unresolved(result, "全省均值缺失")
        return
    values = data.org_values(metric, date_text)
    if "超过" in text or "高于" in text:
        count = sum(1 for value in values.values() if value > mean)
    elif "低于" in text:
        count = sum(1 for value in values.values() if value < mean)
    else:
        _unresolved(result, "阈值方向不明确")
        return
    result.claims.append(make_claim("mean", f"{date_text}全省均值", kind="PROVINCE_MEAN", value=mean, unit=data.metrics[metric].unit, metric=metric, role="mean", must_appear=False, evidence=_evidence_factory("PROVINCE_MEAN", "Σ 全省机构事实值 / 机构数", 2, data, metric=metric, date=date_text, scope={"orgs": sorted(data.organizations)})))
    result.claims.append(make_claim("count", "满足条件的家数", kind="COUNT", value=float(count), unit="家", evidence=_evidence_factory("COUNT_ABOVE_MEAN", "count(v 与全省均值比较)", 2, data, metric=metric, date=date_text, scope={"orgs": sorted(data.organizations)}, direction="above" if "超过" in text or "高于" in text else "below")))


def _template_threshold_reg(result: ParseResult) -> None:
    data = result.data
    date_text = _main_date(result)
    if date_text is None:
        return
    if not result.orgs:
        _unresolved(result, "缺少机构")
        return
    hits = [hit for hit in result.metrics if hit.code is not None]
    if not hits:
        _unresolved(result, "指标解析失败")
        return
    metric = hits[0].code  # type: ignore[assignment]
    org = result.orgs[0]
    value = data.value(metric, org.code, date_text)
    text = result.question
    thresholds = [num for num in rules.question_numbers(text) if num != 13.0]
    if not thresholds:
        _unresolved(result, "阈值解析失败")
        return
    threshold = thresholds[0]
    result.category = "THRESHOLD_REG"
    result.claims.append(make_claim("value", f"{org.name}{date_text}的{data.metrics[metric].name}", kind="VALUE", value=value, unit=data.metrics[metric].unit, metric=metric, org=org.code, date=date_text, extras=[{"kind": "THRESHOLD", "value": threshold}], evidence=_evidence_factory("VALUE_LOOKUP", "直接取事实值 fact[date][org][metric]", 2, data, metric=metric, org=org.code, date=date_text)))
    meets = value is not None and value > threshold
    result.claims.append(make_claim("meets", "是否达标", kind="MEETS", value=1.0 if meets else 0.0, note=str(meets), evidence=_evidence_factory("MEETS", "value > threshold", 2, data, dependsOn=["value"], threshold=threshold)))


# ---- 排名 ----

def _template_rank_org(result: ParseResult) -> None:
    data = result.data
    date_text = _main_date(result)
    if date_text is None:
        return
    if not result.orgs:
        _unresolved(result, "缺少机构")
        return
    hits = [hit for hit in result.metrics if hit.code is not None]
    if not hits:
        _unresolved(result, "指标解析失败")
        return
    metric = hits[0].code  # type: ignore[assignment]
    org = result.orgs[0]
    values = data.org_values(metric, date_text)
    if len(values) != len(data.organizations):
        _unresolved(result, "排名事实不完整")
        return
    positions = rules.rank_positions([(code, data.organizations[code].name, value) for code, value in values.items()], metric)
    target = next((item for item in positions if item[0] == org.code), None)
    if target is None:
        _unresolved(result, "目标机构不在排名中")
        return
    result.category = "RANK_ORG"
    result.claims.append(make_claim("rank", f"{org.name}的{data.metrics[metric].name}在{date_text}的排名", kind="RANK", value=float(target[3]), unit="名", metric=metric, org=org.code, date=date_text, role="rank", extras=[{"kind": "TOTAL", "value": float(len(positions))}], evidence=_evidence_factory("RANK_POSITION", "rank(v) 按排名方向，并列同名次", 2, data, metric=metric, org=org.code, date=date_text, scope={"orgs": sorted(data.organizations)})))


def _top_entries(result: ParseResult, metric: str, date_text: str | None, positions: list[tuple[str, str, float, int]], side: str) -> None:
    data = result.data
    if side == "back":
        ordered = sorted(positions, key=lambda item: item[3], reverse=True)[:3]
    else:
        ordered = sorted(positions, key=lambda item: item[3])[:3]
    for index, (org_code, org_name, value, position) in enumerate(ordered, start=1):
        entry = make_claim(f"entry{index}", f"第{position}名：{org_name}", kind="TOP_ENTRY", value=value, unit=data.metrics[metric].unit, metric=metric, org=org_code, date=date_text, evidence=_evidence_factory("TOP_N", "rank(v) 后取前3，并列同名次", 2, data, metric=metric, date=date_text, scope={"orgs": sorted(data.organizations)}, side=side, position=position))
        entry["position"] = position
        result.claims.append(entry)


def _template_rank_top_n(result: ParseResult) -> None:
    data = result.data
    date_text = _main_date(result)
    if date_text is None:
        return
    hits = [hit for hit in result.metrics if hit.code is not None]
    if not hits:
        _unresolved(result, "指标解析失败")
        return
    metric = hits[0].code  # type: ignore[assignment]
    text = result.question
    values = data.org_values(metric, date_text)
    if len(values) != len(data.organizations):
        _unresolved(result, "排名事实不完整")
        return
    positions = rules.rank_positions([(code, data.organizations[code].name, value) for code, value in values.items()], metric)
    result.category = "RANK_TOP_N"
    side = "back" if ("排名最后" in text or "后3" in text or "后三" in text) else "front"
    _top_entries(result, metric, date_text, positions, side)


def _template_rank_extreme(result: ParseResult) -> None:
    data = result.data
    date_text = _main_date(result)
    if date_text is None:
        return
    hits = [hit for hit in result.metrics if hit.code is not None]
    if not hits:
        _unresolved(result, "指标解析失败")
        return
    metric = hits[0].code  # type: ignore[assignment]
    text = result.question
    values = data.org_values(metric, date_text)
    if len(values) != len(data.organizations):
        _unresolved(result, "排名事实不完整")
        return
    positions = rules.rank_positions([(code, data.organizations[code].name, value) for code, value in values.items()], metric)
    result.category = "RANK_EXTREME"
    if "最后一名" in text:
        target = sorted(positions, key=lambda item: item[3], reverse=True)[0]
        side_label = "最后一名"
    elif "最低" in text or "最少" in text:
        target = min(positions, key=lambda item: item[2])
        side_label = "最低"
    elif "最高" in text or "排第一" in text:
        target = max(positions, key=lambda item: item[2])
        side_label = "第一" if "排第一" in text else "最高"
    else:
        _unresolved(result, "极值方向不明确")
        return
    result.claims.append(make_claim("winner", f"{side_label}：{target[1]}", kind="WINNER", value=target[2], unit=data.metrics[metric].unit, metric=metric, org=target[0], date=date_text, evidence=_evidence_factory("RANK_EXTREME", "按排名方向取极值机构", 2, data, metric=metric, date=date_text, scope={"orgs": sorted(data.organizations)}, extreme="min" if ("最低" in text or "最少" in text) else "max", mode="rank" if "最后一名" in text else "value")))
    result.claims.append(make_claim("position", "名次", kind="RANK", value=float(target[3]), unit="名", role="rank", extras=[{"kind": "TOTAL", "value": float(len(positions))}], must_appear=False, evidence=_evidence_factory("RANK_POSITION", "rank(v) 按排名方向，并列同名次", 2, data, metric=metric, org=target[0], date=date_text, scope={"orgs": sorted(data.organizations)})))


def _template_value_plus_rank(result: ParseResult) -> None:
    data = result.data
    date_text = _main_date(result)
    if date_text is None:
        return
    if not result.orgs:
        _unresolved(result, "缺少机构")
        return
    hits = [hit for hit in result.metrics if hit.code is not None]
    if not hits:
        _unresolved(result, "指标解析失败")
        return
    metric = hits[0].code  # type: ignore[assignment]
    org = result.orgs[0]
    value = data.value(metric, org.code, date_text)
    values = data.org_values(metric, date_text)
    if len(values) != len(data.organizations):
        _unresolved(result, "排名事实不完整")
        return
    positions = rules.rank_positions([(code, data.organizations[code].name, v) for code, v in values.items()], metric)
    target = next((item for item in positions if item[0] == org.code), None)
    result.category = "RANK_VALUE_PLUS"
    result.claims.append(make_claim("value", f"{org.name}{date_text}的{data.metrics[metric].name}", kind="VALUE", value=value, unit=data.metrics[metric].unit, metric=metric, org=org.code, date=date_text, evidence=_evidence_factory("VALUE_LOOKUP", "直接取事实值 fact[date][org][metric]", 2, data, metric=metric, org=org.code, date=date_text)))
    if target is not None:
        result.claims.append(make_claim("rank", "全省排名", kind="RANK", value=float(target[3]), unit="名", metric=metric, org=org.code, date=date_text, role="rank", extras=[{"kind": "TOTAL", "value": float(len(positions))}], evidence=_evidence_factory("RANK_POSITION", "rank(v) 按排名方向，并列同名次", 2, data, metric=metric, org=org.code, date=date_text, scope={"orgs": sorted(data.organizations)})))


def _template_rank_3dim(result: ParseResult) -> None:
    data = result.data
    date_text = _main_date(result)
    if date_text is None:
        return
    if not result.orgs:
        _unresolved(result, "缺少机构")
        return
    org = result.orgs[0]
    text = result.question
    result.category = "RANK_3DIM"
    dims = []
    if "规模" in text and "贷款" in text:
        dims.append(("规模", "ZB002"))
    if "质量" in text and "不良率" in text:
        dims.append(("质量", "ZB013"))
    if "效益" in text and "净利润" in text:
        dims.append(("效益", "ZB011"))
    if not dims:
        _unresolved(result, "三维映射解析失败")
        return
    for label, metric in dims:
        values = data.org_values(metric, date_text)
        if len(values) != len(data.organizations):
            _unresolved(result, "排名事实不完整")
            return
        positions = rules.rank_positions([(code, data.organizations[code].name, v) for code, v in values.items()], metric)
        target = next((item for item in positions if item[0] == org.code), None)
        result.claims.append(make_claim(f"rank_{label}", f"{label}（{data.metrics[metric].name}）排名", kind="RANK", value=float(target[3]) if target else None, unit="名", metric=metric, org=org.code, date=date_text, role="rank", evidence=_evidence_factory("RANK_POSITION", "rank(v) 按排名方向，并列同名次", 2, data, metric=metric, org=org.code, date=date_text, scope={"orgs": sorted(data.organizations)})))


# ---- 变化 ----

def _delta_claims(result: ParseResult, metric: str, org_code: str, main_date: str, baseline: str, comparison_type: str, wants_pct: bool) -> None:
    data = result.data
    current = data.value(metric, org_code, main_date)
    base = data.value(metric, org_code, baseline)
    if current is None or base is None:
        _unresolved(result, f"缺失基期或当前值（{comparison_type}）")
        return
    if wants_pct and data.metrics[metric].is_rate:
        _unresolved(result, "比率类指标不做增幅计算（衍生维度说明）")
        return
    if wants_pct:
        pct = (current - base) * 100.0 / base
        result.claims.append(make_claim("pct", f"{main_date}较{baseline}增幅", kind="PCT_CHANGE", value=pct, unit="%", metric=metric, org=org_code, date=main_date, baseline=baseline, comparison_type=comparison_type, direction=("up" if pct > 0.005 else "down" if pct < -0.005 else "flat"), evidence=_evidence_factory("PCT_CHANGE", "(current-base)/base×100%", 2, data, metric=metric, org=org_code, date=main_date, baseline=baseline, comparisonType=comparison_type)))
    else:
        if data.metrics[metric].is_rate:
            delta = current - base
            result.claims.append(make_claim("delta", f"{main_date}较{baseline}变动（百分点）", kind="DELTA", value=delta, unit="个百分点", metric=metric, org=org_code, date=main_date, baseline=baseline, comparison_type=comparison_type, direction=("up" if delta > 0.005 else "down" if delta < -0.005 else "flat"), evidence=_evidence_factory("DELTA", "current-base（百分点）", 2, data, metric=metric, org=org_code, date=main_date, baseline=baseline, comparisonType=comparison_type)))
            result.claims.append(make_claim("current", "当前值", kind="VALUE", value=current, unit=data.metrics[metric].unit, metric=metric, org=org_code, date=main_date, must_appear=False, evidence=_evidence_factory("VALUE_LOOKUP", "直接取事实值 fact[date][org][metric]", 2, data, metric=metric, org=org_code, date=main_date)))
            result.claims.append(make_claim("base", "基期值", kind="BASE", value=base, unit=data.metrics[metric].unit, metric=metric, org=org_code, date=baseline, must_appear=False, evidence=_evidence_factory("VALUE_LOOKUP", "直接取事实值 fact[date][org][metric]", 2, data, metric=metric, org=org_code, date=baseline)))
        else:
            delta = current - base
            result.claims.append(make_claim("delta", f"{main_date}较{baseline}增量", kind="DELTA", value=delta, unit=data.metrics[metric].unit, metric=metric, org=org_code, date=main_date, baseline=baseline, comparison_type=comparison_type, direction=("up" if delta > 0.005 else "down" if delta < -0.005 else "flat"), evidence=_evidence_factory("DELTA", "current-base", 2, data, metric=metric, org=org_code, date=main_date, baseline=baseline, comparisonType=comparison_type)))


def _resolve_baseline(result: ParseResult) -> tuple[str, str] | None:
    text = result.question
    main = main_date_of(text)
    if main is None:
        return None
    if "较年初" in text or "从年初" in text or "比年初" in text or "和2024年末" in text or "从2024年末" in text:
        return "ytd", rules.YEAR_START
    if "同比" in text or "较去年同期" in text or "较同期" in text:
        return "yoy", rules.same_month_last_year(main)
    if "较上季" in text or "上季度末" in text:
        return "qoq", rules.previous_quarter_end(main)
    if "较上月" in text or "比上月" in text or "比上个月底" in text or "上月月末" in text:
        return "mom", rules.previous_month_end(main)
    return None


def _template_change_delta(result: ParseResult) -> None:
    result.category = "CHANGE_DELTA"
    date_text = _main_date(result)
    if date_text is None:
        return
    baseline = _resolve_baseline(result)
    if baseline is None:
        _unresolved(result, "变化基期无法确定")
        return
    hits = [hit for hit in result.metrics if hit.code is not None]
    if not hits:
        _unresolved(result, "指标解析失败")
        return
    if not result.orgs:
        _unresolved(result, "缺少机构")
        return
    _delta_claims(result, hits[0].code, result.orgs[0].code, date_text, baseline[1], baseline[0], wants_pct=False)  # type: ignore[arg-type]


def _template_change_pct(result: ParseResult) -> None:
    result.category = "CHANGE_PCT"
    date_text = _main_date(result)
    if date_text is None:
        return
    baseline = _resolve_baseline(result)
    if baseline is None:
        _unresolved(result, "增幅基期无法确定")
        return
    hits = [hit for hit in result.metrics if hit.code is not None]
    if not hits:
        _unresolved(result, "指标解析失败")
        return
    if not result.orgs:
        _unresolved(result, "缺少机构")
        return
    _delta_claims(result, hits[0].code, result.orgs[0].code, date_text, baseline[1], baseline[0], wants_pct=True)  # type: ignore[arg-type]


def _template_change_point(result: ParseResult) -> None:
    result.category = "CHANGE_POINT"
    date_text = _main_date(result)
    if date_text is None:
        return
    baseline = _resolve_baseline(result)
    if baseline is None:
        _unresolved(result, "百分点变化基期无法确定")
        return
    hits = [hit for hit in result.metrics if hit.code is not None]
    if not hits:
        _unresolved(result, "指标解析失败")
        return
    if not result.orgs:
        _unresolved(result, "缺少机构")
        return
    metric = hits[0].code  # type: ignore[assignment]
    if not data_is_rate(result.data, metric):
        _unresolved(result, "百分点变化仅适用于比率类指标")
        return
    _delta_claims(result, metric, result.orgs[0].code, date_text, baseline[1], baseline[0], wants_pct=False)


def _template_mom_yoy(result: ParseResult) -> None:
    data = result.data
    result.category = "CHANGE_MOM_YOY"
    date_text = _main_date(result)
    if date_text is None:
        return
    hits = [hit for hit in result.metrics if hit.code is not None]
    if not hits:
        _unresolved(result, "指标解析失败")
        return
    if not result.orgs:
        _unresolved(result, "缺少机构")
        return
    metric = hits[0].code  # type: ignore[assignment]
    org = result.orgs[0].code
    text = result.question
    current = data.value(metric, org, date_text)
    result.claims.append(make_claim("current", f"{org}{date_text}当前值", kind="VALUE", value=current, unit=data.metrics[metric].unit, metric=metric, org=org, date=date_text, role="current", evidence=_evidence_factory("VALUE_LOOKUP", "直接取事实值 fact[date][org][metric]", 2, data, metric=metric, org=org, date=date_text)))
    for comparison_type, keyword in (("mom", "环比"), ("yoy", "同比")):
        if keyword in text or (comparison_type == "mom" and ("较上月" in text or "比上月" in text)) or (comparison_type == "yoy" and "较去年同期" in text):
            baseline = rules.resolve_relative_baseline(date_text, comparison_type)
            if baseline is None or not data.exists_date(baseline):
                _unresolved(result, f"缺失{comparison_type}基期")
                return
            base = data.value(metric, org, baseline)
            pct = (current - base) * 100.0 / base if (current is not None and base) else None
            if pct is None:
                _unresolved(result, f"{comparison_type}事实缺失")
                return
            result.claims.append(make_claim(f"pct_{comparison_type}", f"{date_text}较{baseline}{comparison_type}增幅", kind="PCT_CHANGE", value=pct, unit="%", metric=metric, org=org, date=date_text, baseline=baseline, comparison_type=comparison_type, direction=("up" if pct > 0.005 else "down" if pct < -0.005 else "flat"), evidence=_evidence_factory("PCT_CHANGE", "(current-base)/base×100%", 2, data, metric=metric, org=org, date=date_text, baseline=baseline, comparisonType=comparison_type)))
            result.claims.append(make_claim(f"base_{comparison_type}", f"基期值（{baseline}）", kind="BASE", value=base, unit=data.metrics[metric].unit, metric=metric, org=org, date=baseline, comparison_type=comparison_type, must_appear=False, evidence=_evidence_factory("VALUE_LOOKUP", "直接取事实值 fact[date][org][metric]", 2, data, metric=metric, org=org, date=baseline)))


def _template_direction_multi(result: ParseResult) -> None:
    data = result.data
    result.category = "CHANGE_DIRECTION"
    if not result.orgs:
        _unresolved(result, "缺少机构")
        return
    org = result.orgs[0]
    dates = [iso for _pos, iso in extract_all_dates(result.question)]
    if len(dates) < 2:
        _unresolved(result, "起止日期无法确定")
        return
    start, end = dates[0], dates[1]
    text = result.question
    mapping = {"存款": "ZB001", "贷款": "ZB002", "不良率": "ZB013", "净利润": "ZB011"}
    selected = [code for keyword, code in mapping.items() if keyword in text]
    if not selected:
        _unresolved(result, "指标解析失败")
        return
    for metric in selected:
        start_value = data.value(metric, org.code, start)
        end_value = data.value(metric, org.code, end)
        if start_value is None or end_value is None:
            _unresolved(result, "变动方向事实缺失")
            return
        direction = "up" if end_value - start_value > 0.005 else "down" if end_value - start_value < -0.005 else "flat"
        result.claims.append(make_claim(f"base_{metric}", f"{org.name}{start}的{data.metrics[metric].name}", kind="BASE", value=start_value, unit=data.metrics[metric].unit, metric=metric, org=org.code, date=start, evidence=_evidence_factory("VALUE_LOOKUP", "直接取事实值 fact[date][org][metric]", 2, data, metric=metric, org=org.code, date=start)))
        result.claims.append(make_claim(f"current_{metric}", f"{org.name}{end}的{data.metrics[metric].name}", kind="CURRENT", value=end_value, unit=data.metrics[metric].unit, metric=metric, org=org.code, date=end, evidence=_evidence_factory("VALUE_LOOKUP", "直接取事实值 fact[date][org][metric]", 2, data, metric=metric, org=org.code, date=end)))
        result.claims.append(make_claim(f"direction_{metric}", f"{data.metrics[metric].name}变动方向", kind="TREND", value={"up": 1.0, "down": -1.0, "flat": 0.0}[direction], direction=direction, evidence=_evidence_factory("TREND_DELTA", "sign(end-start)", 2, data, dependsOn=[f"base_{metric}", f"current_{metric}"])))


# ---- 逐季趋势 ----

def _template_trend_quarters(result: ParseResult) -> None:
    data = result.data
    result.category = "TREND_QUARTER"
    dates = [iso for _pos, iso in extract_all_dates(result.question)]
    if len(dates) < 2:
        _unresolved(result, "逐季序列无法确定")
        return
    start, end = dates[0], dates[-1]
    series = rules.quarter_series(start, end)
    if not series:
        _unresolved(result, "逐季序列无法确定")
        return
    hits = [hit for hit in result.metrics if hit.code is not None]
    if not hits:
        _unresolved(result, "指标解析失败")
        return
    if not result.orgs:
        _unresolved(result, "缺少机构")
        return
    metric = hits[0].code  # type: ignore[assignment]
    org = result.orgs[0]
    values = data.series(metric, org.code, series)
    if any(value is None for value in values):
        _unresolved(result, "逐季事实缺失")
        return
    for date_text, value in zip(series, values):
        result.claims.append(make_claim(f"q_{date_text}", f"{org.name}{date_text}的{data.metrics[metric].name}", kind="VALUE", value=value, unit=data.metrics[metric].unit, metric=metric, org=org.code, date=date_text, evidence=_evidence_factory("VALUE_LOOKUP", "直接取事实值 fact[date][org][metric]", 2, data, metric=metric, org=org.code, date=date_text)))
    xs = list(range(len(values)))
    mean_x = sum(xs) / len(xs)
    mean_y = sum(values) / len(values)
    slope = sum((x - mean_x) * (values[i] - mean_y) for i, x in enumerate(xs))
    if slope > 1e-9:
        trend = "up"
    elif slope < -1e-9:
        trend = "down"
    else:
        trend = "flat"
    result.claims.append(make_claim("trend", "整体趋势", kind="TREND", value={"up": 1.0, "down": -1.0, "flat": 0.0}[trend], direction=trend, evidence=_evidence_factory("TREND_SLOPE", "sign(最小二乘斜率)", 2, data, dependsOn=[f"q_{date_text}" for date_text in series])))
    if "哪个季度数值最高" in result.question or "哪个季度最高" in result.question:
        max_index = max(range(len(values)), key=lambda i: values[i])
        result.claims.append(make_claim("max", f"数值最高的季度（{series[max_index]}）", kind="EXTREME", value=values[max_index], unit=data.metrics[metric].unit, metric=metric, org=org.code, date=series[max_index], role="extreme", evidence=_evidence_factory("EXTREME_DEPENDS", "max(依赖序列值)", 2, data, dependsOn=[f"q_{date_text}" for date_text in series], extreme="max")))


# ---- 天数 ----

def _template_days_above_avg(result: ParseResult) -> None:
    data = result.data
    result.category = "DAYS_ABOVE_AVG"
    if not result.orgs:
        _unresolved(result, "缺少机构")
        return
    hits = [hit for hit in result.metrics if hit.code is not None]
    if not hits:
        _unresolved(result, "指标解析失败")
        return
    metric = hits[0].code  # type: ignore[assignment]
    org = result.orgs[0]
    year_match = re.search(r"(20\d{2})\s*年\s*全年", result.question)
    if not year_match:
        _unresolved(result, "全年范围无法确定")
        return
    year = int(year_match.group(1))
    stats = data.count_days_above_mean(metric, org.code, f"{year}-01-01", f"{year}-12-31")
    if stats is None:
        _unresolved(result, "天数统计失败")
        return
    result.claims.append(make_claim("count", "高于全省均值的天数", kind="COUNT", value=float(stats["above"]), unit="天", evidence=_evidence_factory("DAYS_ABOVE_MEAN", "count(org值 > 全省均值) 按日", 2, data, metric=metric, org=org.code, dateRange=[f"{year}-01-01", f"{year}-12-31"], scope={"orgs": sorted(data.organizations)}, stat="above")))
    result.claims.append(make_claim("total", "总天数", kind="TOTAL", value=float(stats["total"]), unit="天", role="total", evidence=_evidence_factory("DAYS_ABOVE_MEAN", "范围内有效日期数", 2, data, metric=metric, org=org.code, dateRange=[f"{year}-01-01", f"{year}-12-31"], scope={"orgs": sorted(data.organizations)}, stat="total")))
    result.claims.append(make_claim("pct", "占比", kind="PCT_RATIO", value=stats["pct"], unit="%", rounding=1, role="ratio", evidence=_evidence_factory("DAYS_ABOVE_MEAN", "above/total×100%", 1, data, metric=metric, org=org.code, dateRange=[f"{year}-01-01", f"{year}-12-31"], scope={"orgs": sorted(data.organizations)}, stat="pct")))


# ---- 日均 ----

def _template_daily_mean(result: ParseResult) -> None:
    data = result.data
    result.category = "DAILY_MEAN"
    if not result.orgs:
        _unresolved(result, "缺少机构")
        return
    hits = [hit for hit in result.metrics if hit.code is not None]
    if not hits:
        _unresolved(result, "指标解析失败")
        return
    metric = hits[0].code  # type: ignore[assignment]
    org = result.orgs[0]
    date_range = rules.parse_date_range(result.question)
    if date_range is None:
        _unresolved(result, "日均时间范围无法确定")
        return
    mean = data.range_mean(metric, org.code, date_range[0], date_range[1])
    if mean is None:
        _unresolved(result, "日均事实缺失")
        return
    result.claims.append(make_claim("mean", f"{org.name}{date_range[0]}至{date_range[1]}的{data.metrics[metric].name}日均", kind="MEAN", value=mean, unit=data.metrics[metric].unit, metric=metric, org=org.code, role="mean", evidence=_evidence_factory("RANGE_MEAN", "Σ 范围内日值 / 天数", 2, data, metric=metric, org=org.code, dateRange=[date_range[0], date_range[1]])))


def _template_daily_mean_extreme(result: ParseResult) -> None:
    data = result.data
    result.category = "DAILY_MEAN_EXTREME"
    if not result.orgs:
        _unresolved(result, "缺少机构")
        return
    hits = [hit for hit in result.metrics if hit.code is not None]
    if not hits:
        _unresolved(result, "指标解析失败")
        return
    metric = hits[0].code  # type: ignore[assignment]
    org = result.orgs[0]
    year_match = re.search(r"(20\d{2})\s*年", result.question)
    if not year_match:
        _unresolved(result, "年度范围无法确定")
        return
    year = int(year_match.group(1))
    start, end = f"{year}-01-01", f"{year}-12-31"
    stats = data.range_stats(metric, org.code, start, end)
    if stats is None:
        _unresolved(result, "日均极值事实缺失")
        return
    result.claims.append(make_claim("mean", f"{org.name}{year}年{data.metrics[metric].name}日均值", kind="MEAN", value=stats["mean"], unit=data.metrics[metric].unit, metric=metric, org=org.code, role="mean", evidence=_evidence_factory("RANGE_STATS", "Σ 范围内日值 / 天数", 2, data, metric=metric, org=org.code, dateRange=[start, end])))
    result.claims.append(make_claim("max", "全年最高日", kind="EXTREME", value=stats["max"], unit=data.metrics[metric].unit, metric=metric, org=org.code, role="extreme", evidence=_evidence_factory("EXTREME_RANGE", "max(范围内日值)", 2, data, metric=metric, org=org.code, dateRange=[start, end], extreme="max")))
    result.claims.append(make_claim("min", "全年最低日", kind="EXTREME", value=stats["min"], unit=data.metrics[metric].unit, metric=metric, org=org.code, role="extreme", evidence=_evidence_factory("EXTREME_RANGE", "min(范围内日值)", 2, data, metric=metric, org=org.code, dateRange=[start, end], extreme="min")))


# ---- 年度均值前3后3 / 年度单日极值 ----

def _template_annual_mean_top_bottom(result: ParseResult) -> None:
    data = result.data
    result.category = "RANK_ANNUAL_MEAN"
    hits = [hit for hit in result.metrics if hit.code is not None]
    if not hits:
        _unresolved(result, "指标解析失败")
        return
    metric = hits[0].code  # type: ignore[assignment]
    year_match = re.search(r"(20\d{2})\s*年", result.question)
    if not year_match:
        _unresolved(result, "年度范围无法确定")
        return
    year = int(year_match.group(1))
    start, end = f"{year}-01-01", f"{year}-12-31"
    means: dict[str, float] = {}
    for org_code in data.organizations:
        mean = data.range_mean(metric, org_code, start, end)
        if mean is None:
            _unresolved(result, "年度均值事实缺失")
            return
        means[org_code] = mean
    positions = rules.rank_positions([(code, data.organizations[code].name, value) for code, value in means.items()], metric)
    front = sorted(positions, key=lambda item: item[3])[:3]
    back = sorted(positions, key=lambda item: item[3])[-3:]
    scope_orgs = sorted(data.organizations)
    for index, (org_code, org_name, value, position) in enumerate(front, start=1):
        front_claim = make_claim(f"front{index}", f"前3第{position}名：{org_name}", kind="TOP_ENTRY", value=value, unit=data.metrics[metric].unit, metric=metric, org=org_code, evidence=_evidence_factory("TOP_N_MEAN", "机构范围内均值后 rank 取前3", 2, data, metric=metric, dateRange=[start, end], scope={"orgs": scope_orgs}, side="front", position=position))
        front_claim["position"] = position
        result.claims.append(front_claim)
    for index, (org_code, org_name, value, position) in enumerate(back, start=1):
        back_claim = make_claim(f"back{index}", f"后3第{position}名：{org_name}", kind="TOP_ENTRY", value=value, unit=data.metrics[metric].unit, metric=metric, org=org_code, evidence=_evidence_factory("TOP_N_MEAN", "机构范围内均值后 rank 取后3", 2, data, metric=metric, dateRange=[start, end], scope={"orgs": scope_orgs}, side="back", position=position))
        back_claim["position"] = position
        result.claims.append(back_claim)


def _template_annual_day_extreme(result: ParseResult) -> None:
    data = result.data
    result.category = "RANK_ANNUAL_EXTREME"
    hits = [hit for hit in result.metrics if hit.code is not None]
    if not hits:
        _unresolved(result, "指标解析失败")
        return
    metric = hits[0].code  # type: ignore[assignment]
    year_match = re.search(r"(20\d{2})\s*年", result.question)
    if not year_match:
        _unresolved(result, "年度范围无法确定")
        return
    year = int(year_match.group(1))
    start, end = f"{year}-01-01", f"{year}-12-31"
    maxima: dict[str, float] = {}
    minima: dict[str, float] = {}
    for org_code in data.organizations:
        stats = data.range_stats(metric, org_code, start, end)
        if stats is None:
            _unresolved(result, "年度极值事实缺失")
            return
        maxima[org_code] = stats["max"]
        minima[org_code] = stats["min"]
    max_org = max(maxima, key=maxima.get)
    min_org = min(minima, key=minima.get)
    scope_orgs = sorted(data.organizations)
    result.claims.append(make_claim("max", f"单日最高值：{data.organizations[max_org].name}", kind="EXTREME", value=maxima[max_org], unit=data.metrics[metric].unit, metric=metric, org=max_org, role="extreme", evidence=_evidence_factory("EXTREME_ALL", "max(各机构范围内最大值)", 2, data, metric=metric, dateRange=[start, end], scope={"orgs": scope_orgs}, extreme="max")))
    result.claims.append(make_claim("min", f"单日最低值：{data.organizations[min_org].name}", kind="EXTREME", value=minima[min_org], unit=data.metrics[metric].unit, metric=metric, org=min_org, role="extreme", evidence=_evidence_factory("EXTREME_ALL", "min(各机构范围内最小值)", 2, data, metric=metric, dateRange=[start, end], scope={"orgs": scope_orgs}, extreme="min")))


# ---- 增幅/降幅 Top3 ----

def _template_top3_pct_change(result: ParseResult) -> None:
    data = result.data
    result.category = "RANK_TOP_PCT"
    hits = [hit for hit in result.metrics if hit.code is not None]
    if not hits:
        _unresolved(result, "指标解析失败")
        return
    metric = hits[0].code  # type: ignore[assignment]
    if data.metrics[metric].is_rate:
        _unresolved(result, "比率类指标不做增幅计算（衍生维度说明）")
        return
    end = main_date_of(result.question)
    baseline = rules.YEAR_START
    if end is None or not data.exists_date(end):
        _unresolved(result, "增幅截止日期缺失")
        return
    changes: list[tuple[str, str, float]] = []
    for org_code in data.organizations:
        current = data.value(metric, org_code, end)
        base = data.value(metric, org_code, baseline)
        if current is None or base is None:
            _unresolved(result, "增幅事实缺失")
            return
        changes.append((org_code, data.organizations[org_code].name, (current - base) * 100.0 / base))
    changes.sort(key=lambda item: (-item[2], item[0]))
    for index, (org_code, org_name, pct) in enumerate(changes[:3], start=1):
        entry = make_claim(f"entry{index}", f"增幅第{index}名：{org_name}", kind="TOP_ENTRY", value=pct, unit="%", metric=metric, org=org_code, evidence=_evidence_factory("TOP_N_PCT_CHANGE", "(end-baseline)/baseline×100% 降序取前3", 2, data, metric=metric, date=end, baseline=baseline, scope={"orgs": sorted(data.organizations)}, position=index))
        entry["position"] = index
        result.claims.append(entry)


def _template_top3_decline(result: ParseResult) -> None:
    data = result.data
    result.category = "RANK_TOP_DECLINE"
    hits = [hit for hit in result.metrics if hit.code is not None]
    if not hits:
        _unresolved(result, "指标解析失败")
        return
    metric = hits[0].code  # type: ignore[assignment]
    end = main_date_of(result.question)
    baseline = rules.YEAR_START
    if end is None or not data.exists_date(end):
        _unresolved(result, "降幅截止日期缺失")
        return
    declines: list[tuple[str, str, float]] = []
    for org_code in data.organizations:
        current = data.value(metric, org_code, end)
        base = data.value(metric, org_code, baseline)
        if current is None or base is None:
            _unresolved(result, "降幅事实缺失")
            return
        declines.append((org_code, data.organizations[org_code].name, base - current))
    declines.sort(key=lambda item: (-item[2], item[0]))
    for index, (org_code, org_name, decline) in enumerate(declines[:3], start=1):
        entry = make_claim(f"entry{index}", f"降幅第{index}名：{org_name}", kind="TOP_ENTRY", value=decline, unit="个百分点", metric=metric, org=org_code, direction="down", evidence=_evidence_factory("TOP_N_DECLINE", "baseline-end（百分点）降序取前3", 2, data, metric=metric, date=end, baseline=baseline, scope={"orgs": sorted(data.organizations)}, position=index))
        entry["position"] = index
        result.claims.append(entry)


# ---- 排名变化 ----

def _template_rank_change(result: ParseResult) -> None:
    data = result.data
    result.category = "RANK_CHANGE"
    if not result.orgs:
        _unresolved(result, "缺少机构")
        return
    org = result.orgs[0]
    end = main_date_of(result.question)
    baseline = rules.YEAR_START
    if end is None or not data.exists_date(end):
        _unresolved(result, "排名变化截止日期缺失")
        return
    text = result.question
    mapping = {"存款": "ZB001", "贷款": "ZB002", "不良率": "ZB013", "净利润": "ZB011"}
    selected = [code for keyword, code in mapping.items() if keyword in text]
    if not selected:
        _unresolved(result, "指标解析失败")
        return
    for metric in selected:
        rank_pairs: list[int] = []
        for day in (baseline, end):
            values = data.org_values(metric, day)
            if len(values) != len(data.organizations):
                _unresolved(result, "排名事实不完整")
                return
            positions = rules.rank_positions([(code, data.organizations[code].name, value) for code, value in values.items()], metric)
            target = next((item for item in positions if item[0] == org.code), None)
            if target is None:
                _unresolved(result, "目标机构不在排名中")
                return
            rank_pairs.append(target[3])
        change = abs(rank_pairs[1] - rank_pairs[0])
        scope_orgs = sorted(data.organizations)
        result.claims.append(make_claim(f"rank_base_{metric}", f"{data.metrics[metric].name}在{baseline}的排名", kind="RANK", value=float(rank_pairs[0]), unit="名", metric=metric, org=org.code, date=baseline, role="rank", must_appear=False, evidence=_evidence_factory("RANK_POSITION", "rank(v) 按排名方向，并列同名次", 2, data, metric=metric, org=org.code, date=baseline, scope={"orgs": scope_orgs})))
        result.claims.append(make_claim(f"rank_end_{metric}", f"{data.metrics[metric].name}在{end}的排名", kind="RANK", value=float(rank_pairs[1]), unit="名", metric=metric, org=org.code, date=end, role="rank", must_appear=False, evidence=_evidence_factory("RANK_POSITION", "rank(v) 按排名方向，并列同名次", 2, data, metric=metric, org=org.code, date=end, scope={"orgs": scope_orgs})))
        result.claims.append(make_claim(f"change_{metric}", f"{data.metrics[metric].name}排名变化", kind="DELTA_ABS", value=float(change), unit="名", evidence=_evidence_factory("DELTA_ABS", "|rank_end-rank_base|", 2, data, dependsOn=[f"rank_base_{metric}", f"rank_end_{metric}"])))


# ---- 双条件 / 风险列表 / 逾期-不良 / 不良+逾期 / 盈利能力 ----

def _template_simultaneous(result: ParseResult) -> None:
    data = result.data
    result.category = "SIMULTANEOUS"
    date_text = _main_date(result)
    if date_text is None:
        return
    text = result.question
    if "不良率" in text and "拨备覆盖率" in text:
        metric_a, metric_b = "ZB013", "ZB015"
        mean_a = data.province_mean(metric_a, date_text)
        mean_b = data.province_mean(metric_b, date_text)
        if mean_a is None or mean_b is None:
            _unresolved(result, "全省均值缺失")
            return
        matching: list[str] = []
        for org_code in data.organizations:
            va = data.value(metric_a, org_code, date_text)
            vb = data.value(metric_b, org_code, date_text)
            if va is not None and vb is not None and va < mean_a and vb > mean_b:
                matching.append(org_code)
        for index, org_code in enumerate(matching, start=1):
            va = data.value(metric_a, org_code, date_text)
            vb = data.value(metric_b, org_code, date_text)
            result.claims.append(make_claim(f"org{index}_a", f"{data.organizations[org_code].name}的不良率", kind="VALUE", value=va, unit="%", metric=metric_a, org=org_code, date=date_text, evidence=_evidence_factory("VALUE_LOOKUP", "直接取事实值 fact[date][org][metric]", 2, data, metric=metric_a, org=org_code, date=date_text)))
            result.claims.append(make_claim(f"org{index}_b", f"{data.organizations[org_code].name}的拨备覆盖率", kind="VALUE", value=vb, unit="%", metric=metric_b, org=org_code, date=date_text, evidence=_evidence_factory("VALUE_LOOKUP", "直接取事实值 fact[date][org][metric]", 2, data, metric=metric_b, org=org_code, date=date_text)))
        result.claims.append(make_claim("count", "满足条件的家数", kind="COUNT", value=float(len(matching)), unit="家", evidence=_evidence_factory("COUNT_CONDITIONS", "count(va<均值a 且 vb>均值b)", 2, data, metricA=metric_a, metricB=metric_b, date=date_text, scope={"orgs": sorted(data.organizations)})))
    else:
        _unresolved(result, "双条件指标结构不匹配")


def _template_risk_list(result: ParseResult) -> None:
    data = result.data
    result.category = "RISK_LIST"
    date_text = _main_date(result)
    if date_text is None:
        return
    if not result.orgs:
        _unresolved(result, "缺少机构")
        return
    org = result.orgs[0]
    text = result.question
    mapping = {
        "不良率": "ZB013",
        "不良贷款率": "ZB013",
        "拨备覆盖率": "ZB015",
        "拨备": "ZB015",
        "逾期率": "ZB017",
        "逾期贷款率": "ZB017",
        "资本充足率": "ZB016",
    }
    selected: list[str] = []
    for keyword, code in mapping.items():
        if keyword in text and code not in selected:
            selected.append(code)
    if len(selected) < 2:
        _unresolved(result, "风险指标列表解析失败")
        return
    for metric in selected:
        _value_claim(result, data.metrics[metric].name, f"{org.name}{date_text}的{data.metrics[metric].name}", metric, org.code, date_text)


def _template_extreme_diff(result: ParseResult) -> None:
    data = result.data
    result.category = "EXTREME_DIFF"
    date_text = _main_date(result)
    if date_text is None:
        return
    if not result.orgs:
        _unresolved(result, "缺少机构")
        return
    org = result.orgs[0]
    va = data.value("ZB017", org.code, date_text)
    vb = data.value("ZB013", org.code, date_text)
    diff = va - vb if (va is not None and vb is not None) else None
    result.claims.append(make_claim("a", f"{org.name}{date_text}的逾期贷款率", kind="VALUE", value=va, unit="%", metric="ZB017", org=org.code, date=date_text, evidence=_evidence_factory("VALUE_LOOKUP", "直接取事实值 fact[date][org][metric]", 2, data, metric="ZB017", org=org.code, date=date_text)))
    result.claims.append(make_claim("b", f"{org.name}{date_text}的不良贷款率", kind="VALUE", value=vb, unit="%", metric="ZB013", org=org.code, date=date_text, evidence=_evidence_factory("VALUE_LOOKUP", "直接取事实值 fact[date][org][metric]", 2, data, metric="ZB013", org=org.code, date=date_text)))
    result.claims.append(make_claim("diff", "逾期率-不良率", kind="DELTA", value=diff, unit="个百分点", direction=("up" if (diff or 0) > 0.005 else "down" if (diff or 0) < -0.005 else "flat"), evidence=_evidence_factory("DELTA_DEPENDS", "a-b", 2, data, dependsOn=["a", "b"])))


def _template_ratio_sum(result: ParseResult) -> None:
    data = result.data
    result.category = "RATIO_SUM"
    date_text = _main_date(result)
    if date_text is None:
        return
    if not result.orgs:
        _unresolved(result, "缺少机构")
        return
    org = result.orgs[0]
    va = data.value("ZB013", org.code, date_text)
    vb = data.value("ZB017", org.code, date_text)
    result.claims.append(make_claim("a", f"{org.name}{date_text}的不良率", kind="VALUE", value=va, unit="%", metric="ZB013", org=org.code, date=date_text, evidence=_evidence_factory("VALUE_LOOKUP", "直接取事实值 fact[date][org][metric]", 2, data, metric="ZB013", org=org.code, date=date_text)))
    result.claims.append(make_claim("b", f"{org.name}{date_text}的逾期率", kind="VALUE", value=vb, unit="%", metric="ZB017", org=org.code, date=date_text, evidence=_evidence_factory("VALUE_LOOKUP", "直接取事实值 fact[date][org][metric]", 2, data, metric="ZB017", org=org.code, date=date_text)))
    result.claims.append(make_claim("sum", "合计", kind="SUM", value=(va + vb) if (va is not None and vb is not None) else None, unit="%", role="sum", evidence=_evidence_factory("SUM_DEPENDS", "a+b", 2, data, dependsOn=["a", "b"])))


def _explicit_metric_hits(text: str) -> list[rules.MetricHit]:
    """从问题文本提取明确列出的指标集合（含衍生占位），按出现顺序去重。

    澄清题使用“待评价指标集合：……”段；旧格式题从全文提取明确命名的指标。
    指标只由问题文本决定，绝不回读答案。
    """
    if "待评价指标集合：" in text:
        segment = text.split("待评价指标集合：", 1)[1].split("。", 1)[0]
    else:
        segment = text
    return _dedup_hits([hit for _position, hit in rules.extract_metrics_ordered(segment)])


def _dedup_hits(hits: list[rules.MetricHit]) -> list[rules.MetricHit]:
    """按 (指标编号或衍生占位名) 去重，保持首次出现顺序。"""
    seen: set[str] = set()
    deduped: list[rules.MetricHit] = []
    for hit in hits:
        identity = hit.code if hit.code is not None else hit.matched_text
        if identity in seen:
            continue
        seen.add(identity)
        deduped.append(hit)
    return deduped


def _derived_org_values(
    data: "WorkbookData", numerator: str, denominator: str, date_text: str, scale: float = 1.0, percent: bool = False
) -> dict[str, float] | None:
    """全省各机构衍生比率值；任一机构缺分子/分母/分母为 0 时返回 None。"""
    values: dict[str, float] = {}
    for org in data.organizations:
        nv = data.value(numerator, org, date_text)
        dv = data.value(denominator, org, date_text)
        if nv is None or dv is None or dv == 0:
            return None
        value = nv * scale / dv
        if percent:
            value *= 100.0
        values[org] = value
    return values


def _template_performance_explicit(result: ParseResult) -> None:
    """题目明确列出指标集合的排名较好/较差题。

    指标集合从问题文本“待评价指标集合：……”段（或问题中明确命名的指标）
    提取；判定规则固定为衍生维度说明：表现较好=前三、表现较差=后四，
    排名方向由指标定义决定。除排名外，为答案中可能出现的点值、较年初
    基期/增量、全省均值生成完整可重算 claims（全部 mustAppear=False，
    指标集合仍只由问题文本决定）。
    """
    data = result.data
    result.category = "PERFORMANCE_EXPLICIT"
    date_text = _main_date(result)
    if date_text is None:
        return
    if not result.orgs:
        _unresolved(result, "缺少机构")
        return
    org = result.orgs[0]
    hits = _explicit_metric_hits(result.question)
    if not hits:
        _unresolved(result, "指标解析失败（问题未明确列出待评价指标集合）")
        return
    scope_orgs = sorted(data.organizations)
    rank_keys: list[str] = []
    good: list[str] = []
    bad: list[str] = []
    for hit in hits:
        if hit.derived is not None:
            numerator, denominator = hit.derived
            unit, scale = _derived_unit_scale(denominator)
            nv = data.value(numerator, org.code, date_text)
            dv = data.value(denominator, org.code, date_text)
            value = nv * scale / dv if (nv is not None and dv) else None
            if value is not None and unit == "%":
                value *= 100.0
            result.claims.append(make_claim(
                f"value_{hit.matched_text}", f"{org.name}{date_text}的{hit.matched_text}",
                kind="RATIO_PERCENT", value=value, unit=unit, metric=None, org=org.code, date=date_text,
                must_appear=False, role="ratio" if unit == "%" else None,
                evidence=_evidence_factory("RATIO", f"{numerator}*{scale}/{denominator}", 2, data,
                    numerator=numerator, denominator=denominator, org=org.code, date=date_text,
                    scale=scale, percent=unit == "%"),
            ))
            values = _derived_org_values(data, numerator, denominator, date_text, scale, unit == "%")
            if values is None:
                _unresolved(result, "排名事实不完整")
                return
            positions = rules.rank_positions([(code, data.organizations[code].name, v) for code, v in values.items()], numerator)
            target = next((item for item in positions if item[0] == org.code), None)
            if target is None:
                _unresolved(result, "目标机构不在排名中")
                return
            result.claims.append(make_claim(
                f"rank_{hit.matched_text}", f"{org.name}的{hit.matched_text}排名",
                kind="RANK", value=float(target[3]), unit="名", metric=None, org=org.code, date=date_text,
                role="rank", must_appear=False, extras=[{"kind": "TOTAL", "value": float(len(positions))}],
                evidence=_evidence_factory("RANK_POSITION", "rank(derived) 按排名方向，并列同名次", 2, data,
                    org=org.code, date=date_text, scope={"orgs": scope_orgs},
                    derived={"numerator": numerator, "denominator": denominator, "scale": scale, "percent": unit == "%"}),
            ))
            rank_keys.append(f"rank_{hit.matched_text}")
            if target[3] <= rules.GOOD_RANK_CUTOFF:
                good.append(hit.matched_text)
            if target[3] >= len(positions) - rules.BAD_TAIL_SIZE + 1:
                bad.append(hit.matched_text)
            continue
        metric = hit.code  # type: ignore[assignment]
        if metric not in data.metrics:
            _unresolved(result, f"指标定义缺失：{metric}")
            return
        metric_name = data.metrics[metric].name
        value = data.value(metric, org.code, date_text)
        result.claims.append(make_claim(f"value_{metric}", f"{org.name}{date_text}的{metric_name}", kind="VALUE", value=value, unit=data.metrics[metric].unit, metric=metric, org=org.code, date=date_text, must_appear=False, evidence=_evidence_factory("VALUE_LOOKUP", "直接取事实值 fact[date][org][metric]", 2, data, metric=metric, org=org.code, date=date_text)))
        base_value = data.value(metric, org.code, rules.YEAR_START)
        if base_value is not None:
            result.claims.append(make_claim(f"base_{metric}", f"{org.name}{rules.YEAR_START}的{metric_name}（较年初基期值）", kind="VALUE", value=base_value, unit=data.metrics[metric].unit, metric=metric, org=org.code, date=rules.YEAR_START, must_appear=False, evidence=_evidence_factory("VALUE_LOOKUP", "直接取事实值 fact[date][org][metric]", 2, data, metric=metric, org=org.code, date=rules.YEAR_START)))
            delta = value - base_value if value is not None else None
            result.claims.append(make_claim(f"delta_{metric}", f"{org.name}{metric_name}较年初增量", kind="DELTA", value=delta, unit=data.metrics[metric].unit, metric=metric, org=org.code, date=date_text, baseline=rules.YEAR_START, comparison_type="ytd", direction=("up" if (delta or 0) > 0.005 else "down" if (delta or 0) < -0.005 else "flat"), must_appear=False, evidence=_evidence_factory("DELTA", "current-base（较年初）", 2, data, metric=metric, org=org.code, date=date_text, baseline=rules.YEAR_START, comparisonType="ytd")))
        values = data.org_values(metric, date_text)
        if len(values) != len(data.organizations):
            _unresolved(result, "排名事实不完整")
            return
        positions = rules.rank_positions([(code, data.organizations[code].name, v) for code, v in values.items()], metric)
        target = next((item for item in positions if item[0] == org.code), None)
        if target is None:
            _unresolved(result, "目标机构不在排名中")
            return
        result.claims.append(make_claim(f"rank_{metric}", f"{org.name}的{metric_name}排名", kind="RANK", value=float(target[3]), unit="名", metric=metric, org=org.code, date=date_text, role="rank", must_appear=False, extras=[{"kind": "TOTAL", "value": float(len(positions))}], evidence=_evidence_factory("RANK_POSITION", "rank(v) 按排名方向，并列同名次", 2, data, metric=metric, org=org.code, date=date_text, scope={"orgs": scope_orgs})))
        rank_keys.append(f"rank_{metric}")
        if target[3] <= rules.GOOD_RANK_CUTOFF:
            good.append(metric_name)
        if target[3] >= len(positions) - rules.BAD_TAIL_SIZE + 1:
            bad.append(metric_name)
        mean = data.province_mean(metric, date_text)
        if mean is not None:
            result.claims.append(make_claim(f"mean_{metric}", f"{date_text}全省均值（{metric_name}）", kind="PROVINCE_MEAN", value=mean, unit=data.metrics[metric].unit, metric=metric, role="mean", must_appear=False, evidence=_evidence_factory("PROVINCE_MEAN", "Σ 全省机构事实值 / 机构数", 2, data, metric=metric, date=date_text, scope={"orgs": scope_orgs})))
    result.claims.append(make_claim("good_set", "表现较好集合（前三）", kind="SET_GOOD", value=float(len(good)), note=",".join(good), evidence=_evidence_factory("SET_GOOD", "count(rank<=3)", 2, data, dependsOn=rank_keys)))
    result.claims.append(make_claim("bad_set", "表现较差集合（后四）", kind="SET_BAD", value=float(len(bad)), note=",".join(bad), evidence=_evidence_factory("SET_BAD", "count(rank>=总机构数-3)", 2, data, dependsOn=rank_keys, total=float(len(data.organizations)))))


_DIMENSION_NAMES = ("规模", "资产质量", "盈利能力")
_DIMENSION_MARKER = "维度与指标映射："


def _parse_dimension_mapping(text: str) -> dict[str, list[rules.MetricHit]] | None:
    """从澄清后问题文本解析 规模/资产质量/盈利能力 三个维度的指标映射。

    返回 {维度: [MetricHit]}；任何维度缺指标、维度名不符、指标跨维度重复时
    返回 None（拒绝修改并保持 UNRESOLVED，不猜测）。
    """
    if _DIMENSION_MARKER not in text:
        return None
    segment = text.split(_DIMENSION_MARKER, 1)[1]
    mapping: dict[str, list[rules.MetricHit]] = {}
    for part in segment.split("；"):
        part = part.strip().rstrip("。")
        if not part:
            continue
        if "=" not in part:
            return None
        dim, metrics_text = part.split("=", 1)
        dim = dim.strip()
        if dim not in _DIMENSION_NAMES or dim in mapping:
            return None
        hits = _dedup_hits([hit for _position, hit in rules.extract_metrics_ordered(metrics_text)])
        if not hits:
            return None
        mapping[dim] = hits
    if set(mapping) != set(_DIMENSION_NAMES):
        return None
    seen: set[str] = set()
    for hits in mapping.values():
        for hit in hits:
            identity = hit.code if hit.code is not None else hit.matched_text
            if identity in seen:
                return None
            seen.add(identity)
    return mapping


def _template_dimension_3_clarified(result: ParseResult) -> None:
    """澄清后的三维度题：按问题文本中的维度与指标映射逐指标生成值/排名 claims。"""
    data = result.data
    result.category = "DIMENSION_3"
    date_text = _main_date(result)
    if date_text is None:
        return
    if not result.orgs:
        _unresolved(result, "缺少机构")
        return
    org = result.orgs[0]
    mapping = _parse_dimension_mapping(result.question)
    if mapping is None:
        _unresolved(result, "维度指标映射解析失败")
        return
    for dim, hits in mapping.items():
        for hit in hits:
            if hit.derived is not None:
                numerator, denominator = hit.derived
                unit, scale = _derived_unit_scale(denominator)
                nv = data.value(numerator, org.code, date_text)
                dv = data.value(denominator, org.code, date_text)
                value = nv * scale / dv if (nv is not None and dv) else None
                if value is not None and unit == "%":
                    value *= 100.0
                result.claims.append(make_claim(
                    f"value_{dim}_{hit.matched_text}", f"{org.name}{date_text}的{hit.matched_text}",
                    kind="RATIO_PERCENT", value=value, unit=unit, metric=None, org=org.code, date=date_text,
                    role="ratio" if unit == "%" else None,
                    evidence=_evidence_factory("RATIO", f"{numerator}*{scale}/{denominator}", 2, data,
                        numerator=numerator, denominator=denominator, org=org.code, date=date_text,
                        scale=scale, percent=unit == "%"),
                ))
                values = _derived_org_values(data, numerator, denominator, date_text, scale, unit == "%")
                if values is None:
                    _unresolved(result, "排名事实不完整")
                    return
                positions = rules.rank_positions([(code, data.organizations[code].name, v) for code, v in values.items()], numerator)
                target = next((item for item in positions if item[0] == org.code), None)
                if target is None:
                    _unresolved(result, "目标机构不在排名中")
                    return
                rank_value = float(target[3])
                result.claims.append(make_claim(
                    f"rank_{dim}_{hit.matched_text}", f"{org.name}的{hit.matched_text}排名",
                    kind="RANK", value=rank_value, unit="名", metric=None, org=org.code, date=date_text,
                    role="rank", must_appear=False,
                    evidence=_evidence_factory("RANK_POSITION", "rank(derived) 按排名方向，并列同名次", 2, data,
                        org=org.code, date=date_text, scope={"orgs": sorted(data.organizations)},
                        derived={"numerator": numerator, "denominator": denominator, "scale": scale, "percent": unit == "%"}),
                ))
                continue
            metric = hit.code  # type: ignore[assignment]
            if metric not in data.metrics:
                _unresolved(result, f"指标定义缺失：{metric}")
                return
            metric_name = data.metrics[metric].name
            _value_claim(result, f"value_{dim}_{hit.matched_text}", f"{org.name}{date_text}的{metric_name}", metric, org.code, date_text)
            values = data.org_values(metric, date_text)
            if len(values) != len(data.organizations):
                _unresolved(result, "排名事实不完整")
                return
            positions = rules.rank_positions([(code, data.organizations[code].name, v) for code, v in values.items()], metric)
            target = next((item for item in positions if item[0] == org.code), None)
            if target is None:
                _unresolved(result, "目标机构不在排名中")
                return
            rank_value = float(target[3])
            result.claims.append(make_claim(f"rank_{dim}_{hit.matched_text}", f"{org.name}的{metric_name}排名", kind="RANK", value=rank_value, unit="名", metric=metric, org=org.code, date=date_text, role="rank", must_appear=False, extras=[{"kind": "TOTAL", "value": float(len(data.organizations))}], evidence=_evidence_factory("RANK_POSITION", "rank(v) 按排名方向，并列同名次", 2, data, metric=metric, org=org.code, date=date_text, scope={"orgs": sorted(data.organizations)})))


def _template_profit_eval(result: ParseResult) -> None:
    data = result.data
    result.category = "PROFIT_EVAL"
    date_text = _main_date(result)
    if date_text is None:
        return
    if not result.orgs:
        _unresolved(result, "缺少机构")
        return
    org = result.orgs[0]
    text = result.question
    if "净利润" not in text and "成本收入比" not in text:
        _unresolved(result, "盈利能力指标解析失败")
        return
    metrics: list[str] = []
    if "净利润" in text:
        metrics.append("ZB011")
    if "成本收入比" in text:
        metrics.append("ZB012")
    if "收入结构" in text or "净利息收入" in text or "中间业务收入" in text:
        metrics.extend(["ZB008", "ZB007"])
    for metric in metrics:
        _value_claim(result, data.metrics[metric].name, f"{org.name}{date_text}的{data.metrics[metric].name}", metric, org.code, date_text)
        values = data.org_values(metric, date_text)
        if len(values) == len(data.organizations):
            positions = rules.rank_positions([(code, data.organizations[code].name, v) for code, v in values.items()], metric)
            target = next((item for item in positions if item[0] == org.code), None)
            if target is not None:
                result.claims.append(make_claim(f"rank_{metric}", f"{data.metrics[metric].name}全省排名", kind="RANK", value=float(target[3]), unit="名", metric=metric, org=org.code, date=date_text, role="rank", must_appear=False, evidence=_evidence_factory("RANK_POSITION", "rank(v) 按排名方向，并列同名次", 2, data, metric=metric, org=org.code, date=date_text, scope={"orgs": sorted(data.organizations)})))
    if "较年初" in text and "净利润" in text:
        current = data.value("ZB011", org.code, date_text)
        base = data.value("ZB011", org.code, rules.YEAR_START)
        delta = current - base if (current is not None and base is not None) else None
        result.claims.append(make_claim("ytd_delta", f"{org.name}净利润较年初增量", kind="DELTA", value=delta, unit=data.metrics["ZB011"].unit, metric="ZB011", org=org.code, date=date_text, baseline=rules.YEAR_START, comparison_type="ytd", direction=("up" if (delta or 0) > 0.005 else "down" if (delta or 0) < -0.005 else "flat"), evidence=_evidence_factory("DELTA", "current-base（较年初）", 2, data, metric="ZB011", org=org.code, date=date_text, baseline=rules.YEAR_START, comparisonType="ytd")))


def data_is_rate(data: WorkbookData, metric: str) -> bool:
    return metric in rules.RATE_METRICS


# --------------------------------------------------------------------------- 核对


def _direction_matches(claim: dict[str, Any], token: rules.AnswerToken) -> bool:
    direction = claim.get("direction")
    if direction is None:
        return True
    value = claim.get("value")
    if value is not None and abs(value) < 0.005:
        return True
    if token.direction == "diff":
        # “相差/差额”表述无方向性
        return True
    if token.negative:
        # 显式负号以符号为准（旧答案存在“高出-0.2”等措辞与数值符号矛盾的情形）
        return (value or 0) < -0.005
    if direction == "up":
        return token.direction in ("up", "flat")
    if direction == "down":
        return token.direction in ("down", "flat")
    return True


def _role_matches(claim: dict[str, Any], token: rules.AnswerToken) -> bool:
    role = claim.get("role")
    if role == "rank":
        return token.role == "rank"
    if token.role is None:
        return True
    if role is None:
        return True
    return token.role == role


def _org_matches(claim: dict[str, Any], token: rules.AnswerToken) -> bool:
    org = claim.get("org")
    if org is None:
        return True
    if token.org is None:
        return True
    return rules.ORG_CODE_BY_NAME.get(token.org) == org


def _metric_matches(claim: dict[str, Any], token: rules.AnswerToken) -> bool:
    metric = claim.get("metric")
    if metric is None:
        return True
    if token.metric is None:
        return True
    token_code = rules.METRIC_CODE_BY_ALIAS.get(token.metric)
    return token_code == metric or token.metric == claim.get("metricName")


def _top_entry_value_matches(claim: dict[str, Any], token: rules.AnswerToken) -> bool:
    """TOP_ENTRY 的值侧匹配：机构 + 数值 + 单位（名次由顺序匹配处理）。"""
    if claim.get("value") is None:
        return False
    if not rules.values_match(claim["value"], token.value, claim.get("rounding", 2)):
        return False
    if not rules.units_compatible(claim.get("unit"), token.unit):
        return False
    if claim.get("org") is not None and token.org is not None:
        if rules.ORG_CODE_BY_NAME.get(token.org) != claim["org"]:
            return False
    return True


def _match_top_entries(claims: list[dict[str, Any]], tokens: list[rules.AnswerToken]) -> set[int]:
    """按答案中出现顺序逐项匹配 TOP_ENTRY 的机构、名次和值。

    - 答案显式给出名次（第N名）时，名次 token 必须与条目名次一致；
    - 无显式名次时，出现顺序即名次顺序，逐项严格匹配；
    - 并列条目（相同名次）组内允许任意顺序，组间顺序仍严格；
    - 任一条目缺失或错位（TopN 截断 / 逆序 / 值错）即停止，剩余条目保持未匹配。
    返回已被条目消耗的 token 索引集合。
    """
    entry_indices = [index for index, claim in enumerate(claims) if claim["kind"] == "TOP_ENTRY"]
    if not entry_indices:
        return set()
    groups: list[list[int]] = []
    for index in entry_indices:
        position = claims[index]["position"]
        if groups and claims[groups[-1][0]]["position"] == position:
            groups[-1].append(index)
        else:
            groups.append([index])
    matched: set[int] = set()
    cursor = 0
    while cursor < len(tokens) and tokens[cursor].role in ("top前", "top后"):
        cursor += 1
    for group in groups:
        position = claims[group[0]]["position"]
        # 结构词（前N名/后N名）可出现在任意两组之间（前3...后3...）
        while cursor < len(tokens) and tokens[cursor].role in ("top前", "top后"):
            cursor += 1
        explicit_rank = False
        if cursor < len(tokens) and tokens[cursor].role == "rank":
            if tokens[cursor].value == position:
                matched.add(cursor)
                cursor += 1
                explicit_rank = True
            else:
                break  # 显式名次与预期不符 -> 顺序错误
        if explicit_rank:
            # 显式名次：窗口延伸到下一个不同名次的名次 token（并列组的同名次名次 token 属于组内）
            end = cursor
            while end < len(tokens) and not (tokens[end].role == "rank" and tokens[end].value != position):
                end += 1
            window = list(range(cursor, end))
            for idx in window:
                if tokens[idx].role == "rank":
                    matched.add(idx)
        else:
            # 无显式名次：出现顺序即名次顺序，窗口正好容纳本组条目
            end = min(cursor + len(group), len(tokens))
            window = list(range(cursor, end))
        value_window = [idx for idx in window if tokens[idx].role != "rank"]
        if len(value_window) < len(group):
            break  # 条目被截断/不足
        if len(group) == 1:
            idx = value_window[0]
            if _top_entry_value_matches(claims[group[0]], tokens[idx]):
                matched.add(idx)
                claims[group[0]]["matched"] = True
                claims[group[0]]["matchNote"] = tokens[idx].raw
                cursor = idx + 1
            else:
                break
        else:
            remaining_window = value_window
            for claim_index in group:
                found = next((idx for idx in remaining_window if _top_entry_value_matches(claims[claim_index], tokens[idx])), None)
                if found is None:
                    return matched
                matched.add(found)
                claims[claim_index]["matched"] = True
                claims[claim_index]["matchNote"] = tokens[found].raw
                remaining_window.remove(found)
            cursor = end
    return matched


def match_claims_to_tokens(claims: list[dict[str, Any]], tokens: list[rules.AnswerToken], answer_text: str, question_numbers: list[float]) -> tuple[list[rules.AnswerToken], list[str]]:
    matched_token_indices = _match_top_entries(claims, tokens)
    unmatched = [token for index, token in enumerate(tokens) if index not in matched_token_indices]
    notes: list[str] = []
    for claim in claims:
        if claim["kind"] == "TOP_ENTRY":
            continue  # 已由 _match_top_entries 按出现顺序逐项匹配
        if claim.get("value") is None:
            continue
        if claim["kind"] == "TREND":
            direction = claim.get("direction")
            words = {"up": ("上升", "增长", "升高", "向好"), "down": ("下降", "回落", "走低", "下滑"), "flat": ("持平", "平稳", "不变")}.get(direction, ())
            if any(word in answer_text for word in words):
                claim["matched"] = True
            continue
        if claim["kind"] == "MEETS":
            positive = claim.get("value") == 1.0
            words = ("达标", "满足", "符合", "超过", "高于", "是") if positive else ("否", "未", "不", "没有", "低于")
            if any(word in answer_text for word in words):
                claim["matched"] = True
            continue
        if claim["kind"] == "SET_GOOD":
            if "较好" in answer_text:
                claim["matched"] = True
            continue
        if claim["kind"] == "SET_BAD":
            if "较差" in answer_text or "无" in answer_text:
                claim["matched"] = True
            continue
        target_index = -1
        for index, token in enumerate(unmatched):
            if not rules.values_match(claim.get("value"), token.value, claim.get("rounding", 2)):
                continue
            if not rules.units_compatible(claim.get("unit"), token.unit):
                continue
            if not _direction_matches(claim, token):
                continue
            if not _role_matches(claim, token):
                continue
            if not _org_matches(claim, token):
                continue
            if not _metric_matches(claim, token):
                continue
            target_index = index
            break
        if target_index >= 0:
            claim["matched"] = True
            claim["matchNote"] = unmatched[target_index].raw
            unmatched.pop(target_index)
        for extra in claim.get("extras", []):
            if extra.get("kind") == "THRESHOLD":
                for index, token in enumerate(unmatched):
                    if rules.values_match(extra["value"], token.value, 2) and (token.unit == "%" or token.unit is None):
                        extra["matched"] = True
                        unmatched.pop(index)
                        break
            elif extra.get("kind") == "TOTAL":
                for index, token in enumerate(unmatched):
                    if token.role == "total" and rules.values_match(extra["value"], token.value, 0):
                        extra["matched"] = True
                        unmatched.pop(index)
                        break
    remaining: list[rules.AnswerToken] = []
    for token in unmatched:
        if token.role in ("top前", "top后"):
            notes.append(f"token {token.raw} 为 TopN 结构词（前N名/后N名）")
            continue
        attributed = False
        for constant in question_numbers:
            if rules.values_match(constant, token.value, 2):
                attributed = True
                notes.append(f"token {token.raw} 归因于问题常数 {constant}")
                break
        if not attributed:
            for claim in claims:
                if claim.get("value") is not None and rules.values_match(claim.get("value"), token.value, claim.get("rounding", 2)) and rules.units_compatible(claim.get("unit"), token.unit):
                    if not _org_matches(claim, token):
                        continue
                    attributed = True
                    notes.append(f"token {token.raw} 归因于已匹配声明（重复提及）{claim['label']}")
                    break
        if not attributed:
            remaining.append(token)
    return remaining, notes


# --------------------------------------------------------------------------- 扩展核对


def extended_check(data: WorkbookData, result: ParseResult, tokens: list[rules.AnswerToken]) -> tuple[list[rules.AnswerToken], list[str]]:
    remaining: list[rules.AnswerToken] = []
    notes: list[str] = []
    date_text = main_date_of(result.question)
    fallback_org = result.orgs[0].code if result.orgs else None
    for token in tokens:
        ok = False
        if token.role == "rank" and token.metric is not None and date_text is not None:
            metric = rules.METRIC_CODE_BY_ALIAS.get(token.metric)
            org = rules.ORG_CODE_BY_NAME.get(token.org) if token.org else fallback_org
            if metric and org:
                values = data.org_values(metric, date_text)
                if len(values) == len(data.organizations):
                    positions = rules.rank_positions([(code, data.organizations[code].name, value) for code, value in values.items()], metric)
                    target = next((item for item in positions if item[0] == org), None)
                    if target is not None and rules.values_match(float(target[3]), token.value, 0):
                        ok = True
                        notes.append(f"token {token.raw} 扩展核对：{token.org or fallback_org} {token.metric} 排名第{target[3]}名")
        if not ok and "均值" in token.context and token.metric is not None and date_text is not None:
            metric = rules.METRIC_CODE_BY_ALIAS.get(token.metric)
            if metric:
                mean = data.province_mean(metric, date_text)
                if mean is not None and rules.values_match(mean, token.value, 2):
                    ok = True
                    notes.append(f"token {token.raw} 扩展核对：{token.metric} 全省均值")
        if not ok and "较年初" in token.context and token.metric is not None and date_text is not None:
            metric = rules.METRIC_CODE_BY_ALIAS.get(token.metric)
            org = rules.ORG_CODE_BY_NAME.get(token.org) if token.org else fallback_org
            if metric and org:
                current = data.value(metric, org, date_text)
                base = data.value(metric, org, rules.YEAR_START)
                if current is not None and base is not None:
                    delta = current - base
                    if rules.values_match(abs(delta), token.value, 2):
                        ok = True
                        notes.append(f"token {token.raw} 扩展核对：{token.org or fallback_org} {token.metric} 较年初增量")
        if not ok and token.pair_prev is not None and date_text is not None and token.metric is not None:
            metric = rules.METRIC_CODE_BY_ALIAS.get(token.metric)
            org = rules.ORG_CODE_BY_NAME.get(token.org) if token.org else fallback_org
            if metric and org:
                base_value = data.value(metric, org, rules.YEAR_START)
                current_value = data.value(metric, org, date_text)
                if base_value is not None and current_value is not None and rules.values_match(base_value, float(token.pair_prev), 2) and rules.values_match(current_value, token.value, 2):
                    ok = True
                    notes.append(f"token {token.raw} 扩展核对：{token.org} {token.metric} 较年初从{token.pair_prev}到{token.raw}")
        if not ok:
            remaining.append(token)
    return remaining, notes


# --------------------------------------------------------------------------- 输出


def build_corrected_answer(claims: list[dict[str, Any]]) -> str:
    parts: list[str] = []
    for claim in claims:
        if claim.get("value") is None:
            continue
        if claim["kind"] in ("TREND", "MEETS"):
            continue
        if not claim.get("mustAppear", True):
            continue
        parts.append(f"{claim['label']}：{claim_display(claim)}")
    return "；".join(parts)


def write_review_xlsx(path: Path, reviews: list[dict[str, Any]]) -> None:
    workbook = Workbook()
    # 固定文档时间戳，保证输出确定性（两次审查 xlsx 哈希一致）
    workbook.properties.created = datetime(2000, 1, 1, 0, 0, 0)
    workbook.properties.modified = datetime(2000, 1, 1, 0, 0, 0)
    sheet = workbook.active
    sheet.title = "review"
    headers = ["id", "split", "difficulty", "status", "category", "fullEvidence", "question", "answerText", "correctedAnswerText", "claims", "auditErrors"]
    sheet.append(headers)
    for review in reviews:
        sheet.append(
            [
                review["id"],
                review["split"],
                review["difficulty"],
                review["status"],
                review["category"],
                review["fullEvidence"],
                review["question"],
                review["answerText"],
                review.get("correctedAnswerText"),
                json.dumps(review["claims"], ensure_ascii=False),
                json.dumps(review["auditErrors"], ensure_ascii=False),
            ]
        )
    workbook.save(path)
    _fix_zip_timestamps(path)


def _fix_zip_timestamps(path: Path) -> None:
    """重写 xlsx zip 条目时间戳为固定值，保证跨秒运行输出字节一致（确定性）。"""
    fixed_ts = (2000, 1, 1, 0, 0, 0)
    tmp_path = path.with_name(path.name + ".fixed")
    try:
        with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(tmp_path, "w", compression=zipfile.ZIP_DEFLATED) as target:
            for info in source.infolist():
                new_info = zipfile.ZipInfo(info.filename, date_time=fixed_ts)
                new_info.compress_type = info.compress_type
                new_info.external_attr = info.external_attr
                target.writestr(new_info, source.read(info.filename))
        tmp_path.replace(path)
    finally:
        if tmp_path.exists():
            tmp_path.unlink()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for block in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest().upper()


# --------------------------------------------------------------------------- 候选 manifest 哈希边车


def validate_manifest_sidecar(manifest_path: Path) -> dict[str, Any]:
    """验证候选 manifest 的哈希边车（<manifest 文件名>.sha256）。

    边车必须与 manifest 同目录存在，内容严格符合生成器契约
    "<UPPER_SHA256>  <manifest 文件名>\\n"（64 位大写 SHA-256、两个空格、
    manifest 基名、单行），且摘要必须等于 manifest 当前原始字节的 SHA-256。
    缺失 / 无法读取 / 格式非法 / 文件名不符 / 摘要不匹配一律 fail closed ——
    这是候选审查前对 manifest 字节完整性的最终锚点，先于任何 JSON 解析与
    字段语义验证。
    """
    errors: list[str] = []
    sidecar_path = manifest_path.with_name(manifest_path.name + ".sha256")
    if not sidecar_path.is_file():
        errors.append(f"候选 manifest 哈希边车缺失: {sidecar_path.name}")
        return {"valid": False, "errors": errors}
    try:
        content = sidecar_path.read_text(encoding="utf-8")
    except (OSError, UnicodeDecodeError) as exc:
        errors.append(f"候选 manifest 哈希边车无法读取: {exc}")
        return {"valid": False, "errors": errors}
    match = re.fullmatch(r"([0-9A-F]{64})  ([^\n]+)\n", content)
    if match is None:
        errors.append(
            f"候选 manifest 哈希边车格式非法（应为 '<UPPER_SHA256>  {manifest_path.name}' 单行）: {content!r}"
        )
        return {"valid": False, "errors": errors}
    digest, filename = match.group(1), match.group(2)
    if filename != manifest_path.name:
        errors.append(f"候选 manifest 哈希边车文件名不符: {filename!r} != {manifest_path.name!r}")
        return {"valid": False, "errors": errors}
    actual = sha256_file(manifest_path)
    if digest != actual:
        errors.append(f"候选 manifest 哈希边车摘要不匹配（{digest} != 实际 {actual}）")
    return {"valid": not errors, "errors": errors}


def write_manifest_with_sidecar(manifest_path: Path, manifest: dict[str, Any]) -> None:
    """同目录临时文件 + replace 原子写回 manifest 及其 .sha256 哈希边车。

    先写 manifest 与边车的临时文件，再依次 replace 落位；任一步失败都不会
    留下半写的目标文件。边车摘要按实际写入的 manifest 字节计算，写回成功后
    边车必然与最终 manifest 字节一致。
    """
    payload = json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True) + "\n"
    manifest_bytes = payload.encode("utf-8")
    digest = hashlib.sha256(manifest_bytes).hexdigest().upper()
    sidecar_path = manifest_path.with_name(manifest_path.name + ".sha256")
    sidecar_bytes = f"{digest}  {manifest_path.name}\n".encode("utf-8")
    tmp_manifest = manifest_path.with_name(manifest_path.name + ".tmp")
    tmp_sidecar = sidecar_path.with_name(sidecar_path.name + ".tmp")
    try:
        tmp_manifest.write_bytes(manifest_bytes)
        tmp_sidecar.write_bytes(sidecar_bytes)
        tmp_manifest.replace(manifest_path)
        tmp_sidecar.replace(sidecar_path)
    finally:
        for tmp_path in (tmp_manifest, tmp_sidecar):
            if tmp_path.exists():
                tmp_path.unlink()


# --------------------------------------------------------------------------- 审查主流程


def compute_canonical_ready(
    reviews: list[dict[str, Any]],
    source_sha256: str,
    expected_total: int = 200,
    expected_splits: dict[str, int] | None = None,
    expected_source_sha: str = SOURCE_SHA256_EXPECTED,
) -> bool:
    """canonicalReady 门控：全部条件同时满足才为 True。"""
    if expected_splits is None:
        expected_splits = {"train": 120, "dev": 40, "test": 40}
    split_counts = Counter(review["split"] for review in reviews)
    status_counts = Counter(review["status"] for review in reviews)
    return (
        len(reviews) == expected_total
        and len({review["id"] for review in reviews}) == expected_total
        and all(split_counts.get(split, 0) == count for split, count in expected_splits.items())
        and status_counts.get("UNRESOLVED", 0) == 0
        and all(not review["auditErrors"] for review in reviews)
        and sum(1 for review in reviews if review["fullEvidence"]) == expected_total
        and all(review["correctedAnswerText"] for review in reviews if review["status"] == "CORRECTED")
        and source_sha256 == expected_source_sha
    )


def candidate_audit_passes(reviews: list[dict[str, Any]], expected: dict[str, Any] | None = None) -> tuple[bool, list[str]]:
    """候选可用门控：就绪不变量与 manifest 意愿无关 —— 候选必须 100% VERIFIED、
    零 CORRECTED/UNRESOLVED、证据全量完整且零错误、split/status 各计数和等于
    totalRecords；实际审查结果同时必须与已验证 manifest 的 expectedAudit 逐项
    一致（本版本真实候选为 199 条 / 119+40+40 / 199 VERIFIED / 0 CORRECTED /
    0 UNRESOLVED / evidenceComplete=199 / evidenceErrors=0）。缺少 expectedAudit
    或任何一项不一致即 fail closed（candidateReady=false，不按真实 ID 硬编码）。
    """
    reasons: list[str] = []
    if expected is None:
        return False, ["候选 manifest 缺少 expectedAudit（候选就绪必须由已验证 manifest 推导）"]
    total_records = expected.get("totalRecords")
    split_counts = Counter(review["split"] for review in reviews)
    status_counts = Counter(review["status"] for review in reviews)
    # 归一化：零计数状态/划分也必须出现在比较与输出中（Counter 会省略零键）
    split_counts = {split: split_counts.get(split, 0) for split in ("train", "dev", "test")}
    status_counts = {status: status_counts.get(status, 0) for status in ("VERIFIED", "CORRECTED", "UNRESOLVED")}
    unique_ids = len({review["id"] for review in reviews})
    evidence_complete = sum(1 for review in reviews if review["fullEvidence"])
    evidence_errors = sum(
        1
        for review in reviews
        if not review["evidenceValidation"]["valid"] or review["evidenceValidation"]["claimErrors"] or review["evidenceValidation"]["errors"]
    )
    if len(reviews) != total_records or unique_ids != total_records:
        reasons.append(f"总数/唯一 ID 不符预期：{len(reviews)} 条 / {unique_ids} 个 != 预期 {total_records}")
    if split_counts != expected.get("splitCounts"):
        reasons.append(f"split 计数不符：{split_counts} != 预期 {expected.get('splitCounts')}")
    if status_counts != expected.get("statusCounts"):
        reasons.append(f"状态分布不符：{status_counts} != 预期 {expected.get('statusCounts')}")
    if evidence_complete != expected.get("evidenceComplete"):
        reasons.append(f"完整证据条数不符：{evidence_complete} != 预期 {expected.get('evidenceComplete')}")
    if evidence_errors != expected.get("evidenceErrors"):
        reasons.append(f"证据错误条数不符：{evidence_errors} != 预期 {expected.get('evidenceErrors')}")
    # 就绪不变量（独立于 manifest 意愿）：任何保留的 CORRECTED/UNRESOLVED 记录、
    # 缺失证据、或 split/status 计数和不等于 totalRecords 的候选一律拒绝
    if status_counts["VERIFIED"] != total_records:
        reasons.append(f"就绪不变量：VERIFIED={status_counts['VERIFIED']} != totalRecords={total_records}")
    if status_counts["CORRECTED"] != 0:
        reasons.append(f"就绪不变量：CORRECTED={status_counts['CORRECTED']} != 0")
    if status_counts["UNRESOLVED"] != 0:
        reasons.append(f"就绪不变量：UNRESOLVED={status_counts['UNRESOLVED']} != 0")
    if sum(split_counts.values()) != total_records:
        reasons.append(f"就绪不变量：split 计数和 {sum(split_counts.values())} != totalRecords={total_records}")
    if sum(status_counts.values()) != total_records:
        reasons.append(f"就绪不变量：status 计数和 {sum(status_counts.values())} != totalRecords={total_records}")
    return not reasons, reasons


def _validate_candidate_manifest(
    manifest: dict[str, Any], candidate_sha256: str, manifest_path: Path, expected_source_sha: str = SOURCE_SHA256_EXPECTED
) -> dict[str, Any]:
    """验证候选 manifest：候选哈希、原始来源哈希、生成器契约、变更账本哈希链、
    totalRecords/splitCounts/expectedAudit 结构。候选就绪预期全部由
    expectedAudit 推导，因此该结构缺失或非法即拒绝。

    候选审查只接受显式 --manifest 且全部验证通过；不提供任何跳过/关闭校验的开关。
    """
    errors: list[str] = []
    if manifest.get("generatorName") != "clarify_ground_truth_contracts":
        errors.append(f"generatorName 不符：{manifest.get('generatorName')!r}")
    version = manifest.get("generatorVersion")
    if not version or not re.fullmatch(r"\d+\.\d+\.\d+", str(version)):
        errors.append(f"generatorVersion 格式非法：{version!r}")
    if manifest.get("sourceSha256") != expected_source_sha:
        errors.append("manifest.sourceSha256 与冻结原始工作簿哈希不一致")
    if manifest.get("candidateSha256") != candidate_sha256:
        errors.append(f"manifest.candidateSha256 与实际候选工作簿哈希不一致（{manifest.get('candidateSha256')} != {candidate_sha256}）")
    change_counts = manifest.get("changeCounts")
    if not isinstance(change_counts, dict):
        errors.append("manifest.changeCounts 缺失")
    else:
        for field in ("answerChanges", "questionClarifications", "questionRemovals", "contractErrors"):
            if not isinstance(change_counts.get(field), int) or change_counts.get(field) < 0:
                errors.append(f"manifest.changeCounts.{field} 非法：{change_counts.get(field)!r}")
    if not isinstance(manifest.get("totalRecords"), int) or manifest.get("totalRecords") < 0:
        errors.append(f"manifest.totalRecords 非法：{manifest.get('totalRecords')!r}")
    split_counts = manifest.get("splitCounts")
    if (
        not isinstance(split_counts, dict)
        or set(split_counts) != {"train", "dev", "test"}
        or any(not isinstance(value, int) or value < 0 for value in split_counts.values())
    ):
        errors.append(f"manifest.splitCounts 非法：{split_counts!r}")
    expected_audit = manifest.get("expectedAudit")
    if not isinstance(expected_audit, dict):
        errors.append("manifest.expectedAudit 缺失或非对象（候选就绪必须由 expectedAudit 推导）")
    else:
        for field in ("totalRecords", "evidenceComplete", "evidenceErrors"):
            if not isinstance(expected_audit.get(field), int) or expected_audit.get(field) < 0:
                errors.append(f"manifest.expectedAudit.{field} 非法：{expected_audit.get(field)!r}")
        expected_splits = expected_audit.get("splitCounts")
        if (
            not isinstance(expected_splits, dict)
            or set(expected_splits) != {"train", "dev", "test"}
            or any(not isinstance(value, int) or value < 0 for value in expected_splits.values())
        ):
            errors.append(f"manifest.expectedAudit.splitCounts 非法：{expected_splits!r}")
        expected_statuses = expected_audit.get("statusCounts")
        if (
            not isinstance(expected_statuses, dict)
            or set(expected_statuses) != {"VERIFIED", "CORRECTED", "UNRESOLVED"}
            or any(not isinstance(value, int) or value < 0 for value in expected_statuses.values())
        ):
            errors.append(f"manifest.expectedAudit.statusCounts 非法：{expected_statuses!r}")
        if expected_audit.get("totalRecords") != manifest.get("totalRecords"):
            errors.append("manifest.expectedAudit.totalRecords 与 manifest.totalRecords 不一致")
        if expected_audit.get("splitCounts") != manifest.get("splitCounts"):
            errors.append("manifest.expectedAudit.splitCounts 与 manifest.splitCounts 不一致")
        # 就绪不变量（独立于 manifest 意愿）：候选必须 100% VERIFIED、零
        # CORRECTED/UNRESOLVED、证据全量完整且零错误；split/status 各计数和
        # 必须等于 totalRecords。预期或祝福任何保留错误/未决记录的 manifest 一律拒绝。
        total_records = manifest.get("totalRecords")
        expected_splits = expected_audit.get("splitCounts") or {}
        expected_statuses = expected_audit.get("statusCounts") or {}
        if sum(expected_splits.values()) != total_records:
            errors.append(f"manifest.expectedAudit.splitCounts 计数和 {sum(expected_splits.values())} != totalRecords={total_records}")
        if sum(expected_statuses.values()) != total_records:
            errors.append(f"manifest.expectedAudit.statusCounts 计数和 {sum(expected_statuses.values())} != totalRecords={total_records}")
        if expected_statuses.get("VERIFIED") != total_records:
            errors.append(f"manifest.expectedAudit.statusCounts.VERIFIED={expected_statuses.get('VERIFIED')} != totalRecords={total_records}（候选必须全部 VERIFIED）")
        if expected_statuses.get("CORRECTED") != 0:
            errors.append(f"manifest.expectedAudit.statusCounts.CORRECTED={expected_statuses.get('CORRECTED')} != 0（候选不得保留 CORRECTED 记录）")
        if expected_statuses.get("UNRESOLVED") != 0:
            errors.append(f"manifest.expectedAudit.statusCounts.UNRESOLVED={expected_statuses.get('UNRESOLVED')} != 0（候选不得保留 UNRESOLVED 记录）")
        if expected_audit.get("evidenceComplete") != total_records:
            errors.append(f"manifest.expectedAudit.evidenceComplete={expected_audit.get('evidenceComplete')} != totalRecords={total_records}")
        if expected_audit.get("evidenceErrors") != 0:
            errors.append(f"manifest.expectedAudit.evidenceErrors={expected_audit.get('evidenceErrors')} != 0")
    if not isinstance(manifest.get("candidateReady"), bool):
        errors.append("manifest.candidateReady 缺失或非布尔")
    ledger_sha = manifest.get("changeLedgerSha256")
    ledger_path = manifest_path.with_name("contract-change-ledger.json")
    if not isinstance(ledger_sha, str) or not ledger_sha:
        errors.append("manifest.changeLedgerSha256 缺失")
    elif not ledger_path.is_file() or sha256_file(ledger_path) != ledger_sha:
        errors.append("contract-change-ledger.json 与 manifest.changeLedgerSha256 不一致")
    # 变更账本内容契约：生成器名/版本、声明计数与条目数、唯一 ID/动作兼容性、
    # 动作计数与 changeCounts 一致、contractErrors 与 manifest 一致。
    # 篡改账本动作、重复条目或计数（即使同步重算哈希）也会在这里 fail closed。
    if ledger_path.is_file():
        try:
            ledger = json.loads(ledger_path.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError) as exc:
            errors.append(f"contract-change-ledger.json 无法解析：{exc}")
            ledger = None
        if ledger is not None:
            if ledger.get("generatorName") != manifest.get("generatorName"):
                errors.append(f"ledger.generatorName={ledger.get('generatorName')!r} 与 manifest.generatorName={manifest.get('generatorName')!r} 不一致")
            if ledger.get("generatorVersion") != manifest.get("generatorVersion"):
                errors.append(f"ledger.generatorVersion={ledger.get('generatorVersion')!r} 与 manifest.generatorVersion={manifest.get('generatorVersion')!r} 不一致")
            entries = ledger.get("entries")
            if not isinstance(entries, list):
                errors.append("ledger.entries 缺失或非列表")
            else:
                if ledger.get("count") != len(entries):
                    errors.append(f"ledger.count={ledger.get('count')} != entries 条目数 {len(entries)}")
                action_counts: Counter = Counter()
                seen_ids: set[str] = set()
                for entry in entries:
                    if not isinstance(entry, dict) or not isinstance(entry.get("id"), str) or not entry["id"]:
                        errors.append("ledger.entries 含非法条目（缺 id 或 id 非字符串）")
                        continue
                    entry_id = entry["id"]
                    if entry_id in seen_ids:
                        errors.append(f"ledger.entries 存在重复 ID：{entry_id}")
                    seen_ids.add(entry_id)
                    change_type = entry.get("changeType")
                    if change_type not in ("ANSWER_CORRECTION", "QUESTION_CLARIFICATION", "QUESTION_REMOVAL"):
                        errors.append(f"ledger.entries 含非法 changeType：{change_type!r}（{entry_id}）")
                        continue
                    action_counts[change_type] += 1
                    if change_type == "QUESTION_REMOVAL":
                        if not isinstance(entry.get("removedAnswerSha256"), str):
                            errors.append(f"QUESTION_REMOVAL 条目缺少 removedAnswerSha256：{entry_id}")
                    else:
                        for field in ("oldTextSha256", "newTextSha256"):
                            if not isinstance(entry.get(field), str):
                                errors.append(f"{change_type} 条目缺少 {field}：{entry_id}")
                change_counts = manifest.get("changeCounts")
                if isinstance(change_counts, dict):
                    for field, change_type in (
                        ("answerChanges", "ANSWER_CORRECTION"),
                        ("questionClarifications", "QUESTION_CLARIFICATION"),
                        ("questionRemovals", "QUESTION_REMOVAL"),
                    ):
                        if action_counts.get(change_type, 0) != change_counts.get(field):
                            errors.append(
                                f"ledger 动作计数 {change_type}={action_counts.get(change_type, 0)} != manifest.changeCounts.{field}={change_counts.get(field)}"
                            )
            ledger_contract_errors = ledger.get("contractErrors")
            if not isinstance(ledger_contract_errors, list):
                errors.append("ledger.contractErrors 缺失或非列表")
            elif manifest.get("contractErrors") != ledger_contract_errors:
                errors.append("ledger.contractErrors 与 manifest.contractErrors 不一致")
            elif isinstance(change_counts, dict) and len(ledger_contract_errors) != change_counts.get("contractErrors"):
                errors.append(
                    f"ledger.contractErrors 数量 {len(ledger_contract_errors)} != manifest.changeCounts.contractErrors={change_counts.get('contractErrors')}"
                )
    return {"valid": not errors, "errors": errors}


def run_audit(
    workbook_path: Path,
    output_dir: Path,
    all_splits: bool,
    expected_total: int | None = None,
    expected_splits: dict[str, int] | None = None,
    expected_source_sha: str = SOURCE_SHA256_EXPECTED,
    manifest_path: Path | None = None,
) -> dict[str, Any]:
    global data_metrics, data_org_names
    data = WorkbookData(workbook_path)
    manifest: dict[str, Any] | None = None
    manifest_validation: dict[str, Any] = {"valid": True, "errors": []}
    manifest_sidecar_validation: dict[str, Any] = {"valid": True, "errors": []}
    if manifest_path is not None:
        # 候选审查模式：不接受任意哈希；必须先通过哈希边车验证 manifest 字节
        # 完整性，再验证候选哈希、原始来源哈希与生成器契约。
        if not manifest_path.is_file():
            raise ValueError(f"候选 manifest 不存在: {manifest_path}")
        manifest_sidecar_validation = validate_manifest_sidecar(manifest_path)
        if not manifest_sidecar_validation["valid"]:
            raise ValueError(f"候选 manifest 哈希边车验证失败: {'；'.join(manifest_sidecar_validation['errors'])}")
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        manifest_validation = _validate_candidate_manifest(manifest, data.source_sha256, manifest_path, expected_source_sha)
        if not manifest_validation["valid"]:
            raise ValueError(f"候选 manifest 验证失败: {'；'.join(manifest_validation['errors'])}")
    if manifest_path is None and data.source_sha256 != expected_source_sha:
        # 候选审查模式下，工作簿哈希由 manifest.candidateSha256 验证；
        # expected_source_sha 只用于校验 manifest.sourceSha256（冻结原始哈希）。
        raise ValueError(f"源文件哈希不匹配：{data.source_sha256} != {expected_source_sha}")
    data_metrics = {code: metric.name for code, metric in data.metrics.items()}
    data_org_names = {code: org.name for code, org in data.organizations.items()}
    rules.ORG_CODE_BY_NAME.update({org.name: code for code, org in data.organizations.items()})

    reviews: list[dict[str, Any]] = []
    for question in data.questions:
        parsed = parse_question(question["question"], data)
        finalize_evidence(parsed.claims, data)
        tokens = rules.tokenize_answer(question["answerText"])
        unmatched, notes = match_claims_to_tokens(parsed.claims, tokens, question["answerText"], rules.question_numbers(question["question"]))
        unmatched, extended_notes = extended_check(data, parsed, unmatched)
        notes.extend(extended_notes)

        all_values = all(claim.get("value") is not None for claim in parsed.claims)
        # auditErrors 只统计审查器异常或无完整证据的程序审计错误：
        # 解析歧义 / 声明无法重算 / 证据验证失败；合法 CORRECTED（旧答案数值、顺序
        # 或覆盖缺口可被完整证据定位）不算程序错误，UNRESOLVED 由
        # statusCounts/unresolvedIds 单独统计。
        audit_errors: list[str] = []
        if parsed.unresolved_reason:
            audit_errors.append(f"解析歧义：{parsed.unresolved_reason}")
        for claim in parsed.claims:
            if claim.get("value") is None and claim.get("mustAppear", True):
                audit_errors.append(f"声明无法重算：{claim['label']}（{claim.get('note') or '事实缺失'}）")

        evidence_validation = validate_review_evidence(parsed.claims, data)
        evidence_errors: list[str] = []
        for key, result in evidence_validation["claimResults"].items():
            if not result["ok"]:
                for error in result["errors"]:
                    evidence_errors.append(f"证据校验失败（{key}）：{error}")
        for error in evidence_validation["errors"]:
            evidence_errors.append(f"证据校验失败：{error}")
        evidence_valid = evidence_validation["valid"] and not evidence_errors

        if parsed.ambiguous:
            status = "UNRESOLVED"
        elif not all_values:
            status = "UNRESOLVED"
        elif not evidence_valid:
            # 证据验证失败：记录必须 UNRESOLVED，不得标成 CORRECTED/VERIFIED
            status = "UNRESOLVED"
            audit_errors.extend(evidence_errors)
        else:
            missing = [claim["label"] for claim in parsed.claims if claim.get("mustAppear", True) and not claim.get("matched") and claim.get("value") is not None and claim["kind"] not in ("TREND", "MEETS")]
            if unmatched or missing:
                status = "CORRECTED"
                for label in missing:
                    notes.append(f"旧答案未覆盖声明：{label}")
                for token in unmatched:
                    notes.append(f"旧答案存在无法核对的数值：{token.raw}")
            else:
                status = "VERIFIED"

        # fullEvidence 只来自证据验证结果，绝不从 status 推导：
        # 解析歧义/缺值记录缺少可复核声明集，证据不完整；其余记录在全部
        # mustAppear claim 证据验证通过时视为完整证据。
        full_evidence = (not parsed.ambiguous and all_values and evidence_valid)

        review = {
            "id": question["id"],
            "split": question["split"],
            "difficulty": question["difficulty"],
            "rowNumber": question["rowNumber"],
            "status": status,
            "category": parsed.category,
            "fullEvidence": full_evidence,
            "question": question["question"],
            "answerText": question["answerText"],
            "correctedAnswerText": build_corrected_answer(parsed.claims) if status == "CORRECTED" else None,
            "claims": parsed.claims,
            "auditErrors": audit_errors,
            "unresolvedReason": parsed.unresolved_reason,
            "matchNotes": notes,
            "evidenceValidation": {
                "valid": evidence_valid,
                "ruleVersion": EVIDENCE_RULES_VERSION,
                "claimErrors": {key: result["errors"] for key, result in evidence_validation["claimResults"].items() if not result["ok"]},
                "errors": evidence_validation["errors"],
            },
        }
        reviews.append(review)

    reviews.sort(key=lambda item: item["rowNumber"])
    split_counts = Counter(review["split"] for review in reviews)
    status_counts = Counter(review["status"] for review in reviews)
    # 归一化：零计数状态/划分也写入输出（Counter 会省略零键，导致
    # {VERIFIED: 199} 之类的分布缺少 CORRECTED/UNRESOLVED 零键）
    split_counts = {split: split_counts.get(split, 0) for split in ("train", "dev", "test")}
    status_counts = {status: status_counts.get(status, 0) for status in ("VERIFIED", "CORRECTED", "UNRESOLVED")}
    full_evidence = sum(1 for review in reviews if review["fullEvidence"])
    audit_error_count = sum(1 for review in reviews if review["auditErrors"])
    evidence_error_count = sum(1 for review in reviews if not review["evidenceValidation"]["valid"] or review["evidenceValidation"]["claimErrors"] or review["evidenceValidation"]["errors"])

    corrected_ledger = [
        {
            "id": review["id"], "split": review["split"], "difficulty": review["difficulty"],
            "question": review["question"], "answerText": review["answerText"],
            "correctedAnswerText": review["correctedAnswerText"], "auditErrors": review["auditErrors"],
            "claims": review["claims"],
        }
        for review in reviews
        if review["status"] == "CORRECTED"
    ]

    canonical_ready = compute_canonical_ready(
        reviews,
        data.source_sha256,
        expected_total=expected_total if expected_total is not None else 200,
        expected_splits=expected_splits,
        # 候选工作簿永远不得 canonical：必须与冻结原始哈希一致才可能 canonical
        expected_source_sha=SOURCE_SHA256_EXPECTED if manifest_path is not None else expected_source_sha,
    )

    output_dir.mkdir(parents=True, exist_ok=True)
    (output_dir / "review.ndjson").write_text(
        "".join(
            json.dumps(
                {key: review[key] for key in ("id", "split", "difficulty", "status", "category", "fullEvidence", "question", "answerText", "correctedAnswerText", "claims", "auditErrors", "unresolvedReason", "matchNotes", "evidenceValidation")},
                ensure_ascii=False, sort_keys=True,
            )
            + "\n"
            for review in reviews
        ),
        encoding="utf-8",
    )
    (output_dir / "correction-ledger.json").write_text(
        json.dumps({"ruleVersion": rules.RULES_VERSION, "count": len(corrected_ledger), "corrections": corrected_ledger}, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    output_name = "canonical-corrected.xlsx" if canonical_ready else "candidate-reviewed.xlsx"
    write_review_xlsx(output_dir / output_name, reviews)
    # 输出完整性：先写文件全部存在且非空，纳入 canonicalReady 门控
    # （audit-summary.json 在 summary 组装后写入，不在此处检查）
    required_outputs = ["review.ndjson", "correction-ledger.json", output_name]
    outputs_complete = all((output_dir / name).is_file() and (output_dir / name).stat().st_size > 0 for name in required_outputs)
    canonical_ready = canonical_ready and outputs_complete

    # 候选审查模式：candidateReady 只由候选审查结果决定，预期 total/split/status/
    # 证据完整性全部从已验证 manifest 的 expectedAudit 推导（不依赖任何真实 ID）；
    # 实际审查与预期逐项一致才为 true，任何不一致保持 false 并报告差异。
    # 候选永远不得 canonical（canonicalReady 已因候选哈希 ≠ 冻结原始哈希而恒为 false）。
    candidate_ready = False
    candidate_reasons: list[str] = []
    if manifest is not None:
        candidate_ready, candidate_reasons = candidate_audit_passes(reviews, manifest.get("expectedAudit"))
        manifest["candidateReady"] = candidate_ready
        manifest["candidateAudit"] = {
            "statusCounts": dict(status_counts),
            "splitCounts": dict(split_counts),
            "candidateReady": candidate_ready,
            "reasons": candidate_reasons,
        }
        write_manifest_with_sidecar(manifest_path, manifest)  # type: ignore[arg-type]

    summary = {
        "ruleVersion": rules.RULES_VERSION,
        "evidenceRuleVersion": EVIDENCE_RULES_VERSION,
        "sourceSha256": data.source_sha256,
        "sourceSha256Match": data.source_sha256 == SOURCE_SHA256_EXPECTED,
        "manifestValidation": manifest_validation,
        "manifestSidecarValidation": manifest_sidecar_validation,
        "candidateReady": candidate_ready,
        "candidateReadyReasons": candidate_reasons,
        "totalRecords": len(reviews),
        "uniqueIds": len({review["id"] for review in reviews}),
        "splitCounts": dict(split_counts),
        "statusCounts": dict(status_counts),
        "fullEvidence": full_evidence,
        "evidenceCompleteCount": full_evidence,
        "evidenceErrorCount": evidence_error_count,
        "auditErrors": audit_error_count,
        "unresolvedIds": [review["id"] for review in reviews if review["status"] == "UNRESOLVED"],
        "correctedIds": [review["id"] for review in reviews if review["status"] == "CORRECTED"],
        "verifiedIds": [review["id"] for review in reviews if review["status"] == "VERIFIED"],
        "canonicalReady": canonical_ready,
        "outputFile": output_name,
        "outputSha256": sha256_file(output_dir / output_name),
    }
    (output_dir / "audit-summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )
    return summary


def main() -> None:
    parser = argparse.ArgumentParser(description="第二版银行 NL2SQL Ground Truth 全量审查器")
    parser.add_argument("--workbook", type=Path, required=True, help="冻结的源工作簿（或 --manifest 指定时的候选工作簿）")
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--all-splits", action="store_true")
    parser.add_argument("--manifest", type=Path, default=None, help="候选 manifest（候选审查模式；验证候选哈希/源哈希/生成器契约）")
    args = parser.parse_args()
    summary = run_audit(args.workbook, args.output_dir, args.all_splits, manifest_path=args.manifest)
    print(f"GT audit v2 完成：total={summary['totalRecords']} "
          f"VERIFIED={summary['statusCounts'].get('VERIFIED', 0)} "
          f"CORRECTED={summary['statusCounts'].get('CORRECTED', 0)} "
          f"UNRESOLVED={summary['statusCounts'].get('UNRESOLVED', 0)}")
    print("splitCounts:", json.dumps(summary["splitCounts"], ensure_ascii=False, sort_keys=True))
    print(f"fullEvidence={summary['fullEvidence']} evidenceComplete={summary['evidenceCompleteCount']} evidenceErrors={summary['evidenceErrorCount']} auditErrors={summary['auditErrors']}")
    print(f"evidenceRuleVersion={summary['evidenceRuleVersion']}")
    print(f"canonicalReady={summary['canonicalReady']} output={summary['outputFile']}")
    if args.manifest is not None:
        print(f"candidateReady={summary['candidateReady']}")
        if summary["candidateReadyReasons"]:
            print("candidateReadyReasons:", " | ".join(summary["candidateReadyReasons"]))
    if summary["unresolvedIds"]:
        print("UNRESOLVED ids:", " ".join(summary["unresolvedIds"]))


if __name__ == "__main__":
    main()
