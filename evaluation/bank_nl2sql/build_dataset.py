#!/usr/bin/env python3
"""Build the DATA-02 NL2SQL annotated dataset from frozen source inputs.

The competition workbook remains the source of truth for the official
questions, their source splits, and their expected answers.  Official
records always keep the workbook ``sourceSplit`` as their ``split`` with
``splitReason`` set to ``source_assignment``; template overlap is reported
verbatim in ``manifest.templateOverlap`` and never triggers cross-split
migration.  The separately curated intent data supplies semantic
annotations only.  Curated augmentations are intentionally emitted to a
separate file, so they can never alter the official benchmark score.

An explicit ``--official-manifest`` (the validated official package manifest)
must match the workbook and the manifest-referenced change ledger: the
workbook hash/version, the ledger bytes/hash, the generator name/version, the
change counts and the removed IDs all fail closed on any mismatch.  Only
ledger ``QUESTION_CLARIFICATION`` entries may override a record's metric
contract (``normalizedIntent.metrics``/``derivedMetrics``); ``ANSWER_CORRECTION``
and ordinary records keep their intent annotations untouched.  Without the
manifest the historical strict unknown/missing-intent behavior is kept.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from collections import Counter
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


METRIC_CODE_PATTERN = re.compile(r"^ZB\d{3}$")


QUESTION_SHEET = "问题答案清单"
QUESTION_HEADERS = ("问题编号", "问题类型", "问题难度", "问题描述", "问题结果")
SOURCE_SPLITS = {"训练集": "train", "验证集": "dev", "测试集": "test"}
EVALUATION_SPLITS = ("train", "dev", "test")
INTENT_FILES = ("train.jsonl", "dev.jsonl", "test.jsonl")

SCHEMA: dict[str, Any] = {
    "$schema": "https://json-schema.org/draft/2020-12/schema",
    "title": "EcoMatch bank NL2SQL annotated sample",
    "type": "object",
    "required": [
        "id",
        "source",
        "sourceSplit",
        "split",
        "splitReason",
        "difficulty",
        "question",
        "normalizedIntent",
        "templateGroup",
        "expectedAction",
        "s2sql",
        "sql",
        "sqlFeatures",
        "expected",
        "errorCategory",
    ],
    "properties": {
        "id": {"type": "string", "minLength": 1},
        "sourceSplit": {"type": ["string", "null"], "enum": ["train", "dev", "test", None]},
        "split": {"type": "string", "enum": ["train", "dev", "test"]},
        "expectedAction": {"type": "string", "enum": ["EXECUTE", "CLARIFY", "REFUSE"]},
        "s2sql": {"type": ["string", "null"]},
        "sql": {"type": ["string", "null"]},
        "expected": {
            "type": "object",
            "required": ["answerText", "columns", "rows", "unit", "numericTolerance", "orderSensitive"],
        },
    },
}


class DatasetBuildError(ValueError):
    """Raised when a frozen DATA-02 input violates its contract."""


def _load_official_manifest(path: Path) -> dict[str, Any]:
    """Validate the official manifest structure and return it.

    The manifest was produced by ``promote_ground_truth.py`` after the full
    candidate/audit hash chain passed; this loader only enforces the fields the
    dataset builder relies on and fails closed on any structural violation.
    """
    try:
        manifest = json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError) as exc:
        raise DatasetBuildError(f"官方 manifest 无法解析: {exc}") from exc
    if not isinstance(manifest, dict):
        raise DatasetBuildError("官方 manifest 必须为对象")
    version = manifest.get("datasetVersion")
    if not isinstance(version, str) or not version:
        raise DatasetBuildError("官方 manifest 缺少 datasetVersion")
    if manifest.get("canonicalReady") is not True:
        raise DatasetBuildError("官方 manifest canonicalReady 必须为 true")
    official_count = manifest.get("officialCount")
    if not isinstance(official_count, int) or official_count <= 0:
        raise DatasetBuildError("官方 manifest officialCount 非法")
    source_splits = manifest.get("sourceSplitCounts")
    if (
        not isinstance(source_splits, dict)
        or set(source_splits) != {"train", "dev", "test"}
        or any(not isinstance(value, int) or value < 0 for value in source_splits.values())
        or sum(source_splits.values()) != official_count
    ):
        raise DatasetBuildError("官方 manifest sourceSplitCounts 非法")
    removed_ids = manifest.get("removedIds")
    if not isinstance(removed_ids, list) or any(not isinstance(value, str) or not value for value in removed_ids):
        raise DatasetBuildError("官方 manifest removedIds 非法")
    if len(set(removed_ids)) != len(removed_ids):
        raise DatasetBuildError("官方 manifest removedIds 存在重复")
    workbook_name = manifest.get("groundTruthWorkbook")
    if not isinstance(workbook_name, str) or not workbook_name:
        raise DatasetBuildError("官方 manifest groundTruthWorkbook 非法")
    generator = manifest.get("generator")
    if (
        not isinstance(generator, dict)
        or not isinstance(generator.get("name"), str)
        or not generator["name"]
        or not isinstance(generator.get("version"), str)
        or not generator["version"]
    ):
        raise DatasetBuildError("官方 manifest generator 非法")
    change_ledger = manifest.get("changeLedger")
    ledger_rel = Path(change_ledger) if isinstance(change_ledger, str) and change_ledger else None
    if ledger_rel is None or ledger_rel.is_absolute() or ".." in ledger_rel.parts:
        raise DatasetBuildError("官方 manifest changeLedger 非法")
    change_counts = manifest.get("changeCounts")
    if (
        not isinstance(change_counts, dict)
        or set(change_counts) != {"answerChanges", "questionClarifications", "questionRemovals", "contractErrors"}
        or any(not isinstance(value, int) or value < 0 for value in change_counts.values())
    ):
        raise DatasetBuildError("官方 manifest changeCounts 非法")
    artifact_hashes = manifest.get("artifactSha256")
    workbook_sha256 = (artifact_hashes or {}).get("groundTruthWorkbook")
    if not isinstance(workbook_sha256, str) or len(workbook_sha256) != 64:
        raise DatasetBuildError("官方 manifest artifactSha256.groundTruthWorkbook 非法")
    ledger_sha256 = (artifact_hashes or {}).get("changeLedger")
    if not isinstance(ledger_sha256, str) or len(ledger_sha256) != 64:
        raise DatasetBuildError("官方 manifest artifactSha256.changeLedger 非法")
    return manifest


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for block in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _load_change_ledger(manifest: dict[str, Any], manifest_dir: Path) -> dict[str, dict[str, Any]]:
    """Read and validate the manifest-referenced change ledger (fail closed).

    Returns the ledger's ``QUESTION_CLARIFICATION`` entries keyed by question
    id; only those entries may override the metric contract of a dataset
    record.  The ledger path is resolved relative to the manifest file and its
    bytes must match ``artifactSha256.changeLedger``.
    """
    ledger_path = (manifest_dir / manifest["changeLedger"]).resolve()
    if not ledger_path.is_file():
        raise DatasetBuildError(f"Change ledger does not exist: {ledger_path}")
    actual_sha = _sha256(ledger_path).upper()
    expected_sha = manifest["artifactSha256"]["changeLedger"].upper()
    if actual_sha != expected_sha:
        raise DatasetBuildError(
            f"Change ledger SHA-256 does not match official manifest ({actual_sha} != {expected_sha})"
        )
    try:
        ledger = json.loads(ledger_path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError) as exc:
        raise DatasetBuildError(f"Change ledger 无法解析: {exc}") from exc
    if not isinstance(ledger, dict):
        raise DatasetBuildError("Change ledger 必须为对象")
    generator = manifest["generator"]
    if ledger.get("generatorName") != generator["name"]:
        raise DatasetBuildError(
            f"ledger.generatorName={ledger.get('generatorName')!r} 与 manifest.generator.name={generator['name']!r} 不一致"
        )
    if ledger.get("generatorVersion") != generator["version"]:
        raise DatasetBuildError(
            f"ledger.generatorVersion={ledger.get('generatorVersion')!r} 与 manifest.generator.version={generator['version']!r} 不一致"
        )
    entries = ledger.get("entries")
    if not isinstance(entries, list):
        raise DatasetBuildError("ledger.entries 缺失或非列表")
    if ledger.get("count") != len(entries):
        raise DatasetBuildError(f"ledger.count={ledger.get('count')} != entries 条目数 {len(entries)}")
    contract_errors = ledger.get("contractErrors")
    if not isinstance(contract_errors, list) or len(contract_errors) != manifest["changeCounts"]["contractErrors"]:
        raise DatasetBuildError(
            f"ledger.contractErrors 数量 {len(contract_errors) if isinstance(contract_errors, list) else contract_errors!r} != manifest.changeCounts.contractErrors={manifest['changeCounts']['contractErrors']}"
        )
    action_counts: Counter[str] = Counter()
    seen_ids: set[str] = set()
    clarifications: dict[str, dict[str, Any]] = {}
    removed_ids: list[str] = []
    for entry in entries:
        if not isinstance(entry, dict) or not isinstance(entry.get("id"), str) or not entry["id"]:
            raise DatasetBuildError("ledger.entries 含非法条目（缺 id 或 id 非字符串）")
        entry_id = entry["id"]
        if entry_id in seen_ids:
            raise DatasetBuildError(f"ledger.entries 存在重复 ID：{entry_id}")
        seen_ids.add(entry_id)
        change_type = entry.get("changeType")
        if change_type not in ("ANSWER_CORRECTION", "QUESTION_CLARIFICATION", "QUESTION_REMOVAL"):
            raise DatasetBuildError(f"ledger.entries 含非法 changeType：{change_type!r}（{entry_id}）")
        action_counts[change_type] += 1
        if change_type == "QUESTION_CLARIFICATION":
            metric_codes = entry.get("metricCodes")
            if not isinstance(metric_codes, list) or not metric_codes:
                raise DatasetBuildError(f"{entry_id}: QUESTION_CLARIFICATION metricCodes 非法")
            clarifications[entry_id] = {"metricCodes": metric_codes}
        elif change_type == "QUESTION_REMOVAL":
            removed_ids.append(entry_id)
    for field, change_type in (
        ("answerChanges", "ANSWER_CORRECTION"),
        ("questionClarifications", "QUESTION_CLARIFICATION"),
        ("questionRemovals", "QUESTION_REMOVAL"),
    ):
        if action_counts.get(change_type, 0) != manifest["changeCounts"][field]:
            raise DatasetBuildError(
                f"ledger 动作计数 {change_type}={action_counts.get(change_type, 0)} != manifest.changeCounts.{field}={manifest['changeCounts'][field]}"
            )
    if sorted(removed_ids) != sorted(manifest["removedIds"]):
        raise DatasetBuildError(
            f"ledger QUESTION_REMOVAL IDs {sorted(removed_ids)} 与 manifest.removedIds {sorted(manifest['removedIds'])} 不一致"
        )
    return clarifications


def _as_text(value: Any, field: str, row_number: int) -> str:
    if value is None or not str(value).strip():
        raise DatasetBuildError(f"{QUESTION_SHEET} row {row_number}: {field} must not be empty")
    return str(value).strip()


def _load_workbook_questions(workbook_path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if QUESTION_SHEET not in workbook.sheetnames:
            raise DatasetBuildError(f"Missing worksheet: {QUESTION_SHEET}")
        sheet = workbook[QUESTION_SHEET]
        rows = sheet.iter_rows(values_only=True)
        header = tuple(str(value).strip() if value is not None else "" for value in next(rows, ()))
        if header[: len(QUESTION_HEADERS)] != QUESTION_HEADERS:
            raise DatasetBuildError(f"Unexpected {QUESTION_SHEET} headers: {header}")

        questions: list[dict[str, str]] = []
        seen_ids: set[str] = set()
        for row_number, row in enumerate(rows, start=2):
            if all(value is None for value in row):
                continue
            if len(row) < len(QUESTION_HEADERS):
                raise DatasetBuildError(f"{QUESTION_SHEET} row {row_number}: incomplete row")
            sample_id = _as_text(row[0], "问题编号", row_number)
            if sample_id in seen_ids:
                raise DatasetBuildError(f"Duplicate workbook question id: {sample_id}")
            question_type = _as_text(row[1], "问题类型", row_number)
            if question_type not in SOURCE_SPLITS:
                raise DatasetBuildError(f"Unknown 问题类型 at row {row_number}: {question_type}")
            seen_ids.add(sample_id)
            questions.append(
                {
                    "id": sample_id,
                    "sourceSplit": SOURCE_SPLITS[question_type],
                    "difficulty": _as_text(row[2], "问题难度", row_number),
                    "question": _as_text(row[3], "问题描述", row_number),
                    "answerText": _as_text(row[4], "问题结果", row_number),
                    "rowNumber": str(row_number),
                }
            )
        if not questions:
            raise DatasetBuildError(f"{QUESTION_SHEET} contains no questions")
        return questions
    finally:
        workbook.close()


def _read_jsonl(path: Path) -> Iterable[dict[str, Any]]:
    if not path.is_file():
        raise DatasetBuildError(f"Missing intent input: {path}")
    for line_number, line in enumerate(path.read_text(encoding="utf-8").splitlines(), start=1):
        if not line.strip():
            continue
        try:
            value = json.loads(line)
        except json.JSONDecodeError as error:
            raise DatasetBuildError(f"Invalid JSON in {path}:{line_number}") from error
        if not isinstance(value, dict):
            raise DatasetBuildError(f"Intent sample must be an object: {path}:{line_number}")
        yield value


def _load_intents(intent_root: Path) -> tuple[dict[str, dict[str, Any]], list[dict[str, Any]]]:
    official: dict[str, dict[str, Any]] = {}
    augmentations: list[dict[str, Any]] = []
    seen_ids: set[str] = set()
    for filename in INTENT_FILES:
        for record in _read_jsonl(intent_root / filename):
            sample_id = str(record.get("id", "")).strip()
            if not sample_id:
                raise DatasetBuildError(f"Intent sample in {filename} has no id")
            if sample_id in seen_ids:
                raise DatasetBuildError(f"Duplicate intent id: {sample_id}")
            seen_ids.add(sample_id)
            if record.get("source") == "competition_workbook":
                source_split = record.get("sourceSplit")
                split = record.get("split")
                if source_split not in EVALUATION_SPLITS:
                    raise DatasetBuildError(f"Official intent {sample_id} has invalid sourceSplit: {source_split}")
                if split not in EVALUATION_SPLITS:
                    raise DatasetBuildError(f"Official intent {sample_id} has invalid split: {split}")
                official[sample_id] = record
            else:
                augmentations.append(record)
    return official, augmentations


def _normalized_intent(record: dict[str, Any]) -> dict[str, Any]:
    return {
        "scene": record.get("scene"),
        "intent": record.get("intent"),
        "metrics": record.get("metrics", []),
        "dimensions": record.get("dimensions", []),
        "time": record.get("time", {}),
        "organizations": record.get("organizations", []),
        "filters": record.get("filters", []),
        "linguisticFeatures": record.get("linguisticFeatures", []),
        "clarificationExpected": bool(record.get("clarificationExpected", False)),
        "referenceDate": record.get("referenceDate"),
    }


def _expected(answer_text: str | None) -> dict[str, Any]:
    return {
        "answerText": answer_text,
        "columns": [],
        "rows": [],
        "unit": None,
        "numericTolerance": None,
        "orderSensitive": False,
    }


def _apply_clarification_contract(record: dict[str, Any], contract: dict[str, Any]) -> None:
    """Project a ledger QUESTION_CLARIFICATION metric contract onto a record.

    Only the ledger-declared base metric codes become ``normalizedIntent.metrics``
    (in ledger order) and derived ratio specs become
    ``normalizedIntent.derivedMetrics`` (in ledger order).  The ledger is the
    only source of truth: nothing is guessed from the question text and no
    default metric may leak in.  Every string base code must be a valid
    project metric code (``ZB###`` format) and duplicates are rejected.
    Derived numerator/denominator are the
    underlying data-source metric codes: each must be a distinct valid
    project metric code (``ZB###`` format), but they need not be declared
    among the output base metrics.
    """
    sample_id = record["id"]
    metric_codes: list[str] = []
    derived_specs: list[dict[str, Any]] = []
    for item in contract["metricCodes"]:
        if isinstance(item, str):
            if METRIC_CODE_PATTERN.fullmatch(item) is None:
                raise DatasetBuildError(
                    f"{sample_id}: QUESTION_CLARIFICATION metricCodes 基础指标 {item!r} 不是合法指标代码（应为 ZB### 格式）"
                )
            metric_codes.append(item)
            continue
        if isinstance(item, dict) and isinstance(item.get("derived"), dict):
            derived_specs.append(item)
            continue
        raise DatasetBuildError(f"{sample_id}: QUESTION_CLARIFICATION metricCodes 含非法条目：{item!r}")
    if not metric_codes:
        raise DatasetBuildError(f"{sample_id}: QUESTION_CLARIFICATION 未声明任何基础指标")
    if len(set(metric_codes)) != len(metric_codes):
        raise DatasetBuildError(f"{sample_id}: QUESTION_CLARIFICATION 基础指标重复")
    derived_metrics: list[dict[str, Any]] = []
    for item in derived_specs:
        derived = item["derived"]
        numerator = derived.get("numerator")
        denominator = derived.get("denominator")
        if not isinstance(numerator, str) or not METRIC_CODE_PATTERN.fullmatch(numerator):
            raise DatasetBuildError(
                f"{sample_id}: derived numerator {numerator!r} 不是合法指标代码（应为 ZB### 格式）"
            )
        if not isinstance(denominator, str) or not METRIC_CODE_PATTERN.fullmatch(denominator):
            raise DatasetBuildError(
                f"{sample_id}: derived denominator {denominator!r} 不是合法指标代码（应为 ZB### 格式）"
            )
        if numerator == denominator:
            raise DatasetBuildError(f"{sample_id}: derived numerator 与 denominator 必须不同：{numerator!r}")
        derived_metrics.append(
            {
                "metricCode": f"DERIVED_{numerator}_DIV_{denominator}",
                "numerator": numerator,
                "denominator": denominator,
                "name": item.get("name"),
            }
        )
    normalized = record["normalizedIntent"]
    normalized["metrics"] = [{"code": code} for code in metric_codes]
    normalized["derivedMetrics"] = derived_metrics


def _official_record(question: dict[str, str], intent: dict[str, Any], workbook_name: str) -> dict[str, Any]:
    """Build an official record frozen to the workbook's source split.

    The workbook is the sole source of truth for the official split
    assignment: ``split`` always equals ``sourceSplit`` and ``splitReason``
    is always ``source_assignment``.  Template overlap never migrates an
    official record across splits.
    """
    sample_id = question["id"]
    if intent.get("sourceSplit") != question["sourceSplit"]:
        raise DatasetBuildError(
            f"{sample_id}: workbook source split {question['sourceSplit']} does not match intent sourceSplit {intent.get('sourceSplit')}"
        )
    return {
        "id": sample_id,
        "source": {
            "kind": "competition_workbook",
            "workbook": workbook_name,
            "rowNumber": int(question["rowNumber"]),
            "intentDataset": "bank_intent",
        },
        "sourceSplit": question["sourceSplit"],
        "split": question["sourceSplit"],
        "splitReason": "source_assignment",
        "difficulty": question["difficulty"],
        "question": question["question"],
        "normalizedIntent": _normalized_intent(intent),
        "templateGroup": intent.get("templateGroup"),
        "expectedAction": "EXECUTE",
        "s2sql": None,
        "sql": None,
        "sqlFeatures": [],
        "expected": _expected(question["answerText"]),
        "errorCategory": None,
    }


def _augmentation_record(intent: dict[str, Any]) -> dict[str, Any]:
    sample_id = str(intent["id"])
    split = intent.get("split")
    if split not in EVALUATION_SPLITS:
        raise DatasetBuildError(f"Augmentation {sample_id} has invalid split: {split}")
    clarify = bool(intent.get("clarificationExpected", False))
    return {
        "id": sample_id,
        "source": {"kind": intent.get("source"), "intentDataset": "bank_intent"},
        "sourceSplit": None,
        "split": split,
        "splitReason": "augmentation",
        "difficulty": intent.get("difficulty"),
        "question": intent.get("question"),
        "normalizedIntent": _normalized_intent(intent),
        "templateGroup": intent.get("templateGroup"),
        "expectedAction": "CLARIFY" if clarify else "EXECUTE",
        "s2sql": None,
        "sql": None,
        "sqlFeatures": [],
        "expected": _expected(None if clarify else intent.get("answer")),
        "errorCategory": "AMBIGUOUS_REQUEST" if clarify else None,
    }


def _write_jsonl(path: Path, records: Iterable[dict[str, Any]]) -> None:
    payload = "".join(json.dumps(record, ensure_ascii=False, sort_keys=True) + "\n" for record in records)
    path.write_text(payload, encoding="utf-8")


def _template_overlap(records_by_split: dict[str, list[dict[str, Any]]]) -> dict[str, list[str]]:
    groups = {
        split: {record["templateGroup"] for record in records if record.get("templateGroup")}
        for split, records in records_by_split.items()
    }
    return {
        "trainDev": sorted(groups["train"] & groups["dev"]),
        "trainTest": sorted(groups["train"] & groups["test"]),
        "devTest": sorted(groups["dev"] & groups["test"]),
    }


def build_dataset(
    workbook_path: Path | str,
    intent_root: Path | str,
    output_path: Path | str,
    official_manifest_path: Path | str | None = None,
) -> dict[str, Any]:
    """Create official and augmentation JSONL files and return their manifest.

    With an official manifest, only its ledger-declared ``removedIds`` are
    exempted from the strict unknown-intent check; the workbook hash, file
    name, question count and source split counts must match the manifest and
    removed IDs must be absent from the workbook.  Official records never
    migrate: ``split`` is always the workbook ``sourceSplit``, so
    ``reassignedForTemplateIsolation`` is always empty.
    """

    workbook_path = Path(workbook_path).resolve()
    intent_root = Path(intent_root).resolve()
    output_path = Path(output_path).resolve()
    if not workbook_path.is_file():
        raise DatasetBuildError(f"Workbook does not exist: {workbook_path}")
    official_manifest: dict[str, Any] | None = None
    if official_manifest_path is not None:
        official_manifest_path = Path(official_manifest_path).resolve()
        if not official_manifest_path.is_file():
            raise DatasetBuildError(f"Official manifest does not exist: {official_manifest_path}")
        official_manifest = _load_official_manifest(official_manifest_path)
        expected_workbook = official_manifest["groundTruthWorkbook"]
        if workbook_path.name != expected_workbook:
            raise DatasetBuildError(
                f"Workbook name {workbook_path.name!r} does not match official manifest {expected_workbook!r}"
            )
        actual_sha = _sha256(workbook_path).upper()
        expected_sha = official_manifest["artifactSha256"]["groundTruthWorkbook"].upper()
        if actual_sha != expected_sha:
            raise DatasetBuildError(f"Workbook SHA-256 does not match official manifest ({actual_sha} != {expected_sha})")
    questions = _load_workbook_questions(workbook_path)
    official_intents, raw_augmentations = _load_intents(intent_root)

    question_ids = {question["id"] for question in questions}
    removed_ids: set[str] = set()
    clarification_contracts: dict[str, dict[str, Any]] = {}
    if official_manifest is not None:
        removed_ids = set(official_manifest["removedIds"])
        if official_manifest["officialCount"] != len(questions):
            raise DatasetBuildError(
                f"Workbook question count {len(questions)} does not match official manifest officialCount {official_manifest['officialCount']}"
            )
        actual_source_splits = {split: sum(1 for q in questions if q["sourceSplit"] == split) for split in EVALUATION_SPLITS}
        if actual_source_splits != official_manifest["sourceSplitCounts"]:
            raise DatasetBuildError(
                f"Workbook source split counts {actual_source_splits} do not match official manifest {official_manifest['sourceSplitCounts']}"
            )
        still_present = removed_ids & question_ids
        if still_present:
            raise DatasetBuildError(f"Ledger-removed IDs still present in workbook: {sorted(still_present)}")
        missing_removed_intents = removed_ids - set(official_intents)
        if missing_removed_intents:
            raise DatasetBuildError(f"Ledger-removed IDs absent from intents: {sorted(missing_removed_intents)}")
        clarification_contracts = _load_change_ledger(official_manifest, official_manifest_path.parent)
        missing_clarifications = set(clarification_contracts) - question_ids
        if missing_clarifications:
            raise DatasetBuildError(
                f"Ledger clarification IDs absent from workbook: {sorted(missing_clarifications)}"
            )
    unknown_intents = sorted(set(official_intents) - question_ids - removed_ids)
    missing_intents = sorted(question_ids - set(official_intents))
    if unknown_intents:
        raise DatasetBuildError(f"Official intents absent from workbook: {unknown_intents}")
    if missing_intents:
        raise DatasetBuildError(f"Workbook questions absent from intents: {missing_intents}")

    records_by_split: dict[str, list[dict[str, Any]]] = {split: [] for split in EVALUATION_SPLITS}
    for question in questions:
        record = _official_record(question, official_intents[question["id"]], workbook_path.name)
        contract = clarification_contracts.get(question["id"])
        if contract is not None:
            _apply_clarification_contract(record, contract)
        records_by_split[record["split"]].append(record)
    for records in records_by_split.values():
        records.sort(key=lambda item: item["id"])

    augmentations = sorted((_augmentation_record(record) for record in raw_augmentations), key=lambda item: item["id"])
    source_split_counts = Counter(question["sourceSplit"] for question in questions)
    evaluation_split_counts = {split: len(records_by_split[split]) for split in EVALUATION_SPLITS}
    manifest = {
        "version": official_manifest["datasetVersion"] if official_manifest is not None else "0.1.0",
        "sourceWorkbook": workbook_path.name,
        "sourceSha256": _sha256(workbook_path),
        "officialCount": len(questions),
        "augmentationCount": len(augmentations),
        "sourceSplitCounts": {split: source_split_counts[split] for split in EVALUATION_SPLITS},
        "evaluationSplitCounts": evaluation_split_counts,
        # 官方切分不可变：workbook sourceSplit 是唯一事实，不存在 template 迁移。
        "reassignedForTemplateIsolation": [],
        "templateOverlap": _template_overlap(records_by_split),
    }

    output_path.mkdir(parents=True, exist_ok=True)
    for split in EVALUATION_SPLITS:
        _write_jsonl(output_path / f"{split}.jsonl", records_by_split[split])
    _write_jsonl(output_path / "augmentation.jsonl", augmentations)
    (output_path / "manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )
    (output_path / "schema.json").write_text(
        json.dumps(SCHEMA, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )
    return manifest


def main() -> None:
    parser = argparse.ArgumentParser(description="Build the DATA-02 bank NL2SQL dataset")
    parser.add_argument("workbook", type=Path, help="Competition workbook containing 问题答案清单")
    parser.add_argument("--intent-root", type=Path, required=True, help="Directory containing bank_intent JSONL files")
    parser.add_argument(
        "--official-manifest",
        type=Path,
        default=None,
        help="Validated official package manifest authorizing only ledger-declared removed intent IDs",
    )
    parser.add_argument("--output", type=Path, required=True, help="Output directory for the annotated dataset")
    args = parser.parse_args()
    print(
        json.dumps(
            build_dataset(args.workbook, args.intent_root, args.output, args.official_manifest),
            ensure_ascii=False,
            sort_keys=True,
        )
    )


if __name__ == "__main__":
    main()
