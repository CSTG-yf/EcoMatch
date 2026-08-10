#!/usr/bin/env python3
"""数据驱动的候选修正版 Ground Truth 生成器（只读消费冻结原始工作簿）。

输入：
- --source：冻结原始工作簿（只读，永不覆盖）；
- --review：基线审查的 review.ndjson（逐条状态/类别/证据记录）；
- --exclude-id：用户授权删除的训练集问题编号（可重复；真实编号只在调用时
  提供，禁止硬编码进代码或测试）。

输出（.local-dev/gt-audit/contract-fix/）：
- candidate-corrected-v1.xlsx  仅 10 条 questionText（PERFORMANCE/DIMENSION_3
  题目契约澄清）、5 条 answerText（证据完整的答案修正）与被授权删除的
  1 条问题行（整行移除、其余行按原顺序压实）与原工作簿不同；
- contract-change-ledger.json  逐条变更账本（类型、新旧文本 SHA-256、
  提取的指标代码/维度映射、依据、审查证据引用；删除为 QUESTION_REMOVAL）；
- candidate-manifest.json      源/候选 SHA-256、生成器版本、变更计数
  （含 questionRemovals）、totalRecords/split 计数、expectedAudit
  （候选审查的预期 total/split/status/证据完整性，candidateReady 初始
  false，由候选审查对照 expectedAudit 写回）；
- <文件>.sha256                小型哈希文件。

选择规则（全部数据驱动，禁止按真实 ID 硬编码）：
- 答案修正：review 记录 status=CORRECTED 且 fullEvidence=true 且
  evidenceValidation.valid=true 且 correctedAnswerText 非空；
- 题目澄清：review 记录 status=UNRESOLVED 且 category 为 PERFORMANCE 或
  DIMENSION_3；指标集合/维度映射只从原 answerText 中“被明确命名的指标”
  提取（恢复官方出题意图），答案中的数值绝不参与提取或成为正确性证据；
- 无法唯一提取指标集合/映射的题拒绝修改并记录 contractError（诚实失败）；
- 题目删除（--exclude-id，全部条件同时满足才接受，任一不满足 fail closed）：
  ID 存在于源审查、属于训练集、status=UNRESOLVED 且 fullEvidence=false、
  未被选为答案修正或题目澄清、且未决证据表明缺少 yoy 基期/当前事实
  （同比基期数据缺失导致问题无法复核）；缺失 ID、VERIFIED/CORRECTED、
  dev/test、可澄清歧义、重复请求或无关未决原因一律拒绝。

终端只打印计数与 ID；题目、答案、事实行不打印。
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
import zipfile
from collections import Counter
from datetime import datetime
from itertools import zip_longest
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

import gt_answer_rules as rules

GENERATOR_NAME = "clarify_ground_truth_contracts"
GENERATOR_VERSION = "2.0.0"
CANDIDATE_WORKBOOK = "candidate-corrected-v1.xlsx"
CHANGE_LEDGER = "contract-change-ledger.json"
CANDIDATE_MANIFEST = "candidate-manifest.json"

QUESTION_HEADERS = rules.QUESTION_HEADERS  # (问题编号, 问题类型, 问题难度, 问题描述, 问题结果)
QID_COL, QTYPE_COL, QDIFF_COL, QQUESTION_COL, QANSWER_COL = 1, 2, 3, 4, 5

# DIMENSION_3 旧答案的段标记 -> 澄清后问题中的维度名
_DIMENSION_SEGMENTS = (("规模：", "规模"), ("质量：", "资产质量"), ("效益：", "盈利能力"))


def sha256_text(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest().upper()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for block in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest().upper()


_FIXED_CORE_TIMESTAMP = b"2000-01-01T00:00:00Z"
# openpyxl 保存时强制把 workbook.properties.modified 覆盖为当前 UTC 时间
# （openpyxl/writer/excel.py save_workbook），导致 docProps/core.xml 的
# <dcterms:modified> 每次运行不同 —— 重写 zip 时一并固定，保证字节级确定性。
_CORE_MODIFIED_RE = re.compile(rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)")


def _fix_zip_timestamps(path: Path) -> None:
    """重写 xlsx zip 条目时间戳为固定值，保证跨秒运行输出字节一致（确定性）。"""
    fixed_ts = (2000, 1, 1, 0, 0, 0)
    tmp_path = path.with_name(path.name + ".fixed")
    try:
        with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(tmp_path, "w", compression=zipfile.ZIP_DEFLATED) as target:
            for info in source.infolist():
                new_info = zipfile.ZipInfo(info.filename, date_time=fixed_ts)
                new_info.compress_type = info.compress_type
                new_info.external_attr = info.external_attr
                data = source.read(info.filename)
                if info.filename == "docProps/core.xml":
                    data = _CORE_MODIFIED_RE.sub(lambda m: m.group(1) + _FIXED_CORE_TIMESTAMP + m.group(2), data)
                target.writestr(new_info, data)
        tmp_path.replace(path)
    finally:
        if tmp_path.exists():
            tmp_path.unlink()


# --------------------------------------------------------------------------- 选择门控


def select_answer_corrections(reviews: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """按证据门控选择答案修正：status=CORRECTED 且 fullEvidence=true 且
    evidenceValidation.valid=true 且 correctedAnswerText 非空。"""
    return [
        review
        for review in reviews
        if review["status"] == "CORRECTED"
        and review.get("fullEvidence") is True
        and (review.get("evidenceValidation") or {}).get("valid") is True
        and review.get("correctedAnswerText")
    ]


def select_clarifications(reviews: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """选择题目契约澄清：status=UNRESOLVED 且 category 为 PERFORMANCE 或 DIMENSION_3。"""
    return [
        review
        for review in reviews
        if review["status"] == "UNRESOLVED" and review.get("category") in ("PERFORMANCE", "DIMENSION_3")
    ]


# --------------------------------------------------------------------------- 题目删除门控（--exclude-id，fail closed）

# 缺失/缺失事实信号：未决原因必须明确指向“数据缺失”，而不是“无法确定/有歧义”
_MISSING_FACT_SIGNALS = ("缺失", "缺少", "不存在", "没有", "未提供", "无法获取")


def is_missing_yoy_baseline_fact(reason: str) -> bool:
    """未决原因是否明确表明“缺少 yoy 基期/当前事实”：必须同时包含
    缺失/缺少类事实信号、yoy（同比）上下文与基期/当前值上下文。
    只表达歧义或“基期无法确定/不确定”的未决原因（无缺失信号）一律拒绝。"""
    return (
        "yoy" in reason
        and any(signal in reason for signal in _MISSING_FACT_SIGNALS)
        and ("基期" in reason or "当前值" in reason or "当前事实" in reason)
    )


def select_removals(
    reviews: list[dict[str, Any]],
    exclude_ids: list[str],
    corrections: list[dict[str, Any]],
    clarifications: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """校验 --exclude-id 请求并返回被授权删除的 review 记录。

    全部条件同时满足才接受（数据驱动，不按真实 ID 硬编码）：
    - ID 存在于源审查记录中；
    - split=train；
    - status=UNRESOLVED 且 fullEvidence=false；
    - 未被选为答案修正或题目澄清；
    - unresolvedReason 表明缺少 yoy 基期/当前事实。
    任一不满足（含重复请求）即 fail closed，返回带错误说明的条目，
    调用方必须中止产出，不写任何候选文件。
    """
    by_id = {review["id"]: review for review in reviews}
    correction_ids = {item["id"] for item in corrections}
    clarification_ids = {item["id"] for item in clarifications}
    removals: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    seen: set[str] = set()
    for qid in exclude_ids:
        if qid in seen:
            errors.append({"id": qid, "error": "重复的排除请求（同一 ID 只能请求一次）"})
            continue
        seen.add(qid)
        review = by_id.get(qid)
        if review is None:
            errors.append({"id": qid, "error": "请求的 ID 不存在于源审查记录中"})
            continue
        if review["split"] != "train":
            errors.append({"id": qid, "error": f"仅允许删除训练集记录（实际 split={review['split']}）"})
            continue
        if review["status"] != "UNRESOLVED" or review.get("fullEvidence") is not False:
            errors.append(
                {
                    "id": qid,
                    "error": f"仅允许删除 UNRESOLVED 且 fullEvidence=false 的记录（实际 status={review['status']} fullEvidence={review.get('fullEvidence')}）",
                }
            )
            continue
        if qid in correction_ids or qid in clarification_ids:
            errors.append({"id": qid, "error": "该记录已被选为答案修正或题目澄清，不得同时删除"})
            continue
        reason = review.get("unresolvedReason") or ""
        if not is_missing_yoy_baseline_fact(reason):
            errors.append({"id": qid, "error": f"未决原因不是缺少 yoy 基期/当前事实（{reason}）"})
            continue
        removals.append(review)
    return removals, errors


# --------------------------------------------------------------------------- 指标提取（只读答案文本中的“明确命名的指标”）

def extract_metric_hits(text: str) -> list[Any]:
    """按文本首次出现顺序去重的指标命中（含衍生占位）；数值不参与提取。"""
    seen: set[str] = set()
    deduped: list[Any] = []
    for _position, hit in rules.extract_metrics_ordered(text):
        identity = hit.code if hit.code is not None else hit.matched_text
        if identity in seen:
            continue
        seen.add(identity)
        deduped.append(hit)
    return deduped


def metric_identity(hit: Any) -> str:
    return hit.code if hit.code is not None else hit.matched_text


def metric_code_repr(hit: Any) -> Any:
    """账本用结构化指标表示：基础指标为编号；衍生指标为 {numerator, denominator}。"""
    if hit.derived is not None:
        return {"derived": {"numerator": hit.derived[0], "denominator": hit.derived[1]}, "name": hit.matched_text}
    return hit.code


def extract_performance_metrics(answer_text: str) -> list[Any]:
    """PERFORMANCE：原答案明确命名的指标去重集合（空列表 = 无法唯一提取）。"""
    return extract_metric_hits(answer_text)


def clarify_performance_question(question: str, metric_hits: list[Any]) -> str:
    names = [hit.matched_text for hit in metric_hits]
    return (
        f"{question}待评价指标集合：{'、'.join(names)}。"
        "判定规则：表现较好=全省排名前三，表现较差=全省排名后四；"
        "排名方向由指标定义决定（不良贷款率、逾期贷款率、成本收入比越低越好，其余越高越好）。"
    )


def extract_dimension_mapping(answer_text: str) -> dict[str, list[Any]] | None:
    """DIMENSION_3：按 规模/质量/效益 段提取指标映射。

    返回 {维度: [MetricHit]}；任一维度缺指标、段标记缺失或指标跨维度重复时
    返回 None（拒绝修改，不猜测）。
    """
    positions: list[tuple[str, str]] = []
    for marker, dim in _DIMENSION_SEGMENTS:
        index = answer_text.find(marker)
        if index < 0:
            return None
        positions.append((marker, dim, index))
    mapping: dict[str, list[Any]] = {}
    for i, (marker, dim, start) in enumerate(positions):
        end = positions[i + 1][2] if i + 1 < len(positions) else len(answer_text)
        segment = answer_text[start + len(marker):end]
        hits = extract_metric_hits(segment)
        if not hits:
            return None
        mapping[dim] = hits
    seen: set[str] = set()
    for hits in mapping.values():
        for hit in hits:
            identity = metric_identity(hit)
            if identity in seen:
                return None
            seen.add(identity)
    return mapping


def clarify_dimension_question(question: str, mapping: dict[str, list[Any]]) -> str:
    parts = [f"{dim}={'、'.join(hit.matched_text for hit in hits)}" for dim, hits in mapping.items()]
    return f"{question}维度与指标映射：{'；'.join(parts)}。"


# --------------------------------------------------------------------------- 澄清契约生成

def generate_clarification_contracts(
    reviews: list[dict[str, Any]], review_path: Path
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """从 UNRESOLVED 的 PERFORMANCE/DIMENSION_3 review 生成完整澄清契约。

    返回 (clarified, contract_errors)：无法唯一提取指标集合/映射的题拒绝修改
    并记录 contractError（诚实失败）。生成的契约含 clarifiedQuestion 等完整
    字段，可直接用于候选校验/账本（原始 review 记录没有这些字段）。
    """
    clarifications_raw = select_clarifications(reviews)
    clarified: list[dict[str, Any]] = []
    contract_errors: list[dict[str, Any]] = []
    for review in clarifications_raw:
        if review["category"] == "PERFORMANCE":
            hits = extract_performance_metrics(review["answerText"])
            if not hits:
                contract_errors.append({"id": review["id"], "category": review["category"], "error": "无法从原答案唯一提取指标集合"})
                continue
            clarified.append(
                {
                    "id": review["id"],
                    "split": review["split"],
                    "difficulty": review["difficulty"],
                    "category": review["category"],
                    "question": review["question"],
                    "answerText": review["answerText"],
                    "clarifiedQuestion": clarify_performance_question(review["question"], hits),
                    "metricCodes": [metric_code_repr(hit) for hit in hits],
                    "dimensionMapping": None,
                    "basis": "原答案明确命名的指标去重集合（数值不参与提取）",
                    "evidenceRef": {
                        "reviewFile": str(review_path),
                        "status": review["status"],
                        "fullEvidence": review.get("fullEvidence"),
                        "category": review["category"],
                    },
                }
            )
        else:  # DIMENSION_3
            mapping = extract_dimension_mapping(review["answerText"])
            if mapping is None:
                contract_errors.append({"id": review["id"], "category": review["category"], "error": "无法从原答案唯一提取维度指标映射"})
                continue
            clarified.append(
                {
                    "id": review["id"],
                    "split": review["split"],
                    "difficulty": review["difficulty"],
                    "category": review["category"],
                    "question": review["question"],
                    "answerText": review["answerText"],
                    "clarifiedQuestion": clarify_dimension_question(review["question"], mapping),
                    "metricCodes": [metric_code_repr(hit) for hits in mapping.values() for hit in hits],
                    "dimensionMapping": {dim: [metric_code_repr(hit) for hit in hits] for dim, hits in mapping.items()},
                    "basis": "原答案 规模/质量/效益 段明确命名的指标（每维度≥1 且映射唯一；数值不参与提取）",
                    "evidenceRef": {
                        "reviewFile": str(review_path),
                        "status": review["status"],
                        "fullEvidence": review.get("fullEvidence"),
                        "category": review["category"],
                    },
                }
            )
    return clarified, contract_errors


# --------------------------------------------------------------------------- 候选工作簿

def load_question_rows(workbook_path: Path) -> dict[str, tuple[int, str, str, str]]:
    """读问题答案清单：{问题编号: (行号, 问题类型, 问题难度, 问题文本)}（值原样，不 strip 编号以外的列）。"""
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheet = workbook[rules.SHEET_QUESTION]
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    actual = tuple(str(value).strip() if value is not None else "" for value in rows[0])
    if actual[: len(QUESTION_HEADERS)] != QUESTION_HEADERS:
        raise ValueError(f"{rules.SHEET_QUESTION} 表头不匹配: {actual!r}")
    result: dict[str, tuple[int, str, str, str]] = {}
    for index, row in enumerate(rows[1:], start=2):
        if row[0] is None:
            continue
        qid = str(row[0]).strip()
        question_type = str(row[1]).strip()
        difficulty = str(row[2]).strip()
        question = str(row[3]) if row[3] is not None else ""
        result[qid] = (index, question_type, difficulty, question)
    return result


def build_candidate_workbook(
    source_path: Path,
    out_path: Path,
    corrections: list[dict[str, Any]],
    clarifications: list[dict[str, Any]],
    removals: list[dict[str, Any]],
) -> None:
    """拷贝源工作簿全部单元格值；问题表删除被授权排除的行（其余行按原顺序
    压实，保持相对顺序），仅改写 10 条 questionText 与 5 条 answerText。
    改写按 ID 定位新表行号（删除导致坐标偏移，不得复用源坐标）。"""
    removed_ids = {review["id"] for review in removals}
    source = load_workbook(source_path, read_only=True, data_only=True)
    target = Workbook()
    target.properties.created = datetime(2000, 1, 1, 0, 0, 0)
    target.properties.modified = datetime(2000, 1, 1, 0, 0, 0)
    target.remove(target.active)
    new_qid_to_row: dict[str, int] = {}
    for sheet in source.worksheets:
        sheet_name = str(sheet.title)
        ws = target.create_sheet(sheet_name)
        if sheet_name == rules.SHEET_QUESTION:
            target_row = 1
            for row in sheet.iter_rows():
                if row[0].value is not None and str(row[0].value).strip() in removed_ids:
                    continue
                for cell in row:
                    if cell.value is not None:
                        ws.cell(row=target_row, column=cell.column, value=cell.value)
                if row[0].value is not None:
                    new_qid_to_row[str(row[0].value).strip()] = target_row
                target_row += 1
        else:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        ws.cell(row=cell.row, column=cell.column, value=cell.value)
    source.close()

    qsheet = target[rules.SHEET_QUESTION]
    for correction in corrections:
        row_number = new_qid_to_row[correction["id"]]
        qsheet.cell(row=row_number, column=QANSWER_COL, value=correction["correctedAnswerText"])
    for clarification in clarifications:
        row_number = new_qid_to_row[clarification["id"]]
        qsheet.cell(row=row_number, column=QQUESTION_COL, value=clarification["clarifiedQuestion"])
    target.save(out_path)
    _fix_zip_timestamps(out_path)


def fact_region_digest(workbook_path: Path) -> str:
    """指标数据表数据区域（行序稳定序列化）的 SHA-256，作为源事实区域哈希。

    有界内存流式：逐行更新摘要，不缓存整表（13 万行级别事实表也保持 O(1)
    内存）；序列化字节与旧版先拼完整 payload 再一次性哈希完全一致（行间
    以 \n 分隔、无尾部换行）。
    """
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheet = workbook[rules.SHEET_FACT]
        digest = hashlib.sha256()
        first = True
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if all(value is None for value in row):
                continue
            line = "|".join("" if value is None else str(value) for value in row)
            if not first:
                digest.update(b"\n")
            digest.update(line.encode("utf-8"))
            first = False
    finally:
        workbook.close()
    return digest.hexdigest().upper()


def verify_candidate_workbook(
    source_path: Path,
    candidate_path: Path,
    corrections: list[dict[str, Any]],
    clarifications: list[dict[str, Any]],
    removals: list[dict[str, Any]],
) -> tuple[bool, list[str]]:
    """机器校验（全部数据驱动，不按真实 ID 硬编码）：
    - 非问题表逐格一致（行未被删除，坐标 diff 语义正确），事实区域另做哈希；
    - 问题表按 ID/字段契约校验（行删除导致坐标偏移，坐标 diff 会产生误导，
      因此改为：被授权删除的 ID 必须恰好消失，其余 ID 的顺序与全部字段必须
      与源一致，仅允许 5 条 answerText 与 10 条 questionText 变更且值必须
      等于账本预期值；任何额外删除、重排、重复 ID、未授权字段变更都会被捕获）；
    - split 计数 = 源计数减去被删除记录按 split 的计数。
    单遍流式：全部走 iter_rows，零 Worksheet.cell 随机访问。"""
    errors: list[str] = []
    removed_ids = {review["id"] for review in removals}
    # 预期变更契约：{qid: {列号: 预期值}}（答案修正写问题结果列，澄清写问题描述列）
    changed_contract: dict[str, dict[int, Any]] = {}
    for correction in corrections:
        changed_contract.setdefault(correction["id"], {})[QANSWER_COL] = correction["correctedAnswerText"]
    for clarification in clarifications:
        changed_contract.setdefault(clarification["id"], {})[QQUESTION_COL] = clarification["clarifiedQuestion"]
    source = load_workbook(source_path, read_only=True, data_only=True)
    candidate = load_workbook(candidate_path, read_only=True, data_only=True)
    try:
        if [str(name) for name in source.sheetnames] != [str(name) for name in candidate.sheetnames]:
            errors.append(f"sheet 名称/顺序不一致：{source.sheetnames} != {candidate.sheetnames}")
        source_contract: dict[str, list[Any]] = {}
        source_order: list[str] = []
        for sname in source.sheetnames:
            s_sheet, c_sheet = source[str(sname)], candidate[str(sname)]
            if str(sname) == rules.SHEET_QUESTION:
                source_header = next(s_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
                candidate_header = next(c_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
                if source_header != candidate_header:
                    errors.append(f"{sname} 表头不一致")
                for row in s_sheet.iter_rows(min_row=2, values_only=True):
                    if row[0] is None:
                        continue
                    qid = str(row[0]).strip()
                    source_contract[qid] = list(row)
                    source_order.append(qid)
                candidate_rows: list[tuple[str, list[Any]]] = []
                candidate_splits: dict[str, int] = {}
                for row in c_sheet.iter_rows(min_row=2, values_only=True):
                    if row[0] is None:
                        continue
                    qid = str(row[0]).strip()
                    candidate_rows.append((qid, list(row)))
                    split = rules.SPLIT_MAP.get(str(row[1]).strip())
                    if split is None:
                        errors.append(f"候选非法问题类型：{row[1]!r}（{qid}）")
                    else:
                        candidate_splits[split] = candidate_splits.get(split, 0) + 1
                expected_order = [qid for qid in source_order if qid not in removed_ids]
                actual_order = [qid for qid, _fields in candidate_rows]
                if actual_order != expected_order:
                    errors.append(
                        f"{sname} ID 顺序/集合与源不一致（被授权删除 {sorted(removed_ids)} 之外不得增删、重排或重复）"
                    )
                if len(actual_order) != len(set(actual_order)):
                    errors.append(f"{sname} 候选问题编号重复")
                missing_removed = removed_ids - set(source_order)
                if missing_removed:
                    errors.append(f"被授权删除的 ID 不在源工作簿中：{sorted(missing_removed)}")
                still_present = removed_ids & set(actual_order)
                if still_present:
                    errors.append(f"被授权删除的 ID 仍存在于候选：{sorted(still_present)}")
                candidate_by_id = {qid: fields for qid, fields in candidate_rows}
                diffs: list[tuple[str, int]] = []
                for qid, fields in candidate_rows:
                    for col_index, value in enumerate(fields, start=1):
                        if value != source_contract[qid][col_index - 1]:
                            diffs.append((qid, col_index))
                expected_changes = len(corrections) + len(clarifications)
                if len(diffs) != expected_changes:
                    errors.append(f"{sname} 变更计数不符：实际 {len(diffs)} != 预期 {expected_changes}")
                for qid, col_index in diffs:
                    if col_index not in changed_contract.get(qid, {}):
                        errors.append(f"{sname} 未授权字段变更：{qid} 第{col_index}列")
                for qid, changes in changed_contract.items():
                    fields = candidate_by_id.get(qid)
                    if fields is None:
                        errors.append(f"{sname} 预期变更的 ID 不在候选：{qid}")
                        continue
                    for col_index, expected_value in changes.items():
                        if fields[col_index - 1] != expected_value:
                            errors.append(f"{sname} 预期变更未生效：{qid} 第{col_index}列")
                # split 计数：源中未删除记录按各自类型映射（数据驱动）
                expected_split_counts = {"train": 0, "dev": 0, "test": 0}
                for qid in expected_order:
                    split = rules.SPLIT_MAP.get(str(source_contract[qid][1]).strip())
                    if split is None:
                        errors.append(f"源非法问题类型：{source_contract[qid][1]!r}（{qid}）")
                    else:
                        expected_split_counts[split] = expected_split_counts.get(split, 0) + 1
                if candidate_splits != expected_split_counts:
                    errors.append(f"{sname} split 计数不符：{candidate_splits} != {expected_split_counts}")
            else:
                # 非问题表：行未被删除，语义相同则坐标必须逐格一致
                if s_sheet.max_row != c_sheet.max_row or s_sheet.max_column != c_sheet.max_column:
                    errors.append(f"{sname} 尺寸不一致：({s_sheet.max_row}x{s_sheet.max_column}) != ({c_sheet.max_row}x{c_sheet.max_column})")
                    continue
                non_question_diffs: list[tuple[int, int]] = []
                for row_index, (s_row, c_row) in enumerate(
                    zip_longest(s_sheet.iter_rows(values_only=True), c_sheet.iter_rows(values_only=True), fillvalue=()),
                    start=1,
                ):
                    for col_index, (sv, cv) in enumerate(zip_longest(s_row, c_row, fillvalue=None), start=1):
                        if sv != cv:
                            non_question_diffs.append((row_index, col_index))
                if non_question_diffs:
                    errors.append(f"{sname} 非问题表发生 {len(non_question_diffs)} 处单元格变更：{non_question_diffs[:5]}")
        if fact_region_digest(source_path) != fact_region_digest(candidate_path):
            errors.append("指标数据表区域与源不一致")
    finally:
        source.close()
        candidate.close()
    return not errors, errors


# --------------------------------------------------------------------------- 主流程

def run_generator(
    source_path: Path,
    review_path: Path,
    output_dir: Path,
    expected_source_sha: str = rules.SOURCE_SHA256_EXPECTED,
    exclude_ids: list[str] | None = None,
) -> int:
    exclude_ids = list(exclude_ids or [])
    source_sha = sha256_file(source_path)
    if source_sha != expected_source_sha:
        print(f"错误：源工作簿哈希不匹配 {source_sha} != {expected_source_sha}")
        return 2
    lines = [line for line in review_path.read_text(encoding="utf-8").splitlines() if line.strip()]
    reviews = [json.loads(line) for line in lines]
    corrections = select_answer_corrections(reviews)
    clarified, contract_errors = generate_clarification_contracts(reviews, review_path)

    # 删除门控：任一排除请求未通过资格校验即 fail closed（不产出任何候选文件）
    removals, removal_errors = select_removals(reviews, exclude_ids, corrections, clarified)
    if removal_errors:
        print("错误：排除请求未通过资格门控（fail closed，不产出任何候选文件）")
        for error in removal_errors:
            print(f"  - {error['id']}: {error['error']}")
        return 4

    print(f"answerCorrections={len(corrections)} questionClarifications={len(clarified)} questionRemovals={len(removals)} contractErrors={len(contract_errors)}")
    if corrections:
        print("answerCorrectionIds:", " ".join(item["id"] for item in corrections))
    if clarified:
        print("clarificationIds:", " ".join(item["id"] for item in clarified))
    if removals:
        print("removalIds:", " ".join(item["id"] for item in removals))
    if contract_errors:
        print("contractErrorIds:", " ".join(item["id"] for item in contract_errors))

    # 变更账本（含文本；终端已只打印计数与 ID）
    ledger_entries: list[dict[str, Any]] = []
    for correction in corrections:
        ledger_entries.append(
            {
                "id": correction["id"],
                "changeType": "ANSWER_CORRECTION",
                "category": correction.get("category"),
                "split": correction["split"],
                "difficulty": correction["difficulty"],
                "oldTextSha256": sha256_text(correction["answerText"]),
                "newTextSha256": sha256_text(correction["correctedAnswerText"]),
                "metricCodes": None,
                "dimensionMapping": None,
                "basis": "review 记录 status=CORRECTED 且 fullEvidence=true 且 evidenceValidation.valid=true 且 correctedAnswerText 非空",
                "evidenceRef": {
                    "reviewFile": str(review_path),
                    "status": correction["status"],
                    "fullEvidence": correction.get("fullEvidence"),
                    "evidenceValidationValid": (correction.get("evidenceValidation") or {}).get("valid"),
                },
            }
        )
    for clarification in clarified:
        ledger_entries.append(
            {
                "id": clarification["id"],
                "changeType": "QUESTION_CLARIFICATION",
                "category": clarification["category"],
                "split": clarification["split"],
                "difficulty": clarification["difficulty"],
                "oldTextSha256": sha256_text(clarification["question"]),
                "newTextSha256": sha256_text(clarification["clarifiedQuestion"]),
                "metricCodes": clarification["metricCodes"],
                "dimensionMapping": clarification["dimensionMapping"],
                "basis": clarification["basis"],
                "evidenceRef": clarification["evidenceRef"],
            }
        )
    for removal in removals:
        ledger_entries.append(
            {
                "id": removal["id"],
                "changeType": "QUESTION_REMOVAL",
                "category": removal.get("category"),
                "split": removal["split"],
                "difficulty": removal["difficulty"],
                "oldTextSha256": sha256_text(removal["question"]),
                "newTextSha256": None,
                "removedAnswerSha256": sha256_text(removal["answerText"]),
                "metricCodes": None,
                "dimensionMapping": None,
                "basis": "用户授权删除：训练集 UNRESOLVED 记录缺少 yoy 基期/当前事实，问题无法复核",
                "evidenceRef": {
                    "reviewFile": str(review_path),
                    "status": removal["status"],
                    "fullEvidence": removal.get("fullEvidence"),
                    "unresolvedReason": removal.get("unresolvedReason"),
                },
            }
        )

    output_dir.mkdir(parents=True, exist_ok=True)
    candidate_path = output_dir / CANDIDATE_WORKBOOK
    ledger_path = output_dir / CHANGE_LEDGER
    manifest_path = output_dir / CANDIDATE_MANIFEST

    build_candidate_workbook(source_path, candidate_path, corrections, clarified, removals)
    ok, verify_errors = verify_candidate_workbook(source_path, candidate_path, corrections, clarified, removals)
    if not ok:
        print("错误：候选工作簿校验失败")
        for error in verify_errors:
            print("  -", error)
        return 3

    candidate_sha = sha256_file(candidate_path)
    ledger = {"generatorName": GENERATOR_NAME, "generatorVersion": GENERATOR_VERSION, "count": len(ledger_entries), "entries": ledger_entries, "contractErrors": contract_errors}
    ledger_path.write_text(json.dumps(ledger, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    # split/total：源全量计数减去被删除记录按 split 的计数
    source_question_rows = load_question_rows(source_path)
    source_split_counts = {"train": 0, "dev": 0, "test": 0}
    for _qid, (_row, question_type, _difficulty, _question) in source_question_rows.items():
        split = rules.SPLIT_MAP.get(question_type)
        if split is not None:
            source_split_counts[split] = source_split_counts.get(split, 0) + 1
    removed_split_counts = Counter(review["split"] for review in removals)
    split_counts = {
        split: source_split_counts.get(split, 0) - removed_split_counts.get(split, 0) for split in ("train", "dev", "test")
    }
    total_records = len(source_question_rows) - len(removals)
    # expectedAudit：候选审查的预期（由源审查状态 + 本次变更推导，审查器逐项对照）
    source_status_counts = Counter(review["status"] for review in reviews)
    expected_audit = {
        "totalRecords": total_records,
        "splitCounts": split_counts,
        "statusCounts": {
            "VERIFIED": source_status_counts.get("VERIFIED", 0) + len(corrections) + len(clarified),
            "CORRECTED": 0,
            "UNRESOLVED": source_status_counts.get("UNRESOLVED", 0) - len(clarified) - len(removals),
        },
        "evidenceComplete": total_records,
        "evidenceErrors": 0,
    }
    manifest = {
        "generatorName": GENERATOR_NAME,
        "generatorVersion": GENERATOR_VERSION,
        "sourceSha256": source_sha,
        "candidateSha256": candidate_sha,
        "candidateWorkbook": CANDIDATE_WORKBOOK,
        "totalRecords": total_records,
        "splitCounts": split_counts,
        "changeCounts": {
            "answerChanges": len(corrections),
            "questionClarifications": len(clarified),
            "questionRemovals": len(removals),
            "contractErrors": len(contract_errors),
        },
        "expectedAudit": expected_audit,
        "contractErrors": contract_errors,
        "factRegionSha256": fact_region_digest(source_path),
        "changeLedgerSha256": sha256_file(ledger_path),
        "candidateReady": False,  # 未经候选审查不得标为可用
        "canonicalReady": False,  # 候选永远不得 canonical
    }
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    for artifact in (candidate_path, ledger_path, manifest_path):
        digest = sha256_file(artifact)
        (artifact.with_suffix(artifact.suffix + ".sha256")).write_text(f"{digest}  {artifact.name}\n", encoding="utf-8")

    print(f"candidateSha256={candidate_sha}")
    print(f"candidateReady=false（待候选审查） canonicalReady=false")
    if contract_errors:
        return 1  # 诚实失败：存在拒绝修改的题，候选不可用
    return 0


def main() -> None:
    parser = argparse.ArgumentParser(description="数据驱动的候选修正版 Ground Truth 生成器")
    parser.add_argument("--source", type=Path, required=True, help="冻结原始工作簿（只读）")
    parser.add_argument("--review", type=Path, required=True, help="基线审查 review.ndjson")
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument(
        "--exclude-id",
        action="append",
        default=None,
        metavar="ID",
        help="用户授权删除的训练集问题编号（可重复；真实编号只在调用时提供，禁止硬编码）",
    )
    args = parser.parse_args()
    sys.exit(run_generator(args.source, args.review, args.output_dir, exclude_ids=args.exclude_id))


if __name__ == "__main__":
    main()
