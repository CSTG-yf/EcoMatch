#!/usr/bin/env python3
"""Build reproducible SQLite and H2 bank benchmark databases from the source workbook."""

from __future__ import annotations

import argparse
import hashlib
import json
import sqlite3
import subprocess
import tempfile
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


ORGANIZATION_SHEET = "机构信息表"
METRIC_SHEET = "指标清单表"
FACT_SHEET = "指标数据表"
ORGANIZATION_HEADERS = ("机构编号", "机构名称")
METRIC_HEADERS = ("指标编号", "指标名称", "指标含义", "指标单位")
FACT_HEADERS = ("数据日期", "指标编号", "指标名称", "机构编号", "指标值")

SQLITE_SCHEMA = """
PRAGMA foreign_keys = ON;
CREATE TABLE bank_organization (
    org_code TEXT PRIMARY KEY,
    org_name TEXT NOT NULL UNIQUE
);
CREATE TABLE bank_metric_definition (
    metric_code TEXT PRIMARY KEY,
    metric_name TEXT NOT NULL UNIQUE,
    metric_meaning TEXT NOT NULL,
    metric_unit TEXT NOT NULL
);
CREATE TABLE bank_metric_daily (
    data_date TEXT NOT NULL,
    org_code TEXT NOT NULL,
    metric_code TEXT NOT NULL,
    metric_value NUMERIC NOT NULL,
    PRIMARY KEY (data_date, org_code, metric_code),
    FOREIGN KEY (org_code) REFERENCES bank_organization(org_code),
    FOREIGN KEY (metric_code) REFERENCES bank_metric_definition(metric_code)
);
CREATE INDEX idx_bank_metric_daily_metric_date ON bank_metric_daily(metric_code, data_date);
CREATE INDEX idx_bank_metric_daily_org_date ON bank_metric_daily(org_code, data_date);
"""

H2_SCHEMA = """
CREATE SCHEMA IF NOT EXISTS bank_benchmark;
DROP VIEW IF EXISTS bank_benchmark.bank_metric_daily;
DROP VIEW IF EXISTS bank_benchmark.bank_metric_definition;
DROP VIEW IF EXISTS bank_benchmark.bank_organization;
DROP TABLE IF EXISTS bank_metric_daily;
DROP TABLE IF EXISTS bank_metric_definition;
DROP TABLE IF EXISTS bank_organization;
CREATE TABLE bank_organization (
    org_code VARCHAR(32) PRIMARY KEY,
    org_name VARCHAR(255) NOT NULL UNIQUE
);
CREATE TABLE bank_metric_definition (
    metric_code VARCHAR(32) PRIMARY KEY,
    metric_name VARCHAR(255) NOT NULL UNIQUE,
    metric_meaning VARCHAR(2000) NOT NULL,
    metric_unit VARCHAR(32) NOT NULL
);
CREATE TABLE bank_metric_daily (
    data_date DATE NOT NULL,
    org_code VARCHAR(32) NOT NULL,
    metric_code VARCHAR(32) NOT NULL,
    metric_value DECIMAL(20, 6) NOT NULL,
    PRIMARY KEY (data_date, org_code, metric_code),
    CONSTRAINT fk_bank_metric_daily_org FOREIGN KEY (org_code) REFERENCES bank_organization(org_code),
    CONSTRAINT fk_bank_metric_daily_metric FOREIGN KEY (metric_code) REFERENCES bank_metric_definition(metric_code)
);
CREATE INDEX idx_bank_metric_daily_metric_date ON bank_metric_daily(metric_code, data_date);
CREATE INDEX idx_bank_metric_daily_org_date ON bank_metric_daily(org_code, data_date);
CREATE VIEW bank_benchmark.bank_organization AS SELECT * FROM PUBLIC.bank_organization;
CREATE VIEW bank_benchmark.bank_metric_definition AS SELECT * FROM PUBLIC.bank_metric_definition;
CREATE VIEW bank_benchmark.bank_metric_daily AS SELECT * FROM PUBLIC.bank_metric_daily;
""".strip()


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for block in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _text(value: Any, field: str, row_number: int) -> str:
    if value is None or not str(value).strip():
        raise ValueError(f"{field} is empty at source row {row_number}")
    return str(value).strip()


def _date(value: Any, row_number: int) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    try:
        return date.fromisoformat(_text(value, "数据日期", row_number)).isoformat()
    except ValueError as error:
        raise ValueError(f"invalid 数据日期 at source row {row_number}: {value!r}") from error


def _decimal(value: Any, row_number: int) -> Decimal:
    try:
        return Decimal(str(value)).quantize(Decimal("0.000001"))
    except (InvalidOperation, ValueError) as error:
        raise ValueError(f"invalid 指标值 at source row {row_number}: {value!r}") from error


def _rows(sheet: Any, headers: tuple[str, ...]) -> Iterable[tuple[int, tuple[Any, ...]]]:
    actual = tuple(cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1)))
    if actual[: len(headers)] != headers:
        raise ValueError(f"unexpected headers for {sheet.title}: {actual!r}")
    for row_number, cells in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if all(value is None for value in cells):
            continue
        yield row_number, cells[: len(headers)]


def _read_workbook(workbook_path: Path) -> tuple[list[tuple[str, str]], list[tuple[str, str, str, str]], list[tuple[str, str, str, Decimal]]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        required = {ORGANIZATION_SHEET, METRIC_SHEET, FACT_SHEET}
        missing = required - set(workbook.sheetnames)
        if missing:
            raise ValueError(f"missing workbook sheets: {sorted(missing)}")

        organizations: list[tuple[str, str]] = []
        organization_codes: set[str] = set()
        for row_number, row in _rows(workbook[ORGANIZATION_SHEET], ORGANIZATION_HEADERS):
            code = _text(row[0], "机构编号", row_number)
            name = _text(row[1], "机构名称", row_number)
            if code in organization_codes:
                raise ValueError(f"duplicate 机构编号: {code}")
            organization_codes.add(code)
            organizations.append((code, name))

        metrics: list[tuple[str, str, str, str]] = []
        metric_codes: set[str] = set()
        metric_names: dict[str, str] = {}
        for row_number, row in _rows(workbook[METRIC_SHEET], METRIC_HEADERS):
            code = _text(row[0], "指标编号", row_number)
            name = _text(row[1], "指标名称", row_number)
            meaning = _text(row[2], "指标含义", row_number)
            unit = _text(row[3], "指标单位", row_number)
            if code in metric_codes:
                raise ValueError(f"duplicate 指标编号: {code}")
            metric_codes.add(code)
            metric_names[code] = name
            metrics.append((code, name, meaning, unit))

        facts: list[tuple[str, str, str, Decimal]] = []
        fact_keys: set[tuple[str, str, str]] = set()
        for row_number, row in _rows(workbook[FACT_SHEET], FACT_HEADERS):
            data_date = _date(row[0], row_number)
            metric_code = _text(row[1], "指标编号", row_number)
            metric_name = _text(row[2], "指标名称", row_number)
            organization_code = _text(row[3], "机构编号", row_number)
            metric_value = _decimal(row[4], row_number)
            if organization_code not in organization_codes:
                raise ValueError(f"unknown 机构编号 at source row {row_number}: {organization_code}")
            if metric_code not in metric_codes:
                raise ValueError(f"unknown 指标编号 at source row {row_number}: {metric_code}")
            if metric_names[metric_code] != metric_name:
                raise ValueError(
                    f"指标名称 mismatch at source row {row_number}: {metric_code} -> {metric_name}"
                )
            key = (data_date, organization_code, metric_code)
            if key in fact_keys:
                raise ValueError(f"duplicate fact key at source row {row_number}: {key}")
            fact_keys.add(key)
            facts.append((data_date, organization_code, metric_code, metric_value))
        if not organizations or not metrics or not facts:
            raise ValueError("workbook must contain organization, metric and fact rows")
        return organizations, metrics, facts
    finally:
        workbook.close()


def _write_sqlite(
    database_path: Path,
    organizations: list[tuple[str, str]],
    metrics: list[tuple[str, str, str, str]],
    facts: list[tuple[str, str, str, Decimal]],
) -> None:
    database_path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(prefix="bank-benchmark-", suffix=".sqlite", dir=database_path.parent, delete=False) as temp_file:
        temp_path = Path(temp_file.name)
    try:
        connection = sqlite3.connect(temp_path)
        try:
            connection.executescript(SQLITE_SCHEMA)
            connection.executemany("INSERT INTO bank_organization(org_code, org_name) VALUES (?, ?)", organizations)
            connection.executemany(
                "INSERT INTO bank_metric_definition(metric_code, metric_name, metric_meaning, metric_unit) VALUES (?, ?, ?, ?)",
                metrics,
            )
            connection.executemany(
                "INSERT INTO bank_metric_daily(data_date, org_code, metric_code, metric_value) VALUES (?, ?, ?, ?)",
                [(data_date, org_code, metric_code, str(value)) for data_date, org_code, metric_code, value in facts],
            )
            connection.commit()
        finally:
            connection.close()
        temp_path.replace(database_path)
    finally:
        if temp_path.exists():
            temp_path.unlink()


def _h2_literal(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def _write_h2_script(
    script_path: Path,
    organizations: list[tuple[str, str]],
    metrics: list[tuple[str, str, str, str]],
    facts: list[tuple[str, str, str, Decimal]],
) -> None:
    script_path.parent.mkdir(parents=True, exist_ok=True)
    with script_path.open("w", encoding="utf-8", newline="\n") as output:
        output.write("BEGIN;\n")
        output.write(H2_SCHEMA + "\n")
        for code, name in organizations:
            output.write(
                "INSERT INTO bank_organization(org_code, org_name) VALUES "
                f"({_h2_literal(code)}, {_h2_literal(name)});\n"
            )
        for code, name, meaning, unit in metrics:
            output.write(
                "INSERT INTO bank_metric_definition(metric_code, metric_name, metric_meaning, metric_unit) VALUES "
                f"({_h2_literal(code)}, {_h2_literal(name)}, {_h2_literal(meaning)}, {_h2_literal(unit)});\n"
            )
        for data_date, org_code, metric_code, metric_value in facts:
            output.write(
                "INSERT INTO bank_metric_daily(data_date, org_code, metric_code, metric_value) VALUES "
                f"(DATE {_h2_literal(data_date)}, {_h2_literal(org_code)}, {_h2_literal(metric_code)}, {metric_value});\n"
            )
        output.write("COMMIT;\n")


def build_h2_database(
    h2_script_path: Path,
    h2_database_path: Path,
    java_path: Path,
    h2_jar_path: Path,
) -> str:
    if not java_path.is_file():
        raise FileNotFoundError(f"H2 Java executable not found: {java_path}")
    if not h2_jar_path.is_file():
        raise FileNotFoundError(f"H2 jar not found: {h2_jar_path}")
    h2_database_path.parent.mkdir(parents=True, exist_ok=True)
    jdbc_url = "jdbc:h2:file:" + h2_database_path.resolve().as_posix() + ";DATABASE_TO_UPPER=false"
    subprocess.run(
        [
            str(java_path),
            "-cp",
            str(h2_jar_path),
            "org.h2.tools.RunScript",
            "-url",
            jdbc_url,
            "-user",
            "root",
            "-password",
            "semantic",
            "-script",
            str(h2_script_path),
        ],
        check=True,
        capture_output=True,
        text=True,
    )
    return jdbc_url


def build_database(
    workbook_path: Path,
    sqlite_output_path: Path,
    h2_script_output_path: Path | None = None,
) -> dict[str, Any]:
    workbook_path = Path(workbook_path)
    if not workbook_path.is_file():
        raise FileNotFoundError(f"workbook not found: {workbook_path}")
    organizations, metrics, facts = _read_workbook(workbook_path)
    _write_sqlite(Path(sqlite_output_path), organizations, metrics, facts)
    if h2_script_output_path is not None:
        _write_h2_script(Path(h2_script_output_path), organizations, metrics, facts)

    dates = sorted({row[0] for row in facts})
    return {
        "sourceSha256": _sha256(workbook_path),
        "counts": {
            "organizations": len(organizations),
            "metrics": len(metrics),
            "facts": len(facts),
        },
        "dateRange": {"min": dates[0], "max": dates[-1]},
        "integrityErrors": [],
    }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--sqlite-output", type=Path, required=True)
    parser.add_argument("--h2-script-output", type=Path)
    parser.add_argument("--h2-database-output", type=Path)
    parser.add_argument("--java-path", type=Path)
    parser.add_argument("--h2-jar-path", type=Path)
    args = parser.parse_args()

    if args.h2_database_output and not args.h2_script_output:
        parser.error("--h2-database-output requires --h2-script-output")
    if args.h2_database_output and (not args.java_path or not args.h2_jar_path):
        parser.error("--h2-database-output requires --java-path and --h2-jar-path")

    report = build_database(args.workbook, args.sqlite_output, args.h2_script_output)
    if args.h2_database_output:
        report["h2JdbcUrl"] = build_h2_database(
            args.h2_script_output,
            args.h2_database_output,
            args.java_path,
            args.h2_jar_path,
        )
    print(json.dumps(report, ensure_ascii=False, sort_keys=True))


if __name__ == "__main__":
    main()
