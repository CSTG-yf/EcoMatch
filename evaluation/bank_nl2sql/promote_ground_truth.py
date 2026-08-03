#!/usr/bin/env python3
"""Promote an audited bank NL2SQL ground-truth candidate into an official release.

Inputs
------
- ``--candidate-dir``: generator output directory produced by
  ``clarify_ground_truth_contracts.py`` and audited by ``audit_ground_truth.py``
  (candidate workbook, candidate manifest + synchronized SHA-256 sidecar,
  contract change ledger + sidecar, frozen source workbook).
- ``--audit-dir``: final candidate-audit output directory (audit-summary.json,
  review.ndjson, correction-ledger.json, candidate-reviewed.xlsx).
  ``candidate-reviewed.xlsx`` 是只含 ``review`` 页的审查报告，不是问题答案
  工作簿；发布器只验证它的存在与 ``audit-summary.outputSha256``，绝不从其中
  加载题目答案清单。

Validation (fail closed, no override switches)
----------------------------------------------
- candidate manifest sidecar/bytes, ledger sidecar/bytes, workbook sidecar/bytes;
- manifest contract: generator identity/version, candidateReady=true,
  split/status/evidence expectations (VERIFIED==totalRecords, 0 CORRECTED,
  0 UNRESOLVED, evidenceComplete==totalRecords, evidenceErrors==0), change
  counts consistent with the ledger actions (all expectations derived from the
  sidecar-anchored manifest, never from real IDs);
- workbook: source/candidate question-sheet diff must be exactly the ledger-
  declared removals/clarifications/corrections (texts verified by SHA-256),
  split counts must match the manifest, fact region must be unchanged;
- audit summary: candidateReady=true, auditErrors=0, counts/IDs/evidence match
  the manifest expectations, reviewed report (candidate-reviewed.xlsx) exists
  with SHA-256 equal to audit-summary.outputSha256, audit-summary.verifiedIds
  set equals the candidate workbook ID set;
- evidence chain: review.ndjson records all VERIFIED with full evidence and no
  evidence errors, correction ledger empty, review.ndjson ID set equals the
  candidate workbook ID set — a three-way cross-check across
  audit-summary.verifiedIds, review.ndjson IDs and candidate workbook IDs
  (the reviewed report itself is never parsed for question/answer data).

Output
------
Writes a deterministic versioned package under ``--output``:
- ``bank-nl2sql-ground-truth-v<version>.xlsx`` official ground-truth workbook
  plus ``.sha256`` sidecar (byte-identical to the verified candidate);
- ``official-manifest.json`` (datasetVersion, canonicalReady, officialCount,
  source splits, ledger-derived removed IDs, source/candidate/fact hashes,
  change counts, audit status, artifact hashes) plus ``.sha256`` sidecar;
- copied ``contract-change-ledger.json`` and final ``audit-summary.json`` plus
  sidecars;
- ``CURRENT.json`` current-version pointer in the ``official/`` parent directory.

Nothing except hashes, counts, IDs and file names is printed; no question text,
answer text, fact rows or SQL is ever emitted.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from audit_ground_truth import sha256_file
from clarify_ground_truth_contracts import (
    CANDIDATE_MANIFEST,
    CANDIDATE_WORKBOOK,
    CHANGE_LEDGER,
    fact_region_digest,
    sha256_text,
)
import gt_answer_rules as rules

GENERATOR_NAME = "clarify_ground_truth_contracts"
SOURCE_WORKBOOK = "source.xlsx"
AUDIT_SUMMARY = "audit-summary.json"
REVIEW_NDJSON = "review.ndjson"
CORRECTION_LEDGER = "correction-ledger.json"
REVIEWED_WORKBOOK = "candidate-reviewed.xlsx"
OFFICIAL_MANIFEST = "official-manifest.json"
FINAL_AUDIT_SUMMARY = "final-audit-summary.json"
CURRENT_POINTER = "CURRENT.json"

VERSION_PATTERN = re.compile(r"\d+\.\d+\.\d+")
SIDECAR_PATTERN = re.compile(r"([0-9A-F]{64})  ([^\n]+)\n")


class PromotionError(ValueError):
    """An audited candidate failed a promotion gate; nothing was written."""


def _validate_sidecar(artifact_path: Path, label: str) -> None:
    """Strict sidecar contract: '<UPPER_SHA256>  <artifact name>\\n', digest match.

    Any missing/malformed/mismatched sidecar fails closed before the artifact
    bytes are trusted.
    """
    sidecar_path = artifact_path.with_name(artifact_path.name + ".sha256")
    if not sidecar_path.is_file():
        raise PromotionError(f"{label} 哈希边车缺失: {sidecar_path.name}")
    try:
        content = sidecar_path.read_text(encoding="utf-8")
    except (OSError, UnicodeDecodeError) as exc:
        raise PromotionError(f"{label} 哈希边车无法读取: {exc}") from exc
    match = SIDECAR_PATTERN.fullmatch(content)
    if match is None:
        raise PromotionError(f"{label} 哈希边车格式非法: {content!r}")
    digest, filename = match.group(1), match.group(2)
    if filename != artifact_path.name:
        raise PromotionError(f"{label} 哈希边车文件名不符: {filename!r} != {artifact_path.name!r}")
    actual = sha256_file(artifact_path)
    if digest != actual:
        raise PromotionError(f"{label} 哈希边车摘要不匹配（{digest} != 实际 {actual}）")


def _load_question_rows(workbook_path: Path) -> dict[str, dict[str, Any]]:
    """问题答案清单 -> {id: {split, difficulty, question, answer}}.

    All five fields are stripped exactly like the audit loader so that text
    hashes match the ledger's old/new text digests.
    """
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if rules.SHEET_QUESTION not in workbook.sheetnames:
            raise PromotionError(f"缺少工作表: {rules.SHEET_QUESTION}（{workbook_path.name}）")
        sheet = workbook[rules.SHEET_QUESTION]
        rows = sheet.iter_rows(values_only=True)
        header = tuple(str(value).strip() if value is not None else "" for value in next(rows, ()))
        if header[: len(rules.QUESTION_HEADERS)] != rules.QUESTION_HEADERS:
            raise PromotionError(f"{rules.SHEET_QUESTION} 表头不匹配: {header!r}")
        result: dict[str, dict[str, Any]] = {}
        for row in rows:
            if row[0] is None:
                continue
            qid = str(row[0]).strip()
            if not qid or qid in result:
                raise PromotionError(f"{rules.SHEET_QUESTION} 问题编号缺失或重复: {qid!r}")
            question_type = str(row[1]).strip()
            if question_type not in rules.SPLIT_MAP:
                raise PromotionError(f"{rules.SHEET_QUESTION} 非法问题类型: {question_type!r}（{qid}）")
            result[qid] = {
                "split": rules.SPLIT_MAP[question_type],
                "difficulty": str(row[2]).strip(),
                "question": str(row[3]).strip(),
                "answer": str(row[4]).strip(),
            }
        if not result:
            raise PromotionError(f"{rules.SHEET_QUESTION} 无任何题目（{workbook_path.name}）")
        return result
    finally:
        workbook.close()


def _validate_manifest(manifest: dict[str, Any], version: str, expected_source_sha: str) -> None:
    """Candidate manifest semantic contract (all data-driven from the manifest)."""
    if manifest.get("generatorName") != GENERATOR_NAME:
        raise PromotionError(f"generatorName 不符: {manifest.get('generatorName')!r}")
    if manifest.get("generatorVersion") != version:
        raise PromotionError(
            f"generatorVersion 不符: {manifest.get('generatorVersion')!r} != 版本 {version!r}"
        )
    if manifest.get("candidateReady") is not True:
        raise PromotionError("candidateReady 必须为 true（候选未通过审查）")
    candidate_audit = manifest.get("candidateAudit")
    if not isinstance(candidate_audit, dict) or candidate_audit.get("candidateReady") is not True:
        raise PromotionError("candidateAudit.candidateReady 必须为 true（候选未通过最终审查）")
    if candidate_audit.get("reasons"):
        raise PromotionError(f"candidateAudit 存在拒绝理由: {candidate_audit['reasons']}")
    if manifest.get("sourceSha256") != expected_source_sha:
        raise PromotionError("manifest.sourceSha256 与冻结原始工作簿哈希不一致")
    if manifest.get("candidateWorkbook") != CANDIDATE_WORKBOOK:
        raise PromotionError(f"candidateWorkbook 不符: {manifest.get('candidateWorkbook')!r}")
    total_records = manifest.get("totalRecords")
    split_counts = manifest.get("splitCounts")
    if not isinstance(total_records, int) or total_records <= 0:
        raise PromotionError(f"manifest.totalRecords 非法: {total_records!r}")
    if (
        not isinstance(split_counts, dict)
        or set(split_counts) != {"train", "dev", "test"}
        or any(not isinstance(value, int) or value < 0 for value in split_counts.values())
    ):
        raise PromotionError(f"manifest.splitCounts 非法: {split_counts!r}")
    if sum(split_counts.values()) != total_records:
        raise PromotionError(f"manifest.splitCounts 计数和 {sum(split_counts.values())} != totalRecords={total_records}")
    expected_audit = manifest.get("expectedAudit")
    if not isinstance(expected_audit, dict):
        raise PromotionError("manifest.expectedAudit 缺失（候选就绪必须由已验证 manifest 推导）")
    for field in ("totalRecords", "evidenceComplete", "evidenceErrors"):
        if not isinstance(expected_audit.get(field), int) or expected_audit.get(field) < 0:
            raise PromotionError(f"manifest.expectedAudit.{field} 非法: {expected_audit.get(field)!r}")
    if expected_audit.get("totalRecords") != total_records:
        raise PromotionError("manifest.expectedAudit.totalRecords 与 manifest.totalRecords 不一致")
    if expected_audit.get("splitCounts") != split_counts:
        raise PromotionError("manifest.expectedAudit.splitCounts 与 manifest.splitCounts 不一致")
    status_counts = expected_audit.get("statusCounts")
    if (
        not isinstance(status_counts, dict)
        or set(status_counts) != {"VERIFIED", "CORRECTED", "UNRESOLVED"}
        or status_counts.get("VERIFIED") != total_records
        or status_counts.get("CORRECTED") != 0
        or status_counts.get("UNRESOLVED") != 0
    ):
        raise PromotionError(f"manifest.expectedAudit.statusCounts 非法（必须 VERIFIED={total_records}/0/0）: {status_counts!r}")
    if expected_audit.get("evidenceComplete") != total_records:
        raise PromotionError(f"manifest.expectedAudit.evidenceComplete={expected_audit.get('evidenceComplete')} != totalRecords={total_records}")
    if expected_audit.get("evidenceErrors") != 0:
        raise PromotionError(f"manifest.expectedAudit.evidenceErrors={expected_audit.get('evidenceErrors')} != 0")
    if candidate_audit.get("statusCounts") != status_counts:
        raise PromotionError("candidateAudit.statusCounts 与 expectedAudit 不一致")
    if candidate_audit.get("splitCounts") != split_counts:
        raise PromotionError("candidateAudit.splitCounts 与 expectedAudit 不一致")
    change_counts = manifest.get("changeCounts")
    if (
        not isinstance(change_counts, dict)
        or set(change_counts) != {"answerChanges", "questionClarifications", "questionRemovals", "contractErrors"}
        or any(not isinstance(change_counts[field], int) or change_counts[field] < 0 for field in change_counts)
    ):
        raise PromotionError(f"manifest.changeCounts 非法: {change_counts!r}")
    if not isinstance(manifest.get("contractErrors"), list) or manifest["contractErrors"] != []:
        raise PromotionError(f"manifest.contractErrors 必须为空: {manifest.get('contractErrors')!r}")
    if len(manifest["contractErrors"]) != change_counts["contractErrors"]:
        raise PromotionError(
            f"manifest.contractErrors 数量 {len(manifest['contractErrors'])} != changeCounts.contractErrors={change_counts['contractErrors']}"
        )
    for field in ("candidateSha256", "changeLedgerSha256", "factRegionSha256"):
        if not isinstance(manifest.get(field), str) or not re.fullmatch(r"[0-9A-F]{64}", str(manifest.get(field))):
            raise PromotionError(f"manifest.{field} 非法: {manifest.get(field)!r}")


def _validate_ledger(ledger_path: Path, manifest: dict[str, Any]) -> dict[str, list[str]]:
    """Ledger content contract; returns {removed_ids, correction_ids, clarification_ids}."""
    try:
        ledger = json.loads(ledger_path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError) as exc:
        raise PromotionError(f"contract-change-ledger.json 无法解析: {exc}") from exc
    if not isinstance(ledger, dict):
        raise PromotionError("contract-change-ledger.json 必须为对象")
    if ledger.get("generatorName") != manifest.get("generatorName"):
        raise PromotionError(
            f"ledger.generatorName={ledger.get('generatorName')!r} 与 manifest.generatorName={manifest.get('generatorName')!r} 不一致"
        )
    if ledger.get("generatorVersion") != manifest.get("generatorVersion"):
        raise PromotionError(
            f"ledger.generatorVersion={ledger.get('generatorVersion')!r} 与 manifest.generatorVersion={manifest.get('generatorVersion')!r} 不一致"
        )
    entries = ledger.get("entries")
    if not isinstance(entries, list):
        raise PromotionError("ledger.entries 缺失或非列表")
    if ledger.get("count") != len(entries):
        raise PromotionError(f"ledger.count={ledger.get('count')} != entries 条目数 {len(entries)}")
    if ledger.get("contractErrors") != manifest.get("contractErrors"):
        raise PromotionError("ledger.contractErrors 与 manifest.contractErrors 不一致")
    action_counts: Counter[str] = Counter()
    seen_ids: set[str] = set()
    buckets: dict[str, list[str]] = {"removed": [], "corrected": [], "clarified": []}
    for entry in entries:
        if not isinstance(entry, dict) or not isinstance(entry.get("id"), str) or not entry["id"]:
            raise PromotionError("ledger.entries 含非法条目（缺 id 或 id 非字符串）")
        entry_id = entry["id"]
        if entry_id in seen_ids:
            raise PromotionError(f"ledger.entries 存在重复 ID：{entry_id}")
        seen_ids.add(entry_id)
        change_type = entry.get("changeType")
        if change_type not in ("ANSWER_CORRECTION", "QUESTION_CLARIFICATION", "QUESTION_REMOVAL"):
            raise PromotionError(f"ledger.entries 含非法 changeType：{change_type!r}（{entry_id}）")
        action_counts[change_type] += 1
        if change_type == "QUESTION_REMOVAL":
            buckets["removed"].append(entry_id)
            if not isinstance(entry.get("removedAnswerSha256"), str):
                raise PromotionError(f"QUESTION_REMOVAL 条目缺少 removedAnswerSha256：{entry_id}")
        else:
            for field in ("oldTextSha256", "newTextSha256"):
                if not isinstance(entry.get(field), str):
                    raise PromotionError(f"{change_type} 条目缺少 {field}：{entry_id}")
            buckets["corrected" if change_type == "ANSWER_CORRECTION" else "clarified"].append(entry_id)
    change_counts = manifest.get("changeCounts")
    for field, change_type in (
        ("answerChanges", "ANSWER_CORRECTION"),
        ("questionClarifications", "QUESTION_CLARIFICATION"),
        ("questionRemovals", "QUESTION_REMOVAL"),
    ):
        if action_counts.get(change_type, 0) != change_counts.get(field):
            raise PromotionError(
                f"ledger 动作计数 {change_type}={action_counts.get(change_type, 0)} != manifest.changeCounts.{field}={change_counts.get(field)}"
            )
    return buckets


def _validate_workbook_diff(
    source_rows: dict[str, dict[str, Any]],
    candidate_rows: dict[str, dict[str, Any]],
    buckets: dict[str, list[str]],
    manifest: dict[str, Any],
) -> None:
    """Candidate workbook must differ from the source exactly as the ledger declares."""
    total_records = manifest["totalRecords"]
    removed_ids = set(buckets["removed"])
    source_ids = set(source_rows)
    candidate_ids = set(candidate_rows)
    if len(candidate_rows) != total_records:
        raise PromotionError(f"候选工作簿题数 {len(candidate_rows)} != totalRecords={total_records}")
    missing_removed = removed_ids - source_ids
    if missing_removed:
        raise PromotionError(f"被授权删除的 ID 不在源工作簿中: {sorted(missing_removed)}")
    still_present = removed_ids & candidate_ids
    if still_present:
        raise PromotionError(f"被授权删除的 ID 仍存在于候选: {sorted(still_present)}")
    extra_removed = source_ids - candidate_ids - removed_ids
    if extra_removed:
        raise PromotionError(f"候选缺失未授权删除的 ID: {sorted(extra_removed)}")
    if candidate_ids != source_ids - removed_ids:
        raise PromotionError("候选 ID 集合与源工作簿减去被授权删除 ID 不一致")
    candidate_split_counts = Counter(candidate["split"] for candidate in candidate_rows.values())
    if dict(candidate_split_counts) != manifest.get("splitCounts"):
        raise PromotionError(
            f"候选工作簿 split 计数 {dict(candidate_split_counts)} != manifest.splitCounts {manifest.get('splitCounts')}"
        )
    for qid, candidate in candidate_rows.items():
        source = source_rows[qid]
        if candidate["split"] != source["split"] or candidate["difficulty"] != source["difficulty"]:
            raise PromotionError(f"{qid}: 候选修改了 split/难度（账本未授权）")
        if qid in buckets["corrected"]:
            if candidate["answer"] == source["answer"]:
                raise PromotionError(f"{qid}: ANSWER_CORRECTION 未实际修改答案")
        elif candidate["answer"] != source["answer"]:
            raise PromotionError(f"{qid}: 候选答案变更未在账本声明")
        if qid in buckets["clarified"]:
            if candidate["question"] == source["question"]:
                raise PromotionError(f"{qid}: QUESTION_CLARIFICATION 未实际修改题目")
        elif candidate["question"] != source["question"]:
            raise PromotionError(f"{qid}: 候选题目变更未在账本声明")
    for qid in buckets["corrected"] + buckets["clarified"]:
        if qid not in source_ids or qid not in candidate_ids:
            raise PromotionError(f"{qid}: 账本变更条目与工作簿 ID 集合不一致")


def _validate_ledger_text_hashes(
    ledger_path: Path,
    source_rows: dict[str, dict[str, Any]],
    candidate_rows: dict[str, dict[str, Any]],
) -> None:
    """Every ledger text digest must match the actual source/candidate cell text."""
    ledger = json.loads(ledger_path.read_text(encoding="utf-8"))
    for entry in ledger["entries"]:
        entry_id = entry["id"]
        change_type = entry["changeType"]
        if change_type == "QUESTION_REMOVAL":
            source = source_rows.get(entry_id)
            if source is None:
                raise PromotionError(f"{entry_id}: QUESTION_REMOVAL 目标不在源工作簿")
            if sha256_text(source["question"]) != entry.get("oldTextSha256"):
                raise PromotionError(f"{entry_id}: QUESTION_REMOVAL oldTextSha256 与源题目不符")
            if sha256_text(source["answer"]) != entry.get("removedAnswerSha256"):
                raise PromotionError(f"{entry_id}: QUESTION_REMOVAL removedAnswerSha256 与源答案不符")
            continue
        source = source_rows.get(entry_id)
        candidate = candidate_rows.get(entry_id)
        if source is None or candidate is None:
            raise PromotionError(f"{entry_id}: 账本变更条目不在源/候选工作簿")
        if change_type == "ANSWER_CORRECTION":
            if sha256_text(source["answer"]) != entry.get("oldTextSha256"):
                raise PromotionError(f"{entry_id}: oldTextSha256 与源答案不符")
            expected_new = candidate["answer"]
        else:
            if sha256_text(source["question"]) != entry.get("oldTextSha256"):
                raise PromotionError(f"{entry_id}: oldTextSha256 与源题目不符")
            expected_new = candidate["question"]
        if sha256_text(expected_new) != entry.get("newTextSha256"):
            raise PromotionError(f"{entry_id}: newTextSha256 与候选{('答案' if change_type == 'ANSWER_CORRECTION' else '题目')}不符")


def _validate_audit_summary(
    summary: dict[str, Any],
    manifest: dict[str, Any],
    audit_dir: Path,
    expected_ids: set[str],
    candidate_sha256: str,
) -> None:
    """Final audit summary must match the manifest expectations and its own files.

    candidate-reviewed.xlsx 是只含 ``review`` 页的审查报告，不是问题答案
    工作簿：只验证文件存在且 SHA-256 等于 audit-summary.outputSha256；
    verifiedIds 必须等于候选工作簿 ID 集合，绝不从审查报告加载题目答案。
    """
    total_records = manifest["totalRecords"]
    if summary.get("candidateReady") is not True:
        raise PromotionError("audit-summary.candidateReady 必须为 true")
    if summary.get("candidateReadyReasons"):
        raise PromotionError(f"audit-summary.candidateReadyReasons 非空: {summary['candidateReadyReasons']}")
    if summary.get("auditErrors") != 0:
        raise PromotionError(f"audit-summary.auditErrors={summary.get('auditErrors')} != 0")
    if summary.get("totalRecords") != total_records or summary.get("uniqueIds") != total_records:
        raise PromotionError(
            f"audit-summary total/uniqueIds 不符: {summary.get('totalRecords')}/{summary.get('uniqueIds')} != {total_records}"
        )
    if summary.get("splitCounts") != manifest.get("splitCounts"):
        raise PromotionError(f"audit-summary.splitCounts 与 manifest 不一致: {summary.get('splitCounts')}")
    expected_statuses = manifest.get("expectedAudit", {}).get("statusCounts")
    if summary.get("statusCounts") != expected_statuses:
        raise PromotionError(f"audit-summary.statusCounts 与 manifest 不一致: {summary.get('statusCounts')}")
    if summary.get("evidenceCompleteCount") != total_records:
        raise PromotionError(f"audit-summary.evidenceCompleteCount={summary.get('evidenceCompleteCount')} != {total_records}")
    if summary.get("evidenceErrorCount") != 0 or summary.get("fullEvidence") != total_records:
        raise PromotionError(
            f"audit-summary 证据计数不符: fullEvidence={summary.get('fullEvidence')} evidenceErrorCount={summary.get('evidenceErrorCount')}"
        )
    if not summary.get("manifestValidation", {}).get("valid"):
        raise PromotionError(f"audit-summary.manifestValidation 未通过: {summary.get('manifestValidation')}")
    if not summary.get("manifestSidecarValidation", {}).get("valid"):
        raise PromotionError(f"audit-summary.manifestSidecarValidation 未通过: {summary.get('manifestSidecarValidation')}")
    if summary.get("sourceSha256") != candidate_sha256:
        raise PromotionError("audit-summary.sourceSha256 与候选工作簿哈希不一致")
    if summary.get("outputFile") != REVIEWED_WORKBOOK:
        raise PromotionError(f"audit-summary.outputFile 不符: {summary.get('outputFile')!r}")
    reviewed_path = audit_dir / REVIEWED_WORKBOOK
    if not reviewed_path.is_file():
        raise PromotionError(f"候选审查工作簿缺失: {reviewed_path.name}")
    if summary.get("outputSha256") != sha256_file(reviewed_path):
        raise PromotionError("audit-summary.outputSha256 与 candidate-reviewed.xlsx 实际哈希不一致")
    if set(summary.get("verifiedIds", [])) != expected_ids:
        raise PromotionError("audit-summary.verifiedIds 与候选工作簿 ID 集合不一致")
    if summary.get("correctedIds") != [] or summary.get("unresolvedIds") != []:
        raise PromotionError(
            f"audit-summary correctedIds/unresolvedIds 非空: {summary.get('correctedIds')}/{summary.get('unresolvedIds')}"
        )


def _validate_review_evidence(review_path: Path, manifest: dict[str, Any], expected_ids: set[str]) -> None:
    """Every review record must be VERIFIED with complete, valid evidence."""
    lines = [line for line in review_path.read_text(encoding="utf-8").splitlines() if line.strip()]
    if len(lines) != manifest["totalRecords"]:
        raise PromotionError(f"review.ndjson 条数 {len(lines)} != totalRecords={manifest['totalRecords']}")
    seen_ids: set[str] = set()
    split_counts: Counter[str] = Counter()
    for line_number, line in enumerate(lines, start=1):
        try:
            review = json.loads(line)
        except json.JSONDecodeError as exc:
            raise PromotionError(f"review.ndjson:{line_number} 无法解析") from exc
        if not isinstance(review, dict):
            raise PromotionError(f"review.ndjson:{line_number} 必须为对象")
        if review.get("status") != "VERIFIED":
            raise PromotionError(f"review.ndjson:{line_number} 状态不是 VERIFIED: {review.get('status')!r}")
        if review.get("fullEvidence") is not True:
            raise PromotionError(f"review.ndjson:{line_number} fullEvidence 必须为 true")
        evidence_validation = review.get("evidenceValidation")
        if not isinstance(evidence_validation, dict) or evidence_validation.get("valid") is not True:
            raise PromotionError(f"review.ndjson:{line_number} evidenceValidation 必须有效")
        if review.get("auditErrors"):
            raise PromotionError(f"review.ndjson:{line_number} auditErrors 必须为空")
        review_id = review.get("id")
        split = review.get("split")
        if not isinstance(review_id, str) or not review_id:
            raise PromotionError(f"review.ndjson:{line_number} 缺少有效 id")
        if review_id in seen_ids:
            raise PromotionError(f"review.ndjson 存在重复 ID: {review_id}")
        seen_ids.add(review_id)
        if split not in ("train", "dev", "test"):
            raise PromotionError(f"review.ndjson:{line_number} 非法 split: {split!r}")
        split_counts[split] += 1
    if seen_ids != expected_ids:
        raise PromotionError("review.ndjson ID 集合与候选工作簿 ID 集合不一致")
    if dict(split_counts) != manifest.get("splitCounts"):
        raise PromotionError(f"review.ndjson split 计数与 manifest 不一致: {dict(split_counts)}")


def _sha256_bytes(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest().upper()


def _write_artifact_with_sidecar(destination_path: Path, payload: bytes) -> str:
    """Atomically write bytes + '<UPPER_SHA256>  <name>\\n' sidecar; returns digest."""
    destination_path.parent.mkdir(parents=True, exist_ok=True)
    digest = _sha256_bytes(payload)
    sidecar_path = destination_path.with_name(destination_path.name + ".sha256")
    tmp_payload = destination_path.with_name(destination_path.name + ".tmp")
    tmp_sidecar = sidecar_path.with_name(sidecar_path.name + ".tmp")
    try:
        tmp_payload.write_bytes(payload)
        tmp_sidecar.write_bytes(f"{digest}  {destination_path.name}\n".encode("utf-8"))
        tmp_payload.replace(destination_path)
        tmp_sidecar.replace(sidecar_path)
    finally:
        for tmp_path in (tmp_payload, tmp_sidecar):
            if tmp_path.exists():
                tmp_path.unlink()
    return digest


def promote_ground_truth(
    candidate_dir: Path | str,
    audit_dir: Path | str,
    version: str,
    output_dir: Path | str,
    expected_source_sha: str = rules.SOURCE_SHA256_EXPECTED,
) -> dict[str, Any]:
    """Validate the audited candidate and write the versioned official package."""
    candidate_dir = Path(candidate_dir).resolve()
    audit_dir = Path(audit_dir).resolve()
    output_dir = Path(output_dir).resolve()
    if VERSION_PATTERN.fullmatch(version) is None:
        raise PromotionError(f"版本号必须为 semver（如 2.0.0）: {version!r}")

    # ---- 候选 manifest（哈希边车锚定字节完整性） ----
    manifest_path = candidate_dir / CANDIDATE_MANIFEST
    if not manifest_path.is_file():
        raise PromotionError(f"候选 manifest 缺失: {manifest_path}")
    _validate_sidecar(manifest_path, "候选 manifest")
    try:
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError) as exc:
        raise PromotionError(f"候选 manifest 无法解析: {exc}") from exc
    _validate_manifest(manifest, version, expected_source_sha)
    total_records = manifest["totalRecords"]

    # ---- 候选工作簿（哈希 + 边车） ----
    workbook_path = candidate_dir / CANDIDATE_WORKBOOK
    if not workbook_path.is_file():
        raise PromotionError(f"候选工作簿缺失: {workbook_path}")
    _validate_sidecar(workbook_path, "候选工作簿")
    candidate_sha256 = sha256_file(workbook_path)
    if candidate_sha256 != manifest.get("candidateSha256"):
        raise PromotionError(f"候选工作簿哈希不一致（{candidate_sha256} != {manifest.get('candidateSha256')}）")

    # ---- 变更账本（哈希 + 边车 + 内容契约） ----
    ledger_path = candidate_dir / CHANGE_LEDGER
    if not ledger_path.is_file():
        raise PromotionError(f"变更账本缺失: {ledger_path}")
    _validate_sidecar(ledger_path, "变更账本")
    if sha256_file(ledger_path) != manifest.get("changeLedgerSha256"):
        raise PromotionError("contract-change-ledger.json 与 manifest.changeLedgerSha256 不一致")
    buckets = _validate_ledger(ledger_path, manifest)

    # ---- 冻结原始工作簿（哈希链回到源） ----
    # 标准布局：候选目录内放 source.xlsx；若候选目录缺失（历史候选目录只
    # 放生成产物），回退到候选目录父目录（gt-audit 根）的 source.xlsx。
    # 两条路径都必须通过同一 fail-closed 哈希校验，哈希不匹配一律拒绝。
    source_path = candidate_dir / SOURCE_WORKBOOK
    if not source_path.is_file():
        source_path = candidate_dir.parent / SOURCE_WORKBOOK
    if not source_path.is_file():
        raise PromotionError(f"源工作簿缺失: {source_path}")
    if sha256_file(source_path) != expected_source_sha:
        raise PromotionError(f"源工作簿哈希不匹配 {sha256_file(source_path)} != {expected_source_sha}")

    # ---- 工作簿与账本互证（删除/澄清/修正逐项核对） ----
    source_rows = _load_question_rows(source_path)
    candidate_rows = _load_question_rows(workbook_path)
    _validate_workbook_diff(source_rows, candidate_rows, buckets, manifest)
    _validate_ledger_text_hashes(ledger_path, source_rows, candidate_rows)
    if fact_region_digest(source_path) != manifest.get("factRegionSha256"):
        raise PromotionError("源工作簿事实区域与 manifest.factRegionSha256 不一致")
    if fact_region_digest(workbook_path) != manifest.get("factRegionSha256"):
        raise PromotionError("候选工作簿事实区域与 manifest.factRegionSha256 不一致")

    # ---- 最终审查目录（摘要、逐条证据、输出哈希） ----
    summary_path = audit_dir / AUDIT_SUMMARY
    if not summary_path.is_file():
        raise PromotionError(f"最终审查摘要缺失: {summary_path}")
    try:
        summary = json.loads(summary_path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError) as exc:
        raise PromotionError(f"audit-summary.json 无法解析: {exc}") from exc
    # candidate-reviewed.xlsx 是只含 review 页的审查报告：只验证存在与哈希
    # （outputSha256）。ID 互证以候选工作簿 ID 集合为基准三方一致：
    # audit-summary.verifiedIds、review.ndjson ID 集合、候选工作簿 ID 集合。
    expected_ids = set(candidate_rows)
    _validate_audit_summary(summary, manifest, audit_dir, expected_ids, candidate_sha256)
    review_path = audit_dir / REVIEW_NDJSON
    if not review_path.is_file():
        raise PromotionError(f"审查证据缺失: {review_path}")
    _validate_review_evidence(review_path, manifest, expected_ids)
    correction_ledger_path = audit_dir / CORRECTION_LEDGER
    if not correction_ledger_path.is_file():
        raise PromotionError(f"审查修正账本缺失: {correction_ledger_path}")
    try:
        correction_ledger = json.loads(correction_ledger_path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError) as exc:
        raise PromotionError(f"审查修正账本无法解析: {exc}") from exc
    if not isinstance(correction_ledger, dict) or correction_ledger.get("count") != 0 or correction_ledger.get("corrections") != []:
        raise PromotionError("审查修正账本必须为空（0 CORRECTED）")

    # ---- 写入版本化官方包（确定性字节，无时间戳） ----
    official_workbook_name = f"bank-nl2sql-ground-truth-v{version}.xlsx"
    official_workbook_path = output_dir / official_workbook_name
    # 官方工作簿必须从已验证候选字节生成（同一原子写入+sidecar 机制），
    # 字节原样保留，输出哈希必须与 manifest 锚定的 candidateSha256 完全一致。
    official_workbook_sha256 = _write_artifact_with_sidecar(official_workbook_path, workbook_path.read_bytes())
    if official_workbook_sha256 != candidate_sha256:
        raise PromotionError("官方工作簿复制后哈希不一致")
    # 输入 ledger 已由哈希边车与 manifest.changeLedgerSha256 锚定，必须原样复制
    # 其字节并为目标生成 sidecar；任何 JSON 重序列化都会改变字节从而破坏锚定。
    ledger_sha256 = _write_artifact_with_sidecar(output_dir / CHANGE_LEDGER, ledger_path.read_bytes())
    if ledger_sha256 != manifest.get("changeLedgerSha256"):
        raise PromotionError("官方账本复制后哈希与 manifest 不一致")
    audit_summary_sha256 = _write_artifact_with_sidecar(
        output_dir / FINAL_AUDIT_SUMMARY,
        json.dumps(summary, ensure_ascii=False, indent=2, sort_keys=True).encode("utf-8") + b"\n",
    )

    official_manifest = {
        "datasetVersion": version,
        "canonicalReady": True,
        "officialCount": total_records,
        "sourceSplitCounts": dict(manifest["splitCounts"]),
        "removedIds": sorted(buckets["removed"]),
        "sourceWorkbook": SOURCE_WORKBOOK,
        "sourceSha256": manifest["sourceSha256"],
        "candidateSha256": candidate_sha256,
        "candidateManifestSha256": sha256_file(manifest_path),
        "factRegionSha256": manifest["factRegionSha256"],
        "groundTruthWorkbook": official_workbook_name,
        "generator": {"name": manifest["generatorName"], "version": manifest["generatorVersion"]},
        "changeCounts": dict(manifest["changeCounts"]),
        "auditStatus": {
            "candidateReady": True,
            "evidenceComplete": total_records,
            "evidenceErrors": 0,
            "auditErrors": 0,
            "manifestValidationValid": True,
            "manifestSidecarValidationValid": True,
        },
        "changeLedger": CHANGE_LEDGER,
        "finalAuditSummary": FINAL_AUDIT_SUMMARY,
        "reviewEvidenceSha256": sha256_file(review_path),
        "candidateReviewedSha256": summary.get("outputSha256"),
        "artifactSha256": {
            "groundTruthWorkbook": official_workbook_sha256,
            "changeLedger": ledger_sha256,
            "finalAuditSummary": audit_summary_sha256,
        },
    }
    official_manifest_path = output_dir / OFFICIAL_MANIFEST
    _write_artifact_with_sidecar(
        official_manifest_path,
        json.dumps(official_manifest, ensure_ascii=False, indent=2, sort_keys=True).encode("utf-8") + b"\n",
    )

    current_pointer = {
        "currentVersion": version,
        "directory": version,
        "officialManifest": f"{version}/{OFFICIAL_MANIFEST}",
        "groundTruthWorkbook": f"{version}/{official_workbook_name}",
    }
    current_path = output_dir.parent / CURRENT_POINTER
    current_path.parent.mkdir(parents=True, exist_ok=True)
    current_path.write_text(
        json.dumps(current_pointer, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )
    return official_manifest


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--candidate-dir", type=Path, required=True, help="audited candidate generator output directory")
    parser.add_argument("--audit-dir", type=Path, required=True, help="final candidate-audit output directory")
    parser.add_argument("--version", required=True, help="release version (semver, e.g. 2.0.0)")
    parser.add_argument("--output", type=Path, required=True, help="versioned official package directory")
    args = parser.parse_args()
    try:
        official_manifest = promote_ground_truth(args.candidate_dir, args.audit_dir, args.version, args.output)
    except PromotionError as error:
        print(f"promotion failed: {error}", file=sys.stderr)
        return 2
    print(json.dumps(official_manifest, ensure_ascii=False, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    sys.exit(main())
