#!/usr/bin/env python3
"""Build a bank-import workbook for the isolated synthetic-360 data domain."""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import Workbook


def _read_jsonl(path: Path) -> list[dict[str, object]]:
    return [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]


def build_workbook(release_dir: Path, output: Path) -> dict[str, int]:
    metrics = _read_jsonl(release_dir / "metrics.jsonl")
    organizations = json.loads((release_dir / "organizations.json").read_text(encoding="utf-8"))
    facts = _read_jsonl(release_dir / "facts.jsonl")
    questions = _read_jsonl(release_dir / "questions.jsonl")
    if len(metrics) != 360 or len(organizations) != 13 or len(facts) != 79_560 or len(questions) != 360:
        raise ValueError("synthetic release counts are not 360/13/79560/360")

    workbook = Workbook()
    organization_sheet = workbook.active
    organization_sheet.title = "机构信息表"
    organization_sheet.append(["机构编号", "机构名称"])
    for organization in organizations:
        organization_sheet.append([organization["orgCode"], organization["orgName"]])

    indicator_sheet = workbook.create_sheet("指标清单表")
    indicator_sheet.append(["指标编号", "指标名称", "指标含义", "指标单位"])
    for metric in metrics:
        indicator_sheet.append([
            metric["code"], metric["name"], metric["definition"], metric["unit"]
        ])

    derived_sheet = workbook.create_sheet("衍生维度说明")
    derived_sheet.append(["衍生维度", "衍生口径说明"])
    derived_sheet.append(["synthetic_360", "明确标注为合成数据的360项候选指标点查询测试域。"])

    fact_sheet = workbook.create_sheet("指标数据表")
    fact_sheet.append(["数据日期", "指标编号", "指标名称", "机构编号", "指标值"])
    names = {str(metric["code"]): str(metric["name"]) for metric in metrics}
    for fact in facts:
        fact_sheet.append([
            fact["dataDate"], fact["metricCode"], names[str(fact["metricCode"])],
            fact["orgCode"], fact["metricValue"],
        ])

    question_sheet = workbook.create_sheet("问题答案清单")
    question_sheet.append(["问题编号", "问题类型", "问题难度", "问题描述", "问题结果"])
    for question in questions:
        question_sheet.append([
            question["id"], "指标点查询", "简单", question["question"],
            "SYNTHETIC_STRUCTURED_FACT",
        ])

    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return {"metrics": len(metrics), "organizations": len(organizations), "facts": len(facts), "questions": len(questions)}


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--release-dir", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    print(json.dumps(build_workbook(args.release_dir, args.output), ensure_ascii=False, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
