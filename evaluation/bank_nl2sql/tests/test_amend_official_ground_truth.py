#!/usr/bin/env python3
"""Regression tests for incremental, answer-only official releases."""

from __future__ import annotations

import json
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from amend_official_ground_truth import (  # noqa: E402
    OfficialAmendmentError,
    amend_derived_dataset,
    amend_official_package,
    sha256_file,
    sha256_text,
)
from clarify_ground_truth_contracts import fact_region_digest  # noqa: E402
from build_dataset import DatasetBuildError, _load_official_manifest  # noqa: E402


def _write_json(path: Path, payload: dict[str, object]) -> None:
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def _write_sidecar(path: Path) -> None:
    digest = sha256_file(path)
    path.with_name(path.name + ".sha256").write_text(
        f"{digest}  {path.name}\n",
        encoding="ascii",
    )


def _make_parent(root: Path) -> Path:
    parent = root / "1.0.0"
    parent.mkdir()
    workbook_name = "bank-nl2sql-ground-truth-v1.0.0.xlsx"
    workbook_path = parent / workbook_name
    workbook = Workbook()
    questions = workbook.active
    questions.title = "问题答案清单"
    questions.append(["问题编号", "问题类型", "问题难度", "问题描述", "问题结果"])
    questions.append(["SYN-T001", "训练集", "普通", "训练问题", "旧训练答案42"])
    questions.append(["SYN-V001", "验证集", "普通", "验证问题", "旧验证答案7"])
    questions.append(["SYN-S001", "测试集", "普通", "测试问题", "冻结测试答案9"])
    facts = workbook.create_sheet("指标数据表")
    facts.append(["org_code", "metric_value"])
    facts.append(["ORG001", 42])
    workbook.save(workbook_path)
    workbook.close()
    _write_sidecar(workbook_path)

    ledger = {
        "generatorName": "clarify_ground_truth_contracts",
        "generatorVersion": "2.0.0",
        "count": 0,
        "contractErrors": [],
        "entries": [],
    }
    ledger_path = parent / "contract-change-ledger.json"
    _write_json(ledger_path, ledger)
    _write_sidecar(ledger_path)

    audit_path = parent / "final-audit-summary.json"
    _write_json(audit_path, {"canonicalReady": True, "totalRecords": 3})
    _write_sidecar(audit_path)

    manifest = {
        "artifactSha256": {
            "changeLedger": sha256_file(ledger_path),
            "finalAuditSummary": sha256_file(audit_path),
            "groundTruthWorkbook": sha256_file(workbook_path),
        },
        "auditStatus": {"auditErrors": 0, "candidateReady": True},
        "canonicalReady": True,
        "changeCounts": {
            "answerChanges": 0,
            "contractErrors": 0,
            "questionClarifications": 0,
            "questionRemovals": 0,
        },
        "changeLedger": ledger_path.name,
        "datasetVersion": "1.0.0",
        "factRegionSha256": fact_region_digest(workbook_path),
        "finalAuditSummary": audit_path.name,
        "generator": {"name": "clarify_ground_truth_contracts", "version": "2.0.0"},
        "groundTruthWorkbook": workbook_name,
        "officialCount": 3,
        "removedIds": [],
        "sourceSha256": "B" * 64,
        "sourceSplitCounts": {"dev": 1, "test": 1, "train": 1},
    }
    manifest_path = parent / "official-manifest.json"
    _write_json(manifest_path, manifest)
    _write_sidecar(manifest_path)
    return parent


def _make_spec(path: Path, *, split: str = "train", old_hash: str | None = None) -> Path:
    old_answer = "冻结测试答案9" if split == "test" else "旧训练答案42"
    sample_id = "SYN-S001" if split == "test" else "SYN-T001"
    payload = {
        "schemaVersion": "1.0",
        "parentVersion": "1.0.0",
        "targetVersion": "1.0.1",
        "entries": [
            {
                "id": sample_id,
                "split": split,
                "oldAnswerSha256": old_hash or sha256_text(old_answer),
                "correctedAnswerText": "修正后的训练答案42",
                "reason": "可复核的答案语义修正",
            }
        ],
    }
    _write_json(path, payload)
    return path


class AmendOfficialGroundTruthTest(unittest.TestCase):
    def test_builds_child_package_and_preserves_test_and_fact_regions(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            parent = _make_parent(root)
            spec = _make_spec(root / "amendments.json")
            output = root / "1.0.1"

            manifest = amend_official_package(parent, spec, output, update_current=True)

            self.assertEqual(manifest["datasetVersion"], "1.0.1")
            self.assertEqual(manifest["parent"]["datasetVersion"], "1.0.0")
            self.assertEqual(manifest["answerAmendmentCount"], 1)
            current = json.loads((root / "CURRENT.json").read_text(encoding="utf-8"))
            self.assertEqual(current["currentVersion"], "1.0.1")
            self.assertEqual(current["officialManifest"], "1.0.1/official-manifest.json")
            self.assertNotIn(b"\r\n", (root / "CURRENT.json").read_bytes())
            child_workbook = output / manifest["groundTruthWorkbook"]
            book = load_workbook(child_workbook, read_only=True, data_only=True)
            try:
                question_rows = list(book["问题答案清单"].iter_rows(values_only=True))
                self.assertEqual(question_rows[1][4], "修正后的训练答案42")
                self.assertEqual(question_rows[3][4], "冻结测试答案9")
                self.assertEqual(
                    list(book["指标数据表"].iter_rows(values_only=True)),
                    [("org_code", "metric_value"), ("ORG001", 42)],
                )
            finally:
                book.close()
            for name in (
                manifest["groundTruthWorkbook"],
                manifest["changeLedger"],
                manifest["answerAmendmentLedger"],
                manifest["finalAuditSummary"],
                "official-manifest.json",
            ):
                target = output / name
                sidecar = target.with_name(target.name + ".sha256")
                expected = sidecar.read_text(encoding="ascii").split()[0]
                self.assertEqual(expected, sha256_file(target))
                self.assertNotIn(b"\r\n", sidecar.read_bytes())
                if target.suffix == ".json" and name != manifest["changeLedger"]:
                    self.assertNotIn(b"\r\n", target.read_bytes())

    def test_rejects_test_split_amendments_without_creating_output(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            parent = _make_parent(root)
            spec = _make_spec(root / "amendments.json", split="test")
            output = root / "1.0.1"

            with self.assertRaisesRegex(OfficialAmendmentError, "test"):
                amend_official_package(parent, spec, output)

            self.assertFalse(output.exists())

    def test_rejects_stale_old_answer_hash_without_creating_output(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            parent = _make_parent(root)
            spec = _make_spec(root / "amendments.json", old_hash="0" * 64)
            output = root / "1.0.1"

            with self.assertRaisesRegex(OfficialAmendmentError, "oldAnswerSha256"):
                amend_official_package(parent, spec, output)

            self.assertFalse(output.exists())

    def test_derived_dataset_changes_only_declared_answer_and_keeps_test_bytes(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            parent_package = _make_parent(root)
            spec = _make_spec(root / "amendments.json")
            child_package = root / "1.0.1"
            amend_official_package(parent_package, spec, child_package)

            parent_dataset = root / "dataset-parent"
            parent_dataset.mkdir()
            train_record = {
                "id": "SYN-T001",
                "question": "训练问题",
                "expected": {"answerText": "旧训练答案42", "columns": ["value"], "rows": [[42]]},
            }
            untouched_record = {
                "id": "SYN-T002",
                "question": "保持不变",
                "expected": {"answerText": "答案8", "columns": ["value"], "rows": [[8]]},
            }
            train_lines = [
                json.dumps(train_record, ensure_ascii=False, sort_keys=True),
                json.dumps(untouched_record, ensure_ascii=False, sort_keys=True),
            ]
            (parent_dataset / "train.jsonl").write_text("\n".join(train_lines) + "\n", encoding="utf-8")
            (parent_dataset / "dev.jsonl").write_text("", encoding="utf-8")
            frozen_test = b'{"id":"SYN-S001","opaque":"frozen"}\r\n'
            (parent_dataset / "test.jsonl").write_bytes(frozen_test)
            (parent_dataset / "augmentation.jsonl").write_text("", encoding="utf-8")
            _write_json(parent_dataset / "schema.json", {"type": "object"})
            _write_json(
                parent_dataset / "manifest.json",
                {
                    "version": "1.0.0",
                    "officialCount": 3,
                    "sourceWorkbook": "bank-nl2sql-ground-truth-v1.0.0.xlsx",
                    "sourceSha256": sha256_file(parent_package / "bank-nl2sql-ground-truth-v1.0.0.xlsx").lower(),
                },
            )
            _write_json(parent_dataset / "gold_manifest.json", {"version": "1.0.0", "officialCount": 3})
            child_dataset = root / "dataset-child"

            amend_derived_dataset(parent_dataset, child_package, spec, child_dataset)

            child_lines = (child_dataset / "train.jsonl").read_text(encoding="utf-8").splitlines()
            self.assertEqual(json.loads(child_lines[0])["expected"]["answerText"], "修正后的训练答案42")
            self.assertEqual(child_lines[1], train_lines[1])
            self.assertNotIn(b"\r\n", (child_dataset / "train.jsonl").read_bytes())
            self.assertEqual((child_dataset / "test.jsonl").read_bytes(), frozen_test)
            child_manifest = json.loads((child_dataset / "manifest.json").read_text(encoding="utf-8"))
            self.assertEqual(child_manifest["version"], "1.0.1")
            self.assertEqual(child_manifest["parentVersion"], "1.0.0")
            self.assertEqual(child_manifest["answerAmendment"]["count"], 1)
            child_gold = json.loads((child_dataset / "gold_manifest.json").read_text(encoding="utf-8"))
            self.assertEqual(child_gold["version"], "1.0.1")

    def test_dataset_loader_rejects_tampered_incremental_amendment_ledger(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            parent = _make_parent(root)
            spec = _make_spec(root / "amendments.json")
            child = root / "1.0.1"
            amend_official_package(parent, spec, child)
            manifest_path = child / "official-manifest.json"
            self.assertEqual(_load_official_manifest(manifest_path)["datasetVersion"], "1.0.1")
            ledger_path = child / "answer-amendment-ledger.json"
            ledger_path.write_text(ledger_path.read_text(encoding="utf-8") + " ", encoding="utf-8")

            with self.assertRaisesRegex(DatasetBuildError, "answerAmendmentLedger"):
                _load_official_manifest(manifest_path)


if __name__ == "__main__":
    unittest.main()
