#!/usr/bin/env python3
"""第二版 GT 审查器合成测试。

全部使用合成题目与合成答案，不包含真实题目或真实答案。覆盖：
单点正确/修正、派生比率正确/错误、全机构升降序排名与 TopN 机构顺序、
缺基期、环比同比 comparison_type、单位与百分点、并列多指标完整性、
部分答案不得 VERIFIED、歧义不得猜测、canonical 门控、重复运行确定性、
源/结构完整性。
"""

from __future__ import annotations

import copy
import hashlib
import json
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook
from unittest import mock

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from audit_ground_truth import (  # noqa: E402
    WorkbookData,
    EVIDENCE_RULES_VERSION,
    fact_digest,
    finalize_evidence,
    parse_question,
    run_audit,
    validate_claim_evidence,
    validate_review_evidence,
)
import gt_answer_rules as rules  # noqa: E402

ORG_NAMES = {
    "ORG001": "江苏省A市农商行",
    "ORG002": "江苏省B市农商行",
    "ORG003": "江苏省C市农商行",
    "ORG004": "江苏省D市农商行",
    "ORG005": "江苏省E市农商行",
}
METRICS = [
    ("ZB001", "各项存款余额", "指机构吸收的全部存款期末余额", "亿元"),
    ("ZB002", "各项贷款余额", "指机构发放的全部贷款期末余额", "亿元"),
    ("ZB003", "对公存款余额", "指机构吸收的企业等单位存款期末余额", "亿元"),
    ("ZB004", "个人存款余额", "指机构吸收的个人储蓄存款期末余额", "亿元"),
    ("ZB011", "净利润", "指机构最终盈利", "万元"),
    ("ZB013", "不良贷款率", "指不良贷款余额占各项贷款余额的比例", "%"),
    ("ZB018", "员工人数", "指机构在岗正式员工总数", "人"),
]
DAYS = [
    "2024-12-31",
    "2025-01-31",
    "2025-02-28",
    "2025-03-31",
    "2025-04-30",
    "2026-03-31",
    "2026-04-30",
]
DERIVED_ROWS = [
    ("较年初", "当日值 - 2024年12月31日值"),
    ("较上季", "当日值 - 上季度末值"),
    ("较上月", "当日值 - 上月月末值"),
    ("较同期", "当日值 - 去年同期值"),
    ("全省均值", "13家机构当日值的算数平均值"),
    ("排名", "不良贷款率/逾期贷款率/成本收入比按从低到高排名（越低越好），其余按从高到低排名，使用rank算法"),
    ("增量", "当日值 - 比较日值"),
    ("增幅", "(当日值-比较日值)/比较日值×100%，比率类指标（ZB012/013/015/016/017）不做增幅计算"),
    ("表现较好", "前三"),
    ("表现较差", "后四"),
]


def fact_value(org_code: str, metric_code: str, day_index: int) -> float:
    org_index = int(org_code[-1]) - 1
    base = {
        "ZB001": 100.0 - 10.0 * org_index,
        "ZB002": 90.0 - 10.0 * org_index,
        "ZB003": 40.0 - 5.0 * org_index,
        "ZB004": 60.0 - 5.0 * org_index,
        "ZB011": 500.0 - 100.0 * org_index,
        "ZB013": 0.5 + 0.1 * org_index,
        "ZB018": 100.0 + 10.0 * org_index,
    }[metric_code]
    drift = {"ZB001": 0.5, "ZB002": 0.3, "ZB003": 0.2, "ZB004": 0.3, "ZB011": 10.0, "ZB013": 0.01, "ZB018": 1.0}[metric_code]
    return base + drift * day_index


def build_synthetic_workbook(path: Path, questions: list[tuple[str, str, str, str]], overrides: dict[tuple[str, str, str], float] | None = None) -> Path:
    """写合成工作簿：机构表/指标表/衍生维度说明/事实表/问题答案清单。overrides 覆盖 (日期, 机构, 指标) 事实值。"""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "机构信息表"
    sheet.append(["机构编号", "机构名称"])
    for code, name in ORG_NAMES.items():
        sheet.append([code, name])

    metric_sheet = workbook.create_sheet("指标清单表")
    metric_sheet.append(["指标编号", "指标名称", "指标含义", "指标单位"])
    for row in METRICS:
        metric_sheet.append(list(row))

    derived_sheet = workbook.create_sheet("衍生维度说明")
    derived_sheet.append(["衍生维度", "衍生口径说明"])
    for row in DERIVED_ROWS:
        derived_sheet.append(list(row))

    fact_sheet = workbook.create_sheet("指标数据表")
    fact_sheet.append(["数据日期", "指标编号", "指标名称", "机构编号", "指标值"])
    metric_names = {code: name for code, name, _meaning, _unit in METRICS}
    for day_index, day in enumerate(DAYS):
        for org_code in ORG_NAMES:
            for metric_code, metric_name, _meaning, _unit in METRICS:
                value = fact_value(org_code, metric_code, day_index)
                if overrides and (day, org_code, metric_code) in overrides:
                    value = overrides[(day, org_code, metric_code)]
                fact_sheet.append([day, metric_code, metric_names[metric_code], org_code, value])

    question_sheet = workbook.create_sheet("问题答案清单")
    question_sheet.append(["问题编号", "问题类型", "问题难度", "问题描述", "问题结果"])
    for index, (qid, split, difficulty, question, answer) in enumerate(questions, start=1):
        question_sheet.append([qid or f"SYN-{index:03d}", split, difficulty, question, answer])

    workbook.save(path)
    return path


def run_synthetic(questions: list[tuple[str, str, str, str]], expected_total: int | None = None, expected_splits: dict[str, int] | None = None, overrides: dict[tuple[str, str, str], float] | None = None) -> dict:
    tmp = tempfile.TemporaryDirectory()
    base = Path(tmp.name)
    workbook_path = build_synthetic_workbook(base / "syn.xlsx", questions, overrides=overrides)
    output_dir = base / "out"
    digest = hashlib.sha256(workbook_path.read_bytes()).hexdigest().upper()
    try:
        return run_audit(workbook_path, output_dir, True, expected_total=expected_total, expected_splits=expected_splits, expected_source_sha=digest)
    finally:
        tmp.cleanup()


def run_synthetic_with_reviews(questions: list[tuple[str, str, str, str]], overrides: dict[tuple[str, str, str], float] | None = None) -> tuple[dict, list[dict]]:
    """运行合成审查并返回 (summary, reviews)；review 保留 evidence 供反证测试篡改。"""
    tmp = tempfile.TemporaryDirectory()
    base = Path(tmp.name)
    workbook_path = build_synthetic_workbook(base / "syn.xlsx", questions, overrides=overrides)
    output_dir = base / "out"
    digest = hashlib.sha256(workbook_path.read_bytes()).hexdigest().upper()
    try:
        summary = run_audit(workbook_path, output_dir, True, expected_source_sha=digest)
        reviews = [json.loads(line) for line in (output_dir / "review.ndjson").read_text(encoding="utf-8").splitlines() if line.strip()]
        return summary, reviews
    finally:
        tmp.cleanup()


def synthetic_data(questions: list[tuple[str, str, str, str]]) -> "WorkbookData":
    tmp = tempfile.TemporaryDirectory()
    base = Path(tmp.name)
    workbook_path = build_synthetic_workbook(base / "syn.xlsx", questions)
    data = WorkbookData(workbook_path)
    tmp.cleanup()
    return data


def synthetic_claim(key: str, kind: str, value: float | None, evidence: dict | None) -> dict:
    return {
        "key": key, "label": key, "kind": kind, "value": value, "unit": None,
        "rounding": 2, "metric": None, "metricName": None, "org": None, "orgName": None,
        "date": None, "baseline": None, "comparisonType": None, "direction": None,
        "role": None, "mustAppear": True, "extras": [], "note": None,
        "evidence": evidence, "matched": False, "matchNote": None,
    }


def status_of(summary: dict, qid: str) -> str:
    index = summary["verifiedIds"] + summary["correctedIds"] + summary["unresolvedIds"]
    return "UNRESOLVED" if qid in summary["unresolvedIds"] else "CORRECTED" if qid in summary["correctedIds"] else "VERIFIED"


class TotalRoleOccurrenceLocalTest(unittest.TestCase):
    """total 角色归因必须 occurrence-local：'共N家/天/个' 只标注紧邻的 occurrence，
    '总天数：N天/总天数为N天/总天数N天' 显式标签不被相邻 '占比/比例' 覆盖，
    等值出现不得互相传染。全部为合成文本。"""

    def test_total_role_marks_only_explicit_occurrence(self) -> None:
        # 等值双出现 + 百分比：只有显式 '共13天' 标注的第二个 occurrence 是 total，
        # 第一个不得因为全局等值搜索被标 total。
        tokens = rules.tokenize_answer("高于全省均值13天（共13天），占比100.0%")
        self.assertEqual([t.raw for t in tokens], ["13", "13", "100.0"])
        totals = [t for t in tokens if t.role == "total"]
        self.assertEqual([t.raw for t in totals], ["13"], "只有显式标注的 occurrence 是 total")
        self.assertNotEqual(tokens[0].role, "total", "未标注的等值 occurrence 不得被标 total")
        self.assertEqual(tokens[1].role, "total")
        self.assertEqual(tokens[2].role, "ratio", "百分比保持 ratio 角色")

    def test_total_gong_forms_preserved(self) -> None:
        # 已接受的 '共N天/家/个' 文本形式（含空白）必须保留
        for answer, expected_total_raw in (
            ("共13天", ["13"]),
            ("一共 13 天", ["13"]),
            ("全省农商行共13家，第7名", ["13"]),
            ("共6个", ["6"]),
        ):
            totals = [t for t in rules.tokenize_answer(answer) if t.role == "total"]
            self.assertEqual([t.raw for t in totals], expected_total_raw, answer)

    def test_total_day_label_not_overridden_by_ratio(self) -> None:
        # '总天数：N天' / '总天数为N天' / '总天数N天' 三种形式；
        # 相邻 '占比' 文本不得把显式 total token 覆盖为 ratio。
        for answer in (
            "高于全省均值的天数：4天；总天数：4天；占比：100%",
            "高于全省均值的天数：4天；总天数为4天；占比：100%",
            "高于全省均值的天数：4天；总天数4天；占比：100%",
        ):
            tokens = rules.tokenize_answer(answer)
            totals = [t for t in tokens if t.role == "total"]
            self.assertEqual([t.raw for t in totals], ["4"], answer)
            ratios = [t for t in tokens if t.role == "ratio"]
            self.assertEqual([t.raw for t in ratios], ["100"], answer)

    def test_days_above_avg_with_total_label_verified_end_to_end(self) -> None:
        # 正回归：above=total=4、占比 100% 的合成天数题，答案显式声明 '总天数：4天'。
        # 修复前（无 occurrence-local total 检测）该答案被 CORRECTED；修复后 VERIFIED。
        question = "2025年全年，江苏省A市农商行的各项存款余额有多少天高于全省均值？"
        answer = "高于全省均值的天数：4天；总天数：4天；占比：100%"
        summary = run_synthetic([("TOT1", "训练集", "复杂", question, answer)])
        self.assertEqual(status_of(summary, "TOT1"), "VERIFIED", summary["statusCounts"])

    def test_days_above_avg_without_total_label_stays_corrected(self) -> None:
        # 负回归：缺少显式 total-day 声明时，即使存在等值声明（4天高于均值）和
        # 百分比（100%），仍必须 CORRECTED——不得因等值数值猜测 total。
        question = "2025年全年，江苏省A市农商行的各项存款余额有多少天高于全省均值？"
        answer = "高于全省均值的天数：4天；占比：100%"
        summary = run_synthetic([("TOT2", "训练集", "复杂", question, answer)])
        self.assertEqual(status_of(summary, "TOT2"), "CORRECTED", summary["statusCounts"])


class SyntheticWorkbookTest(unittest.TestCase):
    def test_point_verified(self) -> None:
        summary = run_synthetic(
            [("S1", "训练集", "简单", "江苏省A市农商行在2025-01-31的各项存款余额是多少？", "100.5亿元")]
        )
        self.assertEqual(summary["totalRecords"], 1)
        self.assertEqual(status_of(summary, "S1"), "VERIFIED")
        self.assertEqual(summary["statusCounts"]["VERIFIED"], 1)

    def test_point_corrected(self) -> None:
        summary = run_synthetic(
            [("S1", "训练集", "简单", "江苏省A市农商行在2025-01-31的各项存款余额是多少？", "999亿元")]
        )
        self.assertEqual(status_of(summary, "S1"), "CORRECTED")
        self.assertEqual(summary["statusCounts"]["CORRECTED"], 1)
        self.assertTrue(summary["canonicalReady"] is False)

    def test_ratio_verified_and_corrected(self) -> None:
        # 存贷比 = ZB002/ZB001：ORG001 @ 2025-01-31 = 90.3/100.5 = 89.85%
        questions = [
            ("R1", "训练集", "普通", "合成核验场景：以江苏省A市农商行在2025年1月末的各项贷款余额除以同日各项存款余额，计算该行存贷比。", "89.85%"),
            ("R2", "训练集", "普通", "江苏省B市农商行在2025-01-31的存贷比是多少？", "12.34%"),
        ]
        summary = run_synthetic(questions)
        self.assertEqual(status_of(summary, "R1"), "VERIFIED")
        self.assertEqual(status_of(summary, "R2"), "CORRECTED")

    def test_rank_ascending_and_descending(self) -> None:
        questions = [
            ("A1", "训练集", "简单", "2025-01-31，哪家农商行的不良贷款率最低？", "江苏省A市农商行，0.51%"),
            ("A2", "训练集", "简单", "2025-01-31，哪家农商行的各项存款余额最高？", "江苏省A市农商行，100.5亿元"),
        ]
        summary = run_synthetic(questions)
        self.assertEqual(status_of(summary, "A1"), "VERIFIED")
        self.assertEqual(status_of(summary, "A2"), "VERIFIED")

    def test_topn_order(self) -> None:
        questions = [
            ("T1", "训练集", "普通", "2025-01-31，各项存款余额排名前三的是哪几家？各多少？", "前3名：江苏省A市农商行(100.5亿元)、江苏省B市农商行(90.5亿元)、江苏省C市农商行(80.5亿元)"),
            ("T2", "训练集", "普通", "2025-01-31，各项存款余额排名前三的是哪几家？各多少？", "前3名：江苏省B市农商行(90.5亿元)、江苏省A市农商行(100.5亿元)、江苏省C市农商行(80.5亿元)"),
            ("T3", "训练集", "普通", "2025-01-31，各项存款余额排名前三的是哪几家？各多少？", "前2名：江苏省A市农商行(100.5亿元)、江苏省B市农商行(90.5亿元)"),
        ]
        summary = run_synthetic(questions)
        self.assertEqual(status_of(summary, "T1"), "VERIFIED", "TopN 机构顺序正确应 VERIFIED")
        self.assertEqual(status_of(summary, "T2"), "CORRECTED", "TopN 机构顺序错误不得 VERIFIED")
        self.assertEqual(status_of(summary, "T3"), "CORRECTED", "TopN 截断（只列前2名）不得 VERIFIED")

    def test_topn_tie_order(self) -> None:
        # ORG001/ORG002/ORG003 并列第1（100.5）：并列组内顺序任意，名次/值逐项一致
        overrides = {("2025-01-31", "ORG002", "ZB001"): 100.5, ("2025-01-31", "ORG003", "ZB001"): 100.5}
        questions = [
            ("TE1", "训练集", "普通", "2025-01-31，各项存款余额排名前三的是哪几家？各多少？", "第1名：江苏省A市农商行(100.5亿元)、第1名：江苏省B市农商行(100.5亿元)、第1名：江苏省C市农商行(100.5亿元)"),
            ("TE2", "训练集", "普通", "2025-01-31，各项存款余额排名前三的是哪几家？各多少？", "第1名：江苏省B市农商行(100.5亿元)、第1名：江苏省A市农商行(100.5亿元)、第1名：江苏省C市农商行(100.5亿元)"),
        ]
        summary = run_synthetic(questions, overrides=overrides)
        self.assertEqual(status_of(summary, "TE1"), "VERIFIED", "并列第1顺序任意应 VERIFIED")
        self.assertEqual(status_of(summary, "TE2"), "VERIFIED", "并列第1顺序任意应 VERIFIED")

    def test_missing_baseline(self) -> None:
        # 同比需要去年同期（2024-01-31），合成事实表缺失 -> UNRESOLVED
        summary = run_synthetic(
            [("M1", "训练集", "普通", "江苏省A市农商行的各项存款余额在2025-01-31，同比（较去年同期）变动了多少？", "增加1亿元")]
        )
        self.assertEqual(status_of(summary, "M1"), "UNRESOLVED")
        self.assertIn("M1", summary["unresolvedIds"])

    def test_mom_yoy_comparison_type(self) -> None:
        # C1: 环比基期 2026-03-31、同比基期 2025-04-30 都存在
        # ZB011 ORG001: idx6=560, idx5=550, idx4=540 -> mom=(560-550)/550=1.82%, yoy=(560-540)/540=3.7%
        # C2: 同比基期 2024-01-31 在合成事实表中缺失 -> UNRESOLVED
        questions = [
            ("C1", "训练集", "复杂", "江苏省A市农商行在2026-04-30的净利润环比和同比分别变动了多少？", "环比增长1.82%，同比增长3.7%，当前560万元"),
            ("C2", "训练集", "复杂", "江苏省A市农商行在2025-01-31的净利润环比和同比分别变动了多少？", "环比增长2%，同比增长X%"),
        ]
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            workbook_path = build_synthetic_workbook(base / "syn.xlsx", questions)
            output_dir = base / "out"
            digest = hashlib.sha256(workbook_path.read_bytes()).hexdigest().upper()
            summary = run_audit(workbook_path, output_dir, True, expected_total=2, expected_splits={"train": 2, "dev": 0, "test": 0}, expected_source_sha=digest)
            self.assertEqual(status_of(summary, "C1"), "VERIFIED")
            self.assertEqual(status_of(summary, "C2"), "UNRESOLVED", "缺同比基期不得 VERIFIED/CORRECTED")
            reviews = [json.loads(line) for line in (output_dir / "review.ndjson").read_text(encoding="utf-8").splitlines() if line.strip()]
            c1 = next(review for review in reviews if review["id"] == "C1")
            pct_claims = [claim for claim in c1["claims"] if claim["kind"] == "PCT_CHANGE"]
            comparison_types = {claim["comparisonType"] for claim in pct_claims}
            self.assertEqual(comparison_types, {"mom", "yoy"})
            c2 = next(review for review in reviews if review["id"] == "C2")
            self.assertEqual(c2["status"], "UNRESOLVED")
            self.assertTrue(c2["unresolvedReason"] and "基期" in c2["unresolvedReason"])

    def test_unit_and_point(self) -> None:
        # 不良率为比率类指标，变化以百分点计：ORG002 idx1=0.61, idx0=0.60 -> +0.01 个百分点
        summary = run_synthetic(
            [("P1", "训练集", "普通", "江苏省B市农商行的不良贷款率从年初到2025-01-31变动了几个百分点？", "上升0.01个百分点")]
        )
        self.assertEqual(status_of(summary, "P1"), "VERIFIED")

    def test_identity_tie_multi_metric(self) -> None:
        # 对公+个人 = 各项（ZB003+ZB004=ZB001 恒等式成立）
        questions = [
            ("I1", "训练集", "简单", "2025年1月末，江苏省A市农商行的对公存款加个人存款是不是等于各项存款？差额多少？", "对公40.2+个人60.3=100.5，各项100.5，差额0.0亿"),
            ("I2", "训练集", "简单", "2025年1月末，江苏省A市农商行的对公存款加个人存款是不是等于各项存款？差额多少？", "对公40.2+个人60.3=100.5，各项100.5"),
        ]
        summary = run_synthetic(questions)
        self.assertEqual(status_of(summary, "I1"), "VERIFIED", "恒等式全部声明一致应 VERIFIED")
        self.assertEqual(status_of(summary, "I2"), "CORRECTED", "恒等式缺失差额声明不得 VERIFIED")

    def test_partial_answer_not_verified(self) -> None:
        # 多指标单点：答案只含一个值 -> 不得 VERIFIED
        summary = run_synthetic(
            [("X1", "训练集", "普通", "江苏省A市农商行在2025-01-31的不良贷款率和净利润分别是多少？", "不良贷款率0.51%")]
        )
        self.assertEqual(status_of(summary, "X1"), "CORRECTED")

    def test_ambiguous_not_guessed(self) -> None:
        # 表现较好/较差题未列出指标全集 -> UNRESOLVED（不得猜测）。
        # 必须从逐条 review 记录验证 unresolvedReason 非空，不能仅检查 unresolvedIds。
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            workbook_path = build_synthetic_workbook(
                base / "syn.xlsx",
                [("Y1", "训练集", "复杂", "江苏省A市农商行在2025-01-31的指标中哪些表现较好？哪些表现较差？", "表现较好：无。表现较差：无。")],
            )
            output_dir = base / "out"
            digest = hashlib.sha256(workbook_path.read_bytes()).hexdigest().upper()
            summary = run_audit(workbook_path, output_dir, True, expected_total=1, expected_splits={"train": 1, "dev": 0, "test": 0}, expected_source_sha=digest)
            self.assertIn("Y1", summary["unresolvedIds"])
            reviews = [json.loads(line) for line in (output_dir / "review.ndjson").read_text(encoding="utf-8").splitlines() if line.strip()]
            review = next(item for item in reviews if item["id"] == "Y1")
            self.assertEqual(review["status"], "UNRESOLVED")
            self.assertTrue(review["unresolvedReason"], "UNRESOLVED 记录必须给出非空 unresolvedReason")

    def test_canonical_gate_true(self) -> None:
        questions = [
            ("C1", "训练集", "简单", "江苏省A市农商行在2025-01-31的各项存款余额是多少？", "100.5亿元"),
            ("C2", "训练集", "简单", "江苏省B市农商行在2025-01-31的各项存款余额是多少？", "90.5亿元"),
            ("C3", "训练集", "简单", "江苏省C市农商行在2025-01-31的各项存款余额是多少？", "80.5亿元"),
        ]
        summary = run_synthetic(
            questions,
            expected_total=3,
            expected_splits={"train": 3, "dev": 0, "test": 0},
        )
        self.assertTrue(summary["canonicalReady"])
        self.assertEqual(summary["outputFile"], "canonical-corrected.xlsx")
        self.assertEqual(summary["fullEvidence"], 3)

    def test_canonical_gate_false_on_unresolved(self) -> None:
        summary = run_synthetic(
            [
                ("C1", "训练集", "简单", "江苏省A市农商行在2025-01-31的各项存款余额是多少？", "100.5亿元"),
                ("U1", "训练集", "复杂", "江苏省A市农商行在2025-01-31的指标中哪些表现较好？", "无"),
            ],
            expected_total=2,
            expected_splits={"train": 2, "dev": 0, "test": 0},
        )
        self.assertFalse(summary["canonicalReady"])
        self.assertEqual(summary["outputFile"], "candidate-reviewed.xlsx")

    def test_determinism(self) -> None:
        questions = [
            ("D1", "训练集", "简单", "江苏省A市农商行在2025-01-31的各项存款余额是多少？", "100.5亿元"),
            ("D2", "训练集", "普通", "合成核验场景：以江苏省A市农商行在2025年1月末的各项贷款余额除以同日各项存款余额，计算该行存贷比。", "89.85%"),
        ]
        first = run_synthetic(questions)
        second = run_synthetic(questions)
        self.assertEqual(first["statusCounts"], second["statusCounts"])
        self.assertEqual(first["splitCounts"], second["splitCounts"])
        self.assertEqual(first["outputSha256"], second["outputSha256"])
    def test_source_sha_mismatch(self) -> None:
        tmp = tempfile.TemporaryDirectory()
        base = Path(tmp.name)
        workbook_path = build_synthetic_workbook(
            base / "syn.xlsx",
            [("S1", "训练集", "简单", "江苏省A市农商行在2025-01-31的各项存款余额是多少？", "100.5亿元")],
        )
        # 篡改事实值使源哈希变化：以篡改前的哈希作为期望 -> 必须报错
        original_digest = hashlib.sha256(workbook_path.read_bytes()).hexdigest().upper()
        loaded = load_workbook(workbook_path)
        sheet = loaded["指标数据表"]
        sheet.cell(row=3, column=5, value=999.0)
        loaded.save(workbook_path)
        with self.assertRaises(ValueError):
            run_audit(workbook_path, base / "out", True, expected_source_sha=original_digest)
        tmp.cleanup()

    def test_structure_integrity(self) -> None:
        tmp = tempfile.TemporaryDirectory()
        base = Path(tmp.name)
        workbook_path = build_synthetic_workbook(
            base / "syn.xlsx",
            [("S1", "训练集", "简单", "江苏省A市农商行在2025-01-31的各项存款余额是多少？", "100.5亿元")],
        )
        # 表头错误 -> ValueError
        loaded = load_workbook(workbook_path)
        loaded["问题答案清单"].cell(row=1, column=1, value="错误表头")
        loaded.save(workbook_path)
        with self.assertRaises(ValueError):
            run_audit(workbook_path, base / "out2", True)
        tmp.cleanup()

    def test_output_structure(self) -> None:
        tmp = tempfile.TemporaryDirectory()
        base = Path(tmp.name)
        workbook_path = build_synthetic_workbook(
            base / "syn.xlsx",
            [("S1", "训练集", "简单", "江苏省A市农商行在2025-01-31的各项存款余额是多少？", "100.5亿元")],
        )
        output_dir = base / "out"
        digest = hashlib.sha256(workbook_path.read_bytes()).hexdigest().upper()
        summary = run_audit(workbook_path, output_dir, True, expected_total=1, expected_splits={"train": 1, "dev": 0, "test": 0}, expected_source_sha=digest)
        for filename in ("audit-summary.json", "review.ndjson", "correction-ledger.json", "canonical-corrected.xlsx"):
            self.assertTrue((output_dir / filename).is_file(), filename)
        for key in ("ruleVersion", "sourceSha256", "totalRecords", "uniqueIds", "splitCounts", "statusCounts", "fullEvidence", "auditErrors", "canonicalReady", "outputFile", "outputSha256"):
            self.assertIn(key, summary)
        tmp.cleanup()


# ---------------------------------------------------------------------------
# 证据契约反证测试（全部合成题目/答案/ID；不包含任何真实 TRAIN/VAL/TST 数据）。
# 覆盖：8 类派生声明 evidence 完整性、篡改必失败、事实变更传播、
# dependsOn 缺失/成环、主流程失败注入、summary/输出契约、合成纯度。

POINT_QUESTION = "江苏省A市农商行在2025-01-31的各项存款余额是多少？"
RATIO_QUESTION = "合成核验场景：以江苏省A市农商行在2025年1月末的各项贷款余额除以同日各项存款余额，计算该行存贷比。"
DELTA_QUESTION = "江苏省A市农商行在2025-01-31的各项存款余额比年初变化了多少？"
MOM_YOY_QUESTION = "江苏省A市农商行在2026-04-30的净利润环比和同比分别变动了多少？"
DAILY_MEAN_QUESTION = "江苏省A市农商行在2025年一季度的各项存款余额日均是多少？"
DAYS_ABOVE_QUESTION = "2025年全年，江苏省A市农商行的各项存款余额有多少天高于全省均值？"
TOP_DECLINE_QUESTION = "2025-01-31，不良贷款率下降幅度最大的前3家农商行是哪些？"
TREND_QUESTION = "江苏省A市农商行从2024年第四季度到2025年第一季度的各项存款余额季度变化趋势如何？"


def parse_evidence(question_text: str, overrides: dict[tuple[str, str, str], float] | None = None) -> tuple[object, dict, WorkbookData]:
    """解析单个合成题目，填充并验证 evidence，返回 (parsed, validation, data)。"""
    tmp = tempfile.TemporaryDirectory()
    base = Path(tmp.name)
    workbook_path = build_synthetic_workbook(
        base / "syn.xlsx",
        [("SYN-001", "训练集", "简单", question_text, "合成答案")],
        overrides=overrides,
    )
    data = WorkbookData(workbook_path)
    parsed = parse_question(question_text, data)
    finalize_evidence(parsed.claims, data)
    validation = validate_review_evidence(parsed.claims, data)
    tmp.cleanup()
    return parsed, validation, data


def claim_by_key(claims: list[dict], key: str) -> dict:
    return next(claim for claim in claims if claim["key"] == key)


def assert_evidence_base(testcase: unittest.TestCase, claim: dict, data: WorkbookData, claims_by_key: dict[str, dict] | None = None) -> None:
    """evidence 基础契约：operation/formula/rounding/sourceFactCount/sourceFactsSha256 非空且验证通过。"""
    ev = claim["evidence"]
    testcase.assertTrue(ev.get("operation"), "evidence 必须带 operation")
    testcase.assertTrue(ev.get("formula"), "evidence 必须带 formula")
    testcase.assertEqual(ev.get("rounding"), claim.get("rounding"), "evidence.rounding 必须与 claim.rounding 一致")
    testcase.assertGreaterEqual(ev.get("sourceFactCount"), 1, "sourceFactCount 必须非空")
    testcase.assertTrue(ev.get("sourceFactsSha256"), "sourceFactsSha256 必须非空")
    ok, errors = validate_claim_evidence(claim, data, claims_by_key)
    testcase.assertTrue(ok, errors)


class EvidenceContractTest(unittest.TestCase):
    """对 POINT/RATIO/DELTA/MOM_YOY/DAILY_MEAN/DAYS_ABOVE_MEAN/TOP_N_DECLINE/TREND
    派生声明逐项断言 evidence.operation、事实范围或 dependsOn、formula/rounding、
    sourceFactCount、sourceFactsSha256，且 validate_review_evidence.valid=True。"""

    def test_point_evidence(self) -> None:
        parsed, validation, data = parse_evidence(POINT_QUESTION)
        self.assertTrue(validation["valid"], validation["errors"])
        claim = parsed.claims[0]
        ev = claim["evidence"]
        self.assertEqual(ev["operation"], "VALUE_LOOKUP")
        self.assertEqual(ev["metric"], "ZB001")
        self.assertEqual(ev["org"], "ORG001")
        self.assertEqual(ev["date"], "2025-01-31")
        self.assertEqual(ev["sourceFactCount"], 1)
        self.assertEqual(ev["sourceFactsSha256"], fact_digest({("2025-01-31", "ORG001", "ZB001"): 100.5}))
        assert_evidence_base(self, claim, data)

    def test_ratio_evidence(self) -> None:
        parsed, validation, data = parse_evidence(RATIO_QUESTION)
        self.assertTrue(validation["valid"], validation["errors"])
        claim = claim_by_key(parsed.claims, "存贷比")
        ev = claim["evidence"]
        self.assertEqual(ev["operation"], "RATIO")
        self.assertEqual(ev["numerator"], "ZB002")
        self.assertEqual(ev["denominator"], "ZB001")
        self.assertEqual(ev["org"], "ORG001")
        self.assertEqual(ev["date"], "2025-01-31")
        self.assertEqual(ev["sourceFactCount"], 2)
        self.assertEqual(ev["sourceFactsSha256"], fact_digest({("2025-01-31", "ORG001", "ZB002"): 90.3, ("2025-01-31", "ORG001", "ZB001"): 100.5}))
        assert_evidence_base(self, claim, data)

    def test_delta_evidence(self) -> None:
        parsed, validation, data = parse_evidence(DELTA_QUESTION)
        self.assertTrue(validation["valid"], validation["errors"])
        claim = claim_by_key(parsed.claims, "delta")
        ev = claim["evidence"]
        self.assertEqual(ev["operation"], "DELTA")
        self.assertEqual(ev["metric"], "ZB001")
        self.assertEqual(ev["org"], "ORG001")
        self.assertEqual(ev["date"], "2025-01-31")
        self.assertEqual(ev["baseline"], "2024-12-31")
        self.assertEqual(claim["value"], 0.5)
        self.assertEqual(ev["sourceFactCount"], 2)
        self.assertEqual(ev["sourceFactsSha256"], fact_digest({("2025-01-31", "ORG001", "ZB001"): 100.5, ("2024-12-31", "ORG001", "ZB001"): 100.0}))
        assert_evidence_base(self, claim, data)

    def test_mom_yoy_evidence(self) -> None:
        parsed, validation, data = parse_evidence(MOM_YOY_QUESTION)
        self.assertTrue(validation["valid"], validation["errors"])
        pct_mom = claim_by_key(parsed.claims, "pct_mom")
        pct_yoy = claim_by_key(parsed.claims, "pct_yoy")
        for claim, expected_baseline, expected_type in ((pct_mom, "2026-03-31", "mom"), (pct_yoy, "2025-04-30", "yoy")):
            ev = claim["evidence"]
            self.assertEqual(ev["operation"], "PCT_CHANGE")
            self.assertEqual(ev["metric"], "ZB011")
            self.assertEqual(ev["org"], "ORG001")
            self.assertEqual(ev["date"], "2026-04-30")
            self.assertEqual(ev["baseline"], expected_baseline)
            self.assertEqual(ev["comparisonType"], expected_type)
            self.assertEqual(ev["sourceFactCount"], 2)
            self.assertEqual(ev["sourceFactsSha256"], fact_digest({("2026-04-30", "ORG001", "ZB011"): 560.0, (expected_baseline, "ORG001", "ZB011"): {"2026-03-31": 550.0, "2025-04-30": 540.0}[expected_baseline]}))
            assert_evidence_base(self, claim, data)

    def test_daily_mean_evidence(self) -> None:
        parsed, validation, data = parse_evidence(DAILY_MEAN_QUESTION)
        self.assertTrue(validation["valid"], validation["errors"])
        claim = claim_by_key(parsed.claims, "mean")
        ev = claim["evidence"]
        self.assertEqual(ev["operation"], "RANGE_MEAN")
        self.assertEqual(ev["metric"], "ZB001")
        self.assertEqual(ev["org"], "ORG001")
        self.assertEqual(ev["dateRange"], ["2025-01-31", "2025-03-31"], "DAILY_MEAN 必须带 dateRange")
        self.assertEqual(claim["value"], 101.0)
        self.assertEqual(ev["sourceFactCount"], 3)
        self.assertEqual(ev["sourceFactsSha256"], fact_digest({("2025-01-31", "ORG001", "ZB001"): 100.5, ("2025-02-28", "ORG001", "ZB001"): 101.0, ("2025-03-31", "ORG001", "ZB001"): 101.5}))
        assert_evidence_base(self, claim, data)

    def test_days_above_mean_evidence(self) -> None:
        parsed, validation, data = parse_evidence(DAYS_ABOVE_QUESTION)
        self.assertTrue(validation["valid"], validation["errors"])
        count_claim = claim_by_key(parsed.claims, "count")
        ev = count_claim["evidence"]
        self.assertEqual(ev["operation"], "DAYS_ABOVE_MEAN")
        self.assertEqual(ev["metric"], "ZB001")
        self.assertEqual(ev["org"], "ORG001")
        self.assertEqual(ev["dateRange"], ["2025-01-01", "2025-12-31"])
        self.assertEqual(sorted(ev["scope"]["orgs"]), ["ORG001", "ORG002", "ORG003", "ORG004", "ORG005"])
        self.assertEqual(count_claim["value"], 4.0)
        self.assertEqual(ev["sourceFactCount"], 20, "4 个有效日 × 5 家机构")
        assert_evidence_base(self, count_claim, data)
        total_claim = claim_by_key(parsed.claims, "total")
        pct_claim = claim_by_key(parsed.claims, "pct")
        assert_evidence_base(self, total_claim, data)
        assert_evidence_base(self, pct_claim, data)
        self.assertEqual(pct_claim["evidence"]["rounding"], pct_claim["rounding"])

    def test_top_n_decline_evidence(self) -> None:
        parsed, validation, data = parse_evidence(TOP_DECLINE_QUESTION)
        self.assertTrue(validation["valid"], validation["errors"])
        entry1 = claim_by_key(parsed.claims, "entry1")
        ev = entry1["evidence"]
        self.assertEqual(ev["operation"], "TOP_N_DECLINE")
        self.assertEqual(ev["metric"], "ZB013")
        self.assertEqual(ev["date"], "2025-01-31")
        self.assertEqual(ev["baseline"], "2024-12-31")
        self.assertEqual(sorted(ev["scope"]["orgs"]), ["ORG001", "ORG002", "ORG003", "ORG004", "ORG005"])
        self.assertEqual(entry1.get("position"), 1)
        self.assertEqual(ev["sourceFactCount"], 10, "2 个日期 × 5 家机构")
        assert_evidence_base(self, entry1, data)
        for index in (2, 3):
            assert_evidence_base(self, claim_by_key(parsed.claims, f"entry{index}"), data)

    def test_trend_evidence(self) -> None:
        parsed, validation, data = parse_evidence(TREND_QUESTION)
        self.assertTrue(validation["valid"], validation["errors"])
        trend = claim_by_key(parsed.claims, "trend")
        ev = trend["evidence"]
        self.assertEqual(ev["operation"], "TREND_SLOPE")
        self.assertEqual(ev["dependsOn"], ["q_2024-12-31", "q_2025-03-31"])
        self.assertEqual(trend["value"], 1.0)
        self.assertEqual(trend["direction"], "up")
        self.assertEqual(ev["sourceFactCount"], 2, "依赖两个季度末点值")
        self.assertEqual(ev["sourceFactsSha256"], fact_digest({("2024-12-31", "ORG001", "ZB001"): 100.0, ("2025-03-31", "ORG001", "ZB001"): 101.5}))
        assert_evidence_base(self, trend, data, claims_by_key={claim["key"]: claim for claim in parsed.claims})


class EvidenceTamperingTest(unittest.TestCase):
    """分别删除 operation/dateRange/baseline、篡改 sourceFactsSha256 等，
    validate_claim_evidence 必须返回 false 且给出可识别错误。"""

    def test_tamper_remove_operation(self) -> None:
        parsed, validation, data = parse_evidence(POINT_QUESTION)
        claim = copy.deepcopy(parsed.claims[0])
        del claim["evidence"]["operation"]
        ok, errors = validate_claim_evidence(claim, data)
        self.assertFalse(ok)
        self.assertTrue(any("operation" in error for error in errors), errors)

    def test_tamper_remove_date_range(self) -> None:
        parsed, validation, data = parse_evidence(DAILY_MEAN_QUESTION)
        claim = copy.deepcopy(claim_by_key(parsed.claims, "mean"))
        del claim["evidence"]["dateRange"]
        ok, errors = validate_claim_evidence(claim, data)
        self.assertFalse(ok)
        self.assertTrue(any("dateRange" in error for error in errors), errors)

    def test_tamper_remove_baseline(self) -> None:
        parsed, validation, data = parse_evidence(MOM_YOY_QUESTION)
        for key in ("pct_mom", "pct_yoy"):
            claim = copy.deepcopy(claim_by_key(parsed.claims, key))
            del claim["evidence"]["baseline"]
            ok, errors = validate_claim_evidence(claim, data)
            self.assertFalse(ok, key)
            self.assertTrue(any("baseline" in error for error in errors), (key, errors))
        parsed, validation, data = parse_evidence(DELTA_QUESTION)
        claim = copy.deepcopy(claim_by_key(parsed.claims, "delta"))
        del claim["evidence"]["baseline"]
        ok, errors = validate_claim_evidence(claim, data)
        self.assertFalse(ok)
        self.assertTrue(any("baseline" in error for error in errors), errors)

    def test_tamper_source_facts_digest(self) -> None:
        parsed, validation, data = parse_evidence(POINT_QUESTION)
        claim = copy.deepcopy(parsed.claims[0])
        claim["evidence"]["sourceFactsSha256"] = "DEADBEEF" * 8
        ok, errors = validate_claim_evidence(claim, data)
        self.assertFalse(ok)
        self.assertTrue(any("sourceFactsSha256" in error for error in errors), errors)
        claim = copy.deepcopy(parsed.claims[0])
        claim["evidence"]["sourceFactCount"] = 999
        ok, errors = validate_claim_evidence(claim, data)
        self.assertFalse(ok)
        self.assertTrue(any("sourceFactCount" in error for error in errors), errors)

    def test_tamper_formula_and_rounding(self) -> None:
        parsed, validation, data = parse_evidence(POINT_QUESTION)
        claim = copy.deepcopy(parsed.claims[0])
        del claim["evidence"]["formula"]
        ok, errors = validate_claim_evidence(claim, data)
        self.assertFalse(ok)
        self.assertTrue(any("formula" in error for error in errors), errors)
        claim = copy.deepcopy(parsed.claims[0])
        claim["evidence"]["rounding"] = 5
        ok, errors = validate_claim_evidence(claim, data)
        self.assertFalse(ok)
        self.assertTrue(any("rounding" in error for error in errors), errors)


class FactChangeDetectionTest(unittest.TestCase):
    """用 synthetic workbook overrides 改一个底层事实：相关 claim 的 value 与
    sourceFactsSha256 都变化；无关 claim 的摘要与值不变。"""

    def test_fact_override_propagates_to_related_claims_only(self) -> None:
        questions = [
            ("F1", "训练集", "简单", POINT_QUESTION, "100.5亿元"),
            ("F2", "训练集", "简单", "江苏省E市农商行在2025-01-31的不良贷款率是多少？", "0.91%"),
        ]
        base_summary, base_reviews = run_synthetic_with_reviews(questions)
        alt_summary, alt_reviews = run_synthetic_with_reviews(questions, overrides={("2025-01-31", "ORG001", "ZB001"): 999.0})
        base_f1 = next(review for review in base_reviews if review["id"] == "F1")["claims"][0]
        alt_f1 = next(review for review in alt_reviews if review["id"] == "F1")["claims"][0]
        self.assertEqual(base_f1["value"], 100.5)
        self.assertEqual(alt_f1["value"], 999.0)
        self.assertNotEqual(alt_f1["evidence"]["sourceFactsSha256"], base_f1["evidence"]["sourceFactsSha256"])
        self.assertEqual(alt_f1["evidence"]["sourceFactsSha256"], fact_digest({("2025-01-31", "ORG001", "ZB001"): 999.0}))
        base_f2 = next(review for review in base_reviews if review["id"] == "F2")["claims"][0]
        alt_f2 = next(review for review in alt_reviews if review["id"] == "F2")["claims"][0]
        self.assertEqual(base_f2["value"], 0.91)
        self.assertEqual(alt_f2["value"], 0.91, "无关 claim 的 value 不应变化")
        self.assertEqual(alt_f2["evidence"]["sourceFactsSha256"], base_f2["evidence"]["sourceFactsSha256"], "无关 claim 的摘要不应变化")


class DependsOnIntegrityTest(unittest.TestCase):
    """dependsOn 缺失目标与循环依赖都必须验证失败、错误可识别、无异常/无限递归。"""

    def _data(self) -> WorkbookData:
        tmp = tempfile.TemporaryDirectory()
        base = Path(tmp.name)
        workbook_path = build_synthetic_workbook(base / "syn.xlsx", [("D1", "训练集", "简单", POINT_QUESTION, "100.5亿元")])
        data = WorkbookData(workbook_path)
        tmp.cleanup()
        return data

    def test_missing_depends_on_target_fails(self) -> None:
        data = self._data()
        claim = synthetic_claim("a", "SUM", 1.0, {"operation": "SUM_DEPENDS", "formula": "a+b", "rounding": 2, "dependsOn": ["a", "ghost"]})
        ok, errors = validate_claim_evidence(claim, data, {"a": claim})
        self.assertFalse(ok)
        self.assertTrue(any("dependsOn 目标缺失" in error for error in errors), errors)
        validation = validate_review_evidence([claim], data)
        self.assertFalse(validation["valid"])
        self.assertFalse(validation["claimResults"]["a"]["ok"])
        self.assertTrue(validation["claimResults"]["a"]["errors"])

    def test_depends_on_cycle_fails_without_recursion(self) -> None:
        data = self._data()
        a = synthetic_claim("a", "SUM", 1.0, {"operation": "SUM_DEPENDS", "formula": "b", "rounding": 2, "dependsOn": ["b"]})
        b = synthetic_claim("b", "SUM", 1.0, {"operation": "SUM_DEPENDS", "formula": "a", "rounding": 2, "dependsOn": ["a"]})
        validation = validate_review_evidence([a, b], data)  # 不应抛异常或无限递归
        self.assertFalse(validation["valid"])
        self.assertTrue(any("成环" in error for error in validation["errors"]), validation["errors"])
        self.assertFalse(validation["claimResults"]["a"]["ok"])
        self.assertFalse(validation["claimResults"]["b"]["ok"])
        self.assertTrue(any("成环" in error for error in validation["claimResults"]["a"]["errors"]))


class EvidenceFailureInjectionTest(unittest.TestCase):
    """mock validate_review_evidence 返回 invalid：记录必须 UNRESOLVED、
    fullEvidence=false、auditErrors 非空、evidenceValidation.valid=false、
    canonicalReady=false。"""

    def test_invalid_evidence_forces_unresolved(self) -> None:
        fake = {
            "valid": False,
            "claimResults": {"各项存款余额": {"ok": False, "errors": ["注入的 claim 验证失败"]}},
            "errors": ["注入的 review 验证失败"],
        }
        with mock.patch("audit_ground_truth.validate_review_evidence", return_value=fake):
            summary, reviews = run_synthetic_with_reviews([("FI1", "训练集", "简单", POINT_QUESTION, "100.5亿元")])
        review = reviews[0]
        self.assertEqual(review["status"], "UNRESOLVED")
        self.assertFalse(review["fullEvidence"])
        self.assertTrue(review["auditErrors"], "auditErrors 必须非空")
        self.assertTrue(any("证据校验失败" in error for error in review["auditErrors"]), review["auditErrors"])
        self.assertFalse(review["evidenceValidation"]["valid"])
        self.assertEqual(summary["statusCounts"]["UNRESOLVED"], 1)
        self.assertEqual(summary["fullEvidence"], 0)
        self.assertEqual(summary["evidenceCompleteCount"], 0)
        self.assertEqual(summary["evidenceErrorCount"], 1)
        self.assertFalse(summary["canonicalReady"])


class EvidenceSummaryContractTest(unittest.TestCase):
    """summary 聚合与逐条 review 一致；review.ndjson 每条含 evidenceValidation；
    CORRECTED 可 evidence valid 且不计 evidenceError；UNRESOLVED 单独计数。"""

    def test_summary_and_ndjson_contract(self) -> None:
        questions = [
            ("V1", "训练集", "简单", POINT_QUESTION, "100.5亿元"),
            ("C1", "训练集", "简单", "江苏省B市农商行在2025-01-31的各项存款余额是多少？", "999亿元"),
            ("U1", "训练集", "复杂", "江苏省A市农商行在2025-01-31的净利润环比和同比分别变动了多少？", "无"),
        ]
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            workbook_path = build_synthetic_workbook(base / "syn.xlsx", questions)
            output_dir = base / "out"
            digest = hashlib.sha256(workbook_path.read_bytes()).hexdigest().upper()
            summary = run_audit(workbook_path, output_dir, True, expected_total=3, expected_splits={"train": 3, "dev": 0, "test": 0}, expected_source_sha=digest)
            lines = [line for line in (output_dir / "review.ndjson").read_text(encoding="utf-8").splitlines() if line.strip()]
            reviews = [json.loads(line) for line in lines]
        self.assertEqual(summary["statusCounts"], {"VERIFIED": 1, "CORRECTED": 1, "UNRESOLVED": 1})
        self.assertEqual(summary["evidenceRuleVersion"], EVIDENCE_RULES_VERSION)
        self.assertEqual(summary["evidenceCompleteCount"], sum(1 for review in reviews if review["fullEvidence"]))
        self.assertEqual(summary["evidenceErrorCount"], sum(1 for review in reviews if not review["evidenceValidation"]["valid"] or review["evidenceValidation"]["claimErrors"] or review["evidenceValidation"]["errors"]))
        self.assertEqual(summary["evidenceCompleteCount"], 2, "VERIFIED + CORRECTED 各一")
        self.assertEqual(summary["evidenceErrorCount"], 0)
        self.assertEqual(summary["unresolvedIds"], ["U1"])
        self.assertEqual(len(lines), 3)
        by_id = {review["id"]: review for review in reviews}
        for review in reviews:
            self.assertIn("evidenceValidation", review, "review.ndjson 每条必须包含 evidenceValidation")
            self.assertEqual(review["evidenceValidation"]["ruleVersion"], EVIDENCE_RULES_VERSION)
            if review["fullEvidence"]:
                self.assertTrue(review["evidenceValidation"]["valid"], "fullEvidence 必须来自证据验证")
                self.assertFalse(review["evidenceValidation"]["claimErrors"])
                self.assertFalse(review["evidenceValidation"]["errors"])
        self.assertTrue(by_id["C1"]["evidenceValidation"]["valid"], "CORRECTED 记录可 evidence valid")
        self.assertFalse(by_id["C1"]["evidenceValidation"]["claimErrors"])
        self.assertTrue(by_id["C1"]["fullEvidence"])
        self.assertEqual(by_id["U1"]["status"], "UNRESOLVED")
        self.assertTrue(by_id["U1"]["evidenceValidation"]["valid"], "缺基期的 UNRESOLVED 不产生 evidence 错误")
        self.assertFalse(by_id["U1"]["fullEvidence"])


class SyntheticPurityTest(unittest.TestCase):
    """静态扫描：测试文件中真实 TRAIN/VAL/TST ID、完整真实问题、完整真实答案均为 0。"""

    def test_no_real_ids_questions_or_answers(self) -> None:
        test_source = Path(__file__).read_text(encoding="utf-8")
        real_ids: list[str] = []
        real_questions: list[str] = []
        real_answers: list[str] = []
        source_path = ROOT.parent.parent / ".local-dev" / "gt-audit" / "source.xlsx"
        if source_path.is_file():
            workbook = load_workbook(source_path, read_only=True, data_only=True)
            sheet = workbook["问题答案清单"]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue
                real_ids.append(str(row[0]).strip())
                real_questions.append(str(row[3]).strip())
                real_answers.append(str(row[4]).strip())
        else:
            for filename in ("train.jsonl", "dev.jsonl", "test.jsonl"):
                path = ROOT / filename
                if not path.is_file():
                    continue
                for line in path.read_text(encoding="utf-8").splitlines():
                    if not line.strip():
                        continue
                    record = json.loads(line)
                    real_questions.append(record.get("question", ""))
                    expected = record.get("expected") or {}
                    real_answers.append(expected.get("answerText", ""))
        found_ids = [qid for qid in real_ids if qid in test_source]
        found_questions = [question for question in real_questions if question and question in test_source]
        found_answers = [answer for answer in real_answers if answer and answer in test_source]
        self.assertEqual(found_ids, [], f"测试文件不得包含真实题目编号: {found_ids[:5]}")
        self.assertEqual(found_questions, [], f"测试文件不得包含完整真实问题: {found_questions[:3]}")
        self.assertEqual(found_answers, [], f"测试文件不得包含完整真实答案: {found_answers[:3]}")


if __name__ == "__main__":
    unittest.main()
