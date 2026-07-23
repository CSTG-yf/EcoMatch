#!/usr/bin/env python3
"""Contract tests for the reproducible bank NL2SQL benchmark database."""

from __future__ import annotations

import sqlite3
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from db.build_database import build_database  # noqa: E402
from db.validate_database import validate_database  # noqa: E402


class BuildDatabaseTest(unittest.TestCase):
    def create_workbook(self, path: Path) -> None:
        workbook = Workbook()
        organization_sheet = workbook.active
        organization_sheet.title = "机构信息表"
        organization_sheet.append(["机构编号", "机构名称"])
        organization_sheet.append(["ORG001", "江苏省A市农商行"])
        organization_sheet.append(["ORG002", "江苏省B市农商行"])

        metric_sheet = workbook.create_sheet("指标清单表")
        metric_sheet.append(["指标编号", "指标名称", "指标含义", "指标单位"])
        metric_sheet.append(["ZB001", "各项存款余额", "存款期末余额", "亿元"])
        metric_sheet.append(["ZB013", "不良贷款率", "不良贷款占比", "%"])

        workbook.create_sheet("衍生维度说明").append(["衍生维度", "衍生口径说明"])

        fact_sheet = workbook.create_sheet("指标数据表")
        fact_sheet.append(["数据日期", "指标编号", "指标名称", "机构编号", "指标值"])
        fact_sheet.append(["2024-12-31", "ZB001", "各项存款余额", "ORG001", 41.76])
        fact_sheet.append(["2024-12-31", "ZB013", "不良贷款率", "ORG001", 1.02])
        fact_sheet.append(["2024-12-31", "ZB001", "各项存款余额", "ORG002", 38.50])
        fact_sheet.append(["2024-12-31", "ZB013", "不良贷款率", "ORG002", 1.10])

        question_sheet = workbook.create_sheet("问题答案清单")
        question_sheet.append(["问题编号", "问题类型", "问题难度", "问题描述", "问题结果"])
        question_sheet.append(["TRAIN-S-01", "训练集", "简单", "A行存款余额是多少？", "41.76亿元"])
        workbook.save(path)

    def test_builds_normalized_database_with_integrity_report(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            workbook_path = temp_path / "source.xlsx"
            database_path = temp_path / "benchmark.sqlite"
            self.create_workbook(workbook_path)

            report = build_database(workbook_path, database_path)

            self.assertEqual(report["counts"], {
                "organizations": 2,
                "metrics": 2,
                "facts": 4,
            })
            self.assertEqual(report["dateRange"], {"min": "2024-12-31", "max": "2024-12-31"})
            self.assertEqual(report["integrityErrors"], [])

            validation = validate_database(database_path)
            self.assertEqual(validation["result"], "PASS")
            self.assertEqual(validation["dateCount"], 1)
            self.assertEqual(validation["factsPerDate"], 4)

            connection = sqlite3.connect(database_path)
            try:
                self.assertEqual(connection.execute("SELECT COUNT(*) FROM bank_organization").fetchone()[0], 2)
                self.assertEqual(connection.execute("SELECT COUNT(*) FROM bank_metric_definition").fetchone()[0], 2)
                self.assertEqual(connection.execute("SELECT COUNT(*) FROM bank_metric_daily").fetchone()[0], 4)
                self.assertEqual(
                    connection.execute(
                        "SELECT metric_value FROM bank_metric_daily "
                        "WHERE org_code = 'ORG001' AND metric_code = 'ZB001'"
                    ).fetchone()[0],
                    41.76,
                )
            finally:
                connection.close()

    def test_h2_script_exposes_database_named_compatibility_views(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            workbook_path = temp_path / "source.xlsx"
            database_path = temp_path / "benchmark.sqlite"
            h2_script_path = temp_path / "benchmark-h2.sql"
            self.create_workbook(workbook_path)

            build_database(workbook_path, database_path, h2_script_path)

            h2_script = h2_script_path.read_text(encoding="utf-8")
            self.assertIn("CREATE SCHEMA IF NOT EXISTS bank_benchmark;", h2_script)
            self.assertIn(
                "CREATE VIEW bank_benchmark.bank_metric_daily AS "
                "SELECT * FROM PUBLIC.bank_metric_daily;",
                h2_script,
            )


if __name__ == "__main__":
    unittest.main()
