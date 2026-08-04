#!/usr/bin/env python3
"""候选修正版 GT 生成器、候选审查与官方晋升的纯合成测试。

全部使用合成题目、合成答案与合成 ID，不包含任何真实 TRAIN/VAL/TST 数据。
覆盖：5 条答案修正选择门控、PERFORMANCE 指标提取与重写、DIMENSION_3 映射
提取与重写、答案数值不参与提取、无指标/映射歧义拒绝、候选 cell diff 白名单、
manifest/hash 校验、默认审查器拒绝非原始哈希、manifest 验证候选、确定性、
候选重新审查行为、审查器对澄清题的解析行为，以及候选晋升
（promote_ground_truth）的就绪晋升、确定性输出与全部 fail-closed 门控。
"""

from __future__ import annotations

import copy
import hashlib
import json
import shutil
import sys
import tempfile
import unittest
from collections import Counter
from datetime import date, timedelta
from pathlib import Path
from typing import Any
from unittest import mock

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import audit_ground_truth as audit  # noqa: E402
import clarify_ground_truth_contracts as gen  # noqa: E402
import gt_answer_rules as rules  # noqa: E402
from promote_ground_truth import (  # noqa: E402
    CURRENT_POINTER,
    FINAL_AUDIT_SUMMARY,
    OFFICIAL_MANIFEST,
    PromotionError,
    promote_ground_truth,
)

ORG_NAMES = {
    "ORG001": "江苏省A市农商行",
    "ORG002": "江苏省B市农商行",
    "ORG003": "江苏省C市农商行",
    "ORG004": "江苏省D市农商行",
    "ORG005": "江苏省E市农商行",
}
METRICS = [
    ("ZB001", "各项存款余额", "指机构吸收的全部存款期末余额", "亿元"),
    ("ZB002", "各项贷款余额", "指机构发放的全部贷款期末余额", "亿元"),
    ("ZB011", "净利润", "指机构最终盈利", "万元"),
    ("ZB012", "成本收入比", "指营业支出占营业收入的比例", "%"),
    ("ZB013", "不良贷款率", "指不良贷款余额占各项贷款余额的比例", "%"),
    ("ZB015", "拨备覆盖率", "指贷款损失准备对不良贷款的覆盖率", "%"),
    ("ZB016", "资本充足率", "指资本对风险加权资产的比率", "%"),
    ("ZB017", "逾期贷款率", "指逾期贷款余额占各项贷款余额的比例", "%"),
]
DAYS = ["2024-12-31", "2025-01-31", "2025-02-28", "2025-03-31", "2025-04-30"]
DERIVED_ROWS = [
    ("较年初", "当日值 - 2024年12月31日值"),
    ("较上季", "当日值 - 上季度末值"),
    ("较上月", "当日值 - 上月月末值"),
    ("较同期", "当日值 - 去年同期值"),
    ("全省均值", "13家机构当日值的算数平均值"),
    ("排名", "不良贷款率/逾期贷款率/成本收入比按从低到高排名（越低越好），其余按从高到低排名，使用rank算法"),
    ("增量", "当日值 - 比较日值"),
    ("增幅", "(当日值-比较日值)/比较日值×100%，比率类指标（ZB012/013/015/016/017）不做增幅计算"),
    ("表现较好", "前三"),
    ("表现较差", "后四"),
]

# 大表流式回归用：1,200 个日期 × 5 家机构 × 8 个指标 = 48,000 行事实表
# （约 24 万单元格），远超默认 200 行，用于证明候选校验是单遍流式。
LARGE_DAYS = [(date(2021, 1, 1) + timedelta(days=index)).isoformat() for index in range(1200)]


def fact_value(org_code: str, metric_code: str, day_index: int) -> float:
    org_index = int(org_code[-1]) - 1
    base = {
        "ZB001": 100.0 - 10.0 * org_index,
        "ZB002": 90.0 - 10.0 * org_index,
        "ZB011": 500.0 - 100.0 * org_index,
        "ZB012": 30.0 + 2.0 * org_index,
        "ZB013": 0.5 + 0.1 * org_index,
        "ZB015": 180.0 - 5.0 * org_index,
        "ZB016": 12.0 + 0.1 * org_index,
        "ZB017": 0.3 + 0.05 * org_index,
    }[metric_code]
    drift = {"ZB001": 0.5, "ZB002": 0.3, "ZB011": 10.0, "ZB012": 0.1, "ZB013": 0.01, "ZB015": 0.2, "ZB016": 0.05, "ZB017": 0.01}[metric_code]
    return base + drift * day_index


def build_synthetic_workbook(path: Path, questions: list[tuple[str, str, str, str]], days: list[str] = DAYS) -> Path:
    """写合成工作簿：机构表/指标表/衍生维度说明/事实表/问题答案清单。

    days 可注入大日期集以构造大事实表（流式回归测试用）。
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "机构信息表"
    sheet.append(["机构编号", "机构名称"])
    for code, name in ORG_NAMES.items():
        sheet.append([code, name])

    metric_sheet = workbook.create_sheet("指标清单表")
    metric_sheet.append(["指标编号", "指标名称", "指标含义", "指标单位"])
    for row in METRICS:
        metric_sheet.append(list(row))

    derived_sheet = workbook.create_sheet("衍生维度说明")
    derived_sheet.append(["衍生维度", "衍生口径说明"])
    for row in DERIVED_ROWS:
        derived_sheet.append(list(row))

    fact_sheet = workbook.create_sheet("指标数据表")
    fact_sheet.append(["数据日期", "指标编号", "指标名称", "机构编号", "指标值"])
    metric_names = {code: name for code, name, _meaning, _unit in METRICS}
    for day_index, day in enumerate(days):
        for org_code in ORG_NAMES:
            for metric_code, metric_name, _meaning, _unit in METRICS:
                fact_sheet.append([day, metric_code, metric_names[metric_code], org_code, fact_value(org_code, metric_code, day_index)])

    question_sheet = workbook.create_sheet("问题答案清单")
    question_sheet.append(["问题编号", "问题类型", "问题难度", "问题描述", "问题结果"])
    for index, (qid, split, difficulty, question, answer) in enumerate(questions, start=1):
        question_sheet.append([qid or f"SYN-{index:03d}", split, difficulty, question, answer])

    workbook.save(path)
    return path


def make_review(
    qid: str,
    split: str,
    category: str,
    status: str,
    question: str,
    answer: str,
    corrected: str | None = None,
    full_evidence: bool = True,
    ev_valid: bool = True,
    unresolved_reason: str | None = None,
) -> dict:
    return {
        "id": qid,
        "split": split,
        "difficulty": "普通",
        "status": status,
        "category": category,
        "fullEvidence": full_evidence,
        "question": question,
        "answerText": answer,
        "correctedAnswerText": corrected,
        "claims": [],
        "auditErrors": [] if status != "UNRESOLVED" else ["解析歧义：合成场景"],
        "unresolvedReason": unresolved_reason,
        "matchNotes": [],
        "evidenceValidation": {"valid": ev_valid, "ruleVersion": "1.0.0", "claimErrors": {}, "errors": []},
    }


def build_200_question_workbook(path: Path, special: dict[str, tuple[str, str, str, str, str]], days: list[str] = DAYS) -> Path:
    """200 条合成问题（120/40/40），special 按已存在的 ID 替换其中的题。"""
    questions: list[tuple[str, str, str, str, str]] = []
    for i in range(120):
        questions.append((f"SYN-T{i:03d}", "训练集", "简单", f"合成问题T{i}", f"合成答案{i}"))
    for i in range(40):
        questions.append((f"SYN-V{i:03d}", "验证集", "简单", f"合成问题V{i}", f"合成答案{i}"))
    for i in range(40):
        questions.append((f"SYN-S{i:03d}", "测试集", "简单", f"合成问题S{i}", f"合成答案{i}"))
    by_id = {item[0]: list(item) for item in questions}
    for qid, item in special.items():
        if qid not in by_id:
            raise ValueError(f"special ID 必须是已有 200 题之一：{qid}")
        by_id[qid] = list(item)
    return build_synthetic_workbook(path, [tuple(item) for item in by_id.values()], days=days)


# --------------------------------------------------------------------------- 选择门控


class SelectionGateTest(unittest.TestCase):
    """5 条答案修正选择门控：status/evidence 驱动，不按 ID。"""

    def test_answer_correction_gate(self) -> None:
        good = make_review("SYN-A1", "train", "POINT", "CORRECTED", "合成问题", "旧答案", corrected="新答案", full_evidence=True, ev_valid=True)
        no_full = make_review("SYN-A2", "train", "POINT", "CORRECTED", "合成问题", "旧答案", corrected="新答案", full_evidence=False, ev_valid=True)
        no_valid = make_review("SYN-A3", "train", "POINT", "CORRECTED", "合成问题", "旧答案", corrected="新答案", full_evidence=True, ev_valid=False)
        no_corrected = make_review("SYN-A4", "train", "POINT", "CORRECTED", "合成问题", "旧答案", corrected=None, full_evidence=True, ev_valid=True)
        verified = make_review("SYN-A5", "train", "POINT", "VERIFIED", "合成问题", "答案", corrected=None, full_evidence=True, ev_valid=True)
        selected = gen.select_answer_corrections([good, no_full, no_valid, no_corrected, verified])
        self.assertEqual([item["id"] for item in selected], ["SYN-A1"])
        self.assertTrue(all(item["fullEvidence"] and item["evidenceValidation"]["valid"] and item["correctedAnswerText"] for item in selected))

    def test_clarify_selection_by_status_and_category(self) -> None:
        perf = make_review("SYN-C1", "train", "PERFORMANCE", "UNRESOLVED", "合成问题", "存贷比80%，不良贷款率1.2%。", unresolved_reason="指标全集无法确定")
        dim3 = make_review("SYN-C2", "train", "DIMENSION_3", "UNRESOLVED", "合成问题", "规模：存款100亿。质量：不良贷款率1%。效益：净利润1万。", unresolved_reason="各维度对应指标无法确定")
        other_unresolved = make_review("SYN-C3", "train", "CHANGE_DELTA", "UNRESOLVED", "合成问题", "无", unresolved_reason="缺基期")
        verified = make_review("SYN-C4", "train", "PERFORMANCE", "VERIFIED", "合成问题", "答案")
        selected = gen.select_clarifications([perf, dim3, other_unresolved, verified])
        self.assertEqual(sorted(item["id"] for item in selected), ["SYN-C1", "SYN-C2"])
        self.assertNotIn("SYN-C3", [item["id"] for item in selected], "非 PERFORMANCE/DIMENSION_3 的 UNRESOLVED 不得入选")


# --------------------------------------------------------------------------- PERFORMANCE 提取与重写


PERF_ANSWER = (
    "存贷比80%，不良贷款率1.2%，拨备覆盖率150%，资本充足率12%，逾期贷款率0.5%，净利润200万元，成本收入比35%。"
    "表现较好指标：各项存款余额(第2名)、各项贷款余额(第1名)、不良贷款率(第1名)。表现较差指标：无。"
)


class PerformanceExtractionTest(unittest.TestCase):
    def test_extract_metrics_and_rewrite(self) -> None:
        hits = gen.extract_performance_metrics(PERF_ANSWER)
        codes = [gen.metric_identity(hit) for hit in hits]
        self.assertEqual(
            codes,
            ["存贷比", "ZB013", "ZB015", "ZB016", "ZB017", "ZB011", "ZB012", "ZB001", "ZB002"],
        )
        self.assertEqual(hits[0].derived, ("ZB002", "ZB001"))
        question = "合成机构在2025-01-31的指标中哪些表现较好？哪些表现较差？"
        clarified = gen.clarify_performance_question(question, hits)
        self.assertIn("待评价指标集合：存贷比、不良贷款率、拨备覆盖率、资本充足率、逾期贷款率、净利润、成本收入比、各项存款余额、各项贷款余额。", clarified)
        self.assertIn("表现较好=全省排名前三", clarified)
        self.assertIn("表现较差=全省排名后四", clarified)
        self.assertIn("排名方向由指标定义决定", clarified)
        # 解析器必须能从澄清文本独立重算同一指标集合（不得回读答案）
        reparsed = audit._explicit_metric_hits(clarified)
        self.assertEqual([gen.metric_identity(hit) for hit in reparsed], codes)

    def test_answer_numbers_never_participate_in_extraction(self) -> None:
        alt_answer = PERF_ANSWER.replace("80%", "99%").replace("1.2%", "2.3%").replace("150%", "160%").replace("200万元", "999万元")
        self.assertNotEqual(alt_answer, PERF_ANSWER)
        self.assertEqual(
            [gen.metric_identity(hit) for hit in gen.extract_performance_metrics(alt_answer)],
            [gen.metric_identity(hit) for hit in gen.extract_performance_metrics(PERF_ANSWER)],
            "数值变化不得影响指标提取",
        )

    def test_reject_when_no_unique_metric_set(self) -> None:
        self.assertEqual(gen.extract_performance_metrics("表现较好：无。表现较差：无。"), [])
        self.assertEqual(gen.extract_performance_metrics("较年初上升。"), [])


# --------------------------------------------------------------------------- DIMENSION_3 映射提取与重写

DIM3_ANSWER = (
    "规模：存款100.5亿（第1名），贷款90.3亿，存贷比89.85%。"
    "质量：不良贷款率0.51%（第1名）。"
    "效益：净利润510万元（第1名）。"
)


class DimensionMappingTest(unittest.TestCase):
    def test_extract_mapping_and_rewrite(self) -> None:
        mapping = gen.extract_dimension_mapping(DIM3_ANSWER)
        self.assertIsNotNone(mapping)
        assert mapping is not None
        self.assertEqual([gen.metric_identity(hit) for hit in mapping["规模"]], ["ZB001", "ZB002", "存贷比"])
        self.assertEqual([gen.metric_identity(hit) for hit in mapping["资产质量"]], ["ZB013"])
        self.assertEqual([gen.metric_identity(hit) for hit in mapping["盈利能力"]], ["ZB011"])
        question = "从规模、资产质量、盈利能力三个维度，分别列出合成机构在2025-01-31的各项指标及排名。"
        clarified = gen.clarify_dimension_question(question, mapping)
        self.assertIn("维度与指标映射：规模=存款、贷款、存贷比；资产质量=不良贷款率；盈利能力=净利润。", clarified)
        reparsed = audit._parse_dimension_mapping(clarified)
        self.assertIsNotNone(reparsed)
        assert reparsed is not None
        self.assertEqual(
            {dim: [gen.metric_identity(hit) for hit in hits] for dim, hits in reparsed.items()},
            {dim: [gen.metric_identity(hit) for hit in hits] for dim, hits in mapping.items()},
            "澄清文本必须能独立重算同一映射",
        )

    def test_mapping_unique_and_complete(self) -> None:
        # 缺质量段
        self.assertIsNone(gen.extract_dimension_mapping("规模：存款100亿。效益：净利润1万。"))
        # 指标跨维度重复
        self.assertIsNone(gen.extract_dimension_mapping("规模：存款100亿。质量：存款50亿。效益：净利润1万。"))
        # 维度段无数值内容（无指标名）
        self.assertIsNone(gen.extract_dimension_mapping("规模：。质量：。效益：。"))
        # 答案数值变化不影响映射
        alt = DIM3_ANSWER.replace("100.5", "99.99").replace("90.3", "88.88").replace("0.51", "1.11").replace("510", "300")
        alt_mapping = gen.extract_dimension_mapping(alt)
        assert alt_mapping is not None
        self.assertEqual(
            {dim: [gen.metric_identity(hit) for hit in hits] for dim, hits in alt_mapping.items()},
            {dim: [gen.metric_identity(hit) for hit in hits] for dim, hits in gen.extract_dimension_mapping(DIM3_ANSWER).items()},  # type: ignore[union-attr]
        )


# --------------------------------------------------------------------------- 候选工作簿与 manifest


def write_review_file(path: Path, reviews: list[dict]) -> None:
    path.write_text(
        "".join(json.dumps(review, ensure_ascii=False, sort_keys=True) + "\n" for review in reviews),
        encoding="utf-8",
    )


CORRECTION_ITEMS: dict[str, tuple[str, str, str, str, str]] = {
    "SYN-T000": ("SYN-T000", "训练集", "简单", "合成机构在2025-01-31的各项存款余额是多少？", "旧答案0"),
    "SYN-T001": ("SYN-T001", "训练集", "简单", "合成机构在2025-01-31的各项贷款余额是多少？", "旧答案1"),
    "SYN-T002": ("SYN-T002", "训练集", "简单", "合成机构在2025-01-31的净利润是多少？", "旧答案2"),
    "SYN-T003": ("SYN-T003", "训练集", "简单", "合成机构在2025-01-31的不良贷款率是多少？", "旧答案3"),
    "SYN-T004": ("SYN-T004", "训练集", "简单", "合成机构在2025-01-31的成本收入比是多少？", "旧答案4"),
}
CLARIFY_ITEMS: dict[str, tuple[str, str, str, str, str]] = {
    "SYN-T005": ("SYN-T005", "训练集", "复杂", "合成机构在2025-01-31的指标中哪些表现较好？哪些表现较差？", "旧答案5"),
    "SYN-T006": ("SYN-T006", "训练集", "复杂", "从规模、资产质量、盈利能力三个维度，分别列出合成机构在2025-01-31的各项指标及排名。", "旧答案6"),
    "SYN-V000": ("SYN-V000", "验证集", "复杂", "合成机构在2025-01-31的指标中哪些表现较好？哪些表现较差？", "旧答案7"),
    "SYN-S000": ("SYN-S000", "测试集", "复杂", "合成机构在2025-01-31的指标中哪些表现较好？哪些表现较差？", "旧答案8"),
}


def build_synthetic_scenario(base: Path, days: list[str] = DAYS) -> tuple[Path, Path, dict]:
    """构造完整合成场景：200 条工作簿 + review.ndjson，返回 (workbook, review, ids)。"""
    special = dict(CORRECTION_ITEMS)
    special.update(CLARIFY_ITEMS)
    workbook_path = build_200_question_workbook(base / "source.xlsx", special, days=days)
    reviews: list[dict] = []
    for index, qid in enumerate(CORRECTION_ITEMS):
        reviews.append(
            make_review(
                qid, "train",
                "POINT", "CORRECTED", f"合成问题{index}", f"旧答案{index}",
                corrected=f"新答案{index}", full_evidence=True, ev_valid=True,
            )
        )
    reviews.append(
        make_review(
            "SYN-T005", "train", "PERFORMANCE", "UNRESOLVED",
            CLARIFY_ITEMS["SYN-T005"][3],
            "存贷比80%，不良贷款率1.2%，拨备覆盖率150%，资本充足率12%，逾期贷款率0.5%，净利润200万元，成本收入比35%。"
            "表现较好指标：各项存款余额(第2名)、各项贷款余额(第1名)、不良贷款率(第1名)。表现较差指标：无。",
            unresolved_reason="指标全集无法唯一确定",
        )
    )
    reviews.append(
        make_review(
            "SYN-T006", "train", "DIMENSION_3", "UNRESOLVED",
            CLARIFY_ITEMS["SYN-T006"][3],
            DIM3_ANSWER,
            unresolved_reason="各维度对应指标无法唯一确定",
        )
    )
    reviews.append(
        make_review(
            "SYN-V000", "dev", "PERFORMANCE", "UNRESOLVED",
            CLARIFY_ITEMS["SYN-V000"][3],
            "存贷比81%，不良贷款率0.9%，拨备覆盖率160%，资本充足率11.5%，净利润300万元，成本收入比32%。"
            "表现较好指标：不良贷款率(第2名)。表现较差指标：各项存款余额(第4名)。",
            unresolved_reason="指标全集无法唯一确定",
        )
    )
    reviews.append(
        make_review(
            "SYN-S000", "test", "PERFORMANCE", "UNRESOLVED",
            CLARIFY_ITEMS["SYN-S000"][3],
            "不良贷款率1.1%，拨备覆盖率155%，资本充足率12.5%，净利润250万元，成本收入比33%。"
            "表现较好指标：净利润(第1名)。表现较差指标：无。",
            unresolved_reason="指标全集无法唯一确定",
        )
    )
    # 一条非 PERFORMANCE/DIMENSION_3 的 UNRESOLVED：必须原样保留
    reviews.append(
        make_review("SYN-U01", "train", "CHANGE_DELTA", "UNRESOLVED", "合成问题U", "无", unresolved_reason="缺基期")
    )
    review_path = base / "review.ndjson"
    write_review_file(review_path, reviews)
    return workbook_path, review_path, {"corrections": list(CORRECTION_ITEMS), "clarifications": list(CLARIFY_ITEMS)}


# ---------------------------------------------------------------------------
# 删除场景（16 条变更账本：5 答案修正 + 10 题目澄清 + 1 题目删除）
# 与真实基线分布同构（184 VERIFIED / 5 CORRECTED / 11 UNRESOLVED，其中 1 条
# 可删除）：删除 1 条训练题后候选 199 条全部可审查为 VERIFIED。
# 全部为合成题目/答案/ID，不包含真实数据。

REMOVAL_ID = "SYN-T119"

CORRECTION_ROWS: dict[str, tuple[str, str, str, str, str]] = {
    "SYN-T000": ("SYN-T000", "训练集", "简单", "江苏省A市农商行在2025-01-31的各项存款余额是多少？", "旧答案0"),
    "SYN-T001": ("SYN-T001", "训练集", "简单", "江苏省A市农商行在2025-01-31的各项贷款余额是多少？", "旧答案1"),
    "SYN-T002": ("SYN-T002", "训练集", "简单", "江苏省A市农商行在2025-01-31的净利润是多少？", "旧答案2"),
    "SYN-T003": ("SYN-T003", "训练集", "简单", "江苏省A市农商行在2025-01-31的不良贷款率是多少？", "旧答案3"),
    "SYN-T004": ("SYN-T004", "训练集", "简单", "江苏省A市农商行在2025-01-31的成本收入比是多少？", "旧答案4"),
}
CORRECTED_ANSWERS = {
    "SYN-T000": "100.5亿元",
    "SYN-T001": "90.3亿元",
    "SYN-T002": "510万元",
    "SYN-T003": "0.51%",
    "SYN-T004": "30.1%",
}
PERF_QUESTION = "江苏省A市农商行在2025-01-31的指标中哪些表现较好？哪些表现较差？"
DIM3_QUESTION = "从规模、资产质量、盈利能力三个维度，分别列出江苏省A市农商行在2025-01-31的各项指标及排名。"
REMOVAL_QUESTION = "江苏省A市农商行的各项存款余额在2025-01-31，同比（较去年同期）变动了多少？"
REMOVAL_REASON = "缺失基期或当前值（yoy）"

# 5 条 PERFORMANCE 澄清答案：ORG001 全部指标 2025-01-31 排名第 1
PERFORMANCE_ANSWERS = [
    "各项存款余额(第1名)、不良贷款率(第1名)。表现较好指标：各项存款余额、不良贷款率。表现较差指标：无。",
    "各项贷款余额(第1名)、净利润(第1名)。表现较好指标：各项贷款余额、净利润。表现较差指标：无。",
    "各项存款余额(第1名)、各项贷款余额(第1名)、净利润(第1名)。表现较好指标：各项存款余额、各项贷款余额、净利润。表现较差指标：无。",
    "不良贷款率(第1名)、成本收入比(第1名)。表现较好指标：不良贷款率、成本收入比。表现较差指标：无。",
    "拨备覆盖率(第1名)、资本充足率(第1名)、逾期贷款率(第1名)。表现较好指标：拨备覆盖率、资本充足率、逾期贷款率。表现较差指标：无。",
]
# 5 条 DIMENSION_3 澄清答案：规模/质量/效益 三段各含 ≥1 指标，跨维度不重复
DIMENSION3_ANSWERS = [
    "规模：各项存款余额100.5亿元（第1名），各项贷款余额90.3亿元。质量：不良贷款率0.51%（第1名）。效益：净利润510万元（第1名）。",
    "规模：各项存款余额100.5亿元（第1名）。质量：拨备覆盖率180.2%（第1名）。效益：净利润510万元（第1名）。",
    "规模：各项贷款余额90.3亿元（第1名）。质量：不良贷款率0.51%（第1名）。效益：成本收入比30.1%（第1名）。",
    "规模：各项存款余额100.5亿元（第1名）。质量：资本充足率12.05%（第1名）。效益：净利润510万元（第1名）。",
    "规模：各项存款余额100.5亿元（第1名）。质量：逾期贷款率0.31%（第1名）。效益：净利润510万元（第1名）。",
]
CLARIFY_PERF_IDS = [f"SYN-T{i:03d}" for i in range(5, 10)]
CLARIFY_DIM3_IDS = [f"SYN-T{i:03d}" for i in range(10, 15)]


def _point_rows(count: int) -> list[tuple[str, str, str, str, str]]:
    """合成单点题（机构/指标/日期 三元组互不相同，答案=重算值+单位），
    全部可审查为 VERIFIED（与现有单点题测试同构）。"""
    metric_names = {code: name for code, name, _meaning, _unit in METRICS}
    metric_units = {code: unit for code, _name, _meaning, unit in METRICS}
    rows: list[tuple[str, str, str, str, str]] = []
    for day_index, day in enumerate(DAYS):
        for org_code, org_name in ORG_NAMES.items():
            for metric_code, _name, _meaning, _unit in METRICS:
                if len(rows) >= count:
                    return rows
                value = rules.display_value(fact_value(org_code, metric_code, day_index), 2)
                rows.append((f"{org_name}在{day}的{metric_names[metric_code]}是多少？", f"{value}{metric_units[metric_code]}"))
    return rows


def build_removal_scenario(base: Path) -> tuple[Path, Path, dict]:
    """构造完整删除场景：200 条合成题（120/40/40）+ 覆盖全部 200 条的
    review 文件（184 VERIFIED / 5 CORRECTED / 10 可澄清 UNRESOLVED /
    1 可删除 UNRESOLVED）。返回 (workbook, review, ids)。"""
    questions: list[tuple[str, str, str, str, str]] = []
    for qid in sorted(CORRECTION_ROWS):
        questions.append(CORRECTION_ROWS[qid])
    for index, qid in enumerate(CLARIFY_PERF_IDS):
        questions.append((qid, "训练集", "复杂", PERF_QUESTION, PERFORMANCE_ANSWERS[index]))
    for index, qid in enumerate(CLARIFY_DIM3_IDS):
        questions.append((qid, "训练集", "复杂", DIM3_QUESTION, DIMENSION3_ANSWERS[index]))
    train_point_ids = [f"SYN-T{i:03d}" for i in range(15, 119)]  # 104 条
    dev_point_ids = [f"SYN-V{i:03d}" for i in range(40)]
    test_point_ids = [f"SYN-S{i:03d}" for i in range(40)]
    point_rows = _point_rows(len(train_point_ids) + len(dev_point_ids) + len(test_point_ids))
    for qid, (question, answer) in zip(train_point_ids + dev_point_ids + test_point_ids, point_rows):
        split = "训练集" if qid.startswith("SYN-T") else "验证集" if qid.startswith("SYN-V") else "测试集"
        questions.append((qid, split, "简单", question, answer))
    questions.append((REMOVAL_ID, "训练集", "普通", REMOVAL_QUESTION, "无"))
    workbook_path = build_synthetic_workbook(base / "source.xlsx", questions)

    reviews: list[dict] = []
    for qid in sorted(CORRECTION_ROWS):
        _qid, split, difficulty, question, answer = CORRECTION_ROWS[qid]
        reviews.append(
            make_review(qid, "train", "POINT", "CORRECTED", question, answer, corrected=CORRECTED_ANSWERS[qid], full_evidence=True, ev_valid=True)
        )
    for index, qid in enumerate(CLARIFY_PERF_IDS):
        reviews.append(
            make_review(qid, "train", "PERFORMANCE", "UNRESOLVED", PERF_QUESTION, PERFORMANCE_ANSWERS[index], full_evidence=False, unresolved_reason="指标全集无法从题意唯一确定")
        )
    for index, qid in enumerate(CLARIFY_DIM3_IDS):
        reviews.append(
            make_review(qid, "train", "DIMENSION_3", "UNRESOLVED", DIM3_QUESTION, DIMENSION3_ANSWERS[index], full_evidence=False, unresolved_reason="各维度对应指标无法从题意唯一确定")
        )
    for qid, (question, answer) in zip(train_point_ids + dev_point_ids + test_point_ids, point_rows):
        split = "train" if qid.startswith("SYN-T") else "dev" if qid.startswith("SYN-V") else "test"
        reviews.append(make_review(qid, split, "POINT", "VERIFIED", question, answer, full_evidence=True))
    reviews.append(
        make_review(REMOVAL_ID, "train", "CHANGE_DELTA", "UNRESOLVED", REMOVAL_QUESTION, "无", full_evidence=False, unresolved_reason=REMOVAL_REASON)
    )
    review_path = base / "review.ndjson"
    write_review_file(review_path, reviews)
    return workbook_path, review_path, {"removal": REMOVAL_ID, "corrections": list(CORRECTION_ROWS), "clarifications": CLARIFY_PERF_IDS + CLARIFY_DIM3_IDS}


class CandidateWorkbookTest(unittest.TestCase):
    def _run_scenario(self, base: Path) -> tuple[Path, Path, dict]:
        workbook_path, review_path, ids = build_synthetic_scenario(base)
        source_sha = hashlib.sha256(workbook_path.read_bytes()).hexdigest().upper()
        output_dir = base / "contract-fix"
        exit_code = gen.run_generator(workbook_path, review_path, output_dir, expected_source_sha=source_sha)
        self.assertEqual(exit_code, 0)
        return workbook_path, output_dir, ids

    def test_changed_cell_whitelist_and_counts(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            workbook_path, output_dir, ids = self._run_scenario(base)
            candidate_path = output_dir / gen.CANDIDATE_WORKBOOK
            source = load_workbook(workbook_path, read_only=True, data_only=True)
            candidate = load_workbook(candidate_path, read_only=True, data_only=True)
            try:
                changed: list[tuple[str, int, int]] = []
                for sname in source.sheetnames:
                    s_sheet, c_sheet = source[str(sname)], candidate[str(sname)]
                    for row_index in range(1, s_sheet.max_row + 1):
                        for col_index in range(1, s_sheet.max_column + 1):
                            if s_sheet.cell(row=row_index, column=col_index).value != c_sheet.cell(row=row_index, column=col_index).value:
                                changed.append((str(sname), row_index, col_index))
            finally:
                source.close()
                candidate.close()
            self.assertEqual(len(changed), len(ids["corrections"]) + len(ids["clarifications"]), "精确计数：5 answerText + 4 questionText")
            for sname, row_index, col_index in changed:
                self.assertEqual(sname, rules.SHEET_QUESTION)
                self.assertIn(col_index, (4, 5), "只允许 问题描述/问题结果 列变更")

    def test_determinism(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            workbook_path, review_path, _ids = build_synthetic_scenario(base)
            source_sha = hashlib.sha256(workbook_path.read_bytes()).hexdigest().upper()
            digests: list[str] = []
            for index in range(3):
                output_dir = base / f"out{index}"
                self.assertEqual(gen.run_generator(workbook_path, review_path, output_dir, expected_source_sha=source_sha), 0)
                digests.append(gen.sha256_file(output_dir / gen.CANDIDATE_WORKBOOK))
            self.assertEqual(len(set(digests)), 1, f"候选工作簿连续 3 次生成 SHA-256 必须完全一致：{digests}")
            first_dir, second_dir = base / "out0", base / "out1"
            self.assertEqual(
                (first_dir / gen.CANDIDATE_WORKBOOK).read_bytes(),
                (second_dir / gen.CANDIDATE_WORKBOOK).read_bytes(),
                "候选工作簿两次生成必须逐字节一致",
            )
            self.assertEqual(
                (first_dir / gen.CANDIDATE_MANIFEST).read_text(encoding="utf-8"),
                (second_dir / gen.CANDIDATE_MANIFEST).read_text(encoding="utf-8"),
            )
            self.assertEqual(
                (first_dir / gen.CHANGE_LEDGER).read_text(encoding="utf-8"),
                (second_dir / gen.CHANGE_LEDGER).read_text(encoding="utf-8"),
            )

    def test_manifest_and_hash_files(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            _workbook_path, output_dir, _ids = self._run_scenario(base)
            manifest = json.loads((output_dir / gen.CANDIDATE_MANIFEST).read_text(encoding="utf-8"))
            candidate_path = output_dir / gen.CANDIDATE_WORKBOOK
            self.assertEqual(manifest["candidateSha256"], gen.sha256_file(candidate_path))
            self.assertEqual(manifest["sourceSha256"], gen.sha256_file(_workbook_path))
            self.assertEqual(manifest["changeCounts"], {"answerChanges": 5, "questionClarifications": 4, "questionRemovals": 0, "contractErrors": 0})
            self.assertEqual(manifest["totalRecords"], 200)
            self.assertEqual(manifest["splitCounts"], {"train": 120, "dev": 40, "test": 40})
            expected_audit = manifest["expectedAudit"]
            self.assertEqual(expected_audit["totalRecords"], 200)
            self.assertEqual(expected_audit["splitCounts"], {"train": 120, "dev": 40, "test": 40})
            self.assertEqual(expected_audit["statusCounts"]["UNRESOLVED"], 1, "未删除场景：唯一 UNRESOLVED 保留在预期中")
            self.assertEqual(expected_audit["evidenceComplete"], 200)
            self.assertEqual(expected_audit["evidenceErrors"], 0)
            self.assertFalse(manifest["candidateReady"])
            self.assertFalse(manifest["canonicalReady"])
            self.assertEqual(manifest["changeLedgerSha256"], gen.sha256_file(output_dir / gen.CHANGE_LEDGER))
            ledger = json.loads((output_dir / gen.CHANGE_LEDGER).read_text(encoding="utf-8"))
            self.assertEqual(len(ledger["entries"]), 9)
            for entry in ledger["entries"]:
                self.assertIn("oldTextSha256", entry)
                self.assertIn("newTextSha256", entry)
            for artifact in (gen.CANDIDATE_WORKBOOK, gen.CHANGE_LEDGER, gen.CANDIDATE_MANIFEST):
                digest = gen.sha256_file(output_dir / artifact)
                sha_file = output_dir / (artifact + ".sha256")
                self.assertTrue(sha_file.is_file())
                self.assertIn(digest, sha_file.read_text(encoding="utf-8"))

    def test_ledger_records_metrics_and_mapping(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            _workbook_path, output_dir, ids = self._run_scenario(base)
            ledger = json.loads((output_dir / gen.CHANGE_LEDGER).read_text(encoding="utf-8"))
            by_id = {entry["id"]: entry for entry in ledger["entries"]}
            perf_entry = by_id["SYN-T005"]
            self.assertEqual(perf_entry["changeType"], "QUESTION_CLARIFICATION")
            self.assertEqual(perf_entry["metricCodes"][1:3], ["ZB013", "ZB015"])
            self.assertEqual(
                perf_entry["metricCodes"][0],
                {"derived": {"numerator": "ZB002", "denominator": "ZB001"}, "name": "存贷比"},
                "衍生指标必须以结构化对象 {name, derived:{numerator,denominator}} 写入账本",
            )
            self.assertNotIn(
                "存贷比",
                [code for code in perf_entry["metricCodes"] if isinstance(code, str)],
                "衍生指标不得以普通字符串形式写入账本",
            )
            dim_entry = by_id["SYN-T006"]
            self.assertEqual(dim_entry["dimensionMapping"]["资产质量"], ["ZB013"])
            self.assertEqual(dim_entry["dimensionMapping"]["规模"], ["ZB001", "ZB002", {"derived": {"numerator": "ZB002", "denominator": "ZB001"}, "name": "存贷比"}])
            corr_entry = by_id["SYN-T000"]
            self.assertEqual(corr_entry["changeType"], "ANSWER_CORRECTION")
            self.assertTrue(corr_entry["evidenceRef"]["fullEvidence"])
            self.assertTrue(corr_entry["evidenceRef"]["evidenceValidationValid"])


class ManifestAuditTest(unittest.TestCase):
    def test_default_audit_rejects_non_source_hash(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            workbook_path = build_synthetic_workbook(
                base / "syn.xlsx",
                [("SYN-X1", "训练集", "简单", "合成机构在2025-01-31的各项存款余额是多少？", "100.5亿元")],
            )
            with self.assertRaises(ValueError):
                audit.run_audit(workbook_path, base / "out", True)

    def test_candidate_audit_via_manifest(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            workbook_path, review_path, _ids = build_removal_scenario(base)
            source_sha = hashlib.sha256(workbook_path.read_bytes()).hexdigest().upper()
            output_dir = base / "contract-fix"
            self.assertEqual(gen.run_generator(workbook_path, review_path, output_dir, expected_source_sha=source_sha, exclude_ids=[REMOVAL_ID]), 0)
            candidate_path = output_dir / gen.CANDIDATE_WORKBOOK
            manifest_path = output_dir / gen.CANDIDATE_MANIFEST
            # 默认（无 manifest）拒绝候选：候选哈希 != 冻结原始哈希
            with self.assertRaises(ValueError):
                audit.run_audit(candidate_path, base / "out-default", True)
            # 显式 manifest 验证候选：哈希链 + 生成器契约通过后审查。
            # 合成删除场景：候选 199 条全部 VERIFIED -> candidateReady=true。
            summary = audit.run_audit(candidate_path, base / "out-candidate", True, manifest_path=manifest_path, expected_source_sha=source_sha)
            self.assertEqual(summary["totalRecords"], 199)
            self.assertEqual(summary["splitCounts"], {"train": 119, "dev": 40, "test": 40})
            self.assertEqual(summary["statusCounts"], {"VERIFIED": 199, "CORRECTED": 0, "UNRESOLVED": 0})
            self.assertEqual(summary["evidenceCompleteCount"], 199)
            self.assertEqual(summary["evidenceErrorCount"], 0)
            self.assertFalse(summary["canonicalReady"], "候选永远不得 canonical")
            self.assertTrue(summary["candidateReady"], "合成 199 全 VERIFIED 候选应就绪")
            self.assertEqual(summary["candidateReadyReasons"], [])
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            self.assertTrue(manifest["candidateReady"])
            self.assertIn("candidateAudit", manifest)
            self.assertEqual(manifest["candidateAudit"]["statusCounts"], summary["statusCounts"])
            self.assertEqual(manifest["candidateAudit"]["splitCounts"], summary["splitCounts"])
            self.assertTrue(manifest["candidateAudit"]["candidateReady"])

    def test_manifest_blessing_unresolved_record_rejected(self) -> None:
        # 未删除场景的 expectedAudit 保留 UNRESOLVED=1：候选就绪不变量要求
        # 100% VERIFIED，此类 manifest 必须被拒绝（fail closed）。
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            workbook_path, review_path, _ids = build_synthetic_scenario(base)
            source_sha = hashlib.sha256(workbook_path.read_bytes()).hexdigest().upper()
            output_dir = base / "contract-fix"
            self.assertEqual(gen.run_generator(workbook_path, review_path, output_dir, expected_source_sha=source_sha), 0)
            manifest = json.loads((output_dir / gen.CANDIDATE_MANIFEST).read_text(encoding="utf-8"))
            self.assertEqual(manifest["expectedAudit"]["statusCounts"]["UNRESOLVED"], 1, "未删除场景预期保留 1 条 UNRESOLVED")
            with self.assertRaises(ValueError):
                audit.run_audit(
                    output_dir / gen.CANDIDATE_WORKBOOK,
                    base / "out-blessed",
                    True,
                    manifest_path=output_dir / gen.CANDIDATE_MANIFEST,
                    expected_source_sha=source_sha,
                )

    def test_manifest_validation_rejects_tampering(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            workbook_path, review_path, _ids = build_synthetic_scenario(base)
            source_sha = hashlib.sha256(workbook_path.read_bytes()).hexdigest().upper()
            output_dir = base / "contract-fix"
            self.assertEqual(gen.run_generator(workbook_path, review_path, output_dir, expected_source_sha=source_sha), 0)
            candidate_path = output_dir / gen.CANDIDATE_WORKBOOK
            manifest_path = output_dir / gen.CANDIDATE_MANIFEST
            tampered = json.loads(manifest_path.read_text(encoding="utf-8"))
            tampered["candidateSha256"] = "DEADBEEF" * 8
            manifest_path.write_text(json.dumps(tampered, ensure_ascii=False), encoding="utf-8")
            # 同步重算边车：本测试针对内容级校验（候选哈希不符），
            # 边车字节完整性由 ManifestSidecarTest 单独覆盖
            sidecar_path = manifest_path.with_name(manifest_path.name + ".sha256")
            sidecar_path.write_text(f"{audit.sha256_file(manifest_path)}  {manifest_path.name}\n", encoding="utf-8")
            with self.assertRaises(ValueError):
                audit.run_audit(candidate_path, base / "out-tamper", True, manifest_path=manifest_path, expected_source_sha=source_sha)

    def test_candidate_ready_gate_function(self) -> None:
        def review(qid: str, split: str, status: str, full_evidence: bool = True) -> dict:
            return {
                "id": qid, "split": split, "difficulty": "普通", "status": status,
                "category": "POINT", "fullEvidence": full_evidence,
                "question": "合成问题", "answerText": "合成答案", "correctedAnswerText": None,
                "claims": [], "auditErrors": [], "unresolvedReason": None, "matchNotes": [],
                "evidenceValidation": {"valid": True, "claimErrors": {}, "errors": []},
            }
        expected = {
            "totalRecords": 199,
            "splitCounts": {"train": 119, "dev": 40, "test": 40},
            "statusCounts": {"VERIFIED": 199, "CORRECTED": 0, "UNRESOLVED": 0},
            "evidenceComplete": 199,
            "evidenceErrors": 0,
        }
        # 正例：199 条全部 VERIFIED（119/40/40）且完整证据 -> candidateReady
        passing = [review(f"SYN-{i:03d}", "train" if i < 119 else "dev" if i < 159 else "test", "VERIFIED") for i in range(199)]
        ok, reasons = audit.candidate_audit_passes(passing, expected)
        self.assertTrue(ok, reasons)
        # 缺少 expectedAudit -> fail closed
        ok, reasons = audit.candidate_audit_passes(passing, None)
        self.assertFalse(ok)
        self.assertTrue(any("expectedAudit" in reason for reason in reasons))
        # 多出一条记录（总数/唯一 ID 不符）-> 拒绝
        extra = list(passing)
        extra.append(review("SYN-EXTRA", "train", "VERIFIED"))
        ok, reasons = audit.candidate_audit_passes(extra, expected)
        self.assertFalse(ok)
        self.assertTrue(any("总数" in reason for reason in reasons))
        # 状态分布不符（出现 CORRECTED）-> 拒绝
        with_corrected = list(passing)
        with_corrected[38] = review("SYN-038", "train", "CORRECTED")
        ok, reasons = audit.candidate_audit_passes(with_corrected, expected)
        self.assertFalse(ok)
        self.assertTrue(any("状态分布" in reason for reason in reasons))
        # split 计数不符 -> 拒绝
        bad_split = list(passing)
        bad_split[0] = review("SYN-000", "dev", "VERIFIED")
        ok, reasons = audit.candidate_audit_passes(bad_split, expected)
        self.assertFalse(ok)
        self.assertTrue(any("split" in reason for reason in reasons))
        # 证据错误条数不符 -> 拒绝
        bad_evidence = list(passing)
        bad_evidence[0] = dict(bad_evidence[0])
        bad_evidence[0]["evidenceValidation"] = {"valid": False, "claimErrors": {"注入": ["注入"]}, "errors": ["注入"]}
        ok, reasons = audit.candidate_audit_passes(bad_evidence, expected)
        self.assertFalse(ok)
        self.assertTrue(any("证据错误" in reason for reason in reasons))


# --------------------------------------------------------------------------- 删除资格门控（--exclude-id，fail closed）


class RemovalGateTest(unittest.TestCase):
    """删除资格门控：全部条件同时满足才接受；缺失 ID、重复、VERIFIED/CORRECTED、
    dev/test、可澄清、无关或仅“无法确定”的未决原因一律拒绝。"""

    def _review(self, qid: str = "SYN-D1", split: str = "train", status: str = "UNRESOLVED", full_evidence: bool = False,
                category: str = "CHANGE_DELTA", reason: str = REMOVAL_REASON) -> dict:
        return make_review(qid, split, category, status, REMOVAL_QUESTION, "无", full_evidence=full_evidence, unresolved_reason=reason)

    def test_eligible_removal_accepted(self) -> None:
        removals, errors = gen.select_removals([self._review()], ["SYN-D1"], [], [])
        self.assertEqual(errors, [])
        self.assertEqual([item["id"] for item in removals], ["SYN-D1"])

    def test_missing_id_rejected(self) -> None:
        removals, errors = gen.select_removals([self._review()], ["SYN-GHOST"], [], [])
        self.assertEqual(removals, [])
        self.assertTrue(any("不存在" in error["error"] for error in errors))

    def test_duplicate_request_rejected(self) -> None:
        # 重复请求：fail closed —— 返回错误条目，调用方必须中止产出（不写候选文件）
        removals, errors = gen.select_removals([self._review()], ["SYN-D1", "SYN-D1"], [], [])
        self.assertTrue(any("重复" in error["error"] for error in errors), errors)
        self.assertTrue(errors)

    def test_verified_and_corrected_rejected(self) -> None:
        for status in ("VERIFIED", "CORRECTED"):
            removals, errors = gen.select_removals([self._review(status=status)], ["SYN-D1"], [], [])
            self.assertEqual(removals, [], status)
            self.assertTrue(any("UNRESOLVED" in error["error"] for error in errors), status)

    def test_full_evidence_true_rejected(self) -> None:
        removals, errors = gen.select_removals([self._review(full_evidence=True)], ["SYN-D1"], [], [])
        self.assertEqual(removals, [])
        self.assertTrue(any("fullEvidence" in error["error"] for error in errors))

    def test_dev_and_test_rejected(self) -> None:
        for split in ("dev", "test"):
            removals, errors = gen.select_removals([self._review(split=split)], ["SYN-D1"], [], [])
            self.assertEqual(removals, [], split)
            self.assertTrue(any("仅允许删除训练集" in error["error"] for error in errors), split)

    def test_clarifiable_and_correctable_rejected(self) -> None:
        review = self._review()
        for corrections, clarifications in (([review], []), ([], [review])):
            removals, errors = gen.select_removals([review], ["SYN-D1"], corrections, clarifications)
            self.assertEqual(removals, [])
            self.assertTrue(any("已被选为答案修正或题目澄清" in error["error"] for error in errors))

    def test_unrelated_reason_rejected(self) -> None:
        for reason in ("指标全集无法从题意唯一确定", "各维度对应指标无法从题意唯一确定", "双条件指标结构不匹配", "环比基期缺失"):
            removals, errors = gen.select_removals([self._review(reason=reason)], ["SYN-D1"], [], [])
            self.assertEqual(removals, [], reason)
            self.assertTrue(any("未决原因" in error["error"] for error in errors), reason)

    def test_undetermined_yoy_baseline_rejected(self) -> None:
        # 有 yoy 与基期上下文，但只有“无法确定/不明确”而无缺失/缺少事实信号 -> 拒绝
        for reason in ("yoy 同比基期无法确定", "yoy 基期不确定", "yoy 同比基期不明确", "yoy 基期数据不完整"):
            removals, errors = gen.select_removals([self._review(reason=reason)], ["SYN-D1"], [], [])
            self.assertEqual(removals, [], reason)
            self.assertTrue(any("未决原因" in error["error"] for error in errors), reason)

    def test_missing_fact_signal_variants_accepted(self) -> None:
        for reason in ("缺失基期或当前值（yoy）", "yoy 同比基期数据缺失", "yoy 缺少当前值", "yoy 基期不存在", "yoy 当前值未提供"):
            removals, errors = gen.select_removals([self._review(reason=reason)], ["SYN-D1"], [], [])
            self.assertEqual(errors, [], reason)
            self.assertEqual([item["id"] for item in removals], ["SYN-D1"], reason)


# --------------------------------------------------------------------------- 删除场景端到端（恰好删 1 行 / 顺序字段 / 16 条账本 / manifest 哈希）


class RemovalPipelineTest(unittest.TestCase):
    def _run(self, base: Path) -> tuple[Path, Path, dict]:
        workbook_path, review_path, ids = build_removal_scenario(base)
        source_sha = hashlib.sha256(workbook_path.read_bytes()).hexdigest().upper()
        output_dir = base / "contract-fix"
        exit_code = gen.run_generator(workbook_path, review_path, output_dir, expected_source_sha=source_sha, exclude_ids=[REMOVAL_ID])
        self.assertEqual(exit_code, 0)
        return workbook_path, output_dir, ids

    def test_exact_one_row_removal_and_remaining_order_fields(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            workbook_path, output_dir, ids = self._run(base)
            source = load_workbook(workbook_path, read_only=True, data_only=True)
            candidate = load_workbook(output_dir / gen.CANDIDATE_WORKBOOK, read_only=True, data_only=True)
            try:
                s_rows = [list(row) for row in source[rules.SHEET_QUESTION].iter_rows(min_row=2, values_only=True) if row[0] is not None]
                c_rows = [list(row) for row in candidate[rules.SHEET_QUESTION].iter_rows(min_row=2, values_only=True) if row[0] is not None]
            finally:
                source.close()
                candidate.close()
            self.assertEqual(len(s_rows), 200)
            self.assertEqual(len(c_rows), 199, "恰好删除 1 行")
            self.assertNotIn(REMOVAL_ID, [row[0] for row in c_rows], "被授权删除的 ID 必须消失")
            self.assertEqual([row[0] for row in c_rows], [row[0] for row in s_rows if row[0] != REMOVAL_ID], "剩余 ID 顺序必须与源一致")
            # 字段：除 5 answerText + 10 questionText 白名单变更外逐格一致
            diffs: list[tuple[str, int]] = []
            for s_row, c_row in zip(s_rows, c_rows):
                if s_row[0] == REMOVAL_ID:
                    continue
                for col in range(1, len(rules.QUESTION_HEADERS) + 1):
                    if s_row[col - 1] != c_row[col - 1]:
                        diffs.append((str(s_row[0]), col))
            self.assertEqual(len(diffs), 15, f"白名单变更必须恰好 15 处：{diffs[:5]}")
            for _qid, col in diffs:
                self.assertIn(col, (4, 5), "只允许 问题描述/问题结果 列变更")

    def test_manifest_counts_splits_and_expected_audit_invariant(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            _workbook_path, output_dir, _ids = self._run(base)
            manifest = json.loads((output_dir / gen.CANDIDATE_MANIFEST).read_text(encoding="utf-8"))
            self.assertEqual(manifest["totalRecords"], 199)
            self.assertEqual(manifest["splitCounts"], {"train": 119, "dev": 40, "test": 40})
            self.assertEqual(manifest["changeCounts"], {"answerChanges": 5, "questionClarifications": 10, "questionRemovals": 1, "contractErrors": 0})
            expected_audit = manifest["expectedAudit"]
            self.assertEqual(
                expected_audit,
                {
                    "totalRecords": 199,
                    "splitCounts": {"train": 119, "dev": 40, "test": 40},
                    "statusCounts": {"VERIFIED": 199, "CORRECTED": 0, "UNRESOLVED": 0},
                    "evidenceComplete": 199,
                    "evidenceErrors": 0,
                },
            )
            self.assertFalse(manifest["candidateReady"])
            self.assertFalse(manifest["canonicalReady"])
            self.assertEqual(manifest["sourceSha256"], gen.sha256_file(_workbook_path))
            self.assertEqual(manifest["candidateSha256"], gen.sha256_file(output_dir / gen.CANDIDATE_WORKBOOK))
            self.assertEqual(manifest["changeLedgerSha256"], gen.sha256_file(output_dir / gen.CHANGE_LEDGER))

    def test_ledger_16_entries_and_hashes(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            _workbook_path, output_dir, _ids = self._run(base)
            ledger = json.loads((output_dir / gen.CHANGE_LEDGER).read_text(encoding="utf-8"))
            self.assertEqual(ledger["count"], 16)
            entries = ledger["entries"]
            self.assertEqual(len(entries), 16, "5 修正 + 10 澄清 + 1 删除 = 16 条账本")
            self.assertEqual(len({entry["id"] for entry in entries}), 16, "账本 ID 必须唯一")
            action_counts = Counter(entry["changeType"] for entry in entries)
            self.assertEqual(
                dict(action_counts),
                {"ANSWER_CORRECTION": 5, "QUESTION_CLARIFICATION": 10, "QUESTION_REMOVAL": 1},
            )
            removal_entries = [entry for entry in entries if entry["changeType"] == "QUESTION_REMOVAL"]
            self.assertEqual(len(removal_entries), 1)
            removal = removal_entries[0]
            self.assertEqual(removal["id"], REMOVAL_ID)
            self.assertEqual(removal["split"], "train")
            self.assertEqual(removal["oldTextSha256"], gen.sha256_text(REMOVAL_QUESTION))
            self.assertEqual(removal["removedAnswerSha256"], gen.sha256_text("无"))
            self.assertEqual(removal["newTextSha256"], None)
            for entry in entries:
                self.assertIn("oldTextSha256", entry)
                self.assertIn("newTextSha256", entry)
                self.assertIn("evidenceRef", entry)
            self.assertEqual(ledger["contractErrors"], [])
            for artifact in (gen.CANDIDATE_WORKBOOK, gen.CHANGE_LEDGER, gen.CANDIDATE_MANIFEST):
                digest = gen.sha256_file(output_dir / artifact)
                sha_file = output_dir / (artifact + ".sha256")
                self.assertTrue(sha_file.is_file())
                self.assertIn(digest, sha_file.read_text(encoding="utf-8"))


# --------------------------------------------------------------------------- 候选校验负例（意外删除/重排/重复/字段变更/非问题表变更）


def _rewrite_question_sheet(path: Path, rows: list[tuple[Any, ...]]) -> None:
    """用给定数据行重写问题答案清单（表头保持不变），用于构造篡改候选。"""
    workbook = load_workbook(path)
    sheet = workbook[rules.SHEET_QUESTION]
    sheet.delete_rows(2, sheet.max_row)
    for row in rows:
        sheet.append(list(row))
    workbook.save(path)


class VerifyCandidateNegativeTest(unittest.TestCase):
    """verify_candidate_workbook 必须捕获：意外删除、重排、重复 ID、
    未授权字段变更、非问题表单元格变更（即使行数不变）。
    篡改候选后必须用同一份契约校验（不得重新生成候选，否则会覆盖篡改）。"""

    def _contracts(self, base: Path) -> tuple[Path, Path, list, list, list]:
        workbook_path, review_path, _ids = build_removal_scenario(base)
        source_sha = hashlib.sha256(workbook_path.read_bytes()).hexdigest().upper()
        output_dir = base / "contract-fix"
        self.assertEqual(gen.run_generator(workbook_path, review_path, output_dir, expected_source_sha=source_sha, exclude_ids=[REMOVAL_ID]), 0)
        reviews = [json.loads(line) for line in review_path.read_text(encoding="utf-8").splitlines() if line.strip()]
        corrections = gen.select_answer_corrections(reviews)
        clarified, contract_errors = gen.generate_clarification_contracts(reviews, review_path)
        self.assertEqual(contract_errors, [])
        removals, removal_errors = gen.select_removals(reviews, [REMOVAL_ID], corrections, clarified)
        self.assertEqual(removal_errors, [])
        return workbook_path, output_dir / gen.CANDIDATE_WORKBOOK, corrections, clarified, removals

    def test_unexpected_row_removal_detected(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            source_path, candidate, corrections, clarified, removals = self._contracts(base)
            workbook = load_workbook(candidate)
            sheet = workbook[rules.SHEET_QUESTION]
            rows = [tuple(cell.value for cell in row) for row in sheet.iter_rows(min_row=2) if row[0].value is not None]
            rows = [row for row in rows if row[0] != "SYN-V000"]
            sheet.delete_rows(2, sheet.max_row)
            for row in rows:
                sheet.append(list(row))
            workbook.save(candidate)
            ok, errors = gen.verify_candidate_workbook(source_path, candidate, corrections, clarified, removals)
            self.assertFalse(ok, errors)
            self.assertTrue(errors)

    def test_reorder_detected(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            source_path, candidate, corrections, clarified, removals = self._contracts(base)
            workbook = load_workbook(candidate)
            sheet = workbook[rules.SHEET_QUESTION]
            rows = [tuple(cell.value for cell in row) for row in sheet.iter_rows(min_row=2) if row[0].value is not None]
            rows[0], rows[1] = rows[1], rows[0]
            sheet.delete_rows(2, sheet.max_row)
            for row in rows:
                sheet.append(list(row))
            workbook.save(candidate)
            ok, errors = gen.verify_candidate_workbook(source_path, candidate, corrections, clarified, removals)
            self.assertFalse(ok, errors)
            self.assertTrue(errors)

    def test_duplicate_detected(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            source_path, candidate, corrections, clarified, removals = self._contracts(base)
            workbook = load_workbook(candidate)
            sheet = workbook[rules.SHEET_QUESTION]
            rows = [tuple(cell.value for cell in row) for row in sheet.iter_rows(min_row=2) if row[0].value is not None]
            rows.append(rows[100])
            sheet.delete_rows(2, sheet.max_row)
            for row in rows:
                sheet.append(list(row))
            workbook.save(candidate)
            ok, errors = gen.verify_candidate_workbook(source_path, candidate, corrections, clarified, removals)
            self.assertFalse(ok, errors)
            self.assertTrue(errors)

    def test_field_change_detected(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            source_path, candidate, corrections, clarified, removals = self._contracts(base)
            workbook = load_workbook(candidate)
            sheet = workbook[rules.SHEET_QUESTION]
            for row in sheet.iter_rows(min_row=2):
                if row[0].value == "SYN-V000":
                    row[2].value = "困难"
            workbook.save(candidate)
            ok, errors = gen.verify_candidate_workbook(source_path, candidate, corrections, clarified, removals)
            self.assertFalse(ok, errors)
            self.assertTrue(errors)

    def test_non_question_mutation_detected(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            source_path, candidate, corrections, clarified, removals = self._contracts(base)
            workbook = load_workbook(candidate)
            workbook[rules.SHEET_FACT].cell(row=3, column=5, value=999.0)
            workbook.save(candidate)
            ok, errors = gen.verify_candidate_workbook(source_path, candidate, corrections, clarified, removals)
            self.assertFalse(ok, errors)
            self.assertTrue(any("非问题表" in error for error in errors), errors)


# --------------------------------------------------------------------------- manifest/账本/候选篡改拒绝（全部 fail closed）


class ManifestLedgerTamperTest(unittest.TestCase):
    """篡改 counts、expectedAudit、账本动作、重复条目、split/total 关系、
    候选哈希、账本哈希、候选工作簿 —— 候选审查必须全部拒绝。"""

    def _generated(self, base: Path) -> tuple[Path, Path, Path, str]:
        workbook_path, review_path, _ids = build_removal_scenario(base)
        source_sha = hashlib.sha256(workbook_path.read_bytes()).hexdigest().upper()
        output_dir = base / "contract-fix"
        self.assertEqual(gen.run_generator(workbook_path, review_path, output_dir, expected_source_sha=source_sha, exclude_ids=[REMOVAL_ID]), 0)
        return output_dir / gen.CANDIDATE_WORKBOOK, output_dir / gen.CANDIDATE_MANIFEST, output_dir / gen.CHANGE_LEDGER, source_sha

    def _audit(self, base: Path, candidate: Path, manifest_path: Path, source_sha: str) -> None:
        audit.run_audit(candidate, base / "out", True, manifest_path=manifest_path, expected_source_sha=source_sha)

    def _write_manifest(self, manifest_path: Path, manifest: dict) -> None:
        manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
        # 同步重算边车：本类测试针对内容级校验（counts/expectedAudit/哈希字段），
        # 边车字节完整性由 ManifestSidecarTest 单独覆盖
        sidecar_path = manifest_path.with_name(manifest_path.name + ".sha256")
        sidecar_path.write_text(f"{audit.sha256_file(manifest_path)}  {manifest_path.name}\n", encoding="utf-8")

    def test_tamper_change_counts_rejected(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            candidate, manifest_path, _ledger_path, source_sha = self._generated(base)
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            manifest["changeCounts"]["answerChanges"] = 4
            self._write_manifest(manifest_path, manifest)
            with self.assertRaises(ValueError):
                self._audit(base, candidate, manifest_path, source_sha)

    def test_tamper_expected_audit_blessing_unresolved_rejected(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            candidate, manifest_path, _ledger_path, source_sha = self._generated(base)
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            manifest["expectedAudit"]["statusCounts"] = {"VERIFIED": 198, "CORRECTED": 0, "UNRESOLVED": 1}
            self._write_manifest(manifest_path, manifest)
            with self.assertRaises(ValueError):
                self._audit(base, candidate, manifest_path, source_sha)

    def test_tamper_expected_audit_verified_count_rejected(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            candidate, manifest_path, _ledger_path, source_sha = self._generated(base)
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            manifest["expectedAudit"]["statusCounts"] = {"VERIFIED": 198, "CORRECTED": 1, "UNRESOLVED": 0}
            self._write_manifest(manifest_path, manifest)
            with self.assertRaises(ValueError):
                self._audit(base, candidate, manifest_path, source_sha)

    def test_tamper_split_total_relationship_rejected(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            candidate, manifest_path, _ledger_path, source_sha = self._generated(base)
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            manifest["splitCounts"] = {"train": 120, "dev": 40, "test": 40}
            manifest["expectedAudit"]["splitCounts"] = {"train": 120, "dev": 40, "test": 40}
            self._write_manifest(manifest_path, manifest)
            with self.assertRaises(ValueError):
                self._audit(base, candidate, manifest_path, source_sha)

    def test_tamper_ledger_action_rejected_even_with_recomputed_hash(self) -> None:
        # 篡改账本动作后同步重算 changeLedgerSha256：内容契约（动作计数）仍必须拒绝
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            candidate, manifest_path, ledger_path, source_sha = self._generated(base)
            ledger = json.loads(ledger_path.read_text(encoding="utf-8"))
            ledger["entries"][0]["changeType"] = "QUESTION_CLARIFICATION"
            ledger_path.write_text(json.dumps(ledger, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            manifest["changeLedgerSha256"] = gen.sha256_file(ledger_path)
            self._write_manifest(manifest_path, manifest)
            with self.assertRaises(ValueError):
                self._audit(base, candidate, manifest_path, source_sha)

    def test_tamper_ledger_duplicate_entry_rejected_even_with_recomputed_hash(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            candidate, manifest_path, ledger_path, source_sha = self._generated(base)
            ledger = json.loads(ledger_path.read_text(encoding="utf-8"))
            ledger["entries"].append(copy.deepcopy(ledger["entries"][0]))
            ledger_path.write_text(json.dumps(ledger, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            manifest["changeLedgerSha256"] = gen.sha256_file(ledger_path)
            self._write_manifest(manifest_path, manifest)
            with self.assertRaises(ValueError):
                self._audit(base, candidate, manifest_path, source_sha)

    def test_tamper_ledger_hash_rejected(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            candidate, manifest_path, _ledger_path, source_sha = self._generated(base)
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            manifest["changeLedgerSha256"] = "DEADBEEF" * 8
            self._write_manifest(manifest_path, manifest)
            with self.assertRaises(ValueError):
                self._audit(base, candidate, manifest_path, source_sha)

    def test_tamper_candidate_workbook_rejected(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            candidate, manifest_path, _ledger_path, source_sha = self._generated(base)
            workbook = load_workbook(candidate)
            workbook[rules.SHEET_QUESTION].cell(row=3, column=5, value="被篡改的答案")
            workbook.save(candidate)
            with self.assertRaises(ValueError):
                self._audit(base, candidate, manifest_path, source_sha)

    def test_tamper_candidate_hash_rejected(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            candidate, manifest_path, _ledger_path, source_sha = self._generated(base)
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            manifest["candidateSha256"] = "DEADBEEF" * 8
            self._write_manifest(manifest_path, manifest)
            with self.assertRaises(ValueError):
                self._audit(base, candidate, manifest_path, source_sha)


# --------------------------------------------------------------------------- 候选 manifest 哈希边车（fail closed + 写回同步）


class ManifestSidecarTest(unittest.TestCase):
    """候选 manifest 的 .sha256 边车：审查前必须存在、格式严格为
    '<UPPER_SHA256>  candidate-manifest.json' 单行且摘要匹配 manifest 原始
    字节；缺失/格式非法/文件名不符/摘要不匹配一律 fail closed。审查写回
    candidateReady/candidateAudit 时边车必须同步原子更新，第二次审查必须
    成功且字节保持确定性。"""

    def _generated(self, base: Path) -> tuple[Path, Path, Path, str]:
        workbook_path, review_path, _ids = build_removal_scenario(base)
        source_sha = hashlib.sha256(workbook_path.read_bytes()).hexdigest().upper()
        output_dir = base / "contract-fix"
        self.assertEqual(gen.run_generator(workbook_path, review_path, output_dir, expected_source_sha=source_sha, exclude_ids=[REMOVAL_ID]), 0)
        manifest_path = output_dir / gen.CANDIDATE_MANIFEST
        return output_dir / gen.CANDIDATE_WORKBOOK, manifest_path, manifest_path.with_name(manifest_path.name + ".sha256"), source_sha

    def _audit(self, base: Path, candidate: Path, manifest_path: Path, source_sha: str) -> dict:
        return audit.run_audit(candidate, base / "out", True, manifest_path=manifest_path, expected_source_sha=source_sha)

    def _expected_sidecar(self, manifest_path: Path) -> str:
        return f"{audit.sha256_file(manifest_path)}  {manifest_path.name}\n"

    def test_intact_generated_sidecar_accepted(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            candidate, manifest_path, sidecar_path, source_sha = self._generated(base)
            # 生成器产出的边车严格符合契约，且摘要匹配 manifest 原始字节
            self.assertEqual(sidecar_path.read_text(encoding="utf-8"), self._expected_sidecar(manifest_path))
            validation = audit.validate_manifest_sidecar(manifest_path)
            self.assertTrue(validation["valid"], validation["errors"])
            summary = self._audit(base, candidate, manifest_path, source_sha)
            self.assertTrue(summary["candidateReady"])
            self.assertTrue(summary["manifestSidecarValidation"]["valid"])

    def test_missing_sidecar_rejected(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            candidate, manifest_path, sidecar_path, source_sha = self._generated(base)
            sidecar_path.unlink()
            with self.assertRaises(ValueError) as ctx:
                self._audit(base, candidate, manifest_path, source_sha)
            self.assertIn("边车缺失", str(ctx.exception))

    def test_malformed_sidecar_rejected(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            candidate, manifest_path, sidecar_path, source_sha = self._generated(base)
            digest = audit.sha256_file(manifest_path)
            for label, malformed in (
                ("非哈希内容", "这不是哈希  candidate-manifest.json\n"),
                ("小写摘要", f"{digest.lower()}  {manifest_path.name}\n"),
                ("摘要长度不足", f"{digest[:60]}  {manifest_path.name}\n"),
                ("多余前缀", f"prefix {digest}  {manifest_path.name}\n"),
                ("多余行", f"{digest}  {manifest_path.name}\n附加内容\n"),
                ("缺少换行", f"{digest}  {manifest_path.name}"),
                ("多余空格分隔", f"{digest}   {manifest_path.name}\n"),
            ):
                sidecar_path.write_text(malformed, encoding="utf-8")
                with self.subTest(label=label):
                    with self.assertRaises(ValueError) as ctx:
                        self._audit(base, candidate, manifest_path, source_sha)
                    self.assertIn("哈希边车验证失败", str(ctx.exception))

    def test_wrong_filename_in_sidecar_rejected(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            candidate, manifest_path, sidecar_path, source_sha = self._generated(base)
            sidecar_path.write_text(f"{audit.sha256_file(manifest_path)}  other-file.json\n", encoding="utf-8")
            with self.assertRaises(ValueError) as ctx:
                self._audit(base, candidate, manifest_path, source_sha)
            self.assertIn("文件名不符", str(ctx.exception))

    def test_tampered_manifest_without_sidecar_update_rejected(self) -> None:
        # 篡改 manifest 但不更新边车：摘要不匹配，必须在解析/审查前 fail closed
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            candidate, manifest_path, _sidecar_path, source_sha = self._generated(base)
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            manifest["candidateSha256"] = "DEADBEEF" * 8
            manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
            with self.assertRaises(ValueError) as ctx:
                self._audit(base, candidate, manifest_path, source_sha)
            self.assertIn("摘要不匹配", str(ctx.exception))

    def test_sidecar_digest_mismatch_rejected(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            candidate, manifest_path, sidecar_path, source_sha = self._generated(base)
            sidecar_path.write_text(f"{'DEADBEEF' * 8}  {manifest_path.name}\n", encoding="utf-8")
            with self.assertRaises(ValueError) as ctx:
                self._audit(base, candidate, manifest_path, source_sha)
            self.assertIn("摘要不匹配", str(ctx.exception))

    def test_audit_syncs_manifest_and_sidecar(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            candidate, manifest_path, sidecar_path, source_sha = self._generated(base)
            summary = self._audit(base, candidate, manifest_path, source_sha)
            self.assertTrue(summary["candidateReady"])
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            self.assertTrue(manifest["candidateReady"])
            self.assertIn("candidateAudit", manifest)
            self.assertEqual(manifest["candidateAudit"]["statusCounts"], summary["statusCounts"])
            # 写回后边车必须与最终 manifest 字节一致（格式 + 摘要）
            self.assertEqual(sidecar_path.read_text(encoding="utf-8"), self._expected_sidecar(manifest_path))
            self.assertTrue(audit.validate_manifest_sidecar(manifest_path)["valid"])
            # 同目录临时文件必须清理干净
            self.assertEqual(list(manifest_path.parent.glob("*.tmp")), [])

    def test_rerun_audit_synchronized_deterministic(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            candidate, manifest_path, sidecar_path, source_sha = self._generated(base)
            first = self._audit(base, candidate, manifest_path, source_sha)
            manifest_after_first = manifest_path.read_bytes()
            sidecar_after_first = sidecar_path.read_text(encoding="utf-8")
            self.assertTrue(first["candidateReady"])
            # 同步后的 manifest 再次审查：成功、计数一致且字节保持不变（确定性）
            second = self._audit(base, candidate, manifest_path, source_sha)
            self.assertTrue(second["candidateReady"])
            self.assertEqual(second["statusCounts"], first["statusCounts"])
            self.assertEqual(second["splitCounts"], first["splitCounts"])
            self.assertEqual(manifest_path.read_bytes(), manifest_after_first)
            self.assertEqual(sidecar_path.read_text(encoding="utf-8"), sidecar_after_first)
            self.assertEqual(sidecar_path.read_text(encoding="utf-8"), self._expected_sidecar(manifest_path))


# --------------------------------------------------------------------------- 审查器对澄清题的行为


class AuditClarifiedBehaviorTest(unittest.TestCase):
    def _audit(self, questions: list[tuple[str, str, str, str]]) -> tuple[dict, list[dict]]:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            workbook_path = build_synthetic_workbook(base / "syn.xlsx", questions)
            output_dir = base / "out"
            digest = hashlib.sha256(workbook_path.read_bytes()).hexdigest().upper()
            summary = audit.run_audit(workbook_path, output_dir, True, expected_source_sha=digest)
            reviews = [json.loads(line) for line in (output_dir / "review.ndjson").read_text(encoding="utf-8").splitlines() if line.strip()]
            return summary, reviews

    def status_of(self, summary: dict, qid: str) -> str:
        if qid in summary["unresolvedIds"]:
            return "UNRESOLVED"
        if qid in summary["correctedIds"]:
            return "CORRECTED"
        return "VERIFIED"

    def test_performance_clarified_verified(self) -> None:
        question = (
            "江苏省A市农商行在2025-01-31的指标中哪些表现较好？哪些表现较差？"
            "待评价指标集合：各项存款余额、各项贷款余额、不良贷款率、净利润。"
            "判定规则：表现较好=全省排名前三，表现较差=全省排名后四；"
            "排名方向由指标定义决定（不良贷款率、逾期贷款率、成本收入比越低越好，其余越高越好）。"
        )
        answer = "各项存款余额(第1名)、各项贷款余额(第1名)、不良贷款率(第1名)、净利润(第1名)。表现较好指标：各项存款余额、各项贷款余额、不良贷款率、净利润。表现较差指标：无。"
        summary, reviews = self._audit([("SYN-K1", "训练集", "复杂", question, answer)])
        self.assertEqual(self.status_of(summary, "SYN-K1"), "VERIFIED", summary["statusCounts"])
        review = next(item for item in reviews if item["id"] == "SYN-K1")
        self.assertEqual(review["category"], "PERFORMANCE_EXPLICIT")
        self.assertTrue(review["fullEvidence"])

    def test_performance_clarified_wrong_rank_corrected(self) -> None:
        question = (
            "江苏省A市农商行在2025-01-31的指标中哪些表现较好？哪些表现较差？"
            "待评价指标集合：各项存款余额、不良贷款率。"
            "判定规则：表现较好=全省排名前三，表现较差=全省排名后四；"
            "排名方向由指标定义决定（不良贷款率、逾期贷款率、成本收入比越低越好，其余越高越好）。"
        )
        answer = "各项存款余额(第6名)。表现较好指标：无。表现较差指标：各项存款余额。"
        summary, _ = self._audit([("SYN-K2", "训练集", "复杂", question, answer)])
        self.assertEqual(self.status_of(summary, "SYN-K2"), "CORRECTED", "排名数值与事实不符不得 VERIFIED")

    def test_dimension3_clarified_verified(self) -> None:
        question = (
            "从规模、资产质量、盈利能力三个维度，分别列出江苏省A市农商行在2025-01-31的各项指标及排名。"
            "维度与指标映射：规模=各项存款余额、各项贷款余额；资产质量=不良贷款率；盈利能力=净利润。"
        )
        answer = "规模：各项存款余额100.5亿元（第1名），各项贷款余额90.3亿元。质量：不良贷款率0.51%（第1名）。效益：净利润510万元（第1名）。"
        summary, reviews = self._audit([("SYN-K3", "训练集", "复杂", question, answer)])
        if self.status_of(summary, "SYN-K3") != "VERIFIED":
            review = next(item for item in reviews if item["id"] == "SYN-K3")
            self.fail(
                f"status={review['status']} auditErrors={review['auditErrors']} "
                f"matchNotes={review['matchNotes']} "
                f"claims={[(c['key'], c['kind'], c['value'], c['matched'], c.get('matchNote'), c.get('mustAppear')) for c in review['claims']]}"
            )
        review = next(item for item in reviews if item["id"] == "SYN-K3")
        self.assertEqual(review["category"], "DIMENSION_3")
        self.assertTrue(review["fullEvidence"])

    def test_dimension3_mapping_ambiguity_stays_unresolved(self) -> None:
        question = "从规模、资产质量、盈利能力三个维度，分别列出江苏省A市农商行在2025-01-31的各项指标及排名。"
        summary, _ = self._audit([("SYN-K4", "训练集", "复杂", question, "规模：各项存款余额100亿元。")])
        self.assertEqual(self.status_of(summary, "SYN-K4"), "UNRESOLVED", "无映射标记的旧题必须保持 UNRESOLVED（不猜测）")

    def test_derived_rank_evidence_validates(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            workbook_path = build_synthetic_workbook(
                base / "syn.xlsx",
                [("SYN-K5", "训练集", "简单", "合成机构在2025-01-31的存贷比是多少？", "90.3/100.5")],
            )
            data = audit.WorkbookData(workbook_path)
            evidence = audit._evidence_factory(
                "RANK_POSITION", "rank(derived) 按排名方向，并列同名次", 2, data,
                org="ORG001", date="2025-01-31", scope={"orgs": sorted(data.organizations)},
                derived={"numerator": "ZB002", "denominator": "ZB001", "scale": 1.0, "percent": True},
            )
            claim = audit.make_claim(
                "rank_存贷比", "江苏省A市农商行的存贷比排名", kind="RANK",
                value=1.0, unit="名", metric=None, org="ORG001", date="2025-01-31",
                role="rank", must_appear=False, evidence=evidence,
            )
            ok, errors = audit.validate_claim_evidence(claim, data)
            self.assertTrue(ok, errors)
            self.assertEqual(evidence["sourceFactCount"], 10, "分子分母 × 5 家机构")

    def test_corrected_answer_with_full_date_verified(self) -> None:
        # 修正答案含完整日期（YYYY-MM-DD）：日期残片不得成为无法归因的 token
        question = "江苏省A市农商行在2025-01-31的各项存款余额是多少？"
        answer = "江苏省A市农商行2025-01-31的各项存款余额：100.5亿元"
        summary, _ = self._audit([("SYN-K6", "训练集", "简单", question, answer)])
        self.assertEqual(self.status_of(summary, "SYN-K6"), "VERIFIED", "含完整日期的修正答案必须可核对")


# --------------------------------------------------------------------------- 大表校验单遍流式守卫


class _VerificationAccessGuard:
    """验证路径访问守卫：禁止 Worksheet.cell 随机访问（首次调用即失败），
    并统计 iter_rows 流式访问的单元格总数供线性上界断言使用。

    只读工作表（ReadOnlyWorksheet）在类定义时绑定了 Worksheet 的旧函数对象，
    因此必须同时给两个类打补丁。守卫只包住 verify_candidate_workbook 调用，
    不包住候选生成（生成器的顺序写 cell 不属于校验路径）。
    """

    def __init__(self) -> None:
        self.cell_calls: list[tuple[int, int]] = []
        self.cells_iterated = 0
        self._patchers: list[Any] = []

    def __enter__(self) -> "_VerificationAccessGuard":
        original_cell = Worksheet.cell
        original_iter_rows = Worksheet.iter_rows

        def guarded_cell(sheet: Any, row: int, column: int, value: Any = None) -> Any:
            self.cell_calls.append((row, column))
            raise AssertionError(
                f"大表校验回退到 Worksheet.cell 随机访问：cell(row={row}, column={column})"
                "（校验必须使用 iter_rows 单遍流式）"
            )

        def guarded_iter_rows(
            sheet: Any,
            min_row: int | None = None,
            max_row: int | None = None,
            min_col: int | None = None,
            max_col: int | None = None,
            values_only: bool = False,
        ) -> Any:
            rows = original_iter_rows(sheet, min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=values_only)
            for row in rows:
                self.cells_iterated += len(row) if row is not None else 0
                yield row

        for target in (Worksheet, ReadOnlyWorksheet):
            self._patchers.append(mock.patch.object(target, "cell", guarded_cell))
            self._patchers.append(mock.patch.object(target, "iter_rows", guarded_iter_rows))
        for patcher in self._patchers:
            patcher.start()
        return self

    def __exit__(self, exc_type: Any, exc_value: Any, traceback: Any) -> None:
        for patcher in self._patchers:
            patcher.stop()


def _workbook_total_cells(path: Path) -> int:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return sum(sheet.max_row * sheet.max_column for sheet in workbook.worksheets)
    finally:
        workbook.close()


class StreamingVerificationTest(unittest.TestCase):
    """大表候选校验必须单遍流式 O(总单元格)：零 Worksheet.cell 随机访问、
    遍历量不超过线性上界。纯确定性断言，不用墙钟时间。"""

    def test_large_sheet_verification_is_single_pass_streaming(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            # 48,000 行事实表（1,200 日期 × 5 机构 × 8 指标）≈ 24 万单元格
            workbook_path, review_path, _ids = build_synthetic_scenario(base, days=LARGE_DAYS)
            source_sha = hashlib.sha256(workbook_path.read_bytes()).hexdigest().upper()
            output_dir = base / "contract-fix"
            self.assertEqual(gen.run_generator(workbook_path, review_path, output_dir, expected_source_sha=source_sha), 0)
            candidate_path = output_dir / gen.CANDIDATE_WORKBOOK
            reviews = [json.loads(line) for line in review_path.read_text(encoding="utf-8").splitlines() if line.strip()]
            corrections = gen.select_answer_corrections(reviews)
            # 校验路径需要完整澄清契约（含 clarifiedQuestion 等生成字段），
            # 原始 UNRESOLVED review 记录不包含这些字段 —— 必须走与 run_generator
            # 相同的契约生成函数，而不是把原始 review 记录直接传入。
            clarified, contract_errors = gen.generate_clarification_contracts(reviews, review_path)
            self.assertEqual(contract_errors, [], "合成澄清场景不得产生契约错误")
            total_cells = _workbook_total_cells(workbook_path)
            with _VerificationAccessGuard() as guard:
                ok, errors = gen.verify_candidate_workbook(workbook_path, candidate_path, corrections, clarified, [])
            self.assertTrue(ok, errors)
            self.assertEqual(guard.cell_calls, [], "校验路径不得调用 Worksheet.cell")
            # 流式实现全程访问 ≈ 4×total_cells（源+候选逐表 diff、问题表两遍、
            # 事实表哈希两遍）；10× 常数上界仍排除任何随行数增长的超线性遍历。
            bound = 10 * total_cells + 10000
            self.assertLessEqual(
                guard.cells_iterated, bound,
                f"大表校验遍历 {guard.cells_iterated} 个单元格，超过线性上界 {bound}（O(总单元格) 单遍流式）",
            )


class SyntheticPurityTest(unittest.TestCase):
    """静态扫描：新测试文件不包含真实 ID、完整真实问题或完整真实答案。"""

    def test_no_real_ids_questions_or_answers(self) -> None:
        test_source = Path(__file__).read_text(encoding="utf-8")
        real_ids: list[str] = []
        real_questions: list[str] = []
        real_answers: list[str] = []
        source_path = ROOT.parent.parent / ".local-dev" / "gt-audit" / "source.xlsx"
        if source_path.is_file():
            workbook = load_workbook(source_path, read_only=True, data_only=True)
            sheet = workbook["问题答案清单"]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue
                real_ids.append(str(row[0]).strip())
                real_questions.append(str(row[3]).strip())
                real_answers.append(str(row[4]).strip())
            workbook.close()
        found_ids = [qid for qid in real_ids if qid in test_source]
        found_questions = [question for question in real_questions if question and question in test_source]
        found_answers = [answer for answer in real_answers if answer and answer in test_source]
        self.assertEqual(found_ids, [], f"测试文件不得包含真实题目编号: {found_ids[:5]}")
        self.assertEqual(found_questions, [], f"测试文件不得包含完整真实问题: {found_questions[:3]}")
        self.assertEqual(found_answers, [], f"测试文件不得包含完整真实答案: {found_answers[:3]}")


# ---------------------------------------------------------------------------
# 候选晋升（promote_ground_truth）合成契约测试
#
# 全部为合成题目/答案/ID（SYN-*），复用本文件的合成工作簿构造器。覆盖：
# 就绪候选晋升与确定性；candidateReady=false / 状态分布不是全 VERIFIED /
# 边车缺失或过期 / 审查计数不符 / 账本哈希不匹配 / 重复或额外删除 /
# 版本不符 / 候选与审查输出被篡改 / verifiedIds 或 review.ndjson ID 与候选
# 不一致 / review 证据非 VERIFIED / 工作簿文本与账本哈希不一致 /
# 源哈希链断裂 —— 全部 fail closed。
# candidate-reviewed.xlsx 按真实审计输出构造：只含 review 页的审查报告，
# 发布器只验证其存在与 outputSha256，并以候选工作簿 ID 集合对
# audit-summary.verifiedIds 与 review.ndjson ID 做三方互证。

PROMO_REMOVED_ID = "SYN-T05"
PROMO_CORRECTED_ID = "SYN-T03"
PROMO_CLARIFIED_ID = "SYN-T04"
PROMO_TOTAL = 6
PROMO_SPLITS = {"train": 4, "dev": 1, "test": 1}
PROMO_SOURCE_QUESTIONS: dict[str, tuple[str, str, str, str]] = {
    "SYN-T01": ("训练集", "简单", "合成问题一", "旧答案一"),
    "SYN-T02": ("训练集", "普通", "合成问题二", "旧答案二"),
    PROMO_CORRECTED_ID: ("训练集", "普通", "合成问题三", "旧答案三"),
    PROMO_CLARIFIED_ID: ("训练集", "复杂", "合成问题四", "旧答案四"),
    PROMO_REMOVED_ID: ("训练集", "普通", "合成问题五", "旧答案五"),
    "SYN-V01": ("验证集", "简单", "合成问题六", "旧答案六"),
    "SYN-S01": ("测试集", "复杂", "合成问题七", "旧答案七"),
}
PROMO_CORRECTED_ANSWER = "100.5亿元"
PROMO_CLARIFIED_QUESTION = "江苏省A市农商行在2025-01-31的各项存款余额、各项贷款余额、净利润分别是多少？"


def _promo_sha256(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest().upper()


def _promo_write_sidecar(artifact_path: Path) -> None:
    digest = _promo_sha256(artifact_path.read_bytes())
    artifact_path.with_name(artifact_path.name + ".sha256").write_text(
        f"{digest}  {artifact_path.name}\n", encoding="utf-8"
    )


def _promo_ledger_entries() -> list[dict]:
    return [
        {
            "id": PROMO_CORRECTED_ID,
            "changeType": "ANSWER_CORRECTION",
            "category": "POINT",
            "split": "train",
            "difficulty": "普通",
            "oldTextSha256": gen.sha256_text(PROMO_SOURCE_QUESTIONS[PROMO_CORRECTED_ID][3]),
            "newTextSha256": gen.sha256_text(PROMO_CORRECTED_ANSWER),
            "metricCodes": None,
            "dimensionMapping": None,
            "basis": "合成 basis",
            "evidenceRef": {"reviewFile": "review.ndjson", "status": "CORRECTED", "fullEvidence": True, "evidenceValidationValid": True},
        },
        {
            "id": PROMO_CLARIFIED_ID,
            "changeType": "QUESTION_CLARIFICATION",
            "category": "PERFORMANCE",
            "split": "train",
            "difficulty": "复杂",
            "oldTextSha256": gen.sha256_text(PROMO_SOURCE_QUESTIONS[PROMO_CLARIFIED_ID][2]),
            "newTextSha256": gen.sha256_text(PROMO_CLARIFIED_QUESTION),
            "metricCodes": ["ZB001", "ZB002", "ZB011"],
            "dimensionMapping": None,
            "basis": "合成 basis",
            "evidenceRef": {"reviewFile": "review.ndjson", "status": "UNRESOLVED", "fullEvidence": False},
        },
        {
            "id": PROMO_REMOVED_ID,
            "changeType": "QUESTION_REMOVAL",
            "category": "CHANGE_DELTA",
            "split": "train",
            "difficulty": "普通",
            "oldTextSha256": gen.sha256_text(PROMO_SOURCE_QUESTIONS[PROMO_REMOVED_ID][2]),
            "newTextSha256": None,
            "removedAnswerSha256": gen.sha256_text(PROMO_SOURCE_QUESTIONS[PROMO_REMOVED_ID][3]),
            "metricCodes": None,
            "dimensionMapping": None,
            "basis": "合成删除",
            "evidenceRef": {"reviewFile": "review.ndjson", "status": "UNRESOLVED", "fullEvidence": False},
        },
    ]


class PromotionScenario:
    """完整合成候选目录 + 最终审查目录（全部数据驱动，无真实内容）。"""

    def __init__(self, base: Path, version: str = "2.0.0") -> None:
        self.base = base
        self.version = version
        self.candidate_dir = base / "candidate"
        self.audit_dir = base / "audit"
        self.candidate_dir.mkdir()
        self.audit_dir.mkdir()

        self.source_path = build_synthetic_workbook(
            self.candidate_dir / "source.xlsx",
            [(qid, *fields) for qid, fields in PROMO_SOURCE_QUESTIONS.items()],
        )
        candidate_questions = [
            (qid, *fields) for qid, fields in PROMO_SOURCE_QUESTIONS.items() if qid != PROMO_REMOVED_ID
        ]
        candidate_questions = [
            (
                qid,
                split,
                difficulty,
                PROMO_CLARIFIED_QUESTION if qid == PROMO_CLARIFIED_ID else question,
                PROMO_CORRECTED_ANSWER if qid == PROMO_CORRECTED_ID else answer,
            )
            for qid, split, difficulty, question, answer in candidate_questions
        ]
        self.candidate_path = build_synthetic_workbook(self.candidate_dir / gen.CANDIDATE_WORKBOOK, candidate_questions)

        ledger = {
            "generatorName": gen.GENERATOR_NAME,
            "generatorVersion": version,
            "count": 3,
            "entries": _promo_ledger_entries(),
            "contractErrors": [],
        }
        self.ledger_path = self.candidate_dir / gen.CHANGE_LEDGER
        # 故意以非规范格式（indent=4 且不排序键）写入：manifest.changeLedgerSha256
        # 锚定的是实际字节，发布器必须原样复制，任何 JSON 重序列化都会破坏锚定。
        self.ledger_path.write_text(json.dumps(ledger, ensure_ascii=False, indent=4) + "\n", encoding="utf-8")

        source_sha = _promo_sha256(self.source_path.read_bytes())
        candidate_sha = _promo_sha256(self.candidate_path.read_bytes())
        manifest = {
            "generatorName": gen.GENERATOR_NAME,
            "generatorVersion": version,
            "sourceSha256": source_sha,
            "candidateSha256": candidate_sha,
            "candidateWorkbook": gen.CANDIDATE_WORKBOOK,
            "totalRecords": PROMO_TOTAL,
            "splitCounts": PROMO_SPLITS,
            "changeCounts": {"answerChanges": 1, "questionClarifications": 1, "questionRemovals": 1, "contractErrors": 0},
            "expectedAudit": {
                "totalRecords": PROMO_TOTAL,
                "splitCounts": PROMO_SPLITS,
                "statusCounts": {"VERIFIED": PROMO_TOTAL, "CORRECTED": 0, "UNRESOLVED": 0},
                "evidenceComplete": PROMO_TOTAL,
                "evidenceErrors": 0,
            },
            "contractErrors": [],
            "factRegionSha256": gen.fact_region_digest(self.source_path),
            "changeLedgerSha256": _promo_sha256(self.ledger_path.read_bytes()),
            "candidateReady": True,
            "canonicalReady": False,
            "candidateAudit": {
                "candidateReady": True,
                "reasons": [],
                "splitCounts": PROMO_SPLITS,
                "statusCounts": {"VERIFIED": PROMO_TOTAL, "CORRECTED": 0, "UNRESOLVED": 0},
            },
        }
        self.manifest_path = self.candidate_dir / gen.CANDIDATE_MANIFEST
        self.manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
        for artifact in (self.candidate_path, self.ledger_path, self.manifest_path):
            _promo_write_sidecar(artifact)
        self.source_sha = source_sha
        self.candidate_sha = candidate_sha

        reviews = [
            {
                "id": qid,
                "split": {"训练集": "train", "验证集": "dev", "测试集": "test"}[split],
                "difficulty": difficulty,
                "status": "VERIFIED",
                "category": "POINT",
                "fullEvidence": True,
                "question": "合成问题",
                "answerText": "旧答案",
                "correctedAnswerText": None,
                "claims": [],
                "auditErrors": [],
                "unresolvedReason": None,
                "matchNotes": [],
                "evidenceValidation": {"valid": True, "ruleVersion": "1.0.0", "claimErrors": {}, "errors": []},
            }
            for qid, (split, difficulty, _question, _answer) in PROMO_SOURCE_QUESTIONS.items()
            if qid != PROMO_REMOVED_ID
        ]
        self.review_path = self.audit_dir / "review.ndjson"
        self.review_path.write_text(
            "".join(json.dumps(review, ensure_ascii=False, sort_keys=True) + "\n" for review in reviews), encoding="utf-8"
        )
        # candidate-reviewed.xlsx 与真实审计输出一致：只含 review 页的审查
        # 报告（不含问题答案清单页）。发布器只验证它的存在与 outputSha256，
        # 绝不从其中加载题目答案清单。
        self.reviewed_path = self.audit_dir / "candidate-reviewed.xlsx"
        audit.write_review_xlsx(self.reviewed_path, reviews)
        (self.audit_dir / "correction-ledger.json").write_text(
            json.dumps({"ruleVersion": "2.0.0", "count": 0, "corrections": []}, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
            encoding="utf-8",
        )
        summary = {
            "ruleVersion": "2.0.0",
            "evidenceRuleVersion": "1.0.0",
            "sourceSha256": candidate_sha,
            "sourceSha256Match": False,
            "manifestValidation": {"valid": True, "errors": []},
            "manifestSidecarValidation": {"valid": True, "errors": []},
            "candidateReady": True,
            "candidateReadyReasons": [],
            "totalRecords": PROMO_TOTAL,
            "uniqueIds": PROMO_TOTAL,
            "splitCounts": PROMO_SPLITS,
            "statusCounts": {"VERIFIED": PROMO_TOTAL, "CORRECTED": 0, "UNRESOLVED": 0},
            "fullEvidence": PROMO_TOTAL,
            "evidenceCompleteCount": PROMO_TOTAL,
            "evidenceErrorCount": 0,
            "auditErrors": 0,
            "unresolvedIds": [],
            "correctedIds": [],
            "verifiedIds": [qid for qid in PROMO_SOURCE_QUESTIONS if qid != PROMO_REMOVED_ID],
            "canonicalReady": False,
            "outputFile": "candidate-reviewed.xlsx",
            "outputSha256": _promo_sha256(self.reviewed_path.read_bytes()),
        }
        self.summary_path = self.audit_dir / "audit-summary.json"
        self.summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
        self.summary = summary
        self.output_dir = base / "official" / version

    def rewrite_manifest(self, manifest: dict) -> None:
        self.manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
        _promo_write_sidecar(self.manifest_path)

    def rewrite_ledger(self, ledger: dict) -> None:
        self.ledger_path.write_text(json.dumps(ledger, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
        _promo_write_sidecar(self.ledger_path)

    def promote(self, **kwargs) -> dict:
        return promote_ground_truth(
            self.candidate_dir,
            self.audit_dir,
            kwargs.pop("version", self.version),
            kwargs.pop("output_dir", self.output_dir),
            expected_source_sha=kwargs.pop("expected_source_sha", self.source_sha),
            **kwargs,
        )


class PromoteGroundTruthTest(unittest.TestCase):
    def test_promotes_ready_candidate_and_is_deterministic(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            scenario = PromotionScenario(Path(tmp))
            manifest = scenario.promote()
            self.assertEqual(manifest["datasetVersion"], "2.0.0")
            self.assertTrue(manifest["canonicalReady"])
            self.assertEqual(manifest["officialCount"], PROMO_TOTAL)
            self.assertEqual(manifest["sourceSplitCounts"], PROMO_SPLITS)
            self.assertEqual(manifest["removedIds"], [PROMO_REMOVED_ID])
            self.assertEqual(
                manifest["changeCounts"],
                {"answerChanges": 1, "questionClarifications": 1, "questionRemovals": 1, "contractErrors": 0},
            )
            self.assertEqual(manifest["auditStatus"]["evidenceComplete"], PROMO_TOTAL)
            workbook_path = scenario.output_dir / f"bank-nl2sql-ground-truth-v{scenario.version}.xlsx"
            self.assertTrue(workbook_path.is_file())
            self.assertTrue((scenario.output_dir / gen.CHANGE_LEDGER).is_file())
            self.assertTrue((scenario.output_dir / FINAL_AUDIT_SUMMARY).is_file())
            self.assertTrue((scenario.output_dir / OFFICIAL_MANIFEST).is_file())
            # 输出 ledger 必须原样复制已验证输入字节（不得 JSON 重序列化），
            # 其 SHA-256 必须与 manifest 锚定的 changeLedgerSha256 完全一致。
            ledger_output = scenario.output_dir / gen.CHANGE_LEDGER
            self.assertEqual(ledger_output.read_bytes(), scenario.ledger_path.read_bytes())
            self.assertEqual(
                _promo_sha256(ledger_output.read_bytes()),
                json.loads(scenario.manifest_path.read_text(encoding="utf-8"))["changeLedgerSha256"],
            )
            for artifact in (
                workbook_path,
                scenario.output_dir / gen.CHANGE_LEDGER,
                scenario.output_dir / FINAL_AUDIT_SUMMARY,
                scenario.output_dir / OFFICIAL_MANIFEST,
            ):
                sidecar = artifact.with_name(artifact.name + ".sha256")
                self.assertTrue(sidecar.is_file())
                digest, filename = sidecar.read_text(encoding="utf-8").strip().split("  ")
                self.assertEqual(filename, artifact.name)
                self.assertEqual(digest, _promo_sha256(artifact.read_bytes()))
            self.assertEqual(manifest["artifactSha256"]["groundTruthWorkbook"], _promo_sha256(workbook_path.read_bytes()))
            self.assertEqual(manifest["sourceSha256"], scenario.source_sha)
            self.assertEqual(manifest["candidateSha256"], scenario.candidate_sha)
            pointer = json.loads((scenario.output_dir.parent / CURRENT_POINTER).read_text(encoding="utf-8"))
            self.assertEqual(pointer["currentVersion"], "2.0.0")
            # candidate-reviewed.xlsx 必须是只含 review 页的审查报告（不得含
            # 问题答案清单页），promotion 仍须成功：发布器不得从审查报告加载
            # 题目答案清单。
            reviewed_wb = load_workbook(scenario.reviewed_path, read_only=True)
            self.assertEqual(reviewed_wb.sheetnames, ["review"])
            reviewed_wb.close()

            first = {
                name: (scenario.output_dir / name).read_bytes()
                for name in ("bank-nl2sql-ground-truth-v2.0.0.xlsx", OFFICIAL_MANIFEST, gen.CHANGE_LEDGER, FINAL_AUDIT_SUMMARY)
            }
            scenario.promote()
            second = {
                name: (scenario.output_dir / name).read_bytes()
                for name in ("bank-nl2sql-ground-truth-v2.0.0.xlsx", OFFICIAL_MANIFEST, gen.CHANGE_LEDGER, FINAL_AUDIT_SUMMARY)
            }
            self.assertEqual(first, second)

    def test_fails_closed_when_candidate_not_ready(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            scenario = PromotionScenario(Path(tmp))
            manifest = json.loads(scenario.manifest_path.read_text(encoding="utf-8"))
            manifest["candidateReady"] = False
            scenario.rewrite_manifest(manifest)
            with self.assertRaisesRegex(PromotionError, "candidateReady"):
                scenario.promote()

    def test_fails_closed_when_status_counts_not_all_verified(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            scenario = PromotionScenario(Path(tmp))
            manifest = json.loads(scenario.manifest_path.read_text(encoding="utf-8"))
            manifest["expectedAudit"]["statusCounts"] = {"VERIFIED": PROMO_TOTAL - 1, "CORRECTED": 1, "UNRESOLVED": 0}
            scenario.rewrite_manifest(manifest)
            with self.assertRaisesRegex(PromotionError, "statusCounts"):
                scenario.promote()

    def test_fails_closed_on_missing_or_stale_sidecar(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            scenario = PromotionScenario(Path(tmp))
            (scenario.manifest_path.with_name(scenario.manifest_path.name + ".sha256")).unlink()
            with self.assertRaisesRegex(PromotionError, "边车缺失"):
                scenario.promote()
        with tempfile.TemporaryDirectory() as tmp:
            scenario = PromotionScenario(Path(tmp))
            sidecar = scenario.manifest_path.with_name(scenario.manifest_path.name + ".sha256")
            sidecar.write_text("A" * 64 + f"  {scenario.manifest_path.name}\n", encoding="utf-8")
            with self.assertRaisesRegex(PromotionError, "摘要不匹配"):
                scenario.promote()

    def test_fails_closed_on_wrong_audit_counts(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            scenario = PromotionScenario(Path(tmp))
            scenario.summary["statusCounts"] = {"VERIFIED": PROMO_TOTAL - 1, "CORRECTED": 0, "UNRESOLVED": 1}
            scenario.summary_path.write_text(json.dumps(scenario.summary, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
            with self.assertRaisesRegex(PromotionError, "statusCounts"):
                scenario.promote()
        with tempfile.TemporaryDirectory() as tmp:
            scenario = PromotionScenario(Path(tmp))
            scenario.summary["evidenceErrorCount"] = 1
            scenario.summary_path.write_text(json.dumps(scenario.summary, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
            with self.assertRaisesRegex(PromotionError, "证据计数"):
                scenario.promote()

    def test_fails_closed_on_ledger_hash_mismatch(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            scenario = PromotionScenario(Path(tmp))
            ledger = json.loads(scenario.ledger_path.read_text(encoding="utf-8"))
            ledger["entries"][0]["newTextSha256"] = "B" * 64
            scenario.rewrite_ledger(ledger)  # 边车同步，manifest 哈希未同步
            with self.assertRaisesRegex(PromotionError, "changeLedgerSha256"):
                scenario.promote()

    def test_fails_closed_on_duplicate_removal(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            scenario = PromotionScenario(Path(tmp))
            ledger = json.loads(scenario.ledger_path.read_text(encoding="utf-8"))
            ledger["entries"].append(copy.deepcopy(ledger["entries"][2]))
            ledger["count"] = 4
            scenario.rewrite_ledger(ledger)
            manifest = json.loads(scenario.manifest_path.read_text(encoding="utf-8"))
            manifest["changeLedgerSha256"] = _promo_sha256(scenario.ledger_path.read_bytes())
            manifest["changeCounts"]["questionRemovals"] = 2
            scenario.rewrite_manifest(manifest)
            with self.assertRaisesRegex(PromotionError, "重复 ID"):
                scenario.promote()

    def test_fails_closed_on_extra_removal(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            scenario = PromotionScenario(Path(tmp))
            ledger = json.loads(scenario.ledger_path.read_text(encoding="utf-8"))
            # 把澄清条目改成第二笔删除：目标 ID 仍在候选工作簿中
            ledger["entries"][1]["changeType"] = "QUESTION_REMOVAL"
            ledger["entries"][1]["removedAnswerSha256"] = gen.sha256_text("旧答案四")
            ledger["entries"][1]["newTextSha256"] = None
            scenario.rewrite_ledger(ledger)
            manifest = json.loads(scenario.manifest_path.read_text(encoding="utf-8"))
            manifest["changeLedgerSha256"] = _promo_sha256(scenario.ledger_path.read_bytes())
            manifest["changeCounts"]["questionRemovals"] = 2
            manifest["changeCounts"]["questionClarifications"] = 0
            scenario.rewrite_manifest(manifest)
            with self.assertRaisesRegex(PromotionError, "仍存在于候选"):
                scenario.promote()

    def test_fails_closed_on_incorrect_version(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            scenario = PromotionScenario(Path(tmp))
            with self.assertRaisesRegex(PromotionError, "generatorVersion"):
                scenario.promote(version="1.0.0")

    def test_fails_closed_on_candidate_artifact_tampering(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            scenario = PromotionScenario(Path(tmp))
            workbook = load_workbook(scenario.candidate_path)
            workbook["问题答案清单"].cell(row=2, column=5, value="被篡改的答案")
            workbook.save(scenario.candidate_path)
            _promo_write_sidecar(scenario.candidate_path)
            with self.assertRaisesRegex(PromotionError, "候选工作簿哈希不一致"):
                scenario.promote()

    def test_fails_closed_on_reviewed_output_tampering(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            scenario = PromotionScenario(Path(tmp))
            scenario.reviewed_path.write_bytes(scenario.reviewed_path.read_bytes() + b"tamper")
            with self.assertRaisesRegex(PromotionError, "outputSha256"):
                scenario.promote()

    def test_fails_closed_when_verified_ids_mismatch_candidate(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            scenario = PromotionScenario(Path(tmp))
            # verifiedIds 掺入一个不在候选工作簿中的 ID（已删除题），
            # 集合即与候选工作簿 ID 集合不一致。
            scenario.summary["verifiedIds"] = sorted(set(scenario.summary["verifiedIds"]) | {PROMO_REMOVED_ID})
            scenario.summary_path.write_text(
                json.dumps(scenario.summary, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8"
            )
            with self.assertRaisesRegex(PromotionError, "verifiedIds"):
                scenario.promote()

    def test_fails_closed_when_review_ids_mismatch_candidate(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            scenario = PromotionScenario(Path(tmp))
            lines = [line for line in scenario.review_path.read_text(encoding="utf-8").splitlines() if line.strip()]
            review = json.loads(lines[0])
            review["id"] = "SYN-UNKNOWN-001"
            scenario.review_path.write_text(
                json.dumps(review, ensure_ascii=False, sort_keys=True) + "\n" + "\n".join(lines[1:]) + "\n", encoding="utf-8"
            )
            with self.assertRaisesRegex(PromotionError, "review.ndjson ID 集合"):
                scenario.promote()

    def test_fails_closed_on_review_evidence_not_verified(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            scenario = PromotionScenario(Path(tmp))
            lines = [line for line in scenario.review_path.read_text(encoding="utf-8").splitlines() if line.strip()]
            review = json.loads(lines[0])
            review["status"] = "UNRESOLVED"
            scenario.review_path.write_text(
                json.dumps(review, ensure_ascii=False, sort_keys=True) + "\n" + "\n".join(lines[1:]) + "\n", encoding="utf-8"
            )
            with self.assertRaisesRegex(PromotionError, "VERIFIED"):
                scenario.promote()

    def test_fails_closed_on_workbook_ledger_text_mismatch(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            scenario = PromotionScenario(Path(tmp))
            # 全链同步篡改（工作簿答案、边车、manifest 哈希、审查摘要），唯独账本不动
            workbook = load_workbook(scenario.candidate_path)
            workbook["问题答案清单"].cell(row=4, column=5, value="未被账本声明的新答案")
            workbook.save(scenario.candidate_path)
            _promo_write_sidecar(scenario.candidate_path)
            manifest = json.loads(scenario.manifest_path.read_text(encoding="utf-8"))
            manifest["candidateSha256"] = _promo_sha256(scenario.candidate_path.read_bytes())
            scenario.rewrite_manifest(manifest)
            scenario.summary["sourceSha256"] = manifest["candidateSha256"]
            scenario.reviewed_path.write_bytes(scenario.candidate_path.read_bytes())
            scenario.summary["outputSha256"] = _promo_sha256(scenario.reviewed_path.read_bytes())
            scenario.summary_path.write_text(json.dumps(scenario.summary, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
            with self.assertRaisesRegex(PromotionError, "newTextSha256"):
                scenario.promote()

    def test_fails_closed_when_source_hash_chain_breaks(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            scenario = PromotionScenario(Path(tmp))
            scenario.source_path.write_bytes(scenario.source_path.read_bytes() + b"tamper")
            with self.assertRaisesRegex(PromotionError, "源工作簿哈希不匹配"):
                scenario.promote()

    def assert_promotion_rejects_clarification_metric_codes(self, metric_codes: Any, expected_error: str) -> None:
        """把 QUESTION_CLARIFICATION 条目 metricCodes 改为非法值并同步哈希后 promote。"""
        with tempfile.TemporaryDirectory() as tmp:
            scenario = PromotionScenario(Path(tmp))
            ledger = json.loads(scenario.ledger_path.read_text(encoding="utf-8"))
            ledger["entries"][1]["metricCodes"] = metric_codes
            scenario.rewrite_ledger(ledger)
            manifest = json.loads(scenario.manifest_path.read_text(encoding="utf-8"))
            manifest["changeLedgerSha256"] = _promo_sha256(scenario.ledger_path.read_bytes())
            scenario.rewrite_manifest(manifest)
            with self.assertRaisesRegex(PromotionError, expected_error):
                scenario.promote()

    def test_fails_closed_on_clarification_metric_codes_not_a_list_or_empty(self) -> None:
        for metric_codes in ("ZB001", None, {}, []):
            with self.subTest(metric_codes=metric_codes):
                self.assert_promotion_rejects_clarification_metric_codes(metric_codes, "metricCodes 非法")

    def test_fails_closed_on_clarification_invalid_base_code(self) -> None:
        for metric_codes in (["ZB01"], ["zb001"], ["ZB0011"], ["ZB001 "], [""], ["存款余额"]):
            with self.subTest(metric_codes=metric_codes):
                self.assert_promotion_rejects_clarification_metric_codes(metric_codes, "不是合法指标代码")

    def test_fails_closed_on_clarification_duplicate_base_codes(self) -> None:
        self.assert_promotion_rejects_clarification_metric_codes(["ZB001", "ZB001"], "基础指标重复")

    def test_fails_closed_on_clarification_derived_only(self) -> None:
        """仅 derived 无 string base 的契约必须 fail closed（至少一个 base code）。"""
        self.assert_promotion_rejects_clarification_metric_codes(
            [{"derived": {"numerator": "ZB002", "denominator": "ZB001"}, "name": "存贷比"}],
            "未声明任何基础指标",
        )

    def test_fails_closed_on_clarification_invalid_derived_operands(self) -> None:
        cases = [
            ([{"derived": {"numerator": "NOT_A_CODE", "denominator": "ZB001"}}], "derived numerator 'NOT_A_CODE' 不是合法指标代码"),
            ([{"derived": {"numerator": "ZB002", "denominator": "NOT_A_CODE"}}], "derived denominator 'NOT_A_CODE' 不是合法指标代码"),
            ([{"derived": {"numerator": "ZB002", "denominator": "ZB002"}}], "derived numerator 与 denominator 必须不同"),
        ]
        for metric_codes, expected_error in cases:
            with self.subTest(metric_codes=metric_codes):
                self.assert_promotion_rejects_clarification_metric_codes(metric_codes, expected_error)


if __name__ == "__main__":
    unittest.main()
